import asyncio
import inspect
import json
import os
import sys
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from fastapi import HTTPException
import pytest

import routes.internal_users as internal_users
import routes.internal_users_mod.catalogos_clientes as cp_catalogos
import routes.internal_users_mod.users_auth as users_auth
import routes.facturas as facturas_routes
from services.sw_sapien import build_carta_porte_xml


ROOT = Path(__file__).resolve().parents[1]


def _expand_frontend_includes(source):
    import re

    def repl(match):
        rel_path = match.group(1)
        return _expand_frontend_includes((ROOT / "templates" / rel_path).read_text(encoding="utf-8"))

    return re.sub(r"<!--\s*ge-include:\s*([A-Za-z0-9_./-]+\.html)\s*-->", repl, source)


def _assistant_frontend_source():
    template = _expand_frontend_includes((ROOT / "templates" / "asistente_gas_lp.html").read_text(encoding="utf-8"))
    assets = [
        ROOT / "static/css/gas_lp/asistente.css",
        *sorted((ROOT / "static/js/gas_lp/asistente").glob("*.js")),
    ]
    return template + "\n" + "\n".join(path.read_text(encoding="utf-8") for path in assets)


def _conciliacion_frontend_source():
    template = _expand_frontend_includes((ROOT / "templates" / "conciliacion_gas_lp.html").read_text(encoding="utf-8"))
    assets = [
        ROOT / "static/css/gas_lp/conciliacion.css",
        *sorted((ROOT / "static/js/gas_lp/conciliacion").glob("*.js")),
    ]
    return template + "\n" + "\n".join(path.read_text(encoding="utf-8") for path in assets)


def _dump_model(model):
    if hasattr(model, "model_dump"):
        return model.model_dump()
    return model.dict()


def test_conciliacion_template_exposes_erp_tabs_and_own_endpoints():
    html = _conciliacion_frontend_source()

    for label in (
        "Facturas",
        "Facturar Público General",
        "Complementos de pago",
        "Cancelación / Consulta SAT",
    ):
        assert label in html
    assert 'data-tab="exportaciones"' not in html
    assert 'data-section="exportaciones"' not in html
    assert "facturas-export-actions" in html
    assert "Descargar mes" in html
    assert "Descargar día" in html

    assert "/api/internal-auth/gas-lp/conciliacion/summary" in html
    assert "/api/internal-auth/gas-lp/conciliacion/facturar-publico-general" in html
    assert "/api/internal-auth/gas-lp/conciliacion/export-excel" in html
    assert "/api/internal-auth/gas-lp/conciliacion/facilities" in html
    for token in (
        "pubDescuento",
        "pubDescuentoTipo",
        "pubIvaRate",
        "pubSumSubtotal",
        "pubSumDescuento",
        "pubSumIva",
        "pubSumTotal",
        "updatePublicoSummary()",
        "total_pesos",
        "publicoDiscountPerLiter",
        "Realizado por",
    ):
        assert token in html


def test_conciliacion_credito_ppd_exposes_due_tracking():
    html = _conciliacion_frontend_source()

    for token in (
        "Política",
        "Vencidas",
        "Saldo vencido",
        "Peor atraso",
        "function creditStatusForFactura",
        "CLIENTES=d.clientes||[]",
        "creditBadgeHtml(info)",
        "dias_vencidos",
    ):
        assert token in html


def test_conciliacion_exposes_manual_bank_reconciliation_layer():
    html = _conciliacion_frontend_source()
    source = inspect.getsource(internal_users.gas_lp_conciliacion_bank_reconciliation_save)
    normalize_source = inspect.getsource(internal_users._normalize_bank_reconciliation_status)
    summary_source = inspect.getsource(internal_users.gas_lp_conciliacion_summary)

    for token in (
        "Estado banco",
        "Conciliación bancaria manual",
        "bank_reconciliation",
        "bankStatusHtml(f)",
        "openBankReconciliation",
        "/bank-reconciliation",
    ):
        assert token in html

    assert "payment_status" not in source
    assert "saldo_insoluto" not in source
    assert "gas_lp_invoice_bank_reconciliations" in source
    assert "gas_lp_bank_reconciliation_audit_logs" in source
    assert "BANK_RECONCILIATION_TOLERANCE" in normalize_source
    assert "bank_reconciliation" in summary_source


def test_manual_bank_reconciliation_status_suggestion_stays_separate_from_fiscal_status():
    status, difference = internal_users._normalize_bank_reconciliation_status(
        "conciliada",
        Decimal("90.00"),
        Decimal("100.00"),
    )
    assert status == "parcial"
    assert difference == Decimal("-10.00")

    status, difference = internal_users._normalize_bank_reconciliation_status(
        "conciliada",
        Decimal("101.50"),
        Decimal("100.00"),
    )
    assert status == "diferencia"
    assert difference == Decimal("1.50")

    status, difference = internal_users._normalize_bank_reconciliation_status(
        "conciliada",
        Decimal("100.50"),
        Decimal("100.00"),
    )
    assert status == "conciliada"
    assert difference == Decimal("0.50")

    status, difference = internal_users._normalize_bank_reconciliation_status(
        "pendiente",
        Decimal("0.00"),
        Decimal("100.00"),
    )
    assert status == "pendiente"
    assert difference == Decimal("0.00")


def test_gas_lp_discount_type_controls_exist_without_backend_contract_change():
    assistant_html = _assistant_frontend_source()
    conciliacion_html = _conciliacion_frontend_source()

    for token in (
        "descuentoTipo",
        "tab-descuentos",
        "panel-descuentos",
        "clienteDiscountFields",
        "facturaDiscountInfo",
        "inferInvoiceDiscountType",
        "discountPromedioLitro",
        "Promedio descuento por litro",
        "Desc. por litro",
        "Descuento autorizado del cliente aplicado automáticamente",
        "Descuento modificado manualmente para esta factura",
        "updateClientCreditForm",
        "Selecciona el tipo de descuento autorizado",
        "descuento_facturacion",
        "sin_descuento",
        "por_litro",
        "total_pesos",
        "precio_especial",
        "Descuento total en pesos",
        "discountGrossValue",
        "discountPerLiterForPayload",
        "descuento: isTraspaso ? 0 : descuentoPayloadVal",
    ):
        assert token in assistant_html

    assert 'value="total_pesos"' in assistant_html
    assert 'value="por_litro"' in assistant_html
    assert "Promedio por factura" not in assistant_html
    assert "<th>Promedio</th>" not in assistant_html
    assert "discountPromedioFactura" not in assistant_html
    assert "Notas de crédito" not in assistant_html
    assert "Vigencia inicio" not in assistant_html
    assert "Vigencia fin" not in assistant_html
    assert "Notas de descuento" not in assistant_html
    assert 'value="porcentaje"' not in assistant_html

    for token in (
        "pubDescuentoTipo",
        "sin_descuento",
        "por_litro",
        "total_pesos",
        "Descuento total en pesos",
        "publicoDiscountGross",
        "publicoDiscountPerLiter",
        "descuento:preview.descuento_por_litro_backend",
    ):
        assert token in conciliacion_html


def test_gas_lp_cliente_discount_policy_is_stored_in_metadata_without_migration():
    payload = internal_users.GasLpInternalClientePayload(
        rfc="TSJ010101ABC",
        nombre="TORTILLERIA SAN JOSE",
        cp="20000",
        regimen_fiscal="601",
        uso_cfdi="G03",
        descuento_activo=True,
        tipo_descuento_cliente="por_litro",
        descuento_valor=0.5,
    )
    row = internal_users._gas_lp_cliente_row(
        {"owner_user_id": "user-1", "tenant_id": None, "perfil_id": 1, "id": "assistant-1", "display_name": "Asistente"},
        payload,
    )
    normalized = internal_users._normalize_gas_lp_cliente_credit(row)

    assert "descuento_facturacion" in row["metadata"]
    assert row["metadata"]["descuento_facturacion"]["tipo"] == "por_litro"
    assert row["metadata"]["descuento_facturacion"]["valor"] == 0.5
    assert normalized["descuento_facturacion"]["activo"] is True
    assert normalized["descuento_facturacion"]["tipo"] == "por_litro"

    no_discount = internal_users._gas_lp_cliente_row(
        {"owner_user_id": "user-1", "tenant_id": None, "perfil_id": 1, "id": "assistant-1", "display_name": "Asistente"},
        internal_users.GasLpInternalClientePayload(rfc="HCE010101ABC", nombre="HOTEL CENTRAL", cp="20000", regimen_fiscal="601", uso_cfdi="G03"),
    )
    assert no_discount["metadata"]["descuento_facturacion"]["activo"] is False
    assert no_discount["metadata"]["descuento_facturacion"]["tipo"] == "sin_descuento"

    total_discount = internal_users._gas_lp_cliente_row(
        {"owner_user_id": "user-1", "tenant_id": None, "perfil_id": 1, "id": "assistant-1", "display_name": "Asistente"},
        internal_users.GasLpInternalClientePayload(rfc="HCE010101ABC", nombre="HOTEL CENTRAL", cp="20000", regimen_fiscal="601", uso_cfdi="G03", descuento_activo=True, tipo_descuento_cliente="total_pesos", descuento_valor=100),
    )
    assert total_discount["metadata"]["descuento_facturacion"]["tipo"] == "total_pesos"
    assert total_discount["metadata"]["descuento_facturacion"]["valor"] == 100

    special_price = internal_users._gas_lp_cliente_row(
        {"owner_user_id": "user-1", "tenant_id": None, "perfil_id": 1, "id": "assistant-1", "display_name": "Asistente"},
        internal_users.GasLpInternalClientePayload(rfc="HCE010101ABC", nombre="HOTEL CENTRAL", cp="20000", regimen_fiscal="601", uso_cfdi="G03", descuento_activo=True, tipo_descuento_cliente="precio_especial", precio_especial_litro=10.5),
    )
    assert special_price["metadata"]["descuento_facturacion"]["tipo"] == "precio_especial"
    assert special_price["metadata"]["descuento_facturacion"]["precio_especial_litro"] == 10.5


def test_asistente_credito_ppd_dashboard_has_config_shortcut_and_bottom_detail():
    assistant_html = _assistant_frontend_source()

    for token in (
        "configureDashboardClient",
        "dashboard-policy-btn",
        "clienteCreditFields",
        "renderDashboard();",
        '<div class="dashboard-layout">',
        "Facturas del cliente",
    ):
        assert token in assistant_html

    assert "grid-template-columns:1.35fr .85fr" not in assistant_html


def test_asistente_carta_porte_instalaciones_fallback_to_admin_facilities():
    assistant_html = _assistant_frontend_source()

    for token in (
        "function assistantCpInstallationRows()",
        "(FACILITIES || []).forEach",
        "(CATALOGOS.instalaciones || []).forEach",
        "function assistantCpRows(kind)",
        "kind === 'instalaciones' ? assistantCpInstallationRows()",
        "assistantCpRows('instalaciones').filter",
    ):
        assert token in assistant_html


def test_assistant_carta_porte_route_form_is_operational_and_derives_facility_data():
    html = _assistant_frontend_source()
    payload = internal_users._internal_cp_payload(
        "rutas",
        {
            "nombre": "Ags a GDL Principal",
            "origen_facility_id": "10",
            "destino_facility_id": "20",
            "distancia_km": "250,5",
            "tiempo_estimado_minutos": "180",
            "cp_origen": "20000",
            "nombre_origen": "Planta Aguascalientes",
            "localidad_origen": "Aguascalientes",
            "municipio_origen": "Aguascalientes",
            "estado_origen": "Aguascalientes",
            "id_ubicacion_origen": "OR0001",
            "cp_destino": "44100",
            "nombre_destino": "Estacion Guadalajara",
            "localidad_destino": "Guadalajara",
            "municipio_destino": "Guadalajara",
            "estado_destino": "Jalisco",
            "id_ubicacion_destino": "DE0001",
        },
    )

    assert payload["nombre"] == "Ags a GDL Principal"
    assert payload["origen_facility_id"] == 10
    assert payload["destino_facility_id"] == 20
    assert payload["distancia_km"] == 250.5
    assert payload["tiempo_estimado_minutos"] == 180
    assert payload["metadata"]["cp_origen"] == "20000"
    assert payload["metadata"]["localidad_origen"] == "Aguascalientes"
    assert payload["metadata"]["id_ubicacion_destino"] == "DE0001"
    assert payload["metadata"]["vehiculo_default_id"] is None
    assert payload["metadata"]["chofer_default_id"] is None
    assert payload["metadata"]["mercancia_default_id"] is None

    for removed in (
        "CP origen",
        "CP destino",
        "Localidad origen",
        "Localidad destino",
        "Vehículo default opcional",
        "Chofer default opcional",
        "Mercancía default fija",
        "acpr_cpo",
        "acpr_cpd",
        "acpr_veh",
        "acpr_chof",
        "acpr_merc",
        "Configura primero la mercancía Gas LP válida para poder guardar rutas",
    ):
        assert removed not in html

    for required in (
        "cpRouteFacilityPayload('origen', acpr_origen.value)",
        "cpRouteFacilityPayload('destino', acpr_destino.value)",
        "cpFacilityById",
        "Distancia recorrida km",
        "Duración estimada minutos",
        "origen y destino deben ser distintos",
        "cpName('instalaciones', row.origen_facility_id)",
        "cpVehiculo) cpVehiculo.value = '';",
        "cpChofer) cpChofer.value = '';",
    ):
        assert required in html


def test_gas_lp_cliente_credit_policy_is_mirrored_in_metadata():
    payload = internal_users.GasLpInternalClientePayload(
        rfc="CJE861017DB4",
        nombre="CARNICOS DE JEREZ",
        cp="99300",
        regimen_fiscal="601",
        uso_cfdi="G03",
        email="factura@cliente.mx",
        email_adicional_1="contabilidad@cliente.mx",
        credito_habilitado=True,
        dias_credito=60,
        limite_credito=1000,
        credito_notas="Autorizado",
    )
    row = internal_users._gas_lp_cliente_row(
        {"owner_user_id": "00000000-0000-0000-0000-000000000001", "tenant_id": None, "perfil_id": 1, "id": "assistant-1"},
        payload,
    )

    assert row["credito_habilitado"] is True
    assert row["dias_credito"] == 60
    assert row["email_facturacion"] == "factura@cliente.mx"
    assert row["metadata"]["email_facturacion"] == "factura@cliente.mx"
    assert row["metadata"]["invoice_email_additional"] == ["contabilidad@cliente.mx"]
    assert row["metadata"]["credito_ppd"]["dias_credito"] == 60

    fallback = internal_users._normalize_gas_lp_cliente_credit(
        {
            "metadata": {
                "credito_ppd": row["metadata"]["credito_ppd"],
                "email_facturacion": row["metadata"]["email_facturacion"],
                "invoice_email_additional": row["metadata"]["invoice_email_additional"],
            }
        }
    )
    assert fallback["credito_habilitado"] is True
    assert fallback["dias_credito"] == 60
    assert fallback["email_facturacion"] == "factura@cliente.mx"
    assert fallback["metadata"]["invoice_email_additional"] == ["contabilidad@cliente.mx"]


def test_gas_lp_cliente_credit_policy_is_preserved_on_update_payload():
    payload = internal_users.GasLpInternalClientePayload(
        rfc="SCP101217PB4",
        nombre="SEMINARIO CONCILIAR DE LA PURISIMA",
        cp="98600",
        regimen_fiscal="603",
        uso_cfdi="G03",
        email="auregaslux@grupoemurcia.com.mx",
        email_adicional_1="seminariodezacatecas@gmail.com",
        credito_habilitado=True,
        dias_credito=15,
        limite_credito=None,
        credito_notas="Cobranza semanal",
    )
    row = internal_users._gas_lp_cliente_update_row(
        {"owner_user_id": "admin", "tenant_id": "tenant-a", "perfil_id": 7, "id": "assistant-1", "display_name": "ANABEL"},
        payload,
    )

    assert "created_at" not in row
    assert row["credito_habilitado"] is True
    assert row["dias_credito"] == 15
    assert row["email_facturacion"] == "auregaslux@grupoemurcia.com.mx"
    assert row["metadata"]["credito_ppd"]["credito_habilitado"] is True
    assert row["metadata"]["credito_ppd"]["dias_credito"] == 15
    assert row["metadata"]["email_facturacion"] == "auregaslux@grupoemurcia.com.mx"
    assert row["metadata"]["invoice_email_additional"] == ["seminariodezacatecas@gmail.com"]
    assert row["metadata"]["updated_by"] == "ANABEL"


def test_gas_lp_cliente_credit_normalization_prefers_metadata_policy_when_columns_are_stale():
    normalized = internal_users._normalize_gas_lp_cliente_credit(
        {
            "credito_habilitado": False,
            "dias_credito": 0,
            "metadata": {
                "credito_ppd": {
                    "credito_habilitado": True,
                    "dias_credito": 15,
                    "limite_credito": None,
                    "credito_notas": "Autorizado",
                }
            },
        }
    )

    assert normalized["credito_habilitado"] is True
    assert normalized["dias_credito"] == 15
    assert normalized["credito_notas"] == "Autorizado"


def test_gas_lp_cliente_scope_uses_is_null_for_legacy_tenant():
    class Query:
        def __init__(self):
            self.calls = []

        def eq(self, key, value):
            self.calls.append(("eq", key, value))
            return self

        def is_(self, key, value):
            self.calls.append(("is", key, value))
            return self

    q = Query()
    internal_users._gas_lp_clientes_scope_query(
        q,
        {
            "owner_user_id": "00000000-0000-0000-0000-000000000001",
            "tenant_id": None,
            "perfil_id": 1,
        },
    )

    assert ("eq", "user_id", "00000000-0000-0000-0000-000000000001") in q.calls
    assert ("eq", "perfil_id", 1) in q.calls
    assert ("is", "tenant_id", "null") in q.calls
    assert not any(call[0] == "eq" and call[1] == "tenant_id" for call in q.calls)


def test_conciliacion_publico_general_payload_keeps_operational_defaults():
    payload = internal_users.GasLpConciliacionPublicoGeneralPayload(
        litros=80,
        precio_unitario=10.5,
        facility_id=12,
    )
    data = _dump_model(payload)

    assert data["forma_pago"] == "01"
    assert data["metodo_pago"] == "PUE"
    assert data["factura_global"] is False
    assert data["informacion_global_periodicidad"] == "04"
    assert "probar_claves_producto" not in data
    assert "stop_on_success" not in data


def test_publico_general_xml_auto_includes_informacion_global():
    xml, totals = internal_users._build_gas_lp_consumo_xml(
        issuer={
            "rfc": "AGA990907II8",
            "nombre": "AURE GAS",
            "cp": "99300",
            "regimen": "601",
        },
        receptor={
            "rfc": "XAXX010101000",
            "nombre": "PUBLICO EN GENERAL",
            "cp": "99300",
            "regimen_fiscal": "616",
            "uso_cfdi": "S01",
        },
        litros=100,
        precio_unitario=10,
        concepto="LITRO DE GAS LP",
        forma_pago="01",
        metodo_pago="PUE",
        descuento=0,
        iva_rate=0.16,
        serie="P3U30",
        folio="000003",
        comentarios="",
        fecha="2026-06-03T10:00",
        clave_prod_serv="15111510",
        no_identificacion="GLP-LTR",
        unidad="Litro",
        hyp={},
        informacion_global=None,
    )

    assert totals["total"] == 1000.0
    assert '<cfdi:InformacionGlobal Periodicidad="04" Meses="06" Año="2026"/>' in xml
    assert 'Rfc="XAXX010101000" Nombre="PUBLICO EN GENERAL"' in xml


def test_publico_general_xml_detects_accented_receptor_name():
    xml, totals = internal_users._build_gas_lp_consumo_xml(
        issuer={
            "rfc": "AGA990907II8",
            "nombre": "AURE GAS",
            "cp": "99300",
            "regimen": "601",
        },
        receptor={
            "rfc": "XAXX010101000",
            "nombre": "Público en general",
            "cp": "99300",
            "regimen_fiscal": "616",
            "uso_cfdi": "S01",
        },
        litros=100,
        precio_unitario=10,
        concepto="LITRO DE GAS LP",
        forma_pago="01",
        metodo_pago="PUE",
        fecha="2026-06-03T10:00",
        hyp={},
        informacion_global=None,
    )

    assert totals["total"] == 1000.0
    assert '<cfdi:InformacionGlobal Periodicidad="04" Meses="06" Año="2026"/>' in xml


def test_conciliacion_backend_records_origin_and_uses_conciliation_export_columns():
    public_source = inspect.getsource(internal_users.gas_lp_conciliacion_facturar_publico_general)
    export_source = inspect.getsource(internal_users.gas_lp_conciliacion_export_excel)
    complemento_source = inspect.getsource(internal_users.gas_lp_generar_complemento_pago)

    assert '"created_by_area": "conciliacion"' in public_source
    assert '"created_by_area": "conciliacion"' in complemento_source
    assert '"portal": "conciliacion_gas_lp"' in public_source
    assert 'or "Conciliación"' in public_source
    assert '"descuento": totals["descuento"]' in public_source

    for column in (
        "Fecha",
        "Folio de fact",
        "UUID",
        "Cliente",
        "Monto",
        "Litros",
        "Método",
    ):
        assert column in export_source


def test_conciliacion_export_excel_handles_decimal_null_metadata_and_transfer(monkeypatch):
    rows = [
        {
            "id": 1,
            "metadata": {
                "fecha_emision": "2026-06-01T10:00:00",
                "folio": "100",
                "cliente_nombre": "ALFA GAS CLIENTE",
                "total": Decimal("116.50"),
                "metodo_pago": "PPD",
            },
            "volumen_litros": Decimal("50.125"),
            "importe": Decimal("100.43"),
            "created_at": "2026-06-01T10:00:00",
        },
        {
            "id": 2,
            "metadata": None,
            "rfc_receptor": "XAXX010101000",
            "volumen_litros": None,
            "importe": Decimal("25.00"),
            "created_at": "2026-06-02T10:00:00",
        },
        {
            "id": 3,
            "metadata": {
                "fecha_cfdi": "2026-06-03T10:00:00",
                "folio_usuario": "TR-3",
                "cliente_nombre": "TRASPASO",
                "total": Decimal("0"),
                "metodo_pago": "PUE",
                "operation_type": "transfer",
            },
            "volumen_litros": Decimal("10"),
            "importe": None,
            "created_at": "2026-06-03T10:00:00",
        },
    ]
    monkeypatch.setattr(internal_users, "_gas_lp_conciliacion_context", lambda token, perfil_id=None: {"user": {"perfil_id": perfil_id or 7}})
    monkeypatch.setattr(internal_users, "_gas_lp_profile", lambda user, require_module_marker=True: {"id": user["perfil_id"], "nombre": "ALFA GAS", "rfc": "AAA010101AAA"})
    monkeypatch.setattr(internal_users, "get_supabase_admin", lambda: object())
    monkeypatch.setattr(internal_users, "_gas_lp_company_facturas_rows", lambda sb, user, profile, month="", limit=10000, include_carta_porte=True: rows)
    monkeypatch.setattr(internal_users, "_gas_lp_attach_internal_creators", lambda sb, rows: None)

    response = asyncio.run(
        internal_users.gas_lp_conciliacion_export_excel(
            token="token",
            period="2026-06",
            profile_id=7,
        )
    )

    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    wb = load_workbook(BytesIO(response.body))
    ws = wb.active
    assert [cell.value for cell in ws[1]] == [
        "Fecha",
        "Folio de fact",
        "UUID",
        "Cliente",
        "Monto",
        "Litros",
        "Método",
    ]
    assert ws.max_row == 4
    assert ws["A2"].value == "2026-06-01"
    assert ws["B2"].value == "100"
    assert ws["D2"].value == "ALFA GAS CLIENTE"
    assert ws["E2"].value == 116.5
    assert ws["F2"].value == 50.125
    assert ws["G2"].value == "PPD"
    assert ws["D3"].value == "XAXX010101000"
    assert ws["G4"].value == "PUE"


def test_gas_lp_excel_exports_use_previous_compact_column_order(monkeypatch):
    cancelled_uuid = "6d09da66-366b-41f0-b778-0b89ed625b5f"
    vigente_uuid = "aa054375-5a74-44c3-ad8c-e2e1070512c4"
    rows = [
        {
            "id": 1,
            "uuid_sat": cancelled_uuid,
            "status": "Cancelada fiscalmente",
            "metadata": {
                "fecha_emision": "2026-06-01T10:00:00",
                "folio": "000001",
                "cliente_nombre": "J JESUS ROBLES NAVA",
                "total": Decimal("4990.50"),
                "metodo_pago": "PUE",
                "estado_fiscal": "cancelada_fiscalmente",
            },
            "volumen_litros": Decimal("1"),
            "importe": Decimal("4302.16"),
            "created_at": "2026-06-01T10:00:00",
        },
        {
            "id": 2,
            "uuid_sat": vigente_uuid,
            "status": "timbrada",
            "metadata": {
                "fecha_emision": "2026-06-01T11:00:00",
                "folio": "000002",
                "cliente_nombre": "J JESUS ROBLES NAVA",
                "total": Decimal("4540.50"),
                "metodo_pago": "PUE",
            },
            "volumen_litros": Decimal("1"),
            "importe": Decimal("3914.22"),
            "created_at": "2026-06-01T11:00:00",
        },
        {
            "id": 3,
            "uuid_sat": "ppd-uuid",
            "status": "timbrada",
            "metadata": {
                "fecha_emision": "2026-06-01T12:00:00",
                "folio": "000003",
                "cliente_nombre": "CLIENTE CREDITO",
                "total": Decimal("1000.00"),
                "metodo_pago": "PPD",
            },
            "volumen_litros": Decimal("1"),
            "importe": Decimal("862.07"),
            "created_at": "2026-06-01T12:00:00",
        },
    ]

    monkeypatch.setattr(internal_users, "_gas_lp_conciliacion_context", lambda token, perfil_id=None: {"user": {"perfil_id": perfil_id or 7, "tenant_id": "t1"}})
    monkeypatch.setattr(internal_users, "_gas_lp_internal_context", lambda token: {"user": {"perfil_id": 7, "tenant_id": "t1"}})
    monkeypatch.setattr(internal_users, "_gas_lp_profile", lambda user, require_module_marker=False: {"id": user["perfil_id"], "nombre": "ALFA GAS", "rfc": "AAA010101AAA"})
    monkeypatch.setattr(internal_users, "get_supabase_admin", lambda: object())
    monkeypatch.setattr(internal_users, "_gas_lp_company_facturas_rows", lambda sb, user, profile, month="", limit=10000, include_carta_porte=True: rows)
    monkeypatch.setattr(internal_users, "_gas_lp_attach_internal_creators", lambda sb, rows: None)

    conc_response = asyncio.run(
        internal_users.gas_lp_conciliacion_export_excel(
            token="token",
            period="2026-06",
            profile_id=7,
        )
    )
    assistant_response = asyncio.run(
        internal_users.gas_lp_internal_facturas_export_dia(
            token="token",
            fecha="2026-06-01",
        )
    )

    conc_wb = load_workbook(BytesIO(conc_response.body))
    conc_ws = conc_wb.active
    assert [cell.value for cell in conc_ws[1]] == ["Fecha", "Folio de fact", "UUID", "Cliente", "Monto", "Litros", "Método"]

    assistant_wb = load_workbook(BytesIO(assistant_response.body))
    assistant_ws = assistant_wb.active
    headers = [cell.value for cell in assistant_ws[1]]
    assert headers == ["Fecha", "Folio de fact", "UUID", "Cliente", "Monto", "Litros", "Método", "Estado fiscal"]
    uuid_col = headers.index("UUID") + 1
    method_col = headers.index("Método") + 1
    status_col = headers.index("Estado fiscal") + 1
    by_uuid = {
        assistant_ws.cell(row=i, column=uuid_col).value: (
            assistant_ws.cell(row=i, column=method_col).value,
            assistant_ws.cell(row=i, column=status_col).value,
        )
        for i in range(2, assistant_ws.max_row + 1)
    }
    assert by_uuid[cancelled_uuid] == ("PUE", "Cancelada")
    assert by_uuid[vigente_uuid] == ("PUE", "Vigente")
    assert by_uuid["ppd-uuid"] == ("PPD", "Vigente")


def test_assistant_facturas_endpoint_exposes_shared_fiscal_status(monkeypatch):
    cancelled_uuid = "70d87aa2-b633-43c2-8195-0c887a1e0fa6"
    rows = [
        {
            "id": 1,
            "uuid_sat": cancelled_uuid,
            "status": "Cancelada fiscalmente",
            "metadata": {
                "fecha_emision": "2026-06-08T09:47:00",
                "folio": "000001",
                "cliente_nombre": "PUBLICO EN GENERAL",
                "total": 20785.43,
                "metodo_pago": "PUE",
                "estado_fiscal": "cancelada_fiscalmente",
                "cancelacion_acuse": {"ok": True},
            },
            "volumen_litros": 1874.2497,
            "importe": 17918.47,
            "created_at": "2026-06-08T09:47:00",
        },
        {
            "id": 2,
            "uuid_sat": "vigente-uuid",
            "status": "timbrada",
            "metadata": {
                "fecha_emision": "2026-06-08T10:00:00",
                "folio": "000002",
                "cliente_nombre": "PUBLICO EN GENERAL",
                "total": 100.0,
                "metodo_pago": "PUE",
            },
            "volumen_litros": 1,
            "importe": 86.21,
            "created_at": "2026-06-08T10:00:00",
        },
    ]

    monkeypatch.setattr(internal_users, "_gas_lp_internal_context", lambda token: {"user": {"perfil_id": 7, "tenant_id": "t1"}})
    monkeypatch.setattr(internal_users, "_gas_lp_profile", lambda user, require_module_marker=False: {"id": user["perfil_id"], "nombre": "AURE GAS", "rfc": "AGA990907II8"})
    monkeypatch.setattr(internal_users, "get_supabase_admin", lambda: object())
    monkeypatch.setattr(internal_users, "_gas_lp_company_facturas_rows", lambda sb, user, profile, month="", limit=10000: rows)
    monkeypatch.setattr(internal_users, "_gas_lp_attach_internal_creators", lambda sb, rows: None)
    monkeypatch.setattr(internal_users, "_gas_lp_attach_cliente_email_recipients", lambda sb, user, rows: None)
    monkeypatch.setattr(internal_users, "_gas_lp_complementos_por_factura", lambda sb, ids: {})

    response = asyncio.run(internal_users.gas_lp_internal_facturas(token="token", mes="2026-06"))
    payload = json.loads(response.body)
    by_uuid = {row["uuid_sat"]: row["fiscal_status"] for row in payload["facturas"]}

    assert by_uuid[cancelled_uuid]["code"] == "cancelada"
    assert by_uuid[cancelled_uuid]["label"] == "Cancelada"
    assert by_uuid["vigente-uuid"]["code"] == "vigente"
    assert by_uuid["vigente-uuid"]["label"] == "Vigente"


def test_conciliacion_complemento_payload_supports_multiple_ppd_invoices():
    payload = internal_users.GasLpComplementoPagoPayload(
        factura_ids=[10, 11],
        facturas=[{"factura_id": 10, "monto": 100.0}, {"factura_id": 11, "monto": 50.0}],
        referencia="DEP-123",
        banco="BBVA",
    )
    data = _dump_model(payload)

    assert data["forma_pago"] == "03"
    assert data["factura_ids"] == [10, 11]
    assert data["facturas"][0]["monto"] == 100.0
    assert data["referencia"] == "DEP-123"
    assert data["banco"] == "BBVA"


def test_transfer_email_default_contract_keeps_transfer_email_explicit():
    payload = internal_users.GasLpInternalFacturaPayload(
        litros=10,
        precio_unitario=8,
        tipo_operacion="traspaso",
        transfer_email="",
        transfer_email_provided=True,
    )
    data = _dump_model(payload)
    create_source = inspect.getsource(internal_users._gas_lp_internal_crear_factura_impl)
    html = _assistant_frontend_source()

    assert data["transfer_email_provided"] is True
    assert "payload.transfer_email_provided" in create_source
    assert "transfer-email-default" in html
    assert "Guardar como correo predeterminado para traspasos" in html


def test_assistant_facturas_table_shows_fiscal_status_column():
    html = _assistant_frontend_source()

    assert "Estado fiscal" in html
    assert "fiscal_status" in html
    assert "facturaStatusHtml(f)" in html
    assert '<td class="status-cell">${v.statusHtml}</td>' in html
    assert '<td colspan="11">No fue posible cargar facturas. Presiona Actualizar.</td>' in html
    assert '<tr><td colspan="11">${esc(emptyText)}</td></tr>' in html


def test_assistant_invoice_endpoint_returns_controlled_error_on_unexpected_failure(monkeypatch):
    async def boom(_payload, _token):
        raise RuntimeError("unexpected backend failure")

    monkeypatch.setitem(
        internal_users.gas_lp_internal_crear_factura.__globals__,
        "_gas_lp_internal_crear_factura_impl",
        boom,
    )
    payload = internal_users.GasLpInternalFacturaPayload(litros=10, precio_unitario=10, facility_id=4)

    try:
        asyncio.run(internal_users.gas_lp_internal_crear_factura(payload, "tok"))
    except HTTPException as exc:
        assert exc.status_code == 500
        assert exc.detail == "No se pudo completar la operación. Intenta de nuevo o contacta a soporte."
    else:
        raise AssertionError("expected controlled HTTPException")


def test_assistant_invoice_timbrado_does_not_import_transporte_runtime_helpers():
    create_source = inspect.getsource(internal_users._gas_lp_internal_crear_factura_impl)
    cliente_source = inspect.getsource(internal_users._gas_lp_cliente_row)

    assert "from routes.transporte import" not in create_source
    assert "from routes.transporte import" not in cliente_source
    assert "_gas_lp_normalizar_receptor_cfdi" in create_source
    assert "_gas_lp_validar_datos_cfdi_receptor" in create_source


def test_assistant_invoice_duplicate_guard_runs_before_stamp():
    create_source = inspect.getsource(internal_users._gas_lp_internal_crear_factura_impl)
    duplicate_source = inspect.getsource(internal_users._gas_lp_existing_sale_invoice)

    assert "_gas_lp_existing_sale_invoice(sb, user, payload, totals, receptor)" in create_source
    assert "gas_lp_invoice_duplicate" in create_source
    assert "timbrar_cfdi(xml)" in create_source
    assert create_source.index("_gas_lp_existing_sale_invoice(sb, user, payload, totals, receptor)") < create_source.index("timbrar_cfdi(xml)")
    assert "volumen_litros" in duplicate_source
    assert "internal_user_id" in duplicate_source
    assert "target_total" in duplicate_source
    assert "duplicate_window_seconds" in duplicate_source


def test_assistant_invoice_duplicate_guard_only_blocks_immediate_retry(monkeypatch):
    class FakeResult:
        def __init__(self, data):
            self.data = data

    class FakeQuery:
        def __init__(self, rows):
            self.rows = rows

        def select(self, *_args, **_kwargs):
            return self

        def eq(self, *_args, **_kwargs):
            return self

        def order(self, *_args, **_kwargs):
            return self

        def limit(self, *_args, **_kwargs):
            return self

        def execute(self):
            return FakeResult(self.rows)

    class FakeSupabase:
        def __init__(self, rows):
            self.rows = rows

        def table(self, _name):
            return FakeQuery(self.rows)

    fixed_now = datetime(2026, 6, 9, 17, 23, 0, tzinfo=timezone.utc)
    monkeypatch.setitem(internal_users._gas_lp_existing_sale_invoice.__globals__, "_now", lambda: fixed_now)
    payload = internal_users.GasLpInternalFacturaPayload(litros=180, precio_unitario=11.09, facility_id=5)
    totals = {"fecha": "2026-06-09T11:23:00", "total": Decimal("1996.20")}
    receptor = {"rfc": "XAXX010101000", "nombre": "PUBLICO EN GENERAL"}
    user = {"id": 77, "tenant_id": "t1", "perfil_id": 7}
    base_row = {
        "id": 1,
        "uuid_sat": "same-amount-uuid",
        "status": "timbrada",
        "rfc_receptor": "XAXX010101000",
        "volumen_litros": 180,
        "facility_id": 5,
        "metadata": {
            "fecha_emision": "2026-06-09T11:22:00",
            "internal_user_id": 77,
            "origen_facility_id": 5,
            "total": 1996.20,
        },
    }

    old_row = {**base_row, "created_at": (fixed_now - timedelta(seconds=90)).isoformat()}
    immediate_row = {**base_row, "created_at": (fixed_now - timedelta(seconds=5)).isoformat()}

    assert internal_users._gas_lp_existing_sale_invoice(FakeSupabase([old_row]), user, payload, totals, receptor) is None
    assert internal_users._gas_lp_existing_sale_invoice(FakeSupabase([immediate_row]), user, payload, totals, receptor)["id"] == 1


def test_assistant_load_facturas_does_not_pollute_main_invoice_status_by_default():
    html = _assistant_frontend_source()
    load_start = html.index("async function loadFacturas")
    load_end = html.index("async function loadComplementos", load_start)
    load_source = html[load_start:load_end]

    assert "opts.surfaceError" in load_source
    assert "setStatus('facturaMsg'" in load_source


def test_assistant_startup_errors_do_not_override_user_facility_selection():
    html = _assistant_frontend_source()

    assert "let invoiceUserInteractedAt = 0" in html
    assert "function markInvoiceInteraction()" in html
    assert "function canShowStartupInvoiceError(loadStartedAt)" in html
    assert "const loadStartedAt = Date.now();" in html
    assert "critical && canShowStartupInvoiceError(loadStartedAt)" in html
    assert "function onFacilityChange()" in html
    assert "markInvoiceInteraction();" in html


def test_assistant_client_save_feedback_stays_visible_after_form_hides():
    html = _assistant_frontend_source()

    assert 'id="clientesNotice"' in html
    assert "function setClientesFeedback" in html
    assert "setStatus('clientesNotice'" in html
    assert "Guardando cambios del cliente..." in html
    assert "btnGuardarCliente.disabled = true" in html
    assert "clienteFormClientes.classList.add('hide');" in html
    assert "Cliente guardado'} y seleccionado para facturar" in html


def test_assistant_today_invoices_use_backend_date_key_and_current_month():
    html = _assistant_frontend_source()

    assert "month || document.getElementById('facturaMes')?.value || todayKey().slice(0,7)" in html
    assert "await loadFacturas(month || todayKey().slice(0,7))" in html
    assert "f.fecha_factura_key || facturaDateValue(f)" in html
    assert "todayFacturasRows" in html


def test_conciliacion_publico_general_list_uses_backend_date_key():
    html = _conciliacion_frontend_source()
    summary_source = inspect.getsource(internal_users.gas_lp_conciliacion_summary)

    assert 'row["fecha_factura_key"] = _gas_lp_factura_date_key(row)' in summary_source
    assert "function facturaDateKey(f)" in html
    assert "facturaDateKey(f)!==day" in html
    assert "publicNameKey" in html
    assert "isPublicoGeneral(f)&&facturaDateKey(f)===key" in html


def test_assistant_carta_porte_catalog_save_accepts_decimal_comma_and_confirms_visibility():
    html = _assistant_frontend_source()
    payload = internal_users._internal_cp_payload(
        "mercancias",
        {
            "alias": "Gas LP",
            "bienes_transp": "15111510",
            "factor_kg_litro": "0,524",
            "material_peligroso": "1",
            "clave_material_peligroso": "1075",
            "embalaje": "Z01",
        },
    )

    assert payload["factor_kg_litro"] == 0.524
    assert "function cpDecimalValue" in html
    assert "function assistantCpUpsertLocal" in html
    assert "assistantCpRecordFromResponse(kind, saved, p, id)" in html
    assert "El servidor respondió, pero el registro aún no aparece" not in html
    assert "inputmode=\"decimal\" placeholder=\"0.524\"" in html


def test_assistant_carta_porte_driver_form_requires_rfc_and_keeps_curp_internal():
    html = _assistant_frontend_source()
    payload = internal_users._internal_cp_payload(
        "choferes",
        {
            "nombre_completo": "Operador Prueba",
            "rfc": "OPEP850101AB1",
            "curp": "OPEP850101HZSPRR01",
            "licencia_federal": "LIC123",
            "tipo_licencia_federal": "E",
            "tipo_figura_sat": "01",
            "fecha_expedicion_licencia": "2026-01-01",
            "fecha_vencimiento_licencia": "2028-01-01",
        },
    )

    assert payload["nombre"] == "Operador Prueba"
    assert payload["rfc"] == "OPEP850101AB1"
    assert payload["metadata"]["curp"] == "OPEP850101HZSPRR01"
    assert payload["licencia"] == "LIC123"
    assert payload["metadata"]["tipo_licencia"] == "E"
    assert payload["metadata"]["tipo_figura"] == "01"
    assert payload["metadata"]["fecha_expedicion_licencia"] == "2026-01-01"
    assert payload["metadata"]["fecha_vencimiento_licencia"] == "2028-01-01"
    assert "Parte transporte" not in html
    assert "acpc_parte" not in html
    assert "acpc_curp" in html
    assert "RFC Figura SAT" in html
    assert "CURP interna / referencia" in html
    assert "no sustituye RFCFigura" in html
    assert "Expedición licencia" in html
    assert "Vencimiento licencia" in html
    assert "function calcularEstatusLicencia" in html
    assert "renderAssistantCpDriversSummary" in html
    assert "Licencia vencida" in html
    assert "Por vencer" in html
    assert "Sin vencimiento registrado" in html


def test_carta_porte_vehicle_environmental_insurance_aliases_validate():
    origen = {
        "tipo": "origen",
        "id_ubicacion": "OR123456",
        "rfc": "AGA990907II8",
        "nombre": "AURE GAS",
        "codigo_postal": "98470",
        "estado": "ZAC",
        "municipio": "056",
        "calle": "Planta Villa de Cos Aure",
        "facility_nombre": "Planta Villa de Cos Aure",
    }
    destino = {
        "tipo": "destino",
        "id_ubicacion": "DE123456",
        "rfc": "AGA990907II8",
        "nombre": "AURE GAS",
        "codigo_postal": "98659",
        "estado": "ZAC",
        "municipio": "017",
        "calle": "Estacion Zacatecas",
        "facility_nombre": "Estacion Zacatecas",
    }
    vehiculo = facturas_routes._cp_normalize_vehicle_payload({
        "placas": "AC-6116-E",
        "anio": 2021,
        "config_vehicular": "C2",
        "permiso_cre": "TPAF01",
        "metadata": {
            "numero_economico": "AT-69",
            "numero_permiso": "0170SEFICANLE13",
            "peso_bruto_vehicular": 9249,
            "aseguradora": "INBURSA",
            "poliza_seguro": "16211 20025429",
            "aseguraMedAmbiente": "INBURSA",
            "polizaMedAmbiente": "16211 20025429",
        },
    })
    chofer = facturas_routes._cp_normalize_driver_payload({
        "nombre": "ADAN CASTRO HERNANDEZ",
        "rfc": "CAHA800101AB1",
        "licencia": "LFD01127323",
        "metadata": {"tipo_figura": "01"},
    })
    mercancia = {
        "bienes_transp": "15111510",
        "descripcion": "Gas licuado de petroleo",
        "clave_unidad": "LTR",
        "material_peligroso": True,
        "clave_material_peligroso": "1075",
        "embalaje": "Z01",
    }

    assert vehiculo["aseguradora_medio_ambiente"] == "INBURSA"
    assert vehiculo["poliza_medio_ambiente"] == "16211 20025429"
    facturas_routes._cp_validate_catalog_payload(
        origen=origen,
        destino=destino,
        vehiculo=vehiculo,
        chofer=chofer,
        mercancia=mercancia,
        fecha_salida="2026-06-09T12:00:00",
        fecha_llegada="2026-06-09T13:00:00",
        distancia_km=70,
        litros=50,
        peso_kg=26.2,
    )
    vehiculo_metadata_json = facturas_routes._cp_normalize_vehicle_payload({
        "placas": "AC-6116-E",
        "anio": 2021,
        "config_vehicular": "C2",
        "permiso_cre": "TPAF01",
        "metadata_json": {
            "numero_economico": "AT-69",
            "numero_permiso": "0170SEFICANLE13",
            "peso_bruto_vehicular": 9249,
            "aseguradora": "INBURSA",
            "poliza_seguro": "16211 20025429",
            "aseguraMedAmbiente": "INBURSA",
            "polizaMedAmbiente": "16211 20025429",
        },
    })
    assert vehiculo_metadata_json["aseguradora_medio_ambiente"] == "INBURSA"
    assert vehiculo_metadata_json["poliza_medio_ambiente"] == "16211 20025429"
    facturas_routes._cp_validate_catalog_payload(
        origen=origen,
        destino=destino,
        vehiculo=vehiculo_metadata_json,
        chofer=chofer,
        mercancia=mercancia,
        fecha_salida="2026-06-09T12:00:00",
        fecha_llegada="2026-06-09T13:00:00",
        distancia_km=70,
        litros=50,
        peso_kg=26.2,
    )


def test_carta_porte_driver_curp_does_not_replace_required_rfcfigura():
    chofer = facturas_routes._cp_normalize_driver_payload({
        "nombre": "ADAN CASTRO HERNANDEZ",
        "licencia": "LFD01127323",
        "metadata": {"curp": "CAHA800101HZSSRD01", "tipo_figura": "01"},
    })

    assert chofer["curp"] == "CAHA800101HZSSRD01"
    assert chofer["rfc"] == ""
    with pytest.raises(HTTPException) as exc:
        facturas_routes._cp_validate_catalog_payload(
            origen={"tipo": "origen", "id_ubicacion": "OR", "rfc": "AGA990907II8", "nombre": "AURE GAS", "codigo_postal": "98470", "estado": "ZAC", "municipio": "056", "calle": "Origen"},
            destino={"tipo": "destino", "id_ubicacion": "DE", "rfc": "AGA990907II8", "nombre": "AURE GAS", "codigo_postal": "98659", "estado": "ZAC", "municipio": "017", "calle": "Destino"},
            vehiculo={"placas": "AC-6116-E", "config_vehicular": "C2", "permiso_cre": "TPAF01", "numero_permiso": "0170SEFICANLE13", "peso_bruto_vehicular": 9249, "aseguradora": "INBURSA", "poliza_seguro": "16211 20025429", "aseguradora_medio_ambiente": "INBURSA", "poliza_medio_ambiente": "16211 20025429"},
            chofer=chofer,
            mercancia={"bienes_transp": "15111510", "descripcion": "Gas LP", "clave_unidad": "LTR", "material_peligroso": True, "clave_material_peligroso": "1075", "embalaje": "Z01"},
            fecha_salida="2026-06-09T12:00:00",
            fecha_llegada="2026-06-09T13:00:00",
            distancia_km=70,
            litros=50,
            peso_kg=26.2,
        )

    assert "chofer: RFC Figura SAT" in str(exc.value.detail)


def test_assistant_carta_porte_validation_flow_has_modal_and_real_error_text():
    html = _assistant_frontend_source()
    backend_source = inspect.getsource(facturas_routes._generar_carta_porte_for_scope)

    assert "Validar Carta Porte" not in html
    assert "4. Validación de Carta Porte" not in html
    assert "Timbrar Carta Porte" in html
    assert "Confirmar timbrado Carta Porte tipo T" in html
    assert "Timbrar CFDI tipo T" in html
    assert "flow.includes('traspaso')" not in html
    assert "tipo === 'T'" not in html
    assert "No se pudo conectar con el servidor de timbrado" in html
    assert "Failed to fetch" not in html
    timbrar_start = html.index("async function timbrarCartaPorteGasLp")
    timbrar_end = html.index("async function handleCartaPorteAction", timbrar_start)
    assert "confirm(" not in html[timbrar_start:timbrar_end]
    assert "gas_lp_carta_porte_pac_error" in backend_source
    assert "gas_lp_carta_porte_timbrado_start" in backend_source


def test_carta_porte_xml_adds_seconds_to_browser_datetime_values():
    import re

    xml = build_carta_porte_xml(
        {
            "record_uuid": "CP-FECHA",
            "fecha_hora": "2026-06-09T14:32",
            "fecha_salida": "2026-06-09T14:32",
            "fecha_llegada": "2026-06-09T15:32",
            "volumen_litros": 50,
            "importe": 0,
        },
        {"rfc": "AGA990907II8", "nombre": "AURE GAS", "cp": "98470", "regimen": "601"},
        {"rfc": "AGA990907II8", "nombre": "AURE GAS", "cp": "98470", "regimen": "601", "uso_cfdi": "S01"},
        {
            "placas": "AC-6116-E",
            "anio": 2021,
            "config_vehicular": "C2",
            "permiso_cre": "TPAF02",
            "numero_permiso": "A0122865",
            "peso_bruto_vehicular": 12000,
            "aseguradora": "INBURSA",
            "poliza_seguro": "16211 20025429",
            "aseguradora_medio_ambiente": "INBURSA",
            "poliza_medio_ambiente": "16211 20025429",
        },
        tipo_comprobante="T",
        ruta={"distancia_km": 70},
        origen={"id_ubicacion": "OR123456", "rfc": "AGA990907II8", "nombre": "Planta Villa de Cos Aure", "codigo_postal": "98470", "estado": "ZAC", "municipio": "056", "pais": "MEX"},
        destino={"id_ubicacion": "DE123456", "rfc": "AGA990907II8", "nombre": "Estacion Zacatecas", "codigo_postal": "98659", "estado": "ZAC", "municipio": "017", "pais": "MEX"},
        mercancia={"bienes_transp": "15111510", "descripcion": "Gas LP", "clave_unidad": "LTR", "unidad": "Litro", "material_peligroso": True, "clave_material_peligroso": "1075", "embalaje": "Z01", "factor_kg_litro": 0.524},
        chofer={"nombre": "ADAN CASTRO HERNANDEZ", "rfc": "CAHA800101AB1", "licencia": "LFD01127323", "tipo_figura": "01"},
    )

    assert 'TipoDeComprobante="T"' in xml
    assert re.search(r'IdCCP="CCC[0-9a-f]{5}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}"', xml)
    assert 'SubTotal="0" Total="0" Moneda="XXX"' in xml
    assert '<cfdi:Concepto ClaveProdServ="78101800"' in xml
    assert 'BienesTransp="15111510"' in xml
    assert 'Fecha="2026-06-09T14:32:00"' in xml
    assert 'PlacaVM="AC6116E"' in xml
    assert 'PesoBrutoVehicular="12.00"' in xml
    assert 'CodigoPostal="98470" Estado="ZAC" Municipio=' not in xml
    assert 'CodigoPostal="98470" Estado="ZAC"' in xml
    assert 'CodigoPostal="98659" Estado="ZAC" Municipio="017"' in xml
    assert 'PlacaVM="AC-6116-E"' not in xml
    assert 'NumRegIdTrib=' not in xml
    assert 'ResidenciaFiscal=' not in xml
    assert '<cartaporte31:Remolques' not in xml
    assert 'SectorPVE' not in xml
    assert 'FechaHoraSalidaLlegada="2026-06-09T14:32:00"' in xml
    assert 'FechaHoraSalidaLlegada="2026-06-09T15:32:00"' in xml


def test_carta_porte_id_ccp_uses_sat_pattern():
    import re

    generated = facturas_routes._cp_normalize_id_ccp("")
    assert re.fullmatch(r"CCC[0-9a-f]{5}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}", generated)

    normalized = facturas_routes._cp_normalize_id_ccp("12345678-1234-1234-1234-1234567890ab")
    assert normalized == "CCC45678-1234-1234-1234-1234567890ab"


def test_assistant_carta_porte_vehicle_form_uses_numero_economico_as_alias(monkeypatch):
    html = _assistant_frontend_source()
    monkeypatch.setattr(cp_catalogos, "_gas_lp_profile", lambda user: {"id": 123, "tenant_id": None, "rfc": "GLU760309457"})
    payload = internal_users._internal_cp_payload(
        "vehiculos",
        {
            "numero_economico": "AT-96",
            "placa": "ABC1234",
            "anio": "2024",
            "config_vehicular": "C2",
            "permiso_cre": "TPAF03",
            "numero_permiso": "SCT-123456",
            "peso_bruto_vehicular": "12000",
            "aseguradora": "GNP",
            "poliza_seguro": "RC-123",
            "aseguradora_medio_ambiente": "Ambiental MX",
            "poliza_medio_ambiente": "MA-123",
        },
    )
    record = internal_users._internal_cp_response_record(
        "vehiculos",
        {"id": 96, **payload},
        {"owner_user_id": "user-1", "tenant_id": None, "perfil_id": 123},
    )

    assert payload["metadata"]["alias"] == "AT-96"
    assert payload["metadata"]["numero_economico"] == "AT-96"
    assert payload["metadata"]["peso_bruto_vehicular"] == 12000
    assert payload["metadata"]["aseguradora_carga"] == ""
    assert payload["metadata"]["poliza_carga"] == ""
    assert record["id"] == 96
    assert record["perfil_id"] == 123
    assert "acpv_alias" not in html
    assert "Aseguradora carga" not in html
    assert "Póliza carga" not in html
    assert "Peso bruto vehicular SAT" in html
    assert "peso_bruto_vehicular:acpv_pbv.value" in html
    assert "Aseguradora de responsabilidad civil" in html
    assert "Póliza de responsabilidad civil" in html
    assert "Aseguradora de daños al medio ambiente" in html
    assert "Póliza de daños al medio ambiente" in html
    assert "Seguro obligatorio del vehículo." in html
    assert "Requerido para transporte de material peligroso como Gas LP." in html
    assert "numero_economico:acpv_num.value" in html
    assert "payload.numero_economico" in html


def test_carta_porte_catalog_scope_is_company_not_creator(monkeypatch):
    profiles = {
        101: {"id": 101, "tenant_id": "tenant-gas", "rfc": "GLU760309457", "nombre": "GAS LUX"},
        202: {"id": 202, "tenant_id": "tenant-gas", "rfc": "GLU760309457", "nombre": "GAS LUX"},
        303: {"id": 303, "tenant_id": "tenant-gas", "rfc": "OTR010101AAA", "nombre": "OTRA EMPRESA"},
    }

    monkeypatch.setattr(cp_catalogos, "_gas_lp_profile", lambda user: profiles[user["perfil_id"]])
    anabel = {"id": 1, "owner_user_id": "admin-a", "tenant_id": "tenant-gas", "perfil_id": 101}
    karina = {"id": 2, "owner_user_id": "admin-b", "tenant_id": "tenant-gas", "perfil_id": 202}
    otra = {"id": 3, "owner_user_id": "admin-c", "tenant_id": "tenant-gas", "perfil_id": 303}

    row = cp_catalogos._internal_cp_scope_row(anabel, cp_catalogos._internal_cp_payload("choferes", {"nombre": "Operador GAS LUX", "rfc": "OPEP850101AB1"}))
    row["id"] = 55
    scope_karina = cp_catalogos._internal_cp_company_scope(karina)
    scope_otra = cp_catalogos._internal_cp_company_scope(otra)

    assert row["user_id"] == "admin-a"
    assert row["metadata"]["created_by_internal_user_id"] == 1
    assert row["metadata"]["empresa_rfc"] == "GLU760309457"
    assert cp_catalogos._internal_cp_row_company_match(row, scope_karina) is True
    assert cp_catalogos._internal_cp_row_company_match(row, scope_otra) is False


def test_carta_porte_catalog_legacy_rows_still_match_same_profile(monkeypatch):
    monkeypatch.setattr(cp_catalogos, "_gas_lp_profile", lambda user: {"id": user["perfil_id"], "tenant_id": "tenant-gas", "rfc": "GLU760309457"})
    user = {"id": 1, "owner_user_id": "admin-a", "tenant_id": "tenant-gas", "perfil_id": 101}
    scope = cp_catalogos._internal_cp_company_scope(user)

    assert cp_catalogos._internal_cp_row_company_match({"perfil_id": 101, "metadata": {}}, scope) is True
    assert cp_catalogos._internal_cp_row_company_match({"perfil_id": 202, "metadata": {}}, scope) is False


def test_carta_porte_catalog_existing_supabase_rows_are_listed_by_company_rfc(monkeypatch):
    class FakeResult:
        def __init__(self, data):
            self.data = data

    class FakeQuery:
        def __init__(self, rows):
            self.rows = rows

        def select(self, *_args, **_kwargs):
            return self

        def eq(self, *_args, **_kwargs):
            return self

        def is_(self, *_args, **_kwargs):
            return self

        def or_(self, *_args, **_kwargs):
            return self

        def order(self, *_args, **_kwargs):
            return self

        def execute(self):
            return FakeResult(self.rows)

    class FakeSupabase:
        def __init__(self, tables):
            self.tables = tables

        def table(self, name):
            return FakeQuery(self.tables.get(name, []))

    tables = {
        "gas_lp_choferes": [
            {
                "id": 10,
                "user_id": "otro-creador",
                "tenant_id": None,
                "perfil_id": 999,
                "activo": True,
                "nombre": "MARIA JOSE",
                "licencia": "E123",
                "metadata": {"empresa_rfc": " glu760309457 "},
            },
            {
                "id": 11,
                "user_id": "otro-creador",
                "tenant_id": None,
                "perfil_id": 888,
                "activo": True,
                "nombre": "OTRA EMPRESA",
                "metadata": {"empresa_rfc": "OTR010101AAA"},
            },
        ],
        "gas_lp_mercancias_carta_porte": [
            {
                "id": 20,
                "tenant_id": None,
                "perfil_id": 777,
                "activo": True,
                "alias": "Gas LP",
                "metadata": {"rfc_emisor": "GLU760309457"},
            }
        ],
        "gas_lp_vehiculos": [],
        "gas_lp_rutas": [],
        "gas_lp_ubicaciones_carta_porte": [],
    }
    user = {"id": 1, "owner_user_id": "admin-actual", "tenant_id": "tenant-gas", "perfil_id": 101}

    monkeypatch.setattr(users_auth, "_gas_lp_internal_context", lambda _token: {"user": user})
    monkeypatch.setattr(users_auth, "_gas_lp_profile", lambda _user: {"id": 101, "tenant_id": "tenant-gas", "rfc": "GLU760309457", "nombre": "GAS LUX"})
    monkeypatch.setattr(users_auth, "get_supabase_admin", lambda: FakeSupabase(tables))
    monkeypatch.setattr(users_auth, "_internal_cp_facilities", lambda _user: [])

    response = asyncio.run(users_auth.gas_lp_internal_catalogos("tok"))
    payload = json.loads(response.body)

    assert [row["nombre"] for row in payload["choferes"]] == ["MARIA JOSE"]
    assert payload["choferes"][0]["user_id"] == "otro-creador"
    assert [row["alias"] for row in payload["mercancias"]] == ["Gas LP"]


def test_carta_porte_catalog_rfc_matching_normalizes_spaces_and_case(monkeypatch):
    monkeypatch.setattr(cp_catalogos, "_gas_lp_profile", lambda user: {"id": user["perfil_id"], "tenant_id": "tenant-gas", "rfc": "GLU760309457"})
    user = {"id": 1, "owner_user_id": "admin-a", "tenant_id": "tenant-gas", "perfil_id": 101}
    scope = cp_catalogos._internal_cp_company_scope(user)

    assert cp_catalogos._internal_cp_row_company_match({"metadata": {"empresa_rfc": " glu760309457 "}}, scope) is True
    assert cp_catalogos._internal_cp_row_company_match({"metadata": {"rfc_emisor": "RFC GLU760309457"}}, scope) is True
    assert cp_catalogos._internal_cp_row_company_match({"metadata": {"empresa_rfc": "OTR010101AAA"}}, scope) is False
