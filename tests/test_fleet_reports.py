from datetime import date

from openpyxl import load_workbook

from services.fleet_reports import build_fleet_report, fleet_analytics, parse_maintenance_csv


def test_parse_motive_maintenance_preserves_manager_and_mxn():
    content = (
        "Date,Entity,Entity Type,Service Type,Service Name,Notes,Cost,Fleet Manager,Odometer,Engine Hours\n"
        "20/07/2026,87 ALFA AGS,vehicle,,,KIT CLUTCH,2900,SAUL PEREZ (AGS),200508,700\n"
    ).encode()
    rows = parse_maintenance_csv(content)
    assert len(rows) == 1
    assert rows[0]["amount_mxn"] == 2900
    assert rows[0]["submitted_by"] == "SAUL PEREZ (AGS)"
    assert rows[0]["vehicle_number"] == "87 ALFA AGS"


def test_empty_report_has_executive_sheets_without_misleading_empty_tabs(tmp_path):
    payload = build_fleet_report({"expenses": [], "driver_events": [], "speeding": [], "activity": [], "faults": []}, date(2026, 7, 1), date(2026, 7, 20))
    target = tmp_path / "report.xlsx"; target.write_bytes(payload)
    workbook = load_workbook(target, read_only=True)
    assert workbook.sheetnames == ["Dashboard", "Resumen por chofer", "Unidades y diagnóstico"]


def test_report_adds_visual_manager_dashboard_and_driver_summary(tmp_path):
    payload = build_fleet_report({
        "vehicles": [
            {"vehicle_number": "U-ROJA", "current_driver_name": "Ana"},
            {"vehicle_number": "U-SIN-GPS", "current_driver_name": "Luis"},
        ],
        "driver_events": [
            {"vehicle_number": "U-ROJA", "driver_name": "Ana", "primary_behavior": "hard_brake", "severity": "high"}
            for _ in range(3)
        ],
        "speeding": [],
        "faults": [
            {"vehicle_number": "U-ROJA", "code": "P0201", "occurrence_count": 4},
            {"vehicle_number": "U-ROJA", "code": "P0201", "occurrence_count": 2},
            {"vehicle_number": "U-ROJA", "code": "P0351", "occurrence_count": 1},
        ],
        "inspections": [{"vehicle_number": "U-ROJA", "driver_name": "Ana"}],
        "expenses": [], "activity": [],
    }, date(2026, 7, 1), date(2026, 7, 7), "Zacatecas")
    target = tmp_path / "visual-report.xlsx"
    target.write_bytes(payload)
    workbook = load_workbook(target, data_only=True)

    dashboard = workbook["Dashboard"]
    dashboard_values = [cell.value for row in dashboard.iter_rows() for cell in row]
    assert "INFORME EJECUTIVO · FLOTILLA 360" in dashboard_values
    assert "Choferes que requieren capacitación" in dashboard_values
    assert "Unidades sin datos GPS" in dashboard_values
    assert "UNIDADES SIN DATOS GPS" in dashboard_values
    assert "Choferes que realizaron inspecciones" in dashboard_values
    assert "DECISIONES RECOMENDADAS PARA EL GERENTE" in dashboard_values
    assert "Ana" in dashboard_values
    assert any("Capacitación en distancia y frenado preventivo" in str(value) for value in dashboard_values)
    assert dashboard["O2"].value == "Frenado brusco"
    assert dashboard["P2"].value == 3
    assert len(dashboard._charts) == 1
    assert dashboard._charts[0].__class__.__name__ == "DoughnutChart"

    driver_summary = workbook["Resumen por chofer"]
    assert [cell.value for cell in driver_summary[2]] == [
        "Chofer", "Eventos de seguridad", "Excesos de velocidad", "Eventos totales",
        "Conducta principal", "Críticos / altos", "Inspecciones", "Capacitación recomendada", "Prioridad",
    ]
    assert driver_summary["A3"].value == "Ana"
    assert driver_summary["E3"].value == "Frenado brusco"


def test_report_adds_executive_chronology_when_events_exist(tmp_path):
    payload = build_fleet_report({
        "driver_events": [{
            "started_at": "2026-07-10T12:00:00Z", "vehicle_number": "U-1",
            "driver_name": "Ana", "primary_behavior": "hard_brake", "severity": "high",
        }],
        "speeding": [], "faults": [], "expenses": [], "activity": [],
    }, date(2026, 7, 1), date(2026, 7, 20))
    target = tmp_path / "report.xlsx"; target.write_bytes(payload)
    workbook = load_workbook(target, read_only=True)
    assert "Eventos de chofer" in workbook.sheetnames
    sheet = workbook["Eventos de chofer"]
    assert sheet["E2"].value == "Frenado brusco"


def test_analytics_uses_consumed_fuel_and_exact_utilization_rollup():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "U-1"}],
        "metrics": [{"vehicle_number": "U-1", "metric_date": "2026-07-01", "distance_km": 100}],
        "fuel": [{"vehicle_number": "U-1", "quantity_liters": 30, "total_cost": 900}],
        "utilization": [{
            "vehicle_number": "U-1", "engine_hours": 8, "fuel_consumed_liters": 20,
            "utilization_pct": 0.8,
        }],
        "_sync": {"datasets": {"card_expenses": {"status": "unavailable"}}},
    })
    unit = analytics["units"][0]
    assert unit["purchased_liters"] == 30
    assert unit["liters"] == 20
    assert unit["engine_hours"] == 8
    assert unit["utilization_pct"] == 0.8
    assert unit["km_per_liter"] == 5
    assert analytics["totals"]["expense_complete"] is False


def test_analytics_prioritizes_exact_mileage_rollup_without_double_counting():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "U-1"}],
        "mileage": [{"vehicle_number": "U-1", "distance_km": 250}],
        "metrics": [{"vehicle_number": "U-1", "metric_date": "2026-07-01", "distance_km": 100}],
        "activity": [{"vehicle_number": "U-1", "started_at": "2026-07-01", "distance_km": 75}],
    })
    assert analytics["units"][0]["distance_km"] == 250
    assert analytics["totals"]["distance_km"] == 250


def test_analytics_ranks_by_event_count_and_recovers_driver_from_activity():
    analytics = fleet_analytics({
        "vehicles": [
            {"vehicle_number": "U-1", "current_driver_name": ""},
            {"vehicle_number": "U-2", "current_driver_name": ""},
            {"vehicle_number": "SIN-GPS", "current_driver_name": ""},
        ],
        "driver_events": [
            {"vehicle_number": "U-1", "severity": "low", "primary_behavior": "hard_brake"},
            {"vehicle_number": "U-1", "severity": "low", "primary_behavior": "hard_brake"},
            {"vehicle_number": "U-2", "severity": "critical", "primary_behavior": "cell_phone"},
        ],
        "speeding": [{"vehicle_number": "U-1", "severity": "medium"}],
        "activity": [{"vehicle_number": "U-1", "driver_name": "CONDUCTOR REAL", "started_at": "2026-07-20"}],
        "_period_days": 7,
    })

    assert [row["vehicle_number"] for row in analytics["units"]] == ["U-1", "U-2", "SIN-GPS"]
    assert analytics["units"][0]["driver_name"] == "CONDUCTOR REAL"
    assert analytics["behaviors"][0] == {"label": "Frenado brusco", "count": 2}
    assert {"label": "Exceso de velocidad", "count": 1} in analytics["behaviors"]
    assert analytics["totals"]["vehicles_with_data"] == 2
    assert analytics["totals"]["vehicles_without_gps"] == 1
    assert analytics["units"][-1]["coverage_status"] == "Sin datos GPS / revisión manual"
    assert [row["vehicle_number"] for row in analytics["attention_units"]] == ["U-1", "U-2"]
    assert [row["vehicle_number"] for row in analytics["units_without_gps"]] == ["SIN-GPS"]


def test_analytics_attributes_inspections_to_reported_driver_with_unit_fallback():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "U-1", "current_driver_name": "Chofer asignado"}],
        "inspections": [
            {"vehicle_number": "U-1", "driver_name": "Quien inspeccionó"},
            {"vehicle_number": "U-1", "driver_name": ""},
        ],
    })

    assert analytics["inspection_credits"] == [
        {"vehicle_number": "U-1", "driver_name": "Chofer asignado", "inspections": 1},
        {"vehicle_number": "U-1", "driver_name": "Quien inspeccionó", "inspections": 1},
    ]


def test_inspection_driver_without_events_is_visible_and_explains_missing_assignment():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "UTILITARIA MIGUEL", "current_driver_name": ""}],
        "inspections": [{"vehicle_number": "UTILITARIA MIGUEL", "driver_name": "MIGUEL ANGEL"}],
        "driver_events": [], "speeding": [],
    })

    assert analytics["units_without_gps"][0]["driver_name"] == "MIGUEL ANGEL"
    assert analytics["drivers_without_events"] == [{
        "driver_name": "MIGUEL ANGEL",
        "vehicle_number": "UTILITARIA MIGUEL",
        "inspections": 1,
    }]


def test_analytics_attributes_training_to_event_driver_not_only_assigned_unit_driver():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "U-1", "current_driver_name": "Chofer asignado"}],
        "driver_events": [{
            "vehicle_number": "U-1", "driver_name": "Chofer real",
            "primary_behavior": "hard_brake", "severity": "high",
        }],
        "speeding": [],
        "inspections": [{"vehicle_number": "U-1", "driver_name": "Chofer real"}],
    })

    driver = analytics["training_drivers"][0]
    assert driver["driver_name"] == "Chofer real"
    assert driver["top_behavior"] == "Frenado brusco"
    assert driver["training"] == "Capacitación en distancia y frenado preventivo"
    assert driver["inspections"] == 1


def test_excel_keeps_units_out_of_driver_dashboard_and_in_unit_sheet(tmp_path):
    vehicles = [{"vehicle_number": f"U-{index:02d}"} for index in range(1, 14)]
    events = [
        {"vehicle_number": row["vehicle_number"], "primary_behavior": "hard_brake", "severity": "low"}
        for row in vehicles
    ]
    payload = build_fleet_report(
        {"vehicles": vehicles, "driver_events": events, "speeding": [], "faults": [], "activity": []},
        date(2026, 7, 20),
        date(2026, 7, 25),
    )
    target = tmp_path / "report.xlsx"
    target.write_bytes(payload)
    workbook = load_workbook(target, read_only=True, data_only=True)

    dashboard = workbook["Dashboard"]
    dashboard_values = [cell.value for row in dashboard.iter_rows() for cell in row]
    summary_headers = [cell.value for cell in workbook["Unidades y diagnóstico"][2]]
    assert not any("U-13" in str(value) for value in dashboard_values)
    unit_values = [cell.value for row in workbook["Unidades y diagnóstico"].iter_rows() for cell in row]
    assert "U-13" in unit_values
    assert "Índice" not in dashboard_values
    assert "SCORE CONDUCTORES" not in dashboard_values
    assert "Score" not in summary_headers
    assert "Índice de atención" not in summary_headers
    assert "Actividad" not in workbook.sheetnames
    assert "Scorecard conductores" not in workbook.sheetnames


def test_foreign_currency_fuel_is_not_reported_as_mxn_expense():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "AT 35"}],
        "fuel": [{
            "vehicle_number": "AT 35",
            "quantity_liters": 187.95,
            "total_cost": 23.99,
            "currency": "CAD",
        }],
        "expenses": [],
        "_sync": {"datasets": {"card_expenses": {"status": "unavailable"}}},
    })

    assert analytics["totals"]["expenses_mxn"] == 0
    assert analytics["totals"]["expense_available"] is False


def test_closed_faults_and_resolved_defects_are_excluded_from_excel_alerts(tmp_path):
    payload = build_fleet_report({
        "vehicles": [{"vehicle_number": "U-1"}],
        "faults": [
            {"vehicle_number": "U-1", "code": "P-OPEN", "occurrence_count": 2, "status": "open"},
            {"vehicle_number": "U-1", "code": "P-CLOSED", "occurrence_count": 99, "status": "closed"},
        ],
        "defects": [
            {"vehicle_number": "U-1", "title": "Freno abierto", "severity": "major", "status": "open"},
            {"vehicle_number": "U-1", "title": "Freno reparado", "severity": "critical", "status": "resolved", "resolved_at": "2026-08-01"},
        ],
        "driver_events": [], "speeding": [], "expenses": [], "activity": [],
    }, date(2026, 7, 20), date(2026, 8, 4))
    target = tmp_path / "open-only.xlsx"
    target.write_bytes(payload)
    workbook = load_workbook(target, read_only=True, data_only=True)

    assert [row[2] for row in workbook["Diagnóstico PID"].iter_rows(min_row=2, values_only=True)] == ["P-OPEN"]
    assert [row[2] for row in workbook["Defectos accionables"].iter_rows(min_row=2, values_only=True)] == ["Freno abierto"]
    unit_row = next(workbook["Unidades y diagnóstico"].iter_rows(min_row=3, values_only=True))
    assert unit_row[7] == 2
    assert unit_row[17] == 1
