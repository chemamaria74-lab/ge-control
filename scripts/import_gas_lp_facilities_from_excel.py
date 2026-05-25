#!/usr/bin/env python3
"""Import Gas LP facilities from the operational Excel matrix.

The importer is intentionally idempotent and conservative:
- dry-run is the default;
- it requires explicit profile mapping when more than one company/profile exists;
- it updates existing rows by scope + permit/installation key instead of blindly inserting.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from supabase import create_client


FIELD_MAP = {
    "Tipo de Permiso CRE": "tipo_permiso",
    "Actividad SAT": "actividad_sat",
    "Clave Instalación": "clave_instalacion",
    "Núm. Permiso CRE (distribución)": "num_permiso",
    "Descripción Instalación": "descripcion",
    "Temperatura Default (°C)": "temperatura_default",
    "Núm. Tanques": "num_tanques",
    "Núm. Dispensarios": "num_dispensarios",
    "Clave del Tanque": "clave_tanque",
    "Capacidad Total (L)": "cap_total_tanque",
    "Capacidad Operativa (L)": "cap_operativa_tanque",
    "Capacidad Útil (L)": "cap_util_tanque",
    "Fecha Última Calibración": "fecha_calibracion_tanque",
    "Incertidumbre (0-1)": "incertidumbre_medidor",
    "Modelo del Medidor": "modelo_medidor",
    "Número de Serie": "serie_medidor",
    "Vigencia Calibración Medidor": "fecha_calibracion_medidor",
    "Latitud (decimal)": "latitud",
    "Longitud (decimal)": "longitud",
}

NUMERIC_FIELDS = {
    "temperatura_default",
    "cap_total_tanque",
    "cap_operativa_tanque",
    "cap_util_tanque",
    "incertidumbre_medidor",
    "latitud",
    "longitud",
}
INTEGER_FIELDS = {"num_tanques", "num_dispensarios"}


@dataclass
class ImportStats:
    parsed: int = 0
    inserts: int = 0
    updates: int = 0
    skipped: int = 0
    errors: int = 0


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _number(value: Any) -> float | None:
    text = _clean(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _integer(value: Any) -> int | None:
    num = _number(value)
    return int(num) if num is not None else None


def _infer_tipo_permiso(row: dict[str, Any]) -> str:
    explicit = _clean(row.get("tipo_permiso"))
    if explicit:
        return explicit
    permiso = _clean(row.get("num_permiso")).upper()
    if "/EXP/ES/" in permiso:
        return "PER43"
    if "/EXP/AUT/" in permiso:
        return "PER44"
    if "/DIST/PLA/" in permiso:
        return "PER40"
    return "PER40"


def _infer_tipo_instalacion(row: dict[str, Any]) -> str:
    actividad = _clean(row.get("actividad_sat")).upper()
    permiso = _clean(row.get("num_permiso")).upper()
    nombre = _clean(row.get("nombre")).lower()
    if actividad == "DIS" or "/DIST/PLA/" in permiso or nombre.startswith("planta"):
        return "planta"
    return "estacion"


def _validate_coords(row: dict[str, Any]) -> list[str]:
    warnings: list[str] = []
    lat = row.get("latitud")
    lon = row.get("longitud")
    if lat is None and lon is None:
        return warnings
    if lat is None or lon is None:
        warnings.append("coordenadas incompletas")
        return warnings
    if not (14 <= float(lat) <= 33 and -119 <= float(lon) <= -86):
        warnings.append(f"coordenadas fuera de México: {lat}, {lon}")
    return warnings


def parse_excel(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows[0]) < 3:
        raise ValueError("El Excel no tiene formato de matriz de instalaciones.")
    headers = [_clean(c) for c in rows[0]]
    facilities: list[dict[str, Any]] = []
    for col in range(2, len(headers)):
        name = headers[col]
        if not name:
            continue
        raw: dict[str, Any] = {"nombre": name}
        for row in rows[1:]:
            if len(row) < 2:
                continue
            label = _clean(row[1])
            if not label:
                continue
            key = FIELD_MAP.get(label)
            if not key:
                continue
            value = row[col] if col < len(row) else None
            if value is None or _clean(value) == "":
                continue
            if key in NUMERIC_FIELDS:
                raw[key] = _number(value)
            elif key in INTEGER_FIELDS:
                raw[key] = _integer(value)
            else:
                raw[key] = _clean(value)
        raw["tipo_permiso"] = _infer_tipo_permiso(raw)
        raw["modalidad_permiso"] = raw["tipo_permiso"]
        raw["tipo_instalacion"] = _infer_tipo_instalacion(raw)
        raw["caracter"] = "permisionario"
        raw["modulo_propietario"] = "gas_lp"
        if not raw.get("capacidad_tanque") and raw.get("cap_total_tanque") is not None:
            raw["capacidad_tanque"] = raw["cap_total_tanque"]
        raw["_warnings"] = _validate_coords(raw)
        facilities.append(raw)
    return facilities


def load_mapping(path: str | None) -> dict[str, Any]:
    if not path:
        return {}
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def resolve_profile(row: dict[str, Any], mapping: dict[str, Any], profiles: list[dict[str, Any]]) -> int | None:
    name = _clean(row.get("nombre"))
    permiso = _clean(row.get("num_permiso"))
    direct = mapping.get("facility_to_perfil", {})
    if name in direct:
        return int(direct[name])
    permit_map = mapping.get("permit_to_perfil", {})
    if permiso and permiso in permit_map:
        return int(permit_map[permiso])
    for item in mapping.get("regex_to_perfil", []):
        pattern = item.get("pattern", "")
        if pattern and re.search(pattern, name, flags=re.IGNORECASE):
            return int(item["perfil_id"])
    if mapping.get("default_perfil_id"):
        return int(mapping["default_perfil_id"])
    if len(profiles) == 1:
        return int(profiles[0]["id"])
    return None


def row_hash(row: dict[str, Any]) -> str:
    public_row = {k: v for k, v in row.items() if not k.startswith("_")}
    payload = json.dumps(public_row, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def get_profiles(sb, owner_user_id: str) -> list[dict[str, Any]]:
    return (
        sb.table("perfiles_empresa")
        .select("id,nombre,rfc,tenant_id,activo")
        .eq("user_id", owner_user_id)
        .eq("activo", True)
        .execute()
        .data
        or []
    )


def find_existing(sb, owner_user_id: str, perfil_id: int, row: dict[str, Any]) -> dict[str, Any] | None:
    permiso = _clean(row.get("num_permiso"))
    clave = _clean(row.get("clave_instalacion"))
    name = _clean(row.get("nombre"))
    base = sb.table("user_facilities").select("id,import_source_hash").eq("user_id", owner_user_id).eq("perfil_id", perfil_id)
    if permiso:
        rows = base.eq("num_permiso", permiso).limit(1).execute().data or []
        if rows:
            return rows[0]
    if clave:
        rows = (
            sb.table("user_facilities")
            .select("id,import_source_hash")
            .eq("user_id", owner_user_id)
            .eq("perfil_id", perfil_id)
            .eq("clave_instalacion", clave)
            .limit(1)
            .execute()
            .data
            or []
        )
        if rows:
            return rows[0]
    rows = (
        sb.table("user_facilities")
        .select("id,import_source_hash")
        .eq("user_id", owner_user_id)
        .eq("perfil_id", perfil_id)
        .eq("nombre", name)
        .limit(1)
        .execute()
        .data
        or []
    )
    return rows[0] if rows else None


def build_record(row: dict[str, Any], owner_user_id: str, perfil: dict[str, Any], source_file: str, batch_id: str) -> dict[str, Any]:
    clean_row = {k: v for k, v in row.items() if not k.startswith("_")}
    clean_row.update(
        {
            "user_id": owner_user_id,
            "perfil_id": int(perfil["id"]),
            "import_source": "excel_gas_lp_facilities",
            "import_source_file": source_file,
            "import_source_hash": row_hash(row),
            "import_batch_id": batch_id,
            "import_payload": row,
        }
    )
    return clean_row


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Gas LP facilities from Excel into Supabase staging/production.")
    parser.add_argument("excel_path")
    parser.add_argument("--owner-user-id", required=True, help="Supabase auth user id that owns the target Gas LP profiles.")
    parser.add_argument("--profile-map-json", help="JSON mapping facilities/permits/regex to perfil_id.")
    parser.add_argument("--sheet")
    parser.add_argument("--parse-only", action="store_true", help="Only parse and preview the Excel. Does not require Supabase env vars.")
    parser.add_argument("--apply", action="store_true", help="Write changes. Without this flag the script is dry-run only.")
    parser.add_argument("--limit", type=int)
    args = parser.parse_args()

    excel_path = Path(args.excel_path).expanduser().resolve()
    facilities = parse_excel(excel_path, args.sheet)
    if args.limit:
        facilities = facilities[: args.limit]
    if args.parse_only:
        print(json.dumps({
            "apply": False,
            "parse_only": True,
            "stats": {"parsed": len(facilities)},
            "preview": facilities,
        }, ensure_ascii=False, indent=2))
        return 0

    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_SERVICE_KEY")
    if not url or not key:
        print("Faltan SUPABASE_URL y/o SUPABASE_SERVICE_ROLE_KEY.", file=sys.stderr)
        return 2

    mapping = load_mapping(args.profile_map_json)
    sb = create_client(url, key)
    profiles = get_profiles(sb, args.owner_user_id)
    profile_by_id = {int(p["id"]): p for p in profiles}
    batch_id = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    stats = ImportStats(parsed=len(facilities))
    preview: list[dict[str, Any]] = []

    for row in facilities:
        perfil_id = resolve_profile(row, mapping, profiles)
        if not perfil_id or perfil_id not in profile_by_id:
            stats.errors += 1
            preview.append({"nombre": row.get("nombre"), "status": "ERROR", "reason": "sin perfil_id mapeado"})
            continue
        record = build_record(row, args.owner_user_id, profile_by_id[perfil_id], excel_path.name, batch_id)
        existing = find_existing(sb, args.owner_user_id, perfil_id, row)
        action = "update" if existing else "insert"
        warnings = row.get("_warnings") or []
        preview.append(
            {
                "nombre": row.get("nombre"),
                "perfil_id": perfil_id,
                "empresa": profile_by_id[perfil_id].get("nombre"),
                "permiso": row.get("num_permiso"),
                "tipo": row.get("tipo_instalacion"),
                "action": action,
                "warnings": warnings,
            }
        )
        if not args.apply:
            continue
        try:
            if existing:
                sb.table("user_facilities").update(record).eq("id", existing["id"]).execute()
                stats.updates += 1
            else:
                sb.table("user_facilities").insert(record).execute()
                stats.inserts += 1
        except Exception as exc:
            stats.errors += 1
            preview[-1]["status"] = "ERROR"
            preview[-1]["reason"] = str(exc)

    if not args.apply:
        stats.skipped = stats.parsed - stats.errors
    print(json.dumps({"apply": args.apply, "stats": stats.__dict__, "preview": preview}, ensure_ascii=False, indent=2))
    return 1 if stats.errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
