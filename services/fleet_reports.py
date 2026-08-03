from __future__ import annotations

import csv
import hashlib
import io
import re
from collections import Counter, defaultdict
from datetime import date, datetime, time, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import xlsxwriter


def _text(value: Any) -> str:
    return str(value or "").strip()


def _amount(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(Decimal(re.sub(r"[^0-9.\-]", "", str(value))))
    except (InvalidOperation, ValueError):
        return 0.0


def _date(value: Any) -> str | None:
    if isinstance(value, datetime):
        result = value
    elif isinstance(value, date):
        result = datetime.combine(value, time.min)
    else:
        text = _text(value)
        result = None
        for fmt in ("%d/%m/%Y", "%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                result = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if result is None:
            return None
    return result.replace(tzinfo=timezone.utc).isoformat()


def _key(*values: Any) -> str:
    return hashlib.sha256("|".join(_text(value) for value in values).encode()).hexdigest()[:40]


def _maintenance_row(values: list[Any], source: str, row_number: int) -> dict[str, Any] | None:
    values += [None] * (10 - len(values))
    occurred_at = _date(values[0])
    vehicle_number = _text(values[1])
    if not occurred_at or not vehicle_number or vehicle_number.lower().startswith("total"):
        return None
    return {
        "source": source, "source_key": _key(source, row_number, *values[:10]), "occurred_at": occurred_at,
        "vehicle_number": vehicle_number, "expense_type": "mantenimiento", "category": _text(values[3]),
        "description": " · ".join(filter(None, [_text(values[4]), _text(values[5])])), "amount_mxn": _amount(values[6]),
        "submitted_by": _text(values[7]), "odometer_km": _amount(values[8]) or None,
        "engine_hours": _amount(values[9]) or None, "raw_metadata": {"entity_type": _text(values[2])},
    }


def parse_maintenance_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    return [parsed for i, row in enumerate(rows[1:], 2) if (parsed := _maintenance_row(list(row), "motive_maintenance_export", i))]


def parse_expense_workbook(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    output: list[dict[str, Any]] = []
    if "APP Gastos" in workbook.sheetnames:
        sheet = workbook["APP Gastos"]
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            parsed = _maintenance_row(list(row), "motive_maintenance_export", index)
            if parsed:
                output.append(parsed)

    for sheet_name in ("Crdes Gasoln", "Crdes Disel"):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        group_name, zone_name = "", ""
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            values = list(row) + [None] * 11
            route = _text(values[0])
            if route and not values[1] and not values[4]:
                if not group_name:
                    group_name = route
                else:
                    zone_name = route
                continue
            occurred_at = _date(values[1])
            liters = _amount(values[4])
            if not occurred_at or not route or route.lower().startswith("total") or liters <= 0:
                continue
            fuel_type = _text(values[3]) or ("Gasolina" if "Gasoln" in sheet_name else "Diésel")
            output.append({
                "source": "credes_import", "source_key": _key(sheet_name, index, *values[:11]), "occurred_at": occurred_at,
                "vehicle_number": route, "group_name": group_name, "zone_name": zone_name, "expense_type": "autoconsumo",
                "category": "Combustible", "description": f"Autoconsumo {fuel_type}", "fuel_type": fuel_type,
                "quantity_liters": liters, "unit_cost": _amount(values[5]) or None, "amount_mxn": _amount(values[6]),
                "odometer_km": _amount(values[8]) or None,
                "raw_metadata": {"km_inicial": _amount(values[7]) or None, "rendimiento": _amount(values[9]) or None},
            })

    if "Cdres Gasto" in workbook.sheetnames:
        sheet = workbook["Cdres Gasto"]
        for index, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
            values = list(row) + [None] * 7
            zone, route, category = _text(values[0]), _text(values[1]), _text(values[2])
            amount = _amount(values[5])
            if not zone or not route or route.lower().startswith("total") or amount <= 0:
                continue
            occurred = datetime(int(_amount(values[3]) or date.today().year), _month(values[4]), 1, tzinfo=timezone.utc).isoformat()
            output.append({
                "source": "credes_import", "source_key": _key("Cdres Gasto", index, *values[:7]), "occurred_at": occurred,
                "vehicle_number": route, "zone_name": zone, "expense_type": "gasto", "category": category,
                "description": _text(values[6]), "amount_mxn": amount, "submitted_by": "CREDES",
                "raw_metadata": {"year": values[3], "month": _text(values[4])},
            })
    return output


def _month(value: Any) -> int:
    names = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
             "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12}
    return names.get(_text(value).lower(), 1)


BEHAVIOR_LABELS = {
    "seat_belt_violation": "Cinturón de seguridad",
    "cell_phone": "Uso de celular",
    "tailgating": "Seguimiento cercano",
    "close_following": "Seguimiento cercano",
    "stop_sign_violation": "Señal de alto",
    "distraction": "Distracción",
    "near_miss": "Casi colisión",
    "hard_brake": "Frenado brusco",
    "drowsiness": "Somnolencia",
    "hard_corner": "Giro brusco",
    "unsafe_lane_change": "Cambio de carril peligroso",
    "driver_facing_cam_obstruction": "Cámara del conductor obstruida",
    "road_facing_cam_obstruction": "Cámara frontal obstruida",
    "smoking": "Fumar",
    "crash": "Colisión",
    "speeding": "Exceso de velocidad",
}


def behavior_label(value: Any) -> str:
    raw = _text(value)
    key = re.sub(r"[\s\-]+", "_", raw.casefold())
    return BEHAVIOR_LABELS.get(key, raw.replace("_", " ").strip().capitalize() or "Sin clasificar")


def fleet_analytics(data: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    unit_rows: dict[str, dict[str, Any]] = {}
    behaviors: Counter[str] = Counter()
    severity: Counter[str] = Counter()
    daily: Counter[str] = Counter()
    datasets = (data.get("_sync") or {}).get("datasets") or {}
    card_status = datasets.get("card_expenses")
    expense_complete = not (
        isinstance(card_status, dict) and card_status.get("status") == "unavailable"
    )
    has_mxn_fuel = any(
        _text(row.get("currency") or "MXN").upper() == "MXN"
        and float(row.get("total_cost") or 0) != 0
        for row in data.get("fuel", [])
    )
    expense_available = bool(data.get("expenses")) or has_mxn_fuel or expense_complete

    def unit(value: Any) -> dict[str, Any]:
        name = _text(value) or "Sin unidad vinculada"
        return unit_rows.setdefault(name, {
            "vehicle_number": name, "driver_name": "", "security": 0, "speeding": 0,
            "critical_high": 0, "faults": 0, "expense_mxn": 0.0, "liters": 0.0,
            "purchased_liters": 0.0, "fuel_consumption_available": False,
            "maintenance_mxn": 0.0, "distance_km": 0.0, "engine_hours": 0.0,
            "active_days": set(), "inspections": 0, "open_defects": 0,
            "overdue_defects": 0, "score": None, "attention_index": 0,
            "utilization_pct": None, "engine_hours_available": False,
            "telemetry_available": False, "coverage_status": "Sin datos GPS / revisión manual",
        })

    for row in data.get("vehicles", []):
        item = unit(row.get("vehicle_number"))
        item["driver_name"] = _text(row.get("current_driver_name"))

    for row in data.get("driver_events", []):
        item = unit(row.get("vehicle_number"))
        item["security"] += 1
        item["driver_name"] = item["driver_name"] or _text(row.get("driver_name"))
        level = _text(row.get("severity")).casefold()
        severity[level or "sin clasificar"] += 1
        weight = {"critical": 5, "severe": 4, "high": 3, "medium": 2, "low": 1}.get(level, 1)
        item["attention_index"] += weight
        if level in {"critical", "severe", "high"}:
            item["critical_high"] += 1
        label = behavior_label(row.get("primary_behavior") or row.get("event_type"))
        behaviors[label] += 1
        day = _text(row.get("started_at"))[:10]
        if day:
            daily[day] += 1
    for row in data.get("speeding", []):
        item = unit(row.get("vehicle_number"))
        item["speeding"] += 1
        item["driver_name"] = item["driver_name"] or _text(row.get("driver_name"))
        level = _text(row.get("severity")).casefold()
        weight = {"critical": 5, "severe": 4, "high": 3, "medium": 2, "low": 1}.get(level, 2)
        item["attention_index"] += weight
        if level in {"critical", "severe", "high"}:
            item["critical_high"] += 1
        behaviors["Exceso de velocidad"] += 1
        day = _text(row.get("started_at"))[:10]
        if day:
            daily[day] += 1
    for row in data.get("faults", []):
        unit(row.get("vehicle_number"))["faults"] += int(row.get("occurrence_count") or 1)
    for row in data.get("expenses", []):
        item = unit(row.get("vehicle_number"))
        item["expense_mxn"] += float(row.get("amount_mxn") or 0)
        item["purchased_liters"] += float(row.get("quantity_liters") or 0)
        if _text(row.get("expense_type")).casefold() == "mantenimiento":
            item["maintenance_mxn"] += float(row.get("amount_mxn") or 0)
    has_card_expenses = any(_text(row.get("source")).casefold() == "motive_card" for row in data.get("expenses", []))
    for row in data.get("fuel", []):
        item = unit(row.get("vehicle_number"))
        item["purchased_liters"] += float(row.get("quantity_liters") or 0)
        if not has_card_expenses and _text(row.get("currency") or "MXN").upper() == "MXN":
            item["expense_mxn"] += float(row.get("total_cost") or 0)
    # El rollup IFTA es la fuente oficial del periodo. Actividad/métricas solo son
    # respaldo para datos históricos creados antes de habilitar este endpoint.
    has_mileage_rollup = bool(data.get("mileage"))
    for row in data.get("mileage", []):
        unit(row.get("vehicle_number"))["distance_km"] += float(row.get("distance_km") or 0)
    use_activity_distance = not has_mileage_rollup and not data.get("metrics")
    for row in data.get("activity", []):
        item = unit(row.get("vehicle_number"))
        item["driver_name"] = item["driver_name"] or _text(row.get("driver_name"))
        if use_activity_distance:
            item["distance_km"] += float(row.get("distance_km") or 0)
        day = _text(row.get("started_at"))[:10]
        if day:
            item["active_days"].add(day)
    for row in data.get("metrics", []):
        item = unit(row.get("vehicle_number"))
        if not has_mileage_rollup:
            item["distance_km"] += float(row.get("distance_km") or 0)
        if not has_mileage_rollup and float(row.get("distance_km") or 0) > 0:
            item["active_days"].add(_text(row.get("metric_date")))
    for row in data.get("utilization", []):
        item = unit(row.get("vehicle_number"))
        item["engine_hours"] += float(row.get("engine_hours") or 0)
        item["engine_hours_available"] = True
        item["liters"] += float(row.get("fuel_consumed_liters") or 0)
        item["fuel_consumption_available"] = True
        if row.get("utilization_pct") is not None:
            item["utilization_pct"] = float(row["utilization_pct"])
    inspection_credits: dict[tuple[str, str], dict[str, Any]] = {}
    for row in data.get("inspections", []):
        item = unit(row.get("vehicle_number"))
        item["inspections"] += 1
        driver_name = _text(row.get("driver_name")) or item["driver_name"] or "Sin conductor identificado"
        key = (item["vehicle_number"], driver_name)
        credit = inspection_credits.setdefault(key, {
            "vehicle_number": item["vehicle_number"], "driver_name": driver_name, "inspections": 0,
        })
        credit["inspections"] += 1
    for row in data.get("defects", []):
        item = unit(row.get("vehicle_number"))
        if _text(row.get("status")).casefold() in {"open", "pending", "unresolved", "with_defects"}:
            item["open_defects"] += 1
            if bool(row.get("is_overdue")):
                item["overdue_defects"] += 1
    for row in data.get("scorecards", []):
        item = unit(row.get("vehicle_number") or row.get("driver_name"))
        item["driver_name"] = item["driver_name"] or _text(row.get("driver_name"))
        if row.get("performance_score") is not None:
            item["score"] = float(row["performance_score"])

    period_days = max(int(data.get("_period_days") or 1), 1)
    for item in unit_rows.values():
        item["active_days"] = len(item["active_days"])
        if item["utilization_pct"] is None:
            item["utilization_pct"] = min(item["active_days"] / period_days, 1.0)
        item["telemetry_available"] = bool(
            item["security"] or item["speeding"] or item["distance_km"] > 0
            or item["engine_hours_available"] or item["active_days"]
        )
        item["coverage_status"] = "Con datos GPS" if item["telemetry_available"] else "Sin datos GPS / revisión manual"
        item["km_per_liter"] = (
            item["distance_km"] / item["liters"]
            if item["fuel_consumption_available"] and item["liters"] > 0 else None
        )
        item["cost_per_km"] = (
            item["expense_mxn"] / item["distance_km"]
            if expense_available and item["distance_km"] > 0 else None
        )
        item["events_per_1000_km"] = (
            (item["security"] + item["speeding"]) * 1000 / item["distance_km"]
            if item["distance_km"] > 0 else None
        )
    units = sorted(
        unit_rows.values(),
        key=lambda row: (-(row["security"] + row["speeding"]), row["vehicle_number"]),
    )
    totals = {
        "expenses_mxn": sum(row["expense_mxn"] for row in units),
        "maintenance_mxn": sum(row["maintenance_mxn"] for row in units),
        "liters": sum(row["liters"] for row in units),
        "purchased_liters": sum(row["purchased_liters"] for row in units),
        "distance_km": sum(row["distance_km"] for row in units),
        "engine_hours": sum(row["engine_hours"] for row in units),
        "inspections": sum(row["inspections"] for row in units),
        "open_defects": sum(row["open_defects"] for row in units),
        "overdue_defects": sum(row["overdue_defects"] for row in units),
        "vehicles_with_data": sum(1 for row in units if row["telemetry_available"]),
        "vehicles_without_gps": sum(1 for row in units if not row["telemetry_available"]),
    }
    totals["distance_available"] = any(row["distance_km"] > 0 for row in units)
    totals["engine_hours_available"] = any(row["engine_hours_available"] for row in units)
    totals["fuel_consumption_available"] = any(row["fuel_consumption_available"] for row in units)
    totals["km_per_liter"] = (
        totals["distance_km"] / totals["liters"]
        if totals["fuel_consumption_available"] and totals["liters"] else None
    )
    totals["cost_per_km"] = (
        totals["expenses_mxn"] / totals["distance_km"]
        if expense_available and totals["distance_km"] else None
    )
    totals["expense_complete"] = expense_complete
    totals["expense_available"] = expense_available
    score_values = [row["score"] for row in units if row["score"] is not None]
    totals["driver_score"] = sum(score_values) / len(score_values) if score_values else None
    drivers: dict[str, dict[str, Any]] = {}
    for row in units:
        name = row["driver_name"] or "Sin conductor asignado"
        driver = drivers.setdefault(name, {
            "driver_name": name, "units": set(), "security": 0, "speeding": 0,
            "critical_high": 0, "attention_index": 0,
        })
        driver["units"].add(row["vehicle_number"])
        for key in ("security", "speeding", "critical_high", "attention_index"):
            driver[key] += row[key]
    driver_rows = []
    for driver in drivers.values():
        driver["vehicles"] = ", ".join(sorted(driver.pop("units")))
        driver["score"] = max(0.0, 100.0 - float(driver["attention_index"]))
        driver_rows.append(driver)
    driver_rows.sort(key=lambda row: (-(row["security"] + row["speeding"]), row["driver_name"]))
    attention_units = [row for row in units if row["telemetry_available"] and row["security"] + row["speeding"] > 0]
    units_without_gps = [row for row in units if not row["telemetry_available"]]
    inspection_credit_rows = sorted(
        inspection_credits.values(), key=lambda row: (-row["inspections"], row["driver_name"], row["vehicle_number"])
    )
    return {
        "units": units,
        "attention_units": attention_units,
        "units_without_gps": units_without_gps,
        "inspection_credits": inspection_credit_rows,
        "drivers": driver_rows,
        "behaviors": [{"label": name, "count": count} for name, count in behaviors.most_common()],
        "severity": [{"label": name, "count": count} for name, count in severity.most_common()],
        "daily": [{"date": day, "count": daily[day]} for day in sorted(daily)],
        "critical_high": sum(row["critical_high"] for row in units),
        "totals": totals,
    }


def comparison_row(name: str, analytics: dict[str, Any], previous: dict[str, Any] | None = None) -> dict[str, Any]:
    totals = analytics["totals"]
    previous_totals = (previous or {}).get("totals", {})
    events = sum(row["security"] + row["speeding"] for row in analytics["units"])
    previous_events = sum(row["security"] + row["speeding"] for row in (previous or {}).get("units", []))
    return {
        "zone": name or "Toda la flotilla",
        "vehicles": len(analytics["units"]),
        "vehicles_with_data": totals["vehicles_with_data"],
        "vehicles_without_gps": totals["vehicles_without_gps"],
        "driver_score": totals["driver_score"],
        "events": events,
        "critical_high": analytics["critical_high"],
        "distance_km": totals["distance_km"] if totals["distance_available"] else None,
        "engine_hours": totals["engine_hours"] if totals["engine_hours_available"] else None,
        "utilization_pct": (
            sum(row["utilization_pct"] for row in analytics["units"]) / len(analytics["units"])
            if analytics["units"] else 0
        ),
        "expenses_mxn": totals["expenses_mxn"],
        "expense_available": totals["expense_available"],
        "expense_complete": totals["expense_complete"],
        "maintenance_mxn": totals["maintenance_mxn"],
        "km_per_liter": totals["km_per_liter"],
        "cost_per_km": totals["cost_per_km"],
        "inspections": totals["inspections"],
        "open_defects": totals["open_defects"],
        "overdue_defects": totals["overdue_defects"],
        "events_delta_pct": ((events - previous_events) / previous_events) if previous_events else None,
        "expense_delta_pct": (
            (totals["expenses_mxn"] - float(previous_totals.get("expenses_mxn") or 0))
            / float(previous_totals.get("expenses_mxn") or 1)
            if previous_totals.get("expenses_mxn") else None
        ),
    }


def build_fleet_report(data: dict[str, list[dict[str, Any]]], start: date, end: date, group_name: str = "") -> bytes:
    buffer = io.BytesIO()
    workbook = xlsxwriter.Workbook(buffer, {"in_memory": True})
    analytics = fleet_analytics(data)
    title = workbook.add_format({"bold": True, "font_size": 24, "font_color": "#6B1022"})
    subtitle = workbook.add_format({"font_size": 11, "font_color": "#6F665E"})
    section = workbook.add_format({"bold": True, "font_size": 13, "font_color": "#FFFFFF", "bg_color": "#6B1022", "align": "left", "valign": "vcenter"})
    header = workbook.add_format({"bold": True, "bg_color": "#6B1022", "font_color": "#FFFFFF", "border": 1, "align": "center"})
    money = workbook.add_format({"num_format": '$#,##0.00;[Red]-$#,##0.00', "border": 1})
    number = workbook.add_format({"num_format": '#,##0.00', "border": 1})
    cell = workbook.add_format({"border": 1, "valign": "top"})
    date_fmt = workbook.add_format({"num_format": "dd/mm/yyyy hh:mm", "border": 1})
    kpi_label = workbook.add_format({"bold": True, "font_color": "#6F665E", "font_size": 9, "align": "center", "bg_color": "#F7F2EC", "border": 1})
    kpi_value = workbook.add_format({"bold": True, "font_color": "#171513", "font_size": 18, "align": "center", "bg_color": "#FFFFFF", "border": 1})
    note = workbook.add_format({"font_color": "#6F665E", "bg_color": "#FFF8E8", "border": 1, "text_wrap": True, "valign": "vcenter"})

    dashboard = workbook.add_worksheet("Dashboard")
    dashboard.hide_gridlines(2); dashboard.set_tab_color("#6B1022")
    dashboard.set_landscape(); dashboard.fit_to_pages(1, 0); dashboard.set_margins(0.25, 0.25, 0.35, 0.35)
    dashboard.set_header("&LGE CONTROL · FLOTILLA 360&RInforme gerencial")
    dashboard.set_footer("&LInformación gerencial&C&P de &N&RGenerado por GE Control")
    dashboard.set_column("A:A", 2); dashboard.set_column("B:B", 10)
    dashboard.set_column("C:D", 22); dashboard.set_column("E:F", 14)
    dashboard.set_column("G:G", 3); dashboard.set_column("H:H", 9)
    dashboard.set_column("I:J", 21); dashboard.set_column("K:L", 17); dashboard.set_column("M:M", 13)
    dashboard.set_column("O:P", None, None, {"hidden": True})
    dashboard.set_row(0, 10)

    dashboard_title = workbook.add_format({"bold": True, "font_size": 24, "font_color": "#6B1022", "valign": "vcenter"})
    white_title = workbook.add_format({"bold": True, "font_size": 24, "font_color": "#FFFFFF", "bg_color": "#7C1028", "valign": "vcenter"})
    gold_subtitle = workbook.add_format({"font_size": 11, "font_color": "#6F665E"})
    quick_guide = workbook.add_format({"bold": True, "font_size": 10, "font_color": "#FFFFFF", "bg_color": "#7C1028", "border": 1, "border_color": "#B95B70", "text_wrap": True, "valign": "vcenter"})
    dashboard.merge_range("B2:M3", "INFORME EJECUTIVO · FLOTILLA 360", dashboard_title)
    dashboard.merge_range("B4:M4", f"{(group_name or 'Toda la flotilla').upper()}  |  {start:%d/%m/%Y} al {end:%d/%m/%Y}", gold_subtitle)
    dashboard.set_row(1, 28); dashboard.set_row(2, 28); dashboard.set_row(3, 22)
    event_total = len(data.get("driver_events", [])) + len(data.get("speeding", []))
    avg_utilization = (
        sum(row["utilization_pct"] for row in analytics["units"]) / len(analytics["units"])
        if analytics["units"] else 0
    )
    metrics = [
        ("UNIDADES CON DATOS", analytics["totals"]["vehicles_with_data"], False),
        ("SIN DATOS GPS", analytics["totals"]["vehicles_without_gps"], False),
        ("EVENTOS TOTALES", event_total, False),
        ("CRÍTICOS / ALTOS", analytics["critical_high"], False),
        ("INSPECCIONES", analytics["totals"]["inspections"], False),
        ("DEFECTOS ABIERTOS", analytics["totals"]["open_defects"], False),
    ]
    kpi_backgrounds = ["#E6F4EC", "#FCE8E8", "#F7F3EE", "#FCE8E8", "#E6F4EC", "#FFF3D6"]
    kpi_accents = ["#2E7D5B", "#C43B3B", "#7C1028", "#C43B3B", "#2E7D5B", "#D7A43A"]
    for index, (label, value, is_money) in enumerate(metrics):
        column = 1 + index * 2
        label_format = workbook.add_format({"bold": True, "font_color": "#6F6965", "font_size": 9, "align": "center", "bg_color": kpi_backgrounds[index], "border": 1, "border_color": "#DED6CF"})
        value_format = workbook.add_format({"bold": True, "font_color": kpi_accents[index], "font_size": 20, "align": "center", "valign": "vcenter", "bg_color": kpi_backgrounds[index], "border": 1, "border_color": "#DED6CF"})
        dashboard.merge_range(6, column, 6, column + 1, label, label_format)
        display = "NO DISPONIBLE" if value is None else (
            f"{value:.0%}" if is_money == "percent" else f"${value:,.2f}" if is_money else f"{value:,.1f}"
        )
        dashboard.merge_range(7, column, 8, column + 1, display, value_format)

    manager_section = workbook.add_format({"bold": True, "font_size": 13, "font_color": "#FFFFFF", "bg_color": "#7C1028", "valign": "vcenter"})
    light_section = workbook.add_format({"bold": True, "font_size": 13, "font_color": "#202020", "bg_color": "#F7F3EE", "valign": "vcenter"})
    action_headers = workbook.add_format({"bold": True, "font_color": "#6B1022", "bg_color": "#F7F3EE", "border": 1, "valign": "vcenter"})
    action_body = workbook.add_format({"font_size": 10, "font_color": "#202020", "border": 1, "text_wrap": True, "valign": "top"})
    ranking = analytics["attention_units"]
    behavior_rows = analytics["behaviors"][:10]
    top_unit = ranking[0] if ranking else None
    top_behavior = behavior_rows[0] if behavior_rows else None

    with_gps = sorted(
        (row for row in analytics["units"] if row["telemetry_available"]),
        key=lambda row: (-(row["security"] + row["speeding"]), row["vehicle_number"]),
    )
    without_gps = analytics["units_without_gps"]
    inspection_rows = analytics["inspection_credits"]

    dashboard.merge_range("B10:F10", "Unidades con GPS", section)
    dashboard.write_row("B11", ["#", "Unidad", "Conductor", "Eventos", "Cobertura"], header)
    priority_high = workbook.add_format({"font_color": "#C43B3B", "bg_color": "#FCE8E8", "border": 1})
    priority_medium = workbook.add_format({"font_color": "#8A6200", "bg_color": "#FFF3D6", "border": 1})
    priority_low = workbook.add_format({"font_color": "#2E7D5B", "bg_color": "#E6F4EC", "border": 1})
    for row_index, item in enumerate(with_gps, 11):
        events = item["security"] + item["speeding"]
        dashboard.write(row_index, 1, row_index - 10, cell)
        dashboard.write(row_index, 2, item["vehicle_number"], cell)
        dashboard.write(row_index, 3, item["driver_name"] or "Sin conductor asignado", cell)
        dashboard.write(row_index, 4, events, number)
        dashboard.write(row_index, 5, "Con datos GPS", cell)
    if with_gps:
        dashboard.conditional_format(11, 4, 10 + len(with_gps), 4, {"type": "data_bar", "bar_color": "#98243D"})

    dashboard.merge_range("I10:M10", "Unidades sin datos GPS", section)
    dashboard.write_row("I11", ["#", "Unidad", "Conductor", "Estado", "Acción"], header)
    alert_cell = workbook.add_format({"bold": True, "font_color": "#202020", "bg_color": "#FCE8E8", "border": 1, "border_color": "#C43B3B", "text_wrap": True})
    for row_index, item in enumerate(without_gps, 11):
        dashboard.write_row(row_index, 8, [row_index - 10, item["vehicle_number"], item["driver_name"] or "Sin conductor asignado", "Revisión manual", "Revisar GPS"], alert_cell)

    detail_end = max(14, 10 + len(with_gps), 10 + len(without_gps))
    analysis_row = detail_end + 3
    chart_end_row = analysis_row + 18

    dashboard.merge_range(analysis_row, 8, analysis_row, 12, "Inspecciones realizadas por chofer", section)
    dashboard.write_row(analysis_row + 1, 8, ["#", "Unidad", "Chofer", "Tipo", "Inspecciones"], header)
    if inspection_rows:
        for row_index, item in enumerate(inspection_rows, analysis_row + 2):
            dashboard.write_row(row_index, 8, [row_index - analysis_row - 1, item["vehicle_number"], item["driver_name"], "Registrada", item["inspections"]], cell)
    else:
        empty_inspections = workbook.add_format({"italic": True, "font_color": "#6F665E", "bg_color": "#F7F3EE", "border": 1, "align": "center", "valign": "vcenter"})
        dashboard.merge_range(analysis_row + 2, 8, analysis_row + 4, 12, "No se registraron inspecciones en el periodo.", empty_inspections)

    pid_counts: Counter[str] = Counter()
    for fault in data.get("faults", []):
        pid = _text(fault.get("code") or fault.get("code_label")) or "Sin código"
        pid_counts[pid] += int(_amount(fault.get("occurrence_count")) or 1)
    pid_rows = pid_counts.most_common()
    dashboard.write_row("O1", ["PID", "Incidencias"], header)
    for row_index, (pid, count) in enumerate(pid_rows, 1):
        dashboard.write_row(row_index, 14, [pid, count], cell)
    if pid_rows:
        chart = workbook.add_chart({"type": "doughnut"})
        chart.add_series({
            "name": "Incidencias",
            "categories": ["Dashboard", 1, 14, len(pid_rows), 14],
            "values": ["Dashboard", 1, 15, len(pid_rows), 15],
            "data_labels": {"percentage": True, "leader_lines": True},
        })
        chart.set_title({"name": "Incidencias por código PID"})
        chart.set_legend({"position": "right"})
        chart.set_hole_size(58)
        chart.set_style(10)
        chart.set_size({"width": 650, "height": 360})
        chart.set_chartarea({"border": {"none": True}})
        dashboard.insert_chart(analysis_row, 1, chart, {"x_offset": 4, "y_offset": 4})
    else:
        empty_pid = workbook.add_format({"italic": True, "font_color": "#6F665E", "bg_color": "#F7F3EE", "border": 1, "align": "center", "valign": "vcenter"})
        dashboard.merge_range(analysis_row, 1, analysis_row + 5, 6, "Sin incidencias PID registradas en el periodo.", empty_pid)

    decisions_row = chart_end_row + 2
    dashboard.merge_range(decisions_row, 1, decisions_row, 12, "DECISIONES RECOMENDADAS PARA EL GERENTE", manager_section)
    decision_specs = [
        (1, 4, "1 · CORREGIR LA MAYOR EXPOSICIÓN", (
            f"Revisar {top_unit['vehicle_number']}: concentra {top_unit['security'] + top_unit['speeding']:,} eventos en el periodo."
            if top_unit else "No se detectaron unidades con eventos en el periodo."
        )),
        (5, 8, "2 · ATACAR LA CONDUCTA DOMINANTE", (
            f"Aplicar retroalimentación sobre {top_behavior['label']}, con {top_behavior['count']:,} eventos."
            if top_behavior else "No se detectaron conductas de seguridad en el periodo."
        )),
        (9, 12, "3 · REVISAR EL PID DOMINANTE", (
            f"Priorizar {pid_rows[0][0]}, que concentra {pid_rows[0][1]:,} incidencias, y documentar su cierre."
            if pid_rows else "No se registraron incidencias PID en el periodo."
        )),
    ]
    for first_col, last_col, heading, body in decision_specs:
        dashboard.merge_range(decisions_row + 1, first_col, decisions_row + 1, last_col, heading, action_headers)
        dashboard.merge_range(decisions_row + 2, first_col, decisions_row + 4, last_col, body, action_body)
    dashboard.print_area(1, 1, decisions_row + 4, 12)

    supervisor = workbook.add_worksheet("Seguimiento supervisor")
    supervisor.hide_gridlines(2); supervisor.set_tab_color("#D7A43A")
    supervisor.set_landscape(); supervisor.fit_to_pages(1, 0); supervisor.freeze_panes(7, 0)
    supervisor.set_column("A:A", 4); supervisor.set_column("B:B", 20); supervisor.set_column("C:C", 24)
    supervisor.set_column("D:G", 12); supervisor.set_column("H:H", 38); supervisor.set_column("I:I", 20)
    supervisor.set_column("J:J", 18); supervisor.set_column("K:K", 15); supervisor.set_column("L:L", 14); supervisor.set_column("M:M", 28)
    supervisor.merge_range("B2:I3", "SEGUIMIENTO PARA SUPERVISIÓN", white_title)
    supervisor.merge_range("J2:M3", "USO: filtra PRIORIDAD, asigna responsable y registra el avance.", quick_guide)
    supervisor.merge_range("B5:M5", "No necesitas interpretar todo el reporte: atiende primero ROJO, después AMARILLO y documenta el cierre.", light_section)
    follow_headers = ["#", "UNIDAD", "CONDUCTOR", "GPS", "EVENTOS", "CRÍTICOS", "DEFECTOS", "ACCIÓN CLARA", "PRIORIDAD", "RESPONSABLE", "FECHA LÍMITE", "ESTADO", "EVIDENCIA / NOTA"]
    supervisor.write_row(6, 0, follow_headers, header)
    supervisor.set_row(6, 38)
    status_format = workbook.add_format({"font_color": "#C43B3B", "bg_color": "#FCE8E8", "border": 1})
    for row_index, item in enumerate(analytics["units"], 7):
        events = item["security"] + item["speeding"]
        has_gps = item["telemetry_available"]
        if not has_gps:
            action, priority, priority_format = "Revisar GPS y confirmar señal", "ROJO · HOY", priority_high
        elif item["critical_high"] >= 3:
            action, priority, priority_format = "Hablar con conductor y acordar corrección", "ROJO · HOY", priority_high
        elif item["critical_high"] or item["open_defects"] or events >= 15:
            action, priority, priority_format = (
                "Programar reparación y comprobar cierre" if item["open_defects"] and not item["critical_high"]
                else "Hablar con conductor y acordar corrección",
                "AMARILLO · SEMANA", priority_medium,
            )
        else:
            action, priority, priority_format = "Mantener monitoreo", "VERDE · MONITOREAR", priority_low
        supervisor.write_row(row_index, 0, [
            row_index - 6, item["vehicle_number"], item["driver_name"] or "Sin conductor asignado",
            "CON GPS" if has_gps else "SIN GPS", events, item["critical_high"], item["open_defects"], action,
        ], cell)
        supervisor.write(row_index, 8, priority, priority_format)
        supervisor.write_blank(row_index, 9, None, cell); supervisor.write_blank(row_index, 10, None, date_fmt)
        supervisor.write(row_index, 11, "PENDIENTE", status_format); supervisor.write_blank(row_index, 12, None, cell)
        supervisor.set_row(row_index, 32)
    last_supervisor_row = max(7, 6 + len(analytics["units"]))
    supervisor.autofilter(6, 0, last_supervisor_row, 12)
    if analytics["units"]:
        supervisor.data_validation(7, 11, last_supervisor_row, 11, {"validate": "list", "source": ["PENDIENTE", "EN PROCESO", "CERRADO"]})
        supervisor.conditional_format(7, 11, last_supervisor_row, 11, {"type": "text", "criteria": "containing", "value": "EN PROCESO", "format": priority_medium})
        supervisor.conditional_format(7, 11, last_supervisor_row, 11, {"type": "text", "criteria": "containing", "value": "CERRADO", "format": priority_low})
    supervisor.print_area(0, 0, last_supervisor_row, 12)

    summary = workbook.add_worksheet("Resumen por unidad")
    summary.hide_gridlines(2); summary.set_landscape(); summary.fit_to_pages(1, 0)
    summary.freeze_panes(2, 2)
    unit_headers = [
        "Unidad", "Conductor", "Cobertura GPS", "Seguridad", "Velocidad", "Eventos totales",
        "Críticos / altos", "Fallas abiertas", "Gasto total MXN", "Mantenimiento MXN",
        "Consumo L", "Kilómetros", "Horas motor", "Utilización", "km/L", "Costo/km",
        "Inspecciones", "Defectos abiertos", "Defectos vencidos",
    ]
    group_header = workbook.add_format({"bold": True, "font_color": "#FFFFFF", "bg_color": "#3E0A16", "align": "center", "border": 1})
    summary.merge_range("A1:B1", "IDENTIFICACIÓN", group_header)
    summary.merge_range("C1:H1", "SEGURIDAD Y COBERTURA", group_header)
    summary.merge_range("I1:K1", "COSTOS Y COMBUSTIBLE", group_header)
    summary.merge_range("L1:P1", "USO Y EFICIENCIA", group_header)
    summary.merge_range("Q1:S1", "INSPECCIONES", group_header)
    summary.write_row(1, 0, unit_headers, header)
    summary.set_column("A:A", 24); summary.set_column("B:B", 32)
    summary.set_column("C:C", 27); summary.set_column("D:S", 13)
    for row_index, item in enumerate(analytics["units"], 2):
        values = [
            item["vehicle_number"], item["driver_name"] or "Sin conductor asignado", item["coverage_status"],
            item["security"], item["speeding"], item["security"] + item["speeding"],
            item["critical_high"], item["faults"], item["expense_mxn"], item["maintenance_mxn"],
            item["liters"], item["distance_km"], item["engine_hours"], item["utilization_pct"],
            item["km_per_liter"], item["cost_per_km"], item["inspections"],
            item["open_defects"], item["overdue_defects"],
        ]
        for column, value in enumerate(values):
            if value is None:
                summary.write(row_index, column, "No disponible", cell)
            else:
                summary.write(row_index, column, value, money if column in {8, 9, 15} else cell)
    if analytics["units"]:
        last_row = len(analytics["units"]) + 1
        summary.autofilter(1, 0, last_row, len(unit_headers) - 1)
        summary.conditional_format(2, 5, last_row, 5, {"type": "data_bar", "bar_color": "#8B1E34"})

    _sheet_if_rows(workbook, "Gastos", data.get("expenses", []), [
        ("Fecha", "occurred_at"), ("Grupo", "group_name"), ("Zona", "zone_name"), ("Unidad", "vehicle_number"),
        ("Tipo", "expense_type"), ("Categoría", "category"), ("Descripción", "description"), ("Litros", "quantity_liters"),
        ("Importe MXN", "amount_mxn"), ("Origen", "source"), ("Registró", "submitted_by")], header, cell, date_fmt, money)
    _sheet_if_rows(workbook, "Compras combustible", data.get("fuel", []), [
        ("Fecha", "purchased_at"), ("Unidad", "vehicle_number"), ("Combustible", "fuel_type"),
        ("Litros", "quantity_liters"), ("Importe", "total_cost"), ("Moneda", "currency"),
        ("Proveedor", "vendor"), ("Odómetro km", "odometer_km")], header, cell, date_fmt, money)
    _sheet_if_rows(workbook, "Seguridad", data.get("driver_events", []), [
        ("Fecha", "started_at"), ("Unidad", "vehicle_number"), ("Conductor", "driver_name"), ("Evento", "event_type"),
        ("Conducta", "primary_behavior"), ("Gravedad", "severity"), ("Duración s", "duration_seconds")], header, cell, date_fmt, money)
    _sheet_if_rows(workbook, "Exceso velocidad", data.get("speeding", []), [
        ("Fecha", "started_at"), ("Unidad", "vehicle_number"), ("Conductor", "driver_name"), ("Gravedad", "severity"),
        ("Límite km/h", "posted_limit_kph"), ("Máx. superada km/h", "max_over_kph"), ("Promedio km/h", "avg_speed_kph"),
        ("Duración s", "duration_seconds")], header, cell, date_fmt, money)
    _sheet_if_rows(workbook, "Códigos de falla", data.get("faults", []), [
        ("Primera detección", "occurred_at"), ("Unidad", "vehicle_number"), ("Código", "code"), ("Etiqueta", "code_label"),
        ("Descripción", "description"), ("Estado", "status"), ("Ocurrencias", "occurrence_count")], header, cell, date_fmt, money)
    _sheet_if_rows(workbook, "Inspecciones", data.get("inspections", []), [
        ("Fecha", "inspected_at"), ("Unidad", "vehicle_number"), ("Chofer que inspeccionó", "driver_name"), ("Tipo", "inspection_type"),
        ("Estado", "status"), ("Rechazada", "is_rejected"), ("Odómetro km", "odometer_km")], header, cell, date_fmt, money)
    actionable_defects = [
        row for row in data.get("defects", [])
        if _text(row.get("status")).casefold() in {"open", "pending", "unresolved", "with_defects"}
        or _text(row.get("severity")).casefold() in {"major", "critical", "high", "severe"}
    ]
    _sheet_if_rows(workbook, "Defectos accionables", actionable_defects, [
        ("Unidad", "vehicle_number"), ("Categoría", "category"), ("Defecto", "title"),
        ("Gravedad", "severity"), ("Estado", "status"), ("Vencido", "is_overdue"),
        ("Notas", "notes"), ("Resuelto", "resolved_at")], header, cell, date_fmt, money)
    chronology = []
    for row in data.get("driver_events", []):
        chronology.append({
            "date": row.get("started_at"), "vehicle_number": row.get("vehicle_number"),
            "driver_name": row.get("driver_name"), "kind": "Seguridad",
            "detail": behavior_label(row.get("primary_behavior") or row.get("event_type")),
            "severity": row.get("severity"),
        })
    for row in data.get("speeding", []):
        chronology.append({
            "date": row.get("started_at"), "vehicle_number": row.get("vehicle_number"),
            "driver_name": row.get("driver_name"), "kind": "Velocidad",
            "detail": f"Exceso máximo {float(row.get('max_over_kph') or 0):g} km/h",
            "severity": row.get("severity"),
        })
    for row in data.get("faults", []):
        chronology.append({
            "date": row.get("occurred_at"), "vehicle_number": row.get("vehicle_number"),
            "driver_name": "", "kind": "Falla",
            "detail": row.get("code_label") or row.get("code") or "Código de falla",
            "severity": row.get("severity"),
        })
    chronology.sort(key=lambda row: str(row.get("date") or ""), reverse=True)
    _sheet_if_rows(workbook, "Cronología ejecutiva", chronology, [
        ("Fecha", "date"), ("Unidad", "vehicle_number"), ("Conductor", "driver_name"),
        ("Tipo", "kind"), ("Detalle", "detail"), ("Gravedad", "severity"),
    ], header, cell, date_fmt, money)
    workbook.close()
    return buffer.getvalue()


def _sheet_if_rows(workbook: xlsxwriter.Workbook, name: str, rows: list[dict[str, Any]], columns: list[tuple[str, str]], header: Any, cell: Any, date_fmt: Any, money: Any) -> None:
    if rows:
        _sheet(workbook, name, rows, columns, header, cell, date_fmt, money)


def _sheet(workbook: xlsxwriter.Workbook, name: str, rows: list[dict[str, Any]], columns: list[tuple[str, str]], header: Any, cell: Any, date_fmt: Any, money: Any) -> None:
    sheet = workbook.add_worksheet(name[:31])
    sheet.freeze_panes(1, 0); sheet.autofilter(0, 0, max(len(rows), 1), len(columns) - 1)
    widths = {"occurred_at": 20, "started_at": 20, "ended_at": 20, "vehicle_number": 24, "driver_name": 24,
              "description": 42, "location": 38, "origin": 34, "destination": 34, "submitted_by": 25,
              "primary_behavior": 24, "category": 20, "code_label": 25}
    for column, (label, key) in enumerate(columns):
        sheet.write(0, column, label, header); sheet.set_column(column, column, widths.get(key, min(max(len(label) + 3, 13), 22)))
    for row_index, row in enumerate(rows, 1):
        for column, (_, key) in enumerate(columns):
            value = row.get(key)
            if key in {"date", "occurred_at", "started_at", "ended_at", "inspected_at", "resolved_at", "purchased_at"} and value:
                try:
                    sheet.write_datetime(row_index, column, datetime.fromisoformat(str(value).replace("Z", "+00:00")).replace(tzinfo=None), date_fmt)
                    continue
                except ValueError:
                    pass
            sheet.write(row_index, column, value if value is not None else "", money if key == "amount_mxn" else cell)
