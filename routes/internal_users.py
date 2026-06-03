from __future__ import annotations

import hashlib
import hmac
import json
import logging
import os
import secrets
import traceback
import xml.etree.ElementTree as ET
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, timedelta, timezone
from typing import Optional
from xml.sax.saxutils import escape as xml_escape
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Header, HTTPException
from fastapi.responses import JSONResponse, RedirectResponse, Response
from pydantic import BaseModel

from routes.auth import _resolve_active_module_access, obtener_acceso_modulo, verify_token
from routes.perfiles import _tenant_id_for_user, get_perfiles_for_user
from services.database import get_facilities
from services.email_delivery import send_gas_lp_invoice_email
from services.fiscal_pdf import fiscal_pdf_info, generar_pdf_gas_lp_desde_xml
from services.cfdi_cancellation import cancel_cfdi_universal
from services.sw_sapien import sw_runtime_config, timbrar_cfdi
from supabase_config import get_supabase_admin, get_supabase_for_user


router = APIRouter()
logger = logging.getLogger(__name__)

ROLES = {"admin", "operador", "asistente_facturacion", "asistente_operativo", "conciliacion", "planta", "solo_lectura"}
SECTIONS = {"transporte", "gas_lp", "gasolineras"}
MAX_FAILED_ATTEMPTS = 5
LOCK_MINUTES = 15
SESSION_HOURS = 12
GAS_LP_CLAVE_PROD_SERV = "15111510"
GAS_LP_SW_HYP_EXPERIMENTAL_CLAVE = "15101515"
GAS_LP_HYP_SUBPRODUCTO = "SP23"
GAS_LP_HIDRO_CLAVES = {GAS_LP_CLAVE_PROD_SERV, GAS_LP_SW_HYP_EXPERIMENTAL_CLAVE}
GAS_LP_HYP_DIAGNOSTIC_CLAVES = {GAS_LP_CLAVE_PROD_SERV, GAS_LP_SW_HYP_EXPERIMENTAL_CLAVE}
HYP_TIPO_PERMISOS_VALIDOS = {f"PER{i:02d}" for i in range(1, 12)}
GAS_LP_HYP_MODES = {"required", "disabled", "diagnostic"}
GAS_LP_HYP_DEBUG_LOG = os.environ.get("GAS_LP_HYP_DEBUG_LOG", "logs/gas_lp_hyp_pre_timbrado.log")


class InternalUserCreate(BaseModel):
    display_name: str
    section: str
    role: str
    perfil_id: int
    chofer_id: Optional[int] = None
    code: Optional[str] = ""
    pin: Optional[str] = ""
    permissions: Optional[dict] = None


class InternalUserStatus(BaseModel):
    status: str


class InternalUserUpdate(BaseModel):
    role: Optional[str] = None
    display_name: Optional[str] = None


class InternalResetPin(BaseModel):
    pin: Optional[str] = ""


class InternalLogin(BaseModel):
    section: str = "transporte"
    code: str
    pin: str


class InternalLogout(BaseModel):
    token: str


class DetectedLoadAction(BaseModel):
    action: str
    updates: Optional[dict] = None


class GasLpInternalClientePayload(BaseModel):
    rfc: str = "XAXX010101000"
    nombre: str = "PUBLICO EN GENERAL"
    cp: str = ""
    regimen_fiscal: str = "616"
    uso_cfdi: str = "S01"
    email: str = ""
    email_adicional_1: str = ""
    email_adicional_2: str = ""


class GasLpInternalFacturaPayload(BaseModel):
    cliente_id: Optional[int] = None
    publico_general: bool = False
    rfc: str = "XAXX010101000"
    nombre: str = "PUBLICO EN GENERAL"
    cp: str = ""
    regimen_fiscal: str = "616"
    uso_cfdi: str = "S01"
    litros: float
    precio_unitario: float
    concepto: str = "Venta de Gas LP"
    forma_pago: str = "99"
    metodo_pago: str = "PUE"
    descuento: float = 0
    iva_rate: float = 0.16
    serie: str = "AA"
    folio: str = ""
    comentarios: str = ""
    fecha: str = ""
    clave_prod_serv: str = GAS_LP_CLAVE_PROD_SERV
    no_identificacion: str = "GLP-LTR"
    unidad: str = "Litro"
    facility_id: Optional[int] = None
    tipo_operacion: str = "venta"
    destino_facility_id: Optional[int] = None
    generar_carta_porte: bool = False
    vehiculo_id: Optional[int] = None
    chofer_id: Optional[int] = None
    ruta_id: Optional[int] = None
    enviar_correo: bool = True
    transfer_email: str = ""
    transfer_email_provided: bool = False
    factura_global: bool = False
    informacion_global_periodicidad: str = "04"
    informacion_global_meses: str = ""
    informacion_global_anio: Optional[int] = None
    hyp_experimental_diagnostics: bool = False
    hyp_numero_permiso_override: str = ""
    hyp_tipo_permiso_override: str = ""
    hyp_clave_hyp_override: str = ""


class GasLpSendEmailPayload(BaseModel):
    email: str = ""
    email_adicional_1: str = ""
    email_adicional_2: str = ""


class GasLpTransferEmailDefaultPayload(BaseModel):
    email: str = ""


class GasLpHypLCNEDiagnosticPayload(BaseModel):
    facility_id: Optional[int] = None
    facility_ids: Optional[list[int]] = None
    cliente_id: Optional[int] = 7
    litros: float = 475
    precio_unitario: float = 10.08
    forma_pago: str = "03"
    metodo_pago: str = "PUE"
    descuento: float = 0
    iva_rate: float = 0.16
    probar_claves_producto: bool = True
    stop_on_success: bool = True


class GasLpConciliacionPublicoGeneralPayload(BaseModel):
    litros: float
    precio_unitario: float
    facility_id: int
    forma_pago: str = "01"
    metodo_pago: str = "PUE"
    fecha: str = ""
    comentarios: str = ""
    descuento: float = 0
    iva_rate: float = 0.16
    factura_global: bool = False
    informacion_global_periodicidad: str = "04"
    informacion_global_meses: str = ""
    informacion_global_anio: Optional[int] = None


class GasLpComplementoPagoPayload(BaseModel):
    fecha_pago: str = ""
    forma_pago: str = "03"
    monto: Optional[float] = None
    factura_ids: Optional[list[int]] = None
    facturas: Optional[list[dict]] = None
    referencia: str = ""
    banco: str = ""
    notas: str = ""


class GasLpCancelacionPayload(BaseModel):
    motivo: str = "02"
    uuid_sustitucion: str = ""
    notas: str = ""
    tipo: str = "fiscal"


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _now_iso() -> str:
    return _now().isoformat()


def _gas_lp_cfdi_timezone(issuer_cp: str | None = None):
    try:
        return ZoneInfo("America/Mexico_City")
    except Exception:
        return timezone(timedelta(hours=-6))


def _parse_gas_lp_cfdi_fecha(value: str, tz) -> Optional[datetime]:
    raw = str(value or "").strip()
    if not raw:
        return None
    cleaned = raw.replace("\u202f", " ").replace("\xa0", " ").strip()
    candidates = [cleaned]
    if "," in cleaned:
        candidates.append(cleaned.replace(",", ""))
    for candidate in candidates:
        try:
            parsed = datetime.fromisoformat(candidate)
            return parsed.astimezone(tz) if parsed.tzinfo else parsed.replace(tzinfo=tz)
        except Exception:
            pass
    formats = (
        "%d/%m/%Y, %H:%M:%S",
        "%d/%m/%Y, %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    )
    for fmt in formats:
        try:
            return datetime.strptime(cleaned, fmt).replace(tzinfo=tz)
        except Exception:
            continue
    return None


def _gas_lp_cfdi_fecha_actualizada(fecha: str, issuer_cp: str | None = None) -> tuple[str, bool, str]:
    tz = _gas_lp_cfdi_timezone(issuer_cp)
    now_local = datetime.now(tz)
    parsed = _parse_gas_lp_cfdi_fecha(fecha, tz)
    if parsed is None:
        return now_local.strftime("%Y-%m-%dT%H:%M:%S"), bool(str(fecha or "").strip()), "invalid_or_empty"
    min_allowed = now_local - timedelta(hours=72)
    max_allowed = now_local + timedelta(minutes=2)
    if parsed < min_allowed or parsed > max_allowed:
        logger.warning(
            "gas_lp_cfdi_fecha_replaced_out_of_range original=%s parsed=%s now_local=%s min_allowed=%s max_allowed=%s issuer_cp=%s",
            fecha,
            parsed.isoformat(),
            now_local.isoformat(),
            min_allowed.isoformat(),
            max_allowed.isoformat(),
            issuer_cp or "",
        )
        return now_local.strftime("%Y-%m-%dT%H:%M:%S"), True, "out_of_range"
    return parsed.strftime("%Y-%m-%dT%H:%M:%S"), False, "ok"


def _hash_secret(value: str, salt: str | None = None) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", value.encode(), salt.encode(), 120_000).hex()
    return f"pbkdf2_sha256${salt}${digest}"


def _verify_secret(value: str, stored: str) -> bool:
    try:
        algo, salt, digest = stored.split("$", 2)
        if algo != "pbkdf2_sha256":
            return False
        candidate = _hash_secret(value, salt).split("$", 2)[2]
        return hmac.compare_digest(candidate, digest)
    except Exception:
        return False


def _hash_token(token: str) -> str:
    return hashlib.sha256(token.encode()).hexdigest()


def _clean_code(value: str) -> str:
    code = "".join(ch for ch in str(value or "").upper().strip() if ch.isalnum() or ch in {"-", "_"})
    return code[:24]


def _clean_login(value: str) -> str:
    return str(value or "").strip().upper()


def _matches_login(row: dict, login: str, allow_display_name: bool = False) -> bool:
    code = _clean_login(row.get("code"))
    if code == login:
        return True
    if not allow_display_name:
        return False
    display_name = _clean_login(row.get("display_name"))
    return bool(display_name and display_name == login)


def _profile_for_admin(admin_uid: str, perfil_id: int, token: str = "") -> dict:
    if not perfil_id:
        raise HTTPException(400, "Selecciona una empresa activa antes de gestionar usuarios internos.")
    rows = (
        get_supabase_admin()
        .table("perfiles_empresa")
        .select("id,user_id,tenant_id,nombre,rfc,activo")
        .eq("id", perfil_id)
        .eq("user_id", admin_uid)
        .eq("activo", True)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows:
        raise HTTPException(403, "La empresa seleccionada no pertenece a tu usuario o está inactiva.")
    perfil = rows[0]
    tenant_id = perfil.get("tenant_id") or _tenant_id_for_user(admin_uid, access_token=token)
    if not tenant_id:
        raise HTTPException(400, "La empresa activa no tiene tenant asignado.")
    admin_tenant = _tenant_id_for_user(admin_uid, access_token=token)
    if admin_tenant and str(tenant_id) != str(admin_tenant):
        raise HTTPException(403, "La empresa activa no pertenece al tenant del usuario.")
    perfil["tenant_id"] = tenant_id
    return perfil


def _validate_internal_scope(row: dict) -> None:
    if not row.get("tenant_id") or not row.get("perfil_id"):
        raise HTTPException(403, "Usuario interno huérfano: falta empresa asignada. Requiere backfill antes de usarlo.")
    perfiles = (
        get_supabase_admin()
        .table("perfiles_empresa")
        .select("id,tenant_id,activo")
        .eq("id", row.get("perfil_id"))
        .eq("activo", True)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not perfiles:
        raise HTTPException(403, "La empresa asignada al usuario interno está inactiva o no existe.")
    perfil_tenant = str(perfiles[0].get("tenant_id") or "")
    if perfil_tenant and perfil_tenant != str(row.get("tenant_id") or ""):
        raise HTTPException(403, "Usuario interno con tenant/perfil inconsistente. Requiere revisión antes de usarlo.")


def _status_label(status: str) -> str:
    return {
        "sin_sincronizar": "Sin sincronizar",
        "buscando_cfdi": "Buscando CFDI",
        "new": "Nueva carga detectada",
        "pending_confirmation": "Pendiente de confirmar",
        "confirmed": "Carta Porte borrador",
        "carta_porte_created": "Carta Porte borrador",
        "rejected": "Ignorada",
    }.get(status or "", status or "Sin sincronizar")


def _gas_lp_cliente_row(user: dict, payload: GasLpInternalClientePayload) -> dict:
    from routes.transporte import _normalizar_receptor_cfdi, _validar_datos_cfdi_receptor

    rfc = _clean_rfc(payload.rfc)
    nombre = str(payload.nombre or "").strip()
    uso_cfdi = str(payload.uso_cfdi or "S01").strip()
    regimen = str(payload.regimen_fiscal or "616").strip()
    cp = _clean_cp(payload.cp)
    invoice_emails = _invoice_email_recipients(payload.email, payload.email_adicional_1, payload.email_adicional_2)
    email = invoice_emails[0] if invoice_emails else ""
    if not rfc or not nombre:
        raise HTTPException(400, "RFC y nombre del cliente son obligatorios.")
    if rfc == "XAXX010101000":
        profile = _gas_lp_profile(user)
        settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
        issuer = _require_gas_lp_issuer(profile, settings)
        receptor = {
            "rfc": "XAXX010101000",
            "nombre": "PUBLICO EN GENERAL",
            "cp": cp or issuer["cp"],
            "regimen_fiscal": "616",
        }
        uso_cfdi = "S01"
    else:
        receptor = _normalizar_receptor_cfdi(rfc, nombre, cp, regimen)
        _validar_datos_cfdi_receptor(
            receptor["rfc"],
            receptor["regimen_fiscal"],
            receptor["cp"],
            uso_cfdi,
        )
    return {
        "user_id": user.get("owner_user_id"),
        "tenant_id": user.get("tenant_id"),
        "perfil_id": user.get("perfil_id"),
        "source": "assistant_portal",
        "modulo_propietario": "gas_lp",
        "rfc": receptor["rfc"],
        "nombre": receptor["nombre"],
        "cp": receptor["cp"],
        "regimen_fiscal": receptor["regimen_fiscal"],
        "uso_cfdi": uso_cfdi,
        "email": email,
        "email_facturacion": email,
        "activo": True,
        "metadata": {
            "created_by_internal": user.get("id"),
            "created_by": user.get("display_name"),
            "invoice_email_additional": invoice_emails[1:3],
            "email_adicional_1": invoice_emails[1] if len(invoice_emails) > 1 else "",
            "email_adicional_2": invoice_emails[2] if len(invoice_emails) > 2 else "",
        },
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
    }


def _safe_internal_error(action: str, exc: Exception) -> HTTPException:
    logger.exception("%s internal_user failed: %s", action, exc)
    return HTTPException(500, "No se pudo completar la operación. Intenta de nuevo o contacta a soporte.")


def _money(value) -> Decimal:
    return Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _rate(value) -> Decimal:
    return Decimal(str(value or 0)).quantize(Decimal("0.000000"), rounding=ROUND_HALF_UP)


def _gas_lp_internal_context(token: str, *, write: bool = False) -> dict:
    ctx = _internal_session(token, "gas_lp")
    role = (ctx["user"].get("role") or "").lower()
    if write and role not in {"asistente_facturacion", "admin"}:
        raise HTTPException(403, "Tu rol no permite facturar en este portal.")
    return ctx


def _gas_lp_conciliacion_context(token: str, *, write: bool = False, perfil_id: int | None = None) -> dict:
    if str(token or "").count(".") == 2:
        uid = verify_token(token)
        if not uid:
            raise HTTPException(401, "Sesión inválida o expirada.")
        access = _resolve_active_module_access(uid, "gas_lp", access_token=token)
        role = (access.get("role") or "").lower()
        if role not in {"admin", "conciliacion", "asistente_facturacion"}:
            raise HTTPException(403, "Tu usuario no tiene acceso a conciliación Gas LP.")
        if write and role not in {"admin", "conciliacion"}:
            raise HTTPException(403, "Tu rol no permite modificar conciliación.")
        requested_perfil_id = int(perfil_id or 0) or None
        perfil_id = requested_perfil_id or access.get("perfil_id")
        tenant_id = access.get("tenant_id") or _tenant_id_for_user(uid, access_token=token)
        if requested_perfil_id:
            try:
                rows = (
                    get_supabase_for_user(token)
                    .table("perfiles_empresa")
                    .select("id,tenant_id,descripcion")
                    .eq("id", requested_perfil_id)
                    .eq("user_id", uid)
                    .eq("activo", True)
                    .ilike("descripcion", "%[module:gas_lp]%")
                    .limit(1)
                    .execute()
                    .data
                    or []
                )
            except Exception as exc:
                logger.warning("gas_lp_conciliacion_requested_profile_lookup failed user=%s perfil=%s err=%s", uid, requested_perfil_id, exc)
                rows = []
            if not rows:
                raise HTTPException(403, "No tienes acceso a esa empresa Gas LP.")
            perfil_id = rows[0].get("id")
            tenant_id = rows[0].get("tenant_id") or tenant_id
        if not perfil_id:
            try:
                rows = (
                    get_supabase_for_user(token)
                    .table("perfiles_empresa")
                    .select("id,tenant_id,descripcion")
                    .eq("user_id", uid)
                    .eq("activo", True)
                    .ilike("descripcion", "%[module:gas_lp]%")
                    .limit(1)
                    .execute()
                    .data
                    or []
                )
                if rows:
                    perfil_id = rows[0].get("id")
                    tenant_id = tenant_id or rows[0].get("tenant_id")
            except Exception as exc:
                logger.warning("gas_lp_conciliacion_auth_profile_lookup failed user=%s err=%s", uid, exc)
        if not perfil_id:
            raise HTTPException(400, "Selecciona o asigna una empresa Gas LP antes de entrar a conciliación.")
        user = {
            "id": uid,
            "owner_user_id": uid,
            "tenant_id": tenant_id,
            "perfil_id": perfil_id,
            "section": "gas_lp",
            "role": role,
            "display_name": access.get("display_name") or "",
            "status": "active",
        }
        return {"session": {"section": "gas_lp", "role": role, "tenant_id": tenant_id, "perfil_id": perfil_id}, "user": user}
    ctx = _internal_session(token, "gas_lp")
    role = (ctx["user"].get("role") or "").lower()
    if role not in {"conciliacion", "admin", "asistente_facturacion"}:
        raise HTTPException(403, "Tu rol no permite acceder a conciliación.")
    if write and role not in {"conciliacion", "admin"}:
        raise HTTPException(403, "Tu rol no permite modificar conciliación.")
    if perfil_id and int(perfil_id) != int(ctx["user"].get("perfil_id") or 0):
        raise HTTPException(403, "Tu sesión interna sólo puede usar la empresa asignada.")
    return ctx


def _gas_lp_factura_access_context(token: str, *, write: bool = False, perfil_id: int | None = None) -> dict:
    if str(token or "").count(".") == 2:
        return _gas_lp_conciliacion_context(token, write=write, perfil_id=perfil_id)
    base = _internal_session(token, "gas_lp")
    role = (base["user"].get("role") or "").lower()
    if role in {"conciliacion", "admin"}:
        return _gas_lp_conciliacion_context(token, write=write, perfil_id=perfil_id)
    return _gas_lp_internal_context(token, write=write)


def _gas_lp_profile(user: dict, *, require_module_marker: bool = False) -> dict:
    rows = (
        get_supabase_admin()
        .table("perfiles_empresa")
        .select("id,user_id,tenant_id,nombre,rfc,descripcion,activo")
        .eq("id", user.get("perfil_id"))
        .eq("tenant_id", user.get("tenant_id"))
        .eq("activo", True)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows:
        raise HTTPException(403, "La empresa asignada al asistente no está activa o no existe.")
    profile = rows[0]
    if require_module_marker and "[module:gas_lp]" not in str(profile.get("descripcion") or "").lower():
        raise HTTPException(403, "La empresa asignada no pertenece al módulo Gas LP.")
    return profile


def _gas_lp_settings(owner_user_id: str, perfil_id: int) -> dict:
    from routes.settings import _load as load_settings

    return load_settings(owner_user_id, perfil_id)


def _gas_lp_internal_series(user: dict, settings: dict) -> str:
    configured = settings.get("SerieFacturaGasLp") or settings.get("serie_factura_gas_lp")
    if isinstance(settings.get("series_asistente_facturacion"), dict):
        configured = settings["series_asistente_facturacion"].get(str(user.get("id"))) or configured
    if configured:
        serie = str(configured).strip().upper()
    else:
        serie = f"P{user.get('perfil_id') or 0}U{user.get('id') or 0}"
    serie = "".join(ch for ch in serie if ch.isalnum())[:10]
    return serie or "AA"


def _folio_number(value: object) -> int:
    text = str(value or "").strip()
    return int(text) if text.isdigit() else 0


def _gas_lp_next_invoice_folio(sb, user: dict, serie: str) -> str:
    params = {
        "p_user_id": user.get("owner_user_id"),
        "p_tenant_id": user.get("tenant_id"),
        "p_perfil_id": user.get("perfil_id"),
        "p_serie": serie,
    }
    try:
        value = sb.rpc("next_gas_lp_invoice_folio", params).execute().data
        if isinstance(value, list):
            value = value[0] if value else 0
        number = int(value or 0)
        if number > 0:
            return f"{number:06d}"
    except Exception as exc:
        logger.warning("gas_lp_folio_counter_rpc_unavailable: serie=%s err=%s", serie, exc)

    try:
        rows = (
            sb.table("gas_lp_facturas")
            .select("record_uuid,metadata")
            .eq("user_id", user.get("owner_user_id"))
            .eq("tenant_id", user.get("tenant_id"))
            .eq("perfil_id", user.get("perfil_id"))
            .order("created_at", desc=True)
            .limit(500)
            .execute()
            .data
            or []
        )
    except Exception as exc:
        raise _safe_internal_error("gas_lp_next_folio", exc)
    current = 0
    for row in rows:
        md = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
        if str(md.get("serie") or "").strip().upper() != serie:
            continue
        current = max(current, _folio_number(md.get("folio_usuario")), _folio_number(row.get("record_uuid")))
    return f"{current + 1:06d}"


def _clean_rfc(value: str) -> str:
    return "".join(ch for ch in str(value or "").upper().strip() if ch.isalnum() or ch == "&")[:13]


def _clean_cp(value: str) -> str:
    cp = "".join(ch for ch in str(value or "").strip() if ch.isdigit())
    return cp[:5]


def _clean_billing_email(value: str | None) -> str:
    email = str(value or "").strip().lower()
    if not email:
        return ""
    if "@" not in email or " " in email or "," in email or ";" in email or "." not in email.rsplit("@", 1)[-1]:
        raise HTTPException(400, "Correo de facturación inválido.")
    return email


def _clean_billing_emails(value: str | None) -> list[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    emails: list[str] = []
    for part in raw.split(","):
        email = _clean_billing_email(part)
        if email and email not in emails:
            emails.append(email)
    return emails


def _invoice_email_recipients(
    primary: str | None,
    additional_1: str | None = "",
    additional_2: str | None = "",
    *,
    fallback: str | None = "",
) -> list[str]:
    raw_slots = [primary, additional_1, additional_2]
    if not any(str(slot or "").strip() for slot in raw_slots):
        raw_slots = [fallback]
    recipients: list[str] = []
    for raw in raw_slots:
        for email in _clean_billing_emails(raw):
            if email in recipients:
                raise HTTPException(400, "No puedes repetir correos de destinatario.")
            recipients.append(email)
            if len(recipients) > 3:
                raise HTTPException(400, "Máximo 3 correos por factura: 1 principal y 2 adicionales.")
    return recipients


def _invoice_email_metadata(recipients: list[str]) -> dict:
    primary = recipients[0] if recipients else ""
    additional = recipients[1:3]
    return {
        "cliente_email": primary,
        "email_recipients": recipients,
        "email_principal": primary,
        "email_adicional_1": additional[0] if len(additional) > 0 else "",
        "email_adicional_2": additional[1] if len(additional) > 1 else "",
        "email_adicionales": additional,
    }


def _saved_invoice_additional_emails(cliente_row: dict | None) -> list[str]:
    metadata = (cliente_row or {}).get("metadata")
    if not isinstance(metadata, dict):
        return []
    saved = metadata.get("invoice_email_additional") or metadata.get("email_adicionales")
    if isinstance(saved, list):
        return _invoice_email_recipients("", *(saved[:2] + ["", ""])[:2])
    return _invoice_email_recipients("", fallback=str(saved or ""))


def _customer_invoice_recipients(cliente_row: dict | None) -> list[str]:
    primary = (cliente_row or {}).get("email_facturacion") or (cliente_row or {}).get("email") or ""
    additional = _saved_invoice_additional_emails(cliente_row)
    return _invoice_email_recipients(
        primary,
        additional[0] if len(additional) > 0 else "",
        additional[1] if len(additional) > 1 else "",
    )


def _gas_lp_attach_cliente_email_recipients(sb, user: dict, rows: list[dict]) -> None:
    cliente_ids = sorted({
        int((row.get("metadata") or {}).get("cliente_id") or 0)
        for row in rows
        if isinstance(row.get("metadata"), dict) and (row.get("metadata") or {}).get("cliente_id")
    })
    if not cliente_ids:
        return
    try:
        clientes = (
            sb.table("gas_lp_clientes_facturacion")
            .select("*")
            .eq("user_id", user.get("owner_user_id"))
            .eq("tenant_id", user.get("tenant_id"))
            .eq("perfil_id", user.get("perfil_id"))
            .eq("activo", True)
            .execute()
            .data
            or []
        )
    except Exception as exc:
        logger.warning("gas_lp_attach_cliente_email_recipients_failed tenant=%s perfil=%s err=%s", user.get("tenant_id"), user.get("perfil_id"), exc)
        return
    by_id = {int(cliente.get("id") or 0): cliente for cliente in clientes}
    for row in rows:
        md = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
        cliente = by_id.get(int(md.get("cliente_id") or 0))
        if not cliente:
            continue
        try:
            recipients = _customer_invoice_recipients(cliente)
        except HTTPException:
            recipients = []
        row["cliente_email_recipients"] = recipients
        row["cliente_email_principal"] = recipients[0] if recipients else ""
        row["cliente_email_adicional_1"] = recipients[1] if len(recipients) > 1 else ""
        row["cliente_email_adicional_2"] = recipients[2] if len(recipients) > 2 else ""


def _transfer_email_from_settings(settings: dict) -> str:
    metadata = settings.get("metadata") if isinstance(settings.get("metadata"), dict) else {}
    for key in (
        "transfer_email",
        "traspaso_email",
        "operational_email",
        "billing_internal_email",
        "correo_traspaso",
        "correo_operativo",
    ):
        value = settings.get(key) or metadata.get(key)
        if str(value or "").strip():
            return str(value).strip()
    return ""


def _configured_setting(settings: dict, keys: tuple[str, ...]):
    for key in keys:
        value = settings.get(key)
        if value not in {None, ""}:
            return value, True
    return 0, False


def _save_transfer_email_default(owner_user_id: str, perfil_id: int, email: str) -> dict:
    from routes.settings import _load as load_settings, _save as save_settings_data

    clean_emails = _clean_billing_emails(email)
    if not clean_emails:
        raise HTTPException(400, "Captura un correo de traspaso válido para guardar como predeterminado.")
    email_text = ", ".join(clean_emails)
    current = load_settings(owner_user_id, perfil_id)
    metadata = current.get("metadata") if isinstance(current.get("metadata"), dict) else {}
    merged = {
        **current,
        "transfer_email": email_text,
        "metadata": {**metadata, "transfer_email": email_text},
    }
    save_settings_data(owner_user_id, merged, perfil_id)
    return merged


def _clean_clave_prod_serv(value: str | None) -> str:
    return "".join(ch for ch in str(value or "").strip() if ch.isdigit())[:8] or GAS_LP_CLAVE_PROD_SERV


def _infer_hyp_tipo_permiso_from_numero(numero_permiso: str) -> str:
    permiso = str(numero_permiso or "").strip().upper()
    if not permiso:
        return ""
    if permiso.startswith("LP/"):
        if "/DIST/PLA/" in permiso:
            return "PER06"
        if "/EXP/ES/" in permiso:
            return "PER01"
        if "/DIST/REP/" in permiso or "/DIST/AUT/" in permiso or "/EXP/AUT/" in permiso:
            return "PER06"
        if "/COM/" in permiso:
            return "PER06"
    if permiso.startswith("PL/"):
        if "/EXP/ES/MM/" in permiso:
            return "PER04"
        if "/EXP/ES/" in permiso:
            return "PER01"
        if "/DIS/OM/" in permiso:
            return "PER03"
    if permiso.startswith("H/") and "/COM/" in permiso:
        return "PER02"
    if permiso.startswith("CNE/PL/"):
        if "/EXP/ES/MM/" in permiso:
            return "PER08"
        if "/EXP/ES/" in permiso:
            return "PER05"
        if "/DIS/OM/" in permiso:
            return "PER07"
        if "/COM/" in permiso:
            return "PER09"
    if permiso.startswith("CNE/H/") and "/COM/" in permiso:
        return "PER06"
    return ""


def _sw_env_is_test() -> bool:
    return os.environ.get("SW_ENV", "test").strip().lower() not in {"prod", "production", "real"}


def _gas_lp_legacy_hyp_operation(facility: dict) -> str:
    permiso = str(facility.get("num_permiso") or "").upper()
    text = " ".join(
        str(facility.get(k) or "").lower()
        for k in ("tipo_instalacion", "tipo_permiso", "modalidad_permiso", "descripcion", "nombre", "actividad_sat")
    )
    if "/EXP/" in permiso or "expendio" in text or "estacion" in text:
        return "expendio"
    if "/COM/" in permiso or "comercial" in text:
        return "comercializacion"
    return "distribucion"


def _gas_lp_hyp_permit_from_facility(facility: dict) -> tuple[str, str]:
    md = facility.get("metadata") if isinstance(facility.get("metadata"), dict) else {}
    candidates = [
        md.get("hyp_num_permiso"),
        md.get("NumeroPermisoHYP"),
        md.get("permiso_hyp"),
        facility.get("hyp_num_permiso"),
        facility.get("numero_permiso_hyp"),
        facility.get("permiso_hyp"),
        facility.get("permiso_alm"),
        facility.get("num_permiso_hyp"),
        facility.get("num_permiso"),
    ]
    for value in candidates:
        numero = str(value or "").strip().upper()
        tipo = _infer_hyp_tipo_permiso_from_numero(numero)
        if tipo and tipo in HYP_TIPO_PERMISOS_VALIDOS:
            return tipo, numero

    type_candidates = [
        md.get("hyp_tipo_permiso"),
        md.get("TipoPermisoHYP"),
        facility.get("hyp_tipo_permiso"),
        facility.get("tipo_permiso_hyp"),
    ]
    for value in type_candidates:
        raw = str(value or "").strip().upper()
        if raw in HYP_TIPO_PERMISOS_VALIDOS:
            for value_num in candidates:
                numero = str(value_num or "").strip().upper()
                if _infer_hyp_tipo_permiso_from_numero(numero) == raw:
                    return raw, numero

    numero_real = str(facility.get("num_permiso") or "").strip().upper()
    if numero_real.startswith("LP/"):
        return "PER06", numero_real
    return "", ""


def _gas_lp_hyp_from_facility(facility: dict, clave_prod_serv: str) -> dict:
    clave = _clean_clave_prod_serv(clave_prod_serv)
    if clave not in GAS_LP_HIDRO_CLAVES:
        return {}
    tipo_permiso, numero_permiso = _gas_lp_hyp_permit_from_facility(facility)
    if not tipo_permiso or not numero_permiso:
        raise HTTPException(
            400,
            "La clave HYP Gas LP requiere ComplementoConcepto/HidroYPetro. "
            "Configura un permiso HyP/CNE compatible en la instalación. "
            "Para distribución usa una nomenclatura PL/.../DIS/OM/... o CNE/PL/.../DIS/OM/....",
        )
    if tipo_permiso not in HYP_TIPO_PERMISOS_VALIDOS:
        raise HTTPException(400, "El TipoPermiso HYP debe usar una clave SAT PER01-PER11.")
    return {
        "tipo_permiso": tipo_permiso,
        "numero_permiso": numero_permiso,
        "clave_hyp": clave,
        "subproducto_hyp": GAS_LP_HYP_SUBPRODUCTO,
    }


def _gas_lp_hyp_xml_fragment(hyp: dict) -> str:
    if not hyp:
        return ""
    return (
        '<hidrocarburospetroliferos:HidroYPetro '
        'Version="1.0" '
        f'TipoPermiso="{xml_escape(str(hyp.get("tipo_permiso") or ""))}" '
        f'NumeroPermiso="{xml_escape(str(hyp.get("numero_permiso") or ""))}" '
        f'ClaveHYP="{xml_escape(str(hyp.get("clave_hyp") or ""))}" '
        f'SubProductoHYP="{xml_escape(str(hyp.get("subproducto_hyp") or ""))}"/>'
    )


def _gas_lp_lcne_diagnostic_matrix(facility: dict, include_product_alternatives: bool = True) -> list[dict]:
    numero_real = str(facility.get("num_permiso") or "").strip().upper()
    product_keys = [GAS_LP_CLAVE_PROD_SERV]
    if include_product_alternatives:
        product_keys.append(GAS_LP_SW_HYP_EXPERIMENTAL_CLAVE)

    attempts: list[dict] = []
    if "/EXP/ES/" in numero_real:
        match = numero_real.replace("LP/", "", 1) if numero_real.startswith("LP/") else numero_real
        variants = [
            ("real_lp_per01", numero_real, "PER01"),
            ("real_lp_per05", numero_real, "PER05"),
            ("pl_exp_es_per01", f"PL/{match}", "PER01"),
            ("cne_pl_exp_es_per05", f"CNE/PL/{match}", "PER05"),
        ]
    elif "/DIST/PLA/" in numero_real:
        transformed = numero_real
        if numero_real.startswith("LP/"):
            transformed = numero_real.replace("LP/", "PL/", 1).replace("/DIST/PLA/", "/DIS/OM/")
        cne_transformed = transformed
        if transformed.startswith("PL/"):
            cne_transformed = f"CNE/{transformed}"
        variants = [
            ("real_lp_per06", numero_real, "PER06"),
            ("real_lp_per03", numero_real, "PER03"),
            ("pl_dis_om_per03", transformed, "PER03"),
            ("cne_pl_dis_om_per07", cne_transformed, "PER07"),
        ]
    else:
        inferred = _infer_hyp_tipo_permiso_from_numero(numero_real)
        variants = [("real_inferred", numero_real, inferred)] if numero_real and inferred else []

    for product_key in product_keys:
        for label, numero, tipo in variants:
            attempts.append(
                {
                    "label": label,
                    "numero_permiso": numero,
                    "tipo_permiso": tipo,
                    "clave_hyp": product_key,
                }
            )
    return attempts


def _write_gas_lp_hyp_debug_log(payload: dict) -> None:
    try:
        path = os.path.abspath(GAS_LP_HYP_DEBUG_LOG)
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "a", encoding="utf-8") as fh:
            fh.write(json.dumps(payload, ensure_ascii=False, sort_keys=True) + "\n")
    except Exception as exc:
        logger.warning("gas_lp_hyp_debug_log_failed: %s", exc)


def _sw_config_looks_like_sandbox(config: dict) -> bool:
    sw_env = str(config.get("sw_env") or "").strip().lower()
    if sw_env != "production":
        return True
    url_text = " ".join(
        str(config.get(key) or "").strip().lower()
        for key in ("base_url", "token_url", "xml_issue_url", "xml_stamp_url", "json_issue_url")
    )
    return any(marker in url_text for marker in ("test.sw.com.mx", "sandbox", "demo"))


def _gas_lp_hyp_mode() -> str:
    mode = os.environ.get("GAS_LP_HYP_MODE", "disabled").strip().lower()
    if mode not in GAS_LP_HYP_MODES:
        logger.warning("gas_lp_hyp_mode_invalid value=%s fallback=disabled", mode)
        return "disabled"
    return mode


def _require_gas_lp_issuer(profile: dict, settings: dict) -> dict:
    rfc = _clean_rfc(settings.get("RfcContribuyente") or profile.get("rfc") or "")
    name = str(settings.get("DescripcionInstalacion") or profile.get("nombre") or "").strip()
    cp = _clean_cp(settings.get("CodigoPostal") or settings.get("codigo_postal") or "")
    regimen = str(settings.get("RegimenFiscal") or settings.get("regimen_fiscal") or "601").strip()
    if not rfc or not name or not cp:
        raise HTTPException(
            400,
            "Configura RFC, nombre fiscal y código postal de la empresa antes de facturar.",
        )
    return {"rfc": rfc, "nombre": name, "cp": cp, "regimen": regimen or "601"}


def _public_general_receptor(issuer_cp: str) -> dict:
    return {
        "rfc": "XAXX010101000",
        "nombre": "PUBLICO EN GENERAL",
        "cp": issuer_cp,
        "regimen_fiscal": "616",
        "uso_cfdi": "S01",
    }


def _build_gas_lp_consumo_xml(
    *,
    issuer: dict,
    receptor: dict,
    litros,
    precio_unitario,
    concepto: str,
    forma_pago: str,
    metodo_pago: str,
    descuento=0,
    iva_rate=0.16,
    serie: str = "AA",
    folio: str = "",
    comentarios: str = "",
    fecha: str = "",
    clave_prod_serv: str = GAS_LP_CLAVE_PROD_SERV,
    hyp: Optional[dict] = None,
    informacion_global: Optional[dict] = None,
    no_identificacion: str = "GLP-LTR",
    unidad: str = "Litro",
    allow_zero_total: bool = False,
) -> tuple[str, dict]:
    qty = Decimal(str(litros or 0)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
    unit = Decimal(str(precio_unitario or 0)).quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)
    discount_unit = Decimal(str(descuento or 0)).quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)
    tax_rate = Decimal(str(iva_rate if iva_rate not in {None, ""} else 0.16)).quantize(Decimal("0.000000"), rounding=ROUND_HALF_UP)
    if qty <= 0 or (unit < 0 if allow_zero_total else unit <= 0):
        raise HTTPException(400, "Litros y precio unitario deben ser mayores a cero.")
    if discount_unit < 0 or discount_unit > unit:
        raise HTTPException(400, "El descuento por litro debe estar entre $0 y el precio por litro.")
    gross_total = _money(qty * unit)
    discount_gross = _money(qty * discount_unit)
    net_gross = _money(gross_total - discount_gross)
    divisor = Decimal("1.00") + tax_rate
    unit_net = _rate(unit / divisor) if tax_rate > 0 else unit
    subtotal = _money(gross_total / divisor) if tax_rate > 0 else gross_total
    discount_total = _money(discount_gross / divisor) if tax_rate > 0 else discount_gross
    taxable_base = _money(subtotal - discount_total)
    iva = _money(net_gross - taxable_base)
    total = net_gross
    fecha_cfdi, fecha_reemplazada, fecha_reason = _gas_lp_cfdi_fecha_actualizada(fecha, issuer.get("cp"))
    folio_dt = _parse_gas_lp_cfdi_fecha(fecha_cfdi, _gas_lp_cfdi_timezone(issuer.get("cp")))
    folio = (str(folio or "").strip() or (folio_dt or datetime.now(_gas_lp_cfdi_timezone(issuer.get("cp")))).strftime("GLP%Y%m%d%H%M%S"))[:40]
    serie = (str(serie or "AA").strip() or "AA")[:10]
    fecha = fecha_cfdi[:19]
    desc = concepto.strip() or "Venta de Gas LP"
    comments = str(comentarios or "").strip()[:500]
    descuento_comprobante = f' Descuento="{discount_total:.2f}"' if discount_total > 0 else ""
    descuento_concepto = f' Descuento="{discount_total:.2f}"' if discount_total > 0 else ""
    if total <= 0 and not allow_zero_total:
        raise HTTPException(400, "El total de la factura debe ser mayor a cero. Revisa precio y descuento.")
    clave_prod_serv = _clean_clave_prod_serv(clave_prod_serv)
    hyp = hyp or {}
    hyp_ns = ' xmlns:hidrocarburospetroliferos="http://www.sat.gob.mx/hidrocarburospetroliferos"' if hyp else ""
    hyp_schema = " http://www.sat.gob.mx/hidrocarburospetroliferos http://www.sat.gob.mx/sitio_internet/cfd/hidrocarburospetroliferos.xsd" if hyp else ""
    hyp_xml = ""
    if hyp:
        hyp_xml = (
            '<cfdi:ComplementoConcepto>'
            f'{_gas_lp_hyp_xml_fragment(hyp)}'
            '</cfdi:ComplementoConcepto>'
        )
    info_global_xml = ""
    informacion_global = informacion_global or {}
    if informacion_global:
        try:
            fecha_base = datetime.strptime(fecha[:10], "%Y-%m-%d")
        except Exception:
            fecha_base = datetime.now()
        periodicidad = str(informacion_global.get("periodicidad") or "04").strip()[:2] or "04"
        meses = "".join(ch for ch in str(informacion_global.get("meses") or f"{fecha_base.month:02d}") if ch.isdigit())[:2]
        anio = "".join(ch for ch in str(informacion_global.get("anio") or fecha_base.year) if ch.isdigit())[:4]
        meses = meses.zfill(2)
        if len(anio) != 4:
            raise HTTPException(400, "El año de InformaciónGlobal debe usar 4 dígitos.")
        info_global_xml = (
            f'<cfdi:InformacionGlobal Periodicidad="{xml_escape(periodicidad)}" '
            f'Meses="{xml_escape(meses)}" Año="{xml_escape(anio)}"/>'
        )
    no_identificacion = str(no_identificacion or folio).strip()[:100] or folio
    unidad = str(unidad or "Litro").strip()[:20] or "Litro"
    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        f'<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/4"{hyp_ns} '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
        f'xsi:schemaLocation="http://www.sat.gob.mx/cfd/4 http://www.sat.gob.mx/sitio_internet/cfd/4/cfdv40.xsd{hyp_schema}" '
        f'Version="4.0" Serie="{xml_escape(serie)}" Folio="{xml_escape(folio)}" Fecha="{fecha}" FormaPago="{xml_escape(forma_pago or "99")}" '
        f'NoCertificado="" Certificado="" Sello="" SubTotal="{subtotal:.2f}"{descuento_comprobante} Moneda="MXN" Total="{total:.2f}" '
        f'TipoDeComprobante="I" Exportacion="01" MetodoPago="{xml_escape(metodo_pago or "PUE")}" LugarExpedicion="{issuer["cp"]}">'
        f'{info_global_xml}'
        f'<cfdi:Emisor Rfc="{issuer["rfc"]}" Nombre="{xml_escape(issuer["nombre"])}" RegimenFiscal="{issuer["regimen"]}"/>'
        f'<cfdi:Receptor Rfc="{receptor["rfc"]}" Nombre="{xml_escape(receptor["nombre"])}" '
        f'DomicilioFiscalReceptor="{receptor["cp"]}" RegimenFiscalReceptor="{receptor["regimen_fiscal"]}" UsoCFDI="{receptor["uso_cfdi"]}"/>'
        '<cfdi:Conceptos>'
        f'<cfdi:Concepto ClaveProdServ="{clave_prod_serv}" NoIdentificacion="{xml_escape(no_identificacion)}" Cantidad="{qty:.3f}" '
        f'ClaveUnidad="LTR" Unidad="Litro" Descripcion="{xml_escape(desc)}" ValorUnitario="{unit_net:.6f}" '
        f'Importe="{subtotal:.2f}"{descuento_concepto} ObjetoImp="02">'
        '<cfdi:Impuestos><cfdi:Traslados>'
        f'<cfdi:Traslado Base="{taxable_base:.2f}" Impuesto="002" TipoFactor="Tasa" TasaOCuota="{tax_rate:.6f}" Importe="{iva:.2f}"/>'
        '</cfdi:Traslados></cfdi:Impuestos>'
        f'{hyp_xml}'
        '</cfdi:Concepto>'
        '</cfdi:Conceptos>'
        f'<cfdi:Impuestos TotalImpuestosTrasladados="{iva:.2f}"><cfdi:Traslados>'
        f'<cfdi:Traslado Base="{taxable_base:.2f}" Impuesto="002" TipoFactor="Tasa" TasaOCuota="{tax_rate:.6f}" Importe="{iva:.2f}"/>'
        '</cfdi:Traslados></cfdi:Impuestos>'
        f'{f"<cfdi:Addenda><Observaciones>{xml_escape(comments)}</Observaciones></cfdi:Addenda>" if comments else ""}'
        '</cfdi:Comprobante>'
    )
    return xml, {
        "folio": folio,
        "fecha": fecha,
        "fecha_cfdi_reemplazada": fecha_reemplazada,
        "fecha_cfdi_reason": fecha_reason,
        "subtotal": float(subtotal),
        "descuento": float(discount_total),
        "descuento_con_iva": float(discount_gross),
        "iva": float(iva),
        "total": float(total),
        "precio_unitario_con_iva": float(unit),
        "precio_unitario_sin_iva": float(unit_net),
    }


def _xml_local(tag: str) -> str:
    return str(tag or "").split("}", 1)[-1]


def _xml_first(root: ET.Element, local_name: str) -> Optional[ET.Element]:
    for elem in root.iter():
        if _xml_local(elem.tag) == local_name:
            return elem
    return None


def _xml_attr(elem: Optional[ET.Element], name: str, default: str = "") -> str:
    return str(elem.attrib.get(name) or default) if elem is not None else default


def _cfdi_root(xml_content: str) -> ET.Element:
    try:
        return ET.fromstring(str(xml_content or "").encode("utf-8"))
    except Exception as exc:
        raise HTTPException(400, f"XML base inválido para complemento de pago: {exc}") from exc


def _payment_datetime(value: str) -> str:
    raw = str(value or "").strip().replace("Z", "").replace(" ", "T")
    if raw:
        if len(raw) == 16:
            raw += ":00"
        return raw[:19]
    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")


def _factura_payment_info(factura: dict) -> dict:
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    return {
        "metodo_pago": str(md.get("metodo_pago") or "").upper(),
        "forma_pago": str(md.get("forma_pago") or ""),
        "total": _money(md.get("total") or (Decimal(str(factura.get("importe") or 0)) * Decimal("1.16"))),
        "saldo_insoluto": _money(md.get("saldo_insoluto") if md.get("saldo_insoluto") not in {None, ""} else (md.get("total") or (Decimal(str(factura.get("importe") or 0)) * Decimal("1.16")))),
        "payment_status": str(md.get("payment_status") or ("pendiente_complemento" if str(md.get("metodo_pago") or "").upper() == "PPD" else "pagado_pue")),
    }


def _payment_info_json(info: dict) -> dict:
    return {
        **info,
        "total": float(_money(info.get("total"))),
        "saldo_insoluto": float(_money(info.get("saldo_insoluto"))),
    }


def _gas_lp_factura_date_key(factura: dict) -> str:
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    for value in (md.get("fecha_emision"), md.get("fecha_cfdi"), factura.get("fecha_timbrado"), factura.get("created_at")):
        text = str(value or "")
        if len(text) >= 10:
            return text[:10]
    xml_content = str(factura.get("xml_content") or "")
    if xml_content:
        try:
            return _xml_attr(ET.fromstring(xml_content.encode("utf-8")), "Fecha")[:10]
        except Exception:
            return ""
    return ""


def _gas_lp_factura_xml_root(factura: dict) -> Optional[ET.Element]:
    xml_content = str(factura.get("xml_content") or "")
    if not xml_content:
        return None
    try:
        return ET.fromstring(xml_content.encode("utf-8"))
    except Exception:
        return None


def _gas_lp_factura_pac_xml_summary(xml_content: str) -> dict:
    root = None
    try:
        root = ET.fromstring(str(xml_content or "").encode("utf-8"))
    except Exception:
        return {"xml_len": len(str(xml_content or "")), "parse_ok": False}
    concepto = _xml_first(root, "Concepto")
    timbre = _xml_first(root, "TimbreFiscalDigital")
    return {
        "xml_len": len(str(xml_content or "")),
        "parse_ok": True,
        "version": _xml_attr(root, "Version"),
        "serie": _xml_attr(root, "Serie"),
        "folio": _xml_attr(root, "Folio"),
        "fecha": _xml_attr(root, "Fecha"),
        "no_certificado": _xml_attr(root, "NoCertificado"),
        "has_sello": bool(_xml_attr(root, "Sello")),
        "sello_len": len(_xml_attr(root, "Sello")),
        "has_certificado": bool(_xml_attr(root, "Certificado")),
        "certificado_len": len(_xml_attr(root, "Certificado")),
        "clave_prod_serv": _xml_attr(concepto, "ClaveProdServ"),
        "clave_unidad": _xml_attr(concepto, "ClaveUnidad"),
        "unidad": _xml_attr(concepto, "Unidad"),
        "descripcion": _xml_attr(concepto, "Descripcion"),
        "has_hidroypetro": "HidroYPetro" in str(xml_content or ""),
        "has_informacion_global": _xml_first(root, "InformacionGlobal") is not None,
        "uuid": _xml_attr(timbre, "UUID"),
        "rfc_prov_certif": _xml_attr(timbre, "RfcProvCertif"),
        "fecha_timbrado": _xml_attr(timbre, "FechaTimbrado"),
    }


def _gas_lp_factura_folio_label(factura: dict) -> str:
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    serie = str(md.get("serie") or "").strip()
    folio = str(md.get("folio_usuario") or md.get("folio") or factura.get("record_uuid") or "").strip()
    root = _gas_lp_factura_xml_root(factura)
    if root is not None:
        serie = serie or _xml_attr(root, "Serie")
        folio = folio or _xml_attr(root, "Folio")
    label = f"{serie}{folio}" if serie and folio and not str(folio).startswith(serie) else (folio or serie)
    return label or str(factura.get("id") or "")


def _gas_lp_factura_observaciones(factura: dict) -> str:
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    return str(md.get("comentarios") or md.get("observaciones") or "").strip()


def _gas_lp_factura_razon_social(factura: dict) -> str:
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    if md.get("cliente_nombre"):
        return str(md.get("cliente_nombre") or "")
    root = _gas_lp_factura_xml_root(factura)
    receptor = _xml_first(root, "Receptor") if root is not None else None
    return _xml_attr(receptor, "Nombre") or str(factura.get("rfc_receptor") or "")


def _gas_lp_factura_emisor_rfc(factura: dict) -> str:
    root = _gas_lp_factura_xml_root(factura)
    emisor = _xml_first(root, "Emisor") if root is not None else None
    rfc_xml = _clean_rfc(_xml_attr(emisor, "Rfc"))
    if rfc_xml:
        return rfc_xml
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    rfc = _clean_rfc(md.get("rfc_emisor") or md.get("empresa_rfc") or factura.get("rfc_emisor") or "")
    return rfc


def _gas_lp_factura_emisor_nombre(factura: dict) -> str:
    root = _gas_lp_factura_xml_root(factura)
    emisor = _xml_first(root, "Emisor") if root is not None else None
    nombre_xml = str(_xml_attr(emisor, "Nombre") or "").strip()
    if nombre_xml:
        return nombre_xml
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    return str(
        md.get("nombre_emisor")
        or md.get("empresa_nombre")
        or md.get("empresa_asignada_nombre")
        or ""
    ).strip()


def _gas_lp_factura_matches_company(factura: dict, user: dict, profile: dict) -> bool:
    profile_rfc = _clean_rfc(profile.get("rfc") or "")
    factura_rfc = _gas_lp_factura_emisor_rfc(factura)
    if profile_rfc and factura_rfc == profile_rfc:
        return True
    same_scope = str(factura.get("tenant_id") or "") == str(user.get("tenant_id") or "") and str(factura.get("perfil_id") or "") == str(user.get("perfil_id") or "")
    return bool(same_scope and (not profile_rfc or not factura_rfc))


def _gas_lp_factura_visibility_reason(factura: dict, user: dict, profile: dict, month: str) -> str:
    profile_rfc = _clean_rfc(profile.get("rfc") or "")
    factura_rfc = _gas_lp_factura_emisor_rfc(factura)
    same_scope = (
        str(factura.get("tenant_id") or "") == str(user.get("tenant_id") or "")
        and str(factura.get("perfil_id") or "") == str(user.get("perfil_id") or "")
    )
    same_rfc = bool(profile_rfc and factura_rfc == profile_rfc)
    date_key = _gas_lp_factura_date_key(factura)
    reasons = []
    if profile_rfc and factura_rfc and not same_rfc:
        reasons.append(f"issuer_rfc_mismatch: rfc_emisor={factura_rfc}")
    elif not same_scope and not same_rfc:
        reasons.append(
            "company_mismatch"
            f": tenant={factura.get('tenant_id') or ''} perfil={factura.get('perfil_id') or ''}"
            f" rfc_emisor={factura_rfc or ''}"
        )
    if month and not date_key.startswith(month):
        reasons.append(f"month_mismatch: factura_date={date_key or ''}")
    return "; ".join(reasons) or "included"


def _gas_lp_company_rfc(user: dict, profile: dict) -> str:
    rfc = _clean_rfc(profile.get("rfc") or "")
    if rfc:
        return rfc
    try:
        settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
        return _clean_rfc(settings.get("RfcContribuyente") or settings.get("rfc") or "")
    except Exception:
        return ""


def _dedupe_rows_by_id(rows: list[dict]) -> list[dict]:
    seen: set[str] = set()
    deduped: list[dict] = []
    for row in rows:
        rid = str(row.get("id") or "").strip()
        if rid and rid in seen:
            continue
        if rid:
            seen.add(rid)
        deduped.append(row)
    return deduped


def _safe_int_id(value: object) -> int:
    try:
        return int(value or 0)
    except Exception:
        return 0


def _gas_lp_factura_log_row(row: dict, reason: str) -> dict:
    md = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
    return {
        "id": row.get("id"),
        "uuid_sat": row.get("uuid_sat") or row.get("record_uuid") or md.get("uuid_sat") or md.get("uuid"),
        "tenant_id": row.get("tenant_id"),
        "perfil_id": row.get("perfil_id"),
        "user_id": row.get("user_id"),
        "created_by": md.get("created_by") or md.get("created_by_internal_name") or md.get("internal_user_id"),
        "rfc_emisor": row.get("rfc_emisor") or md.get("rfc_emisor") or md.get("empresa_rfc"),
        "rfc_receptor": row.get("rfc_receptor") or md.get("cliente_rfc"),
        "fecha_factura": _gas_lp_factura_date_key(row),
        "fecha_timbrado": row.get("fecha_timbrado"),
        "created_at": row.get("created_at"),
        "status": row.get("status"),
        "reason": reason,
    }


def _gas_lp_log_facturas_visibility(
    *,
    user: dict,
    profile: dict,
    month: str,
    filters: list[dict],
    candidates: list[dict],
    included: list[dict],
) -> None:
    try:
        included_ids = {str(row.get("id") or "").strip() for row in included if row.get("id")}
        excluded = [
            _gas_lp_factura_log_row(row, _gas_lp_factura_visibility_reason(row, user, profile, month))
            for row in candidates
            if str(row.get("id") or "").strip() not in included_ids
        ]
        logger.info(
            "gas_lp_facturas_visibility %s",
            json.dumps(
                {
                    "usuario_actual": {
                        "id": user.get("id"),
                        "nombre": user.get("display_name"),
                        "role": user.get("role"),
                        "tenant_id": user.get("tenant_id"),
                        "perfil_id": user.get("perfil_id"),
                        "owner_user_id": user.get("owner_user_id"),
                    },
                    "empresa_asignada": {
                        "id": profile.get("id"),
                        "nombre": profile.get("nombre"),
                        "rfc": _gas_lp_company_rfc(user, profile),
                        "tenant_id": profile.get("tenant_id"),
                    },
                    "mes_solicitado": month or "",
                    "query_sql_filtros_aplicados": filters,
                    "facturas_candidatas_antes_de_filtros": len(candidates),
                    "facturas_encontradas_despues_de_filtros": len(included),
                    "incluidas": [_gas_lp_factura_log_row(row, "included") for row in included],
                    "excluidas": excluded,
                },
                ensure_ascii=False,
                sort_keys=True,
            ),
        )
    except Exception as exc:
        logger.warning("gas_lp_facturas_visibility_log_failed: %s", exc)


def _gas_lp_company_facturas_rows(
    sb,
    user: dict,
    profile: dict,
    *,
    month: str = "",
    limit: int = 10000,
) -> list[dict]:
    filters = []
    candidate_rows: list[dict] = []
    base_query = sb.table("gas_lp_facturas").select("*").eq("tenant_id", user.get("tenant_id")).eq("perfil_id", user.get("perfil_id"))
    filters.append({"source": "tenant_perfil", "tenant_id": user.get("tenant_id"), "perfil_id": user.get("perfil_id"), "date_filter": "python:fecha_emision|fecha_cfdi|fecha_timbrado|created_at|xml.Fecha"})
    rows = base_query.order("created_at", desc=True).limit(limit).execute().data or []
    candidate_rows.extend(rows)

    profile_rfc = _gas_lp_company_rfc(user, profile)
    match_profile = {**profile, "rfc": profile_rfc or profile.get("rfc")}
    if profile_rfc:
        # Rescata facturas timbradas por otro usuario/perfil de la misma empresa fiscal.
        rfc_rows = []
        try:
            rfc_rows = (
                sb.table("gas_lp_facturas")
                .select("*")
                .eq("tenant_id", user.get("tenant_id"))
                .eq("metadata->>rfc_emisor", profile_rfc)
                .order("created_at", desc=True)
                .limit(limit)
                .execute()
                .data
                or []
            )
            filters.append({"source": "tenant_metadata_rfc_emisor", "tenant_id": user.get("tenant_id"), "metadata.rfc_emisor": profile_rfc, "date_filter": "python:fecha_emision|fecha_cfdi|fecha_timbrado|created_at|xml.Fecha"})
        except Exception as exc:
            logger.warning("gas_lp_facturas_metadata_rfc_lookup_failed rfc=%s err=%s", profile_rfc, exc)
            filters.append({"source": "tenant_metadata_rfc_emisor", "tenant_id": user.get("tenant_id"), "metadata.rfc_emisor": profile_rfc, "error": str(exc)})
        candidate_rows.extend(rfc_rows)

        tenant_scan_limit = max(limit, int(os.environ.get("GAS_LP_FACTURAS_TENANT_SCAN_LIMIT", "10000") or "10000"))
        try:
            company_rows = (
                sb.table("gas_lp_facturas")
                .select("*")
                .eq("tenant_id", user.get("tenant_id"))
                .order("created_at", desc=True)
                .limit(tenant_scan_limit)
                .execute()
                .data
                or []
            )
            filters.append({"source": "tenant_scan_rfc_fallback", "tenant_id": user.get("tenant_id"), "limit": tenant_scan_limit, "match": "same tenant/perfil OR same issuer RFC from row/metadata/xml", "date_filter": "python:fecha_emision|fecha_cfdi|fecha_timbrado|created_at|xml.Fecha"})
        except Exception as exc:
            company_rows = []
            logger.warning("gas_lp_facturas_tenant_scan_failed tenant=%s err=%s", user.get("tenant_id"), exc)
            filters.append({"source": "tenant_scan_rfc_fallback", "tenant_id": user.get("tenant_id"), "limit": tenant_scan_limit, "error": str(exc)})
        candidate_rows.extend(company_rows)
        rows.extend(row for row in rfc_rows if _gas_lp_factura_matches_company(row, user, match_profile))
        rows.extend(row for row in company_rows if _gas_lp_factura_matches_company(row, user, match_profile))

    rows = _dedupe_rows_by_id(rows)
    rows = [row for row in rows if _gas_lp_factura_matches_company(row, user, match_profile)]
    candidate_rows = _dedupe_rows_by_id(candidate_rows)
    if month:
        rows = [row for row in rows if _gas_lp_factura_date_key(row).startswith(month)]
    rows.sort(key=lambda row: str(row.get("created_at") or ""), reverse=True)
    rows = rows[:limit]
    _gas_lp_log_facturas_visibility(user=user, profile={**profile, "rfc": profile_rfc or profile.get("rfc")}, month=month, filters=filters, candidates=candidate_rows, included=rows)
    return rows


def _gas_lp_factura_metodo_pago(factura: dict) -> str:
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    info = _factura_payment_info(factura)
    method = str(info.get("metodo_pago") or md.get("metodo_pago") or "").upper()
    if method in {"PUE", "PPD"}:
        return method
    root = _gas_lp_factura_xml_root(factura)
    return _xml_attr(root, "MetodoPago").upper() if root is not None else method


def _gas_lp_factura_cancelada(factura: dict) -> bool:
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    markers = [
        factura.get("status"),
        md.get("status"),
        md.get("estado_fiscal"),
        md.get("cancelacion_estado_fiscal_label"),
        md.get("cancelacion_status"),
    ]
    for marker in markers:
        text = str(marker or "").strip().lower()
        if any(value in text for value in ("cancelada", "cancelado", "cancelled", "canceled")):
            return True
    return False


def _gas_lp_factura_estado_excel(factura: dict) -> str:
    if _gas_lp_factura_cancelada(factura):
        return "Cancelada"
    try:
        metodo = _gas_lp_factura_metodo_pago(factura)
    except Exception:
        metodo = ""
    return "Vigente - PPD / Crédito" if str(metodo or "").upper() == "PPD" else "Vigente"


def _gas_lp_factura_total_con_iva(factura: dict) -> Decimal:
    root = _gas_lp_factura_xml_root(factura)
    if root is not None:
        total_xml = _xml_attr(root, "Total")
        if total_xml:
            return _money(total_xml)
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    if md.get("total") not in {None, ""}:
        return _money(md.get("total"))
    return _money(Decimal(str(factura.get("importe") or 0)) * Decimal("1.16"))


def _gas_lp_attach_internal_creators(sb, rows: list[dict]) -> None:
    ids = sorted({
        int((row.get("metadata") or {}).get("internal_user_id") or 0)
        for row in rows
        if isinstance(row.get("metadata"), dict) and (row.get("metadata") or {}).get("internal_user_id")
    })
    if not ids:
        return
    try:
        users = (
            sb.table("internal_users")
            .select("id,display_name,code")
            .in_("id", ids)
            .execute()
            .data
            or []
        )
    except Exception:
        users = []
    by_id = {int(user.get("id") or 0): user for user in users}
    for row in rows:
        md = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
        internal_id = int(md.get("internal_user_id") or 0)
        internal_user = by_id.get(internal_id)
        if internal_user:
            row["created_by_internal"] = {
                "id": internal_id,
                "name": internal_user.get("display_name") or internal_user.get("code") or "Asistente",
            }


def _gas_lp_complementos_por_factura(sb, factura_ids: list[int]) -> dict[int, list[dict]]:
    ids = [int(fid) for fid in factura_ids if fid]
    by_factura: dict[int, list[dict]] = {fid: [] for fid in ids}
    if not ids:
        return by_factura
    try:
        rows = (
            sb.table("gas_lp_complementos_pago_facturas")
            .select("*")
            .in_("factura_id", ids)
            .eq("status", "timbrado")
            .order("created_at", desc=True)
            .execute()
            .data
            or []
        )
    except Exception:
        rows = []
    for row in rows:
        fid = int(row.get("factura_id") or 0)
        if fid in by_factura:
            by_factura[fid].append(row)
    return by_factura


def _gas_lp_internal_factura(user: dict, factura_id: int) -> dict:
    profile = _gas_lp_profile(user)
    match_profile = {**profile, "rfc": _gas_lp_company_rfc(user, profile)}
    rows = (
        get_supabase_admin()
        .table("gas_lp_facturas")
        .select("*")
        .eq("id", factura_id)
        .eq("tenant_id", user.get("tenant_id"))
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows or not _gas_lp_factura_matches_company(rows[0], user, match_profile):
        raise HTTPException(404, "Factura no encontrada para esta empresa.")
    return rows[0]


def _build_gas_lp_pago20_multi_xml(*, facturas: list[dict], issuer: dict, fecha_pago: str, forma_pago: str, pagos: dict[int, Decimal]) -> tuple[str, dict]:
    if not facturas:
        raise HTTPException(400, "Selecciona al menos una factura PPD.")
    receptor_ref: dict | None = None
    doctos = []
    total_pago = Decimal("0.00")
    total_base = Decimal("0.00")
    total_iva = Decimal("0.00")
    for factura in facturas:
        root = _cfdi_root(factura.get("xml_content") or "")
        if _xml_attr(root, "TipoDeComprobante") != "I" or _xml_attr(root, "MetodoPago") != "PPD":
            raise HTTPException(400, "Solo se pueden relacionar facturas de ingreso PPD.")
        timbre = _xml_first(root, "TimbreFiscalDigital")
        uuid_rel = _xml_attr(timbre, "UUID") or factura.get("uuid_sat") or ""
        receptor_node = _xml_first(root, "Receptor")
        receptor = {
            "rfc": _xml_attr(receptor_node, "Rfc"),
            "nombre": _xml_attr(receptor_node, "Nombre"),
            "cp": _xml_attr(receptor_node, "DomicilioFiscalReceptor"),
            "regimen": _xml_attr(receptor_node, "RegimenFiscalReceptor"),
        }
        if not uuid_rel or not all(receptor.values()):
            raise HTTPException(400, "Una factura seleccionada no tiene UUID o receptor completo.")
        if receptor_ref is None:
            receptor_ref = receptor
        elif receptor_ref["rfc"] != receptor["rfc"]:
            raise HTTPException(400, "Selecciona facturas del mismo cliente/RFC.")
        fid = int(factura["id"])
        info = _factura_payment_info(factura)
        pagado = _money(pagos[fid])
        saldo_ant = _money(info["saldo_insoluto"])
        if pagado <= 0 or pagado > saldo_ant:
            raise HTTPException(400, "El importe de pago debe ser mayor a cero y no exceder el saldo.")
        saldo = _money(saldo_ant - pagado)
        total_doc = _money(_xml_attr(root, "Total") or info["total"])
        base = _money(pagado / Decimal("1.16"))
        iva = _money(pagado - base)
        total_pago += pagado
        total_base += base
        total_iva += iva
        serie = _xml_attr(root, "Serie")
        folio = _xml_attr(root, "Folio")
        serie_attr = f' Serie="{xml_escape(serie)}"' if serie else ""
        folio_attr = f' Folio="{xml_escape(folio)}"' if folio else ""
        doctos.append({
            "factura_id": fid,
            "uuid_relacionado": uuid_rel,
            "monto": float(pagado),
            "saldo_anterior": float(saldo_ant),
            "saldo_insoluto": float(saldo),
            "parcialidad": 1,
            "xml": (
                f'<pago20:DoctoRelacionado IdDocumento="{xml_escape(uuid_rel)}"{serie_attr}{folio_attr} MonedaDR="MXN" EquivalenciaDR="1" '
                f'NumParcialidad="1" ImpSaldoAnt="{saldo_ant:.2f}" ImpPagado="{pagado:.2f}" ImpSaldoInsoluto="{saldo:.2f}" ObjetoImpDR="02">'
                '<pago20:ImpuestosDR><pago20:TrasladosDR>'
                f'<pago20:TrasladoDR BaseDR="{base:.2f}" ImpuestoDR="002" TipoFactorDR="Tasa" TasaOCuotaDR="0.160000" ImporteDR="{iva:.2f}"/>'
                '</pago20:TrasladosDR></pago20:ImpuestosDR></pago20:DoctoRelacionado>'
            ),
        })
        if total_doc <= 0:
            raise HTTPException(400, "Una factura seleccionada no tiene total válido.")
    receptor = receptor_ref or {}
    fecha_cfdi = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    folio_pago = datetime.now().strftime("%Y%m%d%H%M%S")
    fecha_pago = _payment_datetime(fecha_pago)
    forma_pago = "".join(ch for ch in str(forma_pago or "03") if ch.isdigit())[:2] or "03"
    doctos_xml = "".join(d["xml"] for d in doctos)
    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/4" xmlns:pago20="http://www.sat.gob.mx/Pagos20" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
        'xsi:schemaLocation="http://www.sat.gob.mx/cfd/4 http://www.sat.gob.mx/sitio_internet/cfd/4/cfdv40.xsd http://www.sat.gob.mx/Pagos20 http://www.sat.gob.mx/sitio_internet/cfd/Pagos/Pagos20.xsd" '
        f'Version="4.0" Serie="PAGO" Folio="{folio_pago}" Fecha="{fecha_cfdi}" Sello="" NoCertificado="" Certificado="" SubTotal="0" Moneda="XXX" Total="0" TipoDeComprobante="P" Exportacion="01" LugarExpedicion="{xml_escape(issuer["cp"])}">'
        f'<cfdi:Emisor Rfc="{xml_escape(issuer["rfc"])}" Nombre="{xml_escape(issuer["nombre"])}" RegimenFiscal="{xml_escape(issuer["regimen"])}"/>'
        f'<cfdi:Receptor Rfc="{xml_escape(receptor["rfc"])}" Nombre="{xml_escape(receptor["nombre"])}" DomicilioFiscalReceptor="{xml_escape(receptor["cp"])}" RegimenFiscalReceptor="{xml_escape(receptor["regimen"])}" UsoCFDI="CP01"/>'
        '<cfdi:Conceptos><cfdi:Concepto ClaveProdServ="84111506" Cantidad="1" ClaveUnidad="ACT" Descripcion="Pago" ValorUnitario="0" Importe="0" ObjetoImp="01"/></cfdi:Conceptos>'
        '<cfdi:Complemento><pago20:Pagos Version="2.0">'
        f'<pago20:Totales TotalTrasladosBaseIVA16="{total_base:.2f}" TotalTrasladosImpuestoIVA16="{total_iva:.2f}" MontoTotalPagos="{total_pago:.2f}"/>'
        f'<pago20:Pago FechaPago="{xml_escape(fecha_pago)}" FormaDePagoP="{xml_escape(forma_pago)}" MonedaP="MXN" TipoCambioP="1" Monto="{total_pago:.2f}">'
        f'{doctos_xml}<pago20:ImpuestosP><pago20:TrasladosP><pago20:TrasladoP BaseP="{total_base:.2f}" ImpuestoP="002" TipoFactorP="Tasa" TasaOCuotaP="0.160000" ImporteP="{total_iva:.2f}"/></pago20:TrasladosP></pago20:ImpuestosP>'
        '</pago20:Pago></pago20:Pagos></cfdi:Complemento></cfdi:Comprobante>'
    )
    for d in doctos:
        d.pop("xml", None)
    return xml, {"fecha_pago": fecha_pago, "forma_pago": forma_pago, "monto": float(total_pago), "saldo_insoluto": float(sum(Decimal(str(d["saldo_insoluto"])) for d in doctos)), "facturas": doctos}


def _gas_lp_invoice_scope(user: dict, profile: dict) -> dict:
    return {
        "user_id": user.get("owner_user_id"),
        "tenant_id": user.get("tenant_id"),
        "perfil_id": user.get("perfil_id"),
        "source": "assistant_portal",
        "updated_at": _now_iso(),
    }


def _auth_admin(authorization: str) -> tuple[str, str]:
    if not authorization.startswith("Bearer "):
        raise HTTPException(401, "No autenticado.")
    token = authorization[7:]
    uid = verify_token(token)
    if not uid:
        raise HTTPException(401, "Token inválido o expirado.")
    accesos = [
        obtener_acceso_modulo(uid, "transporte", access_token=token),
        obtener_acceso_modulo(uid, "gas_lp", access_token=token),
        obtener_acceso_modulo(uid, "gasolineras", access_token=token),
    ]
    if not any((a.get("role") or "").lower() == "admin" for a in accesos):
        raise HTTPException(403, "Solo administradores pueden gestionar usuarios internos.")
    return uid, token


def _clean_payload(payload: InternalUserCreate) -> tuple[str, str, str]:
    section = (payload.section or "").strip().lower()
    role = (payload.role or "").strip().lower()
    name = (payload.display_name or "").strip()
    if section not in SECTIONS:
        raise HTTPException(400, "Módulo inválido.")
    if role not in ROLES:
        raise HTTPException(400, "Rol inválido.")
    if not name:
        raise HTTPException(400, "Nombre requerido.")
    if payload.perfil_id <= 0:
        raise HTTPException(400, "perfil_id requerido.")
    if section == "transporte" and role == "operador" and not payload.chofer_id:
        raise HTTPException(400, "El operador de Transporte debe vincularse con un chofer.")
    if section == "gas_lp":
        if not _clean_code(payload.code or ""):
            raise HTTPException(400, "El usuario de asistente Gas LP es obligatorio.")
        if not (payload.pin or "").strip():
            raise HTTPException(400, "La contraseña de asistente Gas LP es obligatoria.")
    return name, section, role


def _candidate_code(section: str, tenant_id: str) -> str:
    tenant_hint = str(tenant_id or "").replace("-", "")[:4].upper() or "GE"
    return f"{section[:2].upper()}-{tenant_hint}-{secrets.token_hex(2).upper()}"


def _create_unique_internal_user(sb, row: dict, requested_code: str = "") -> tuple[dict, str]:
    """
    Crea usuario interno evitando choques de unique constraint.
    Si el admin capturó código manual, no se reemplaza silenciosamente: se responde limpio.
    Si el código es auto, se reintenta con códigos nuevos dentro del mismo tenant/section.
    """
    manual = bool(requested_code)
    last_exc: Exception | None = None
    for attempt in range(8):
        if not manual:
            row["code"] = _candidate_code(row["section"], row["tenant_id"])
        try:
            created = sb.table("internal_users").insert(row).execute().data or [row]
            return created[0], row["code"]
        except Exception as exc:
            last_exc = exc
            text = str(exc).lower()
            duplicated = "duplicate" in text or "unique" in text or "23505" in text
            if manual or not duplicated:
                break
    if manual:
        raise HTTPException(409, "Ese código ya existe para esta empresa/módulo. Usa otro código o deja Auto.")
    raise _safe_internal_error("create", last_exc or Exception("unknown create error"))


def _internal_session(token_plain: str, section: str | None = None) -> dict:
    if not token_plain:
        raise HTTPException(401, "Sesión requerida.")
    sb = get_supabase_admin()
    token_hash = _hash_token(token_plain)
    rows = (
        sb.table("internal_user_sessions")
        .select("*, internal_users(*)")
        .eq("token_hash", token_hash)
        .limit(1)
        .execute()
        .data or []
    )
    if not rows:
        raise HTTPException(401, "Sesión inválida o expirada.")
    session = rows[0]
    if section and (session.get("section") or "") != section:
        raise HTTPException(403, "Sesión no corresponde a este módulo.")
    try:
        expires_at = datetime.fromisoformat(str(session.get("expires_at")).replace("Z", "+00:00"))
    except Exception:
        raise HTTPException(401, "Sesión inválida o expirada.")
    if expires_at <= _now():
        raise HTTPException(401, "Sesión expirada.")
    user = session.get("internal_users") or {}
    if (user.get("status") or "active") != "active":
        raise HTTPException(403, "Usuario interno inactivo.")
    _validate_internal_scope(user)
    return {"session": session, "user": user}


@router.get("/internal-users")
async def list_internal_users(
    section: Optional[str] = None,
    perfil_id: Optional[int] = None,
    authorization: str = Header(default=""),
):
    admin_uid, token = _auth_admin(authorization)
    tenant_id = _tenant_id_for_user(admin_uid, access_token=token)
    sb = get_supabase_for_user(token)
    q = sb.table("internal_users").select("*").eq("tenant_id", tenant_id).eq("owner_user_id", admin_uid)
    if section:
        q = q.eq("section", section.strip().lower())
    if section and section.strip().lower() == "gas_lp":
        if not perfil_id:
            raise HTTPException(400, "Selecciona una empresa Gas LP para ver sus asistentes.")
        perfil = _profile_for_admin(admin_uid, perfil_id, token)
        q = q.eq("perfil_id", perfil["id"])
    elif perfil_id:
        q = q.eq("perfil_id", perfil_id)
    rows = q.order("created_at", desc=True).execute().data or []
    for row in rows:
        row.pop("pin_hash", None)
    return JSONResponse({"ok": True, "users": rows})


@router.post("/internal-users")
async def create_internal_user(payload: InternalUserCreate, authorization: str = Header(default="")):
    admin_uid, token = _auth_admin(authorization)
    name, section, role = _clean_payload(payload)
    perfil = _profile_for_admin(admin_uid, payload.perfil_id, token)
    tenant_id = perfil["tenant_id"]
    requested_code = _clean_code(payload.code or "")
    code = requested_code or _candidate_code(section, tenant_id)
    temp_pin = (payload.pin or "").strip() or f"{secrets.randbelow(900000) + 100000}"
    row = {
        "tenant_id": tenant_id,
        "owner_user_id": admin_uid,
        "perfil_id": perfil["id"],
        "section": section,
        "role": role,
        "display_name": name,
        "code": code,
        "pin_hash": _hash_secret(temp_pin),
        "status": "active",
        "chofer_id": payload.chofer_id,
        "permissions": payload.permissions or {},
        "failed_attempts": 0,
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
    }
    try:
        response, code = _create_unique_internal_user(get_supabase_for_user(token), row, requested_code)
    except Exception as e:
        if isinstance(e, HTTPException):
            raise
        raise _safe_internal_error("create", e)
    response.pop("pin_hash", None)
    return JSONResponse({"ok": True, "user": response, "temporary_pin": temp_pin})


@router.put("/internal-users/{internal_user_id}/status")
async def update_internal_user_status(internal_user_id: int, payload: InternalUserStatus, authorization: str = Header(default="")):
    admin_uid, token = _auth_admin(authorization)
    tenant_id = _tenant_id_for_user(admin_uid, access_token=token)
    status = (payload.status or "").strip().lower()
    if status not in {"active", "inactive", "locked"}:
        raise HTTPException(400, "Estatus inválido.")
    try:
        get_supabase_for_user(token).table("internal_users").update({
            "status": status,
            "updated_at": _now_iso(),
        }).eq("id", internal_user_id).eq("tenant_id", tenant_id).eq("owner_user_id", admin_uid).execute()
    except Exception as e:
        raise _safe_internal_error("status", e)
    return JSONResponse({"ok": True})


@router.put("/internal-users/{internal_user_id}")
async def update_internal_user(internal_user_id: int, payload: InternalUserUpdate, authorization: str = Header(default="")):
    admin_uid, token = _auth_admin(authorization)
    tenant_id = _tenant_id_for_user(admin_uid, access_token=token)
    data = {"updated_at": _now_iso()}
    if payload.role is not None:
        role = (payload.role or "").strip().lower()
        if role not in ROLES:
            raise HTTPException(400, "Rol inválido.")
        data["role"] = role
    if payload.display_name is not None:
        name = (payload.display_name or "").strip()
        if not name:
            raise HTTPException(400, "Nombre requerido.")
        data["display_name"] = name
    try:
        get_supabase_for_user(token).table("internal_users").update(data).eq("id", internal_user_id).eq("tenant_id", tenant_id).eq("owner_user_id", admin_uid).execute()
    except Exception as e:
        raise _safe_internal_error("update", e)
    return JSONResponse({"ok": True})


@router.post("/internal-users/{internal_user_id}/reset-pin")
async def reset_internal_pin(internal_user_id: int, payload: InternalResetPin, authorization: str = Header(default="")):
    admin_uid, token = _auth_admin(authorization)
    tenant_id = _tenant_id_for_user(admin_uid, access_token=token)
    temp_pin = (payload.pin or "").strip() or f"{secrets.randbelow(900000) + 100000}"
    try:
        get_supabase_for_user(token).table("internal_users").update({
            "pin_hash": _hash_secret(temp_pin),
            "failed_attempts": 0,
            "locked_until": None,
            "status": "active",
            "updated_at": _now_iso(),
        }).eq("id", internal_user_id).eq("tenant_id", tenant_id).eq("owner_user_id", admin_uid).execute()
    except Exception as e:
        raise _safe_internal_error("reset_pin", e)
    return JSONResponse({"ok": True, "temporary_pin": temp_pin})


@router.delete("/internal-users/{internal_user_id}")
async def delete_internal_user_safe(internal_user_id: int, authorization: str = Header(default="")):
    admin_uid, token = _auth_admin(authorization)
    tenant_id = _tenant_id_for_user(admin_uid, access_token=token)
    sb = get_supabase_for_user(token)
    try:
        sessions = sb.table("internal_user_sessions").select("id", count="exact").eq("internal_user_id", internal_user_id).limit(1).execute()
        has_history = bool(getattr(sessions, "count", 0) or (sessions.data or []))
        if has_history:
            sb.table("internal_users").update({
                "status": "inactive",
                "updated_at": _now_iso(),
            }).eq("id", internal_user_id).eq("tenant_id", tenant_id).eq("owner_user_id", admin_uid).execute()
            raise HTTPException(409, "Este usuario interno ya tiene historial de acceso. Se desactivó, no se eliminó.")
        sb.table("internal_users").delete().eq("id", internal_user_id).eq("tenant_id", tenant_id).eq("owner_user_id", admin_uid).execute()
    except HTTPException:
        raise
    except Exception as e:
        raise _safe_internal_error("delete", e)
    return JSONResponse({"ok": True})


@router.post("/internal-auth/login")
async def internal_login(payload: InternalLogin):
    section = (payload.section or "").strip().lower()
    login = _clean_login(payload.code)
    if section not in SECTIONS or not login or not payload.pin:
        raise HTTPException(400, "Usuario, contraseña y módulo son obligatorios.")
    sb = get_supabase_admin()
    rows = (
        sb.table("internal_users")
        .select("*")
        .eq("section", section)
        .limit(300)
        .execute()
        .data
        or []
    )
    rows = [row for row in rows if row.get("tenant_id") and row.get("perfil_id")]
    code_rows = [row for row in rows if _matches_login(row, login)]
    fallback_rows = [row for row in rows if _matches_login(row, login, allow_display_name=(section == "gas_lp"))]
    candidates = code_rows or fallback_rows
    rows = candidates[:20]
    if not rows:
        raise HTTPException(401, "Usuario o contraseña incorrectos.")
    user = next((row for row in rows if _verify_secret(payload.pin, row.get("pin_hash") or "")), None)
    if not user:
        user = rows[0]
    if (user.get("status") or "active") != "active":
        raise HTTPException(403, "Usuario interno inactivo.")
    _validate_internal_scope(user)
    locked_until = user.get("locked_until")
    if locked_until:
        try:
            if datetime.fromisoformat(str(locked_until).replace("Z", "+00:00")) > _now():
                raise HTTPException(423, "Usuario bloqueado temporalmente. Intenta más tarde.")
        except HTTPException:
            raise
    if not _verify_secret(payload.pin, user.get("pin_hash") or ""):
        failed = int(user.get("failed_attempts") or 0) + 1
        update = {"failed_attempts": failed, "updated_at": _now_iso()}
        if failed >= MAX_FAILED_ATTEMPTS:
            update["locked_until"] = (_now() + timedelta(minutes=LOCK_MINUTES)).isoformat()
            update["status"] = "locked"
        sb.table("internal_users").update(update).eq("id", user["id"]).execute()
        raise HTTPException(401, "Usuario o contraseña incorrectos.")

    session_token = secrets.token_urlsafe(32)
    expires_at = _now() + timedelta(hours=SESSION_HOURS)
    sb.table("internal_user_sessions").insert({
        "internal_user_id": user["id"],
        "tenant_id": user.get("tenant_id"),
        "perfil_id": user.get("perfil_id"),
        "section": user.get("section"),
        "role": user.get("role"),
        "token_hash": _hash_token(session_token),
        "expires_at": expires_at.isoformat(),
        "created_at": _now_iso(),
    }).execute()
    sb.table("internal_users").update({
        "failed_attempts": 0,
        "locked_until": None,
        "status": "active",
        "last_access_at": _now_iso(),
        "updated_at": _now_iso(),
    }).eq("id", user["id"]).execute()

    result = {
        "ok": True,
        "token": session_token,
        "expires_at": expires_at.isoformat(),
        "section": user.get("section"),
        "role": user.get("role"),
        "perfil_id": user.get("perfil_id"),
        "display_name": user.get("display_name"),
        "tenant_id": user.get("tenant_id"),
        "permissions": user.get("permissions") or {},
    }
    if section == "transporte" and user.get("role") == "operador" and user.get("chofer_id"):
        operator_token = secrets.token_urlsafe(24)
        sb.table("tr_operador_accesos").insert({
            "user_id": user.get("owner_user_id"),
            "perfil_id": user.get("perfil_id"),
            "chofer_id": user.get("chofer_id"),
            "token_hash": _hash_token(operator_token),
            "status": "activo",
            "expires_at": expires_at.isoformat(),
        }).execute()
        result["operator_url"] = f"/operador/transporte?token={operator_token}"
    return JSONResponse(result)


@router.get("/internal-auth/me")
async def internal_me(token: str, section: str | None = None):
    ctx = _internal_session(token, section)
    user = ctx["user"]
    session = ctx["session"]
    return JSONResponse({
        "ok": True,
        "section": user.get("section"),
        "role": user.get("role"),
        "display_name": user.get("display_name"),
        "perfil_id": user.get("perfil_id"),
        "tenant_id": user.get("tenant_id"),
        "permissions": user.get("permissions") or {},
        "expires_at": session.get("expires_at"),
    })


@router.post("/internal-auth/logout")
async def internal_logout(payload: InternalLogout):
    if payload.token:
        try:
            get_supabase_admin().table("internal_user_sessions").delete().eq("token_hash", _hash_token(payload.token)).execute()
        except Exception as e:
            logger.warning("internal logout failed: %s", e)
    return JSONResponse({"ok": True})


@router.get("/internal-auth/gas-lp/summary")
async def gas_lp_internal_summary(token: str):
    ctx = _internal_session(token, "gas_lp")
    user = ctx["user"]
    profile = _gas_lp_profile(user)
    settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
    role = user.get("role") or "solo_lectura"
    role_modules = {
        "asistente_facturacion": [
            {"key": "facturacion", "title": "Facturación", "desc": "CFDI, XML, Excel y reportes fiscales permitidos."},
            {"key": "xml_excel", "title": "XML / Excel", "desc": "Carga y validación de archivos operativos."},
        ],
        "asistente_operativo": [
            {"key": "operacion", "title": "Operación", "desc": "Seguimiento operativo y datos de entregas."},
            {"key": "consulta", "title": "Consultas", "desc": "Consulta de registros del periodo."},
        ],
        "conciliacion": [
            {"key": "conciliacion", "title": "Conciliación", "desc": "Facturas, complementos de pago, consulta y cancelación."},
        ],
        "planta": [
            {"key": "planta", "title": "Captura de planta", "desc": "Inventario, composición y capturas operativas de planta."},
        ],
        "solo_lectura": [
            {"key": "reportes", "title": "Consulta y reportes", "desc": "Lectura de reportes, historial y métricas sin edición."},
        ],
    }
    modules = role_modules.get(role, role_modules["solo_lectura"])
    precio_venta_litro, precio_venta_litro_configurado = _configured_setting(
        settings,
        ("precio_venta_litro", "PrecioVentaLitro", "precio_default_litro", "precio_litro"),
    )
    return JSONResponse({
        "ok": True,
        "assistant": {
            "id": user.get("id"),
            "display_name": user.get("display_name"),
            "role": role,
            "perfil_id": user.get("perfil_id"),
            "tenant_id": user.get("tenant_id"),
            "serie_factura": _gas_lp_internal_series(user, settings),
        },
        "company": {
            "id": profile.get("id"),
            "name": profile.get("nombre"),
            "fiscal_name": str(settings.get("DescripcionInstalacion") or profile.get("nombre") or "").strip(),
            "rfc": profile.get("rfc"),
            "tenant_id": profile.get("tenant_id"),
            "cp": _clean_cp(settings.get("CodigoPostal") or settings.get("codigo_postal") or ""),
            "regimen": str(settings.get("RegimenFiscal") or settings.get("regimen_fiscal") or "601").strip() or "601",
            "precio_venta_litro": precio_venta_litro,
            "precio_venta_litro_configurado": precio_venta_litro_configurado,
            "transfer_email_default": _transfer_email_from_settings(settings),
        },
        "modules": modules,
        "hyp": {
            "mode": _gas_lp_hyp_mode(),
            "warning": "",
        },
        "session": {"expires_at": ctx["session"].get("expires_at"), "hours": SESSION_HOURS},
        "notices": [
            "Este portal no usa cuenta global Supabase Auth.",
            "Los permisos se limitan por empresa, módulo y rol interno.",
        ],
    })


@router.get("/internal-auth/gas-lp/hyp-mode")
async def gas_lp_internal_hyp_mode(token: str):
    _gas_lp_internal_context(token)
    mode = _gas_lp_hyp_mode()
    return JSONResponse({
        "ok": True,
        "mode": mode,
        "warning": "",
    })


@router.get("/internal-auth/gas-lp/facilities")
async def gas_lp_internal_facilities(token: str):
    ctx = _gas_lp_internal_context(token)
    user = ctx["user"]
    rows = get_facilities(user.get("owner_user_id"), "gas_lp", perfil_id=user.get("perfil_id"))
    return JSONResponse({"ok": True, "facilities": rows})


@router.get("/internal-auth/gas-lp/catalogos")
async def gas_lp_internal_catalogos(token: str, modulo: str = "gas_lp"):
    ctx = _gas_lp_internal_context(token)
    user = ctx["user"]
    sb = get_supabase_admin()

    def scoped(table: str):
        return (
            sb.table(table)
            .select("*")
            .eq("user_id", user.get("owner_user_id"))
            .eq("perfil_id", user.get("perfil_id"))
            .eq("activo", True)
        )

    try:
        choferes = scoped("tr_choferes").order("nombre").execute().data or []
    except Exception:
        choferes = []
    try:
        vehiculos = scoped("tr_vehiculos").order("placas").execute().data or []
    except Exception:
        vehiculos = []
    try:
        rutas = scoped("tr_rutas").order("nombre").execute().data or []
    except Exception:
        rutas = []
    return JSONResponse({"ok": True, "modulo": modulo, "choferes": choferes, "vehiculos": vehiculos, "rutas": rutas})


@router.get("/internal-auth/gas-lp/clientes")
async def gas_lp_internal_clientes(token: str):
    ctx = _gas_lp_internal_context(token)
    user = ctx["user"]
    sb = get_supabase_admin()
    try:
        rows = (
            sb.table("gas_lp_clientes_facturacion")
            .select("*")
            .eq("user_id", user.get("owner_user_id"))
            .eq("tenant_id", user.get("tenant_id"))
            .eq("perfil_id", user.get("perfil_id"))
            .eq("activo", True)
            .order("nombre", desc=False)
            .execute()
            .data
            or []
        )
    except Exception as exc:
        raise _safe_internal_error("gas_lp_clientes", exc)
    return JSONResponse({"ok": True, "clientes": rows})


@router.post("/internal-auth/gas-lp/clientes")
async def gas_lp_internal_crear_cliente(payload: GasLpInternalClientePayload, token: str):
    ctx = _gas_lp_internal_context(token, write=True)
    user = ctx["user"]
    row = _gas_lp_cliente_row(user, payload)
    try:
        data = get_supabase_admin().table("gas_lp_clientes_facturacion").insert(row).execute().data or [row]
    except Exception as exc:
        raise _safe_internal_error("gas_lp_crear_cliente", exc)
    return JSONResponse({"ok": True, "cliente": data[0]})


@router.put("/internal-auth/gas-lp/clientes/{cliente_id}")
async def gas_lp_internal_actualizar_cliente(cliente_id: int, payload: GasLpInternalClientePayload, token: str):
    ctx = _gas_lp_internal_context(token, write=True)
    user = ctx["user"]
    row = _gas_lp_cliente_row(user, payload)
    row.pop("created_at", None)
    row["metadata"] = {
        **(row.get("metadata") or {}),
        "updated_by_internal": user.get("id"),
        "updated_by": user.get("display_name"),
    }
    try:
        data = (
            get_supabase_admin()
            .table("gas_lp_clientes_facturacion")
            .update(row)
            .eq("id", cliente_id)
            .eq("user_id", user.get("owner_user_id"))
            .eq("tenant_id", user.get("tenant_id"))
            .eq("perfil_id", user.get("perfil_id"))
            .eq("activo", True)
            .execute()
            .data
            or []
        )
    except Exception as exc:
        raise _safe_internal_error("gas_lp_actualizar_cliente", exc)
    if not data:
        raise HTTPException(404, "Cliente no encontrado para esta empresa.")
    return JSONResponse({"ok": True, "cliente": data[0]})


@router.delete("/internal-auth/gas-lp/clientes/{cliente_id}")
async def gas_lp_internal_eliminar_cliente(cliente_id: int, token: str):
    ctx = _gas_lp_internal_context(token, write=True)
    user = ctx["user"]
    try:
        q = (
            get_supabase_admin()
            .table("gas_lp_clientes_facturacion")
            .update({"activo": False, "updated_at": _now_iso()})
            .eq("id", cliente_id)
            .eq("user_id", user.get("owner_user_id"))
            .eq("tenant_id", user.get("tenant_id"))
            .eq("perfil_id", user.get("perfil_id"))
        )
        data = q.execute().data or []
    except Exception as exc:
        raise _safe_internal_error("gas_lp_eliminar_cliente", exc)
    if not data:
        raise HTTPException(404, "Cliente no encontrado para esta empresa.")
    return JSONResponse({"ok": True, "message": "Cliente eliminado"})


@router.get("/internal-auth/gas-lp/facturas")
async def gas_lp_internal_facturas(token: str, mes: str | None = None):
    ctx = _gas_lp_internal_context(token)
    user = ctx["user"]
    profile = _gas_lp_profile(user)
    sb = get_supabase_admin()
    month = str(mes or "").strip()[:7]
    if len(month) == 7 and month[4] == "-":
        try:
            datetime.strptime(f"{month}-01", "%Y-%m-%d")
        except ValueError:
            month = ""
    else:
        month = ""
    try:
        rows = _gas_lp_company_facturas_rows(sb, user, profile, month=month, limit=10000 if month else 1000)
    except Exception as exc:
        raise _safe_internal_error("gas_lp_facturas", exc)
    _gas_lp_attach_internal_creators(sb, rows)
    _gas_lp_attach_cliente_email_recipients(sb, user, rows)
    comp_by_factura = _gas_lp_complementos_por_factura(sb, [_safe_int_id(r.get("id")) for r in rows if _safe_int_id(r.get("id"))])
    for row in rows:
        row["payment_info"] = _payment_info_json(_factura_payment_info(row))
        comps = comp_by_factura.get(_safe_int_id(row.get("id")), [])
        row["complementos_pago"] = comps
        if comps:
            row["latest_complemento_pago"] = comps[0]
    return JSONResponse({"ok": True, "facturas": rows})


@router.get("/internal-auth/gas-lp/facturas/export-dia")
async def gas_lp_internal_facturas_export_dia(token: str, fecha: str):
    ctx = _gas_lp_internal_context(token)
    user = ctx["user"]
    profile = _gas_lp_profile(user)
    day = str(fecha or "").strip()[:10]
    try:
        datetime.strptime(day, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(400, "Selecciona una fecha válida para exportar.")
    sb = get_supabase_admin()
    try:
        rows = _gas_lp_company_facturas_rows(sb, user, profile, month=day[:7], limit=10000)
    except Exception as exc:
        raise _safe_internal_error("gas_lp_facturas_export_dia", exc)
    rows = [row for row in rows if _gas_lp_factura_date_key(row) == day]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"
    headers = ["Fecha", "Folio de fact", "UUID", "Razón social", "Monto con IVA", "Litros", "PUE o PPD", "Estado"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7A1E2C")
    for row in rows:
        ws.append([
            _gas_lp_factura_date_key(row),
            _gas_lp_factura_folio_label(row),
            row.get("uuid_sat") or "",
            _gas_lp_factura_razon_social(row),
            float(_gas_lp_factura_total_con_iva(row)),
            float(row.get("volumen_litros") or 0),
            _gas_lp_factura_metodo_pago(row),
            _gas_lp_factura_estado_excel(row),
        ])
    for width, column in zip([14, 18, 40, 42, 18, 14, 14, 26], "ABCDEFGH"):
        ws.column_dimensions[column].width = width
    for cell in ws["E"][1:]:
        cell.number_format = '$#,##0.00'
    for cell in ws["F"][1:]:
        cell.number_format = "#,##0.000"
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"facturas_gas_lp_{day}.xlsx"
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/internal-auth/gas-lp/facturas/{factura_id}/xml")
async def gas_lp_internal_factura_xml(factura_id: int, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_factura_access_context(token, perfil_id=perfil_id)
    row = _gas_lp_internal_factura(ctx["user"], factura_id)
    xml_content = row.get("xml_content") or ""
    if not xml_content:
        raise HTTPException(404, "Factura sin XML timbrado.")
    info = fiscal_pdf_info(xml_content, "factura_gas_lp")
    filename = info.filename.replace(".pdf", ".xml")
    return Response(
        content=xml_content,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/internal-auth/gas-lp/facturas/{factura_id}/pac-audit")
async def gas_lp_internal_factura_pac_audit(factura_id: int, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_factura_access_context(token, perfil_id=perfil_id)
    row = _gas_lp_internal_factura(ctx["user"], factura_id)
    sb = get_supabase_admin()
    uuid_sat = str(row.get("uuid_sat") or "").strip()
    xml_content = str(row.get("xml_content") or "")
    xml_summary = _gas_lp_factura_pac_xml_summary(xml_content)
    responses = []
    try:
        query = sb.table("pac_responses").select("*").order("id", desc=True).limit(20)
        if uuid_sat:
            query = query.eq("uuid_sat", uuid_sat)
        responses = query.execute().data or []
    except Exception as exc:
        logger.warning("gas_lp_pac_audit_responses_lookup_failed factura_id=%s uuid=%s err=%s", factura_id, uuid_sat, exc)
        responses = []
    request_ids = sorted({
        int(resp.get("request_id") or 0)
        for resp in responses
        if resp.get("request_id")
    })
    requests_by_id = {}
    if request_ids:
        try:
            reqs = sb.table("pac_requests").select("*").in_("id", request_ids).execute().data or []
            requests_by_id = {int(req.get("id") or 0): req for req in reqs}
        except Exception as exc:
            logger.warning("gas_lp_pac_audit_requests_lookup_failed factura_id=%s uuid=%s err=%s", factura_id, uuid_sat, exc)
    audit = []
    for resp in responses:
        req = requests_by_id.get(int(resp.get("request_id") or 0), {})
        audit.append({
            "pac_response_id": resp.get("id"),
            "pac_request_id": resp.get("request_id"),
            "provider": resp.get("provider") or req.get("provider") or "",
            "environment": req.get("environment") or "",
            "operation": req.get("operation") or "",
            "request_created_at": req.get("created_at") or "",
            "response_created_at": resp.get("created_at") or "",
            "request_payload": req.get("request_payload") or {},
            "response_payload": resp.get("response_payload") or {},
            "status": resp.get("status") or "",
            "error_message": resp.get("error_message") or "",
            "uuid_sat": resp.get("uuid_sat") or "",
            "pdf_url": resp.get("pdf_url") or "",
            "xml_original": resp.get("xml_original") or "",
            "xml_timbrado": resp.get("xml_timbrado") or "",
        })
    return JSONResponse({
        "ok": True,
        "factura": {
            "id": row.get("id"),
            "uuid_sat": uuid_sat,
            "record_uuid": row.get("record_uuid") or "",
            "fecha_timbrado": row.get("fecha_timbrado") or "",
            "created_at": row.get("created_at") or "",
            "source": row.get("source") or "",
            "metadata": row.get("metadata") or {},
        },
        "xml_summary": xml_summary,
        "audit_count": len(audit),
        "audit": audit,
    })


@router.get("/internal-auth/gas-lp/facturas/{factura_id}/pdf")
async def gas_lp_internal_factura_pdf(factura_id: int, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_factura_access_context(token, perfil_id=perfil_id)
    user = ctx["user"]
    row = _gas_lp_internal_factura(user, factura_id)
    pac_pdf_url = str(row.get("pdf_url") or "").strip()
    if pac_pdf_url:
        return RedirectResponse(pac_pdf_url, status_code=302)
    xml_content = row.get("xml_content") or ""
    if not xml_content:
        raise HTTPException(404, "Factura sin XML timbrado para generar PDF.")
    settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
    info = fiscal_pdf_info(xml_content, "factura_gas_lp")
    pdf_bytes = generar_pdf_gas_lp_desde_xml(
        xml_content,
        logo_data_url=settings.get("PdfLogoDataUrl", ""),
        observaciones=_gas_lp_factura_observaciones(row),
    )
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{info.filename}"'},
    )


@router.post("/internal-auth/gas-lp/facturas/{factura_id}/send-email")
async def gas_lp_internal_factura_send_email(factura_id: int, payload: GasLpSendEmailPayload, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_factura_access_context(token, write=True, perfil_id=perfil_id)
    user = ctx["user"]
    profile = _gas_lp_profile(user)
    settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
    issuer = _require_gas_lp_issuer(profile, settings)
    row = _gas_lp_internal_factura(user, factura_id)
    xml_content = str(row.get("xml_content") or "")
    if not xml_content:
        raise HTTPException(400, "La factura no tiene XML timbrado para enviar.")
    uuid_sat = str(row.get("uuid_sat") or "").strip()
    if not uuid_sat:
        raise HTTPException(400, "La factura no tiene UUID timbrado para enviar.")
    md = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
    if md.get("tipo_operacion") == "traspaso":
        fallback_email = md.get("transfer_email") or md.get("transfer_email_sent_to") or row.get("email_destinatario")
    else:
        cliente_rows = []
        cliente_id = int(md.get("cliente_id") or 0)
        if cliente_id:
            try:
                cliente_rows = (
                    get_supabase_admin()
                    .table("gas_lp_clientes_facturacion")
                    .select("*")
                    .eq("id", cliente_id)
                    .eq("user_id", user.get("owner_user_id"))
                    .eq("tenant_id", user.get("tenant_id"))
                    .eq("perfil_id", user.get("perfil_id"))
                    .eq("activo", True)
                    .limit(1)
                    .execute()
                    .data
                    or []
                )
            except Exception as exc:
                logger.warning("gas_lp_factura_send_email_cliente_lookup_failed factura=%s cliente=%s err=%s", factura_id, cliente_id, exc)
        fallback_recipients = _customer_invoice_recipients(cliente_rows[0]) if cliente_rows else _invoice_email_recipients(
            md.get("cliente_email") or row.get("email_destinatario"),
            md.get("email_adicional_1") or "",
            md.get("email_adicional_2") or "",
            fallback=md.get("email_sent_to") or md.get("email_last_attempt_to") or "",
        )
        fallback_email = ", ".join(fallback_recipients)
    recipients = _invoice_email_recipients(payload.email, payload.email_adicional_1, payload.email_adicional_2, fallback=fallback_email)
    recipient = ", ".join(recipients)
    if not recipients:
        raise HTTPException(400, "Captura un correo destino para enviar XML/PDF.")
    try:
        info = fiscal_pdf_info(xml_content, "factura_gas_lp")
        pdf_bytes = generar_pdf_gas_lp_desde_xml(
            xml_content,
            logo_data_url=settings.get("PdfLogoDataUrl", ""),
            observaciones=_gas_lp_factura_observaciones(row),
        )
    except Exception as exc:
        raise _safe_internal_error("gas_lp_factura_send_email_pdf", exc)
    email_results = []
    email_result = None
    for email_to in recipients:
        email_result = send_gas_lp_invoice_email(
            to_email=email_to,
            issuer_name=issuer["nombre"],
            customer_name=str(md.get("cliente_nombre") or row.get("rfc_receptor") or "Cliente"),
            uuid_sat=uuid_sat,
            total=md.get("total") or _gas_lp_factura_total_con_iva(row),
            xml_content=xml_content,
            pdf_bytes=pdf_bytes,
            pdf_filename=info.filename,
            serie_folio=_gas_lp_factura_folio_label(row),
        )
        email_results.append({"to": email_to, **email_result.as_metadata()})
    now_email = _now_iso()
    all_ok = bool(email_results) and all(item.get("ok") for item in email_results)
    first_error = next((str(item.get("error") or "") for item in email_results if not item.get("ok")), "")
    message_ids = ", ".join(str(item.get("message_id") or "") for item in email_results if item.get("message_id"))
    updated_md = {
        **md,
        "email_delivery": email_result.as_metadata() if email_result else {},
        **_invoice_email_metadata(recipients),
        "email_sent_at": now_email if all_ok else md.get("email_sent_at"),
        "email_sent_to": recipient if all_ok else md.get("email_sent_to", recipient),
        "resend_message_id": message_ids if all_ok else md.get("resend_message_id", ""),
        "email_error": "" if all_ok else first_error,
        "email_last_attempt_at": now_email,
        "email_last_attempt_to": recipient,
    }
    if md.get("tipo_operacion") == "traspaso":
        updated_md = {
            **updated_md,
            "transfer_email_delivery": email_results,
            "transfer_email_sent_at": now_email if all_ok else md.get("transfer_email_sent_at"),
            "transfer_email_sent_to": recipient if all_ok else md.get("transfer_email_sent_to", recipient),
            "transfer_email_message_id": message_ids if all_ok else md.get("transfer_email_message_id", ""),
            "transfer_email_error": "" if all_ok else first_error,
        }
    update_payload = {
        "metadata": updated_md,
        "email_enviado": all_ok,
        "email_enviado_at": now_email if all_ok else row.get("email_enviado_at"),
        "email_destinatario": recipient,
        "email_error": "" if all_ok else first_error,
        "updated_at": now_email,
    }
    try:
        updated = (
            get_supabase_admin()
            .table("gas_lp_facturas")
            .update(update_payload)
            .eq("id", factura_id)
            .execute()
            .data
            or []
        )
    except Exception as exc:
        raise _safe_internal_error("gas_lp_factura_send_email_update", exc)
    factura = updated[0] if updated else {**row, **update_payload}
    response = {"ok": all_ok, "factura": factura, "email": email_result.as_metadata() if email_result else {}, "email_results": email_results}
    if not all_ok:
        response["message"] = first_error or "No se pudo enviar el correo."
    return JSONResponse(response, status_code=200 if all_ok else 400)


@router.get("/internal-auth/gas-lp/conciliacion/perfiles")
async def gas_lp_conciliacion_perfiles(token: str):
    ctx = _gas_lp_conciliacion_context(token)
    user = ctx["user"]
    if str(token or "").count(".") == 2:
        perfiles = get_perfiles_for_user(user.get("owner_user_id"), access_token=token, module="gas_lp")
    else:
        profile = _gas_lp_profile(user, require_module_marker=True)
        perfiles = [{"id": profile.get("id"), "nombre": profile.get("nombre"), "rfc": profile.get("rfc"), "descripcion": ""}]
    return JSONResponse({"ok": True, "perfil_id": user.get("perfil_id"), "perfiles": perfiles})


@router.get("/internal-auth/gas-lp/conciliacion/facilities")
async def gas_lp_conciliacion_facilities(token: str, perfil_id: int | None = None):
    ctx = _gas_lp_conciliacion_context(token, perfil_id=perfil_id)
    user = ctx["user"]
    rows = get_facilities(user.get("owner_user_id"), "gas_lp", perfil_id=user.get("perfil_id"))
    return JSONResponse({"ok": True, "facilities": rows})


@router.get("/internal-auth/gas-lp/conciliacion/summary")
async def gas_lp_conciliacion_summary(token: str, periodo: str | None = None, perfil_id: int | None = None):
    ctx = _gas_lp_conciliacion_context(token, perfil_id=perfil_id)
    user = ctx["user"]
    profile = _gas_lp_profile(user, require_module_marker=True)
    month = (periodo or datetime.now().strftime("%Y-%m"))[:7]
    sb = get_supabase_admin()
    try:
        rows = _gas_lp_company_facturas_rows(sb, user, profile, month=month, limit=10000)
    except Exception as exc:
        raise _safe_internal_error("gas_lp_conciliacion_summary", exc)
    _gas_lp_attach_internal_creators(sb, rows)
    comp_by_factura = _gas_lp_complementos_por_factura(sb, [_safe_int_id(r.get("id")) for r in rows if _safe_int_id(r.get("id"))])
    total = credito = publico = complementos_pendientes = 0.0
    for row in rows:
        md = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
        info = _factura_payment_info(row)
        row["payment_info"] = _payment_info_json(info)
        row["issuer_info"] = {
            "rfc": _gas_lp_factura_emisor_rfc(row),
            "nombre": _gas_lp_factura_emisor_nombre(row),
        }
        comps = comp_by_factura.get(_safe_int_id(row.get("id")), [])
        row["complementos_pago"] = comps
        if comps:
            row["latest_complemento_pago"] = comps[0]
        if str(row.get("status") or "").lower().startswith("cancel"):
            continue
        amount = float(info["total"])
        total += amount
        if str(row.get("rfc_receptor") or "").upper() == "XAXX010101000":
            publico += amount
        if info["metodo_pago"] == "PPD" or str(md.get("metodo_pago") or "").upper() == "PPD":
            saldo = float(info["saldo_insoluto"])
            if saldo > 0:
                credito += saldo
                complementos_pendientes += 1
    return JSONResponse({
        "ok": True,
        "periodo": month,
        "company": {"id": profile.get("id"), "name": profile.get("nombre"), "rfc": profile.get("rfc")},
        "kpis": {
            "facturas": len(rows),
            "total_facturado": round(total, 2),
            "credito_estimado": round(credito, 2),
            "publico_general": round(publico, 2),
            "complementos_pendientes": int(complementos_pendientes),
        },
        "facturas": rows,
    })


@router.post("/internal-auth/gas-lp/conciliacion/facturar-publico-general")
async def gas_lp_conciliacion_facturar_publico_general(payload: GasLpConciliacionPublicoGeneralPayload, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_conciliacion_context(token, write=True, perfil_id=perfil_id)
    user = ctx["user"]
    profile = _gas_lp_profile(user, require_module_marker=True)
    settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
    issuer = _require_gas_lp_issuer(profile, settings)
    receptor = _public_general_receptor(issuer["cp"])
    sb = get_supabase_admin()
    serie_factura = _gas_lp_internal_series(user, settings)
    folio_factura = _gas_lp_next_invoice_folio(sb, user, serie_factura)
    facilities_by_id = {
        int(f["id"]): f
        for f in get_facilities(user.get("owner_user_id"), "gas_lp", perfil_id=user.get("perfil_id"))
        if f.get("id") is not None
    }
    origen = facilities_by_id.get(int(payload.facility_id or 0), {})
    if not origen:
        raise HTTPException(400, "Selecciona la instalación origen para timbrar Público en General.")
    hyp_mode = _gas_lp_hyp_mode()
    hyp = {}
    if hyp_mode == "required":
        hyp = _gas_lp_hyp_from_facility(origen, GAS_LP_CLAVE_PROD_SERV)
    informacion_global = None
    if payload.factura_global:
        informacion_global = {
            "periodicidad": payload.informacion_global_periodicidad,
            "meses": payload.informacion_global_meses,
            "anio": payload.informacion_global_anio,
        }
    concepto_cfdi = "LITRO DE GAS LP" if hyp_mode == "disabled" else "Gas licuado de petróleo"
    xml, totals = _build_gas_lp_consumo_xml(
        issuer=issuer,
        receptor=receptor,
        litros=payload.litros,
        precio_unitario=payload.precio_unitario,
        concepto=concepto_cfdi,
        forma_pago=payload.forma_pago,
        metodo_pago=payload.metodo_pago,
        descuento=payload.descuento,
        iva_rate=payload.iva_rate,
        serie=serie_factura,
        folio=folio_factura,
        comentarios=payload.comentarios,
        fecha=payload.fecha,
        clave_prod_serv=GAS_LP_CLAVE_PROD_SERV,
        no_identificacion="GLP-LTR",
        unidad="Litro",
        hyp=hyp,
        informacion_global=informacion_global,
    )
    sw_config = sw_runtime_config()
    if _sw_config_looks_like_sandbox(sw_config):
        raise HTTPException(400, "Este emisor está configurado en modo pruebas. Cambia a producción antes de timbrar CFDI real.")
    resultado = timbrar_cfdi(xml)
    if resultado.get("error"):
        raise HTTPException(400, f"PAC rechazó la factura: {resultado['error']}")
    now = _now_iso()
    row = {
        **_gas_lp_invoice_scope(user, profile),
        "facility_id": payload.facility_id,
        "record_uuid": totals["folio"],
        "uuid_sat": resultado.get("uuid") or "",
        "xml_content": resultado.get("xml_timbrado") or xml,
        "pdf_url": resultado.get("pdf_url") or "",
        "status": "Vigente",
        "fecha_timbrado": now,
        "rfc_receptor": receptor["rfc"],
        "volumen_litros": float(payload.litros),
        "importe": totals["subtotal"],
        "tipo_comprobante": "I",
        "distancia_km": 1,
        "metadata": {
            "portal": "conciliacion_gas_lp",
            "created_by_area": "conciliacion",
            "internal_user_id": user.get("id"),
            "created_by_internal_name": user.get("display_name") or "",
            "created_by": user.get("display_name") or "",
            "empresa_asignada_id": user.get("perfil_id"),
            "empresa_asignada_nombre": profile.get("nombre") or "",
            "empresa_rfc": profile.get("rfc") or "",
            "rfc_emisor": issuer.get("rfc") or "",
            "cliente_id": None,
            "cliente_nombre": receptor["nombre"],
            "cliente_email": "",
            "concepto": "LITRO DE GAS LP",
            "precio_unitario": payload.precio_unitario,
            "descuento_por_litro": payload.descuento,
            "descuento": totals["descuento"],
            "iva_rate": payload.iva_rate,
            "serie": serie_factura,
            "folio_usuario": folio_factura,
            "comentarios": payload.comentarios,
            "fecha_emision": totals["fecha"],
            "clave_prod_serv": GAS_LP_CLAVE_PROD_SERV,
            "gas_lp_hyp_mode": hyp_mode,
            "hidrocarburos_petroliferos": hyp,
            "no_identificacion": "GLP-LTR",
            "unidad": "Litro",
            "metodo_pago": payload.metodo_pago,
            "forma_pago": payload.forma_pago,
            "tipo_operacion": "venta_publico_general",
            "facility_id": payload.facility_id,
            "origen_nombre": origen.get("nombre") or "",
            "payment_status": "pendiente_complemento" if payload.metodo_pago.upper() == "PPD" else "pagado_pue",
            "saldo_insoluto": totals["total"] if payload.metodo_pago.upper() == "PPD" else 0,
            "iva": totals["iva"],
            "total": totals["total"],
        },
        "created_at": now,
    }
    try:
        factura = (sb.table("gas_lp_facturas").insert(row).execute().data or [row])[0]
    except Exception as exc:
        raise _safe_internal_error("gas_lp_conciliacion_facturar_publico_general", exc)
    return JSONResponse({"ok": True, "factura": factura, "totals": totals})


@router.get("/internal-auth/gas-lp/conciliacion/export-excel")
async def gas_lp_conciliacion_export_excel(
    token: str,
    period: str | None = None,
    periodo: str | None = None,
    fecha: str | None = None,
    profile_id: int | None = None,
    perfil_id: int | None = None,
):
    selected_perfil_id = perfil_id if perfil_id is not None else profile_id
    ctx = _gas_lp_conciliacion_context(token, perfil_id=selected_perfil_id)
    user = ctx["user"]
    day = str(fecha or "").strip()[:10]
    month = str(periodo or period or "").strip()[:7]
    if day:
        try:
            datetime.strptime(day, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(400, "Selecciona una fecha válida para exportar.")
        month = day[:7]
    elif len(month) == 7 and month[4] == "-":
        try:
            datetime.strptime(f"{month}-01", "%Y-%m-%d")
        except ValueError:
            month = datetime.now().strftime("%Y-%m")
    else:
        month = datetime.now().strftime("%Y-%m")

    start = datetime.strptime(day or f"{month}-01", "%Y-%m-%d")
    if day:
        end = start + timedelta(days=1)
    elif start.month == 12:
        end = start.replace(year=start.year + 1, month=1)
    else:
        end = start.replace(month=start.month + 1)

    sb = get_supabase_admin()
    try:
        profile = _gas_lp_profile(user, require_module_marker=True)
        rows = _gas_lp_company_facturas_rows(sb, user, profile, month=month, limit=10000)
    except Exception as exc:
        logger.error(
            "conciliacion_export_excel_error profile_id=%s period=%s factura_id=%s exception=%s traceback=%s",
            selected_perfil_id or user.get("perfil_id"),
            month,
            None,
            exc,
            traceback.format_exc(),
        )
        raise _safe_internal_error("gas_lp_conciliacion_export_excel", exc)
    try:
        _gas_lp_attach_internal_creators(sb, rows)
    except Exception as exc:
        logger.error(
            "conciliacion_export_excel_error profile_id=%s period=%s factura_id=%s exception=%s traceback=%s",
            selected_perfil_id or user.get("perfil_id"),
            month,
            None,
            exc,
            traceback.format_exc(),
        )

    if day:
        rows = [row for row in rows if _gas_lp_factura_date_key(row) == day]
    else:
        rows = [row for row in rows if _gas_lp_factura_date_key(row).startswith(month)]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"
    headers = [
        "Fecha",
        "Folio de fact",
        "UUID",
        "Razón social",
        "Monto con IVA",
        "Litros",
        "PUE o PPD",
        "Estado",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7A1E2C")

    def _excel_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return str(value)

    def _excel_number(value) -> float:
        try:
            return float(_money(value))
        except Exception:
            return 0.0

    def _excel_liters(value) -> float:
        try:
            return float(Decimal(str(value or 0)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP))
        except Exception:
            return 0.0

    def _safe_total(row: dict):
        try:
            return _factura_payment_info(row).get("total")
        except Exception:
            try:
                return _gas_lp_factura_total_con_iva(row)
            except Exception:
                return 0

    def _safe_metodo_pago(row: dict) -> str:
        try:
            metodo = _gas_lp_factura_metodo_pago(row)
        except Exception:
            md = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
            metodo = md.get("metodo_pago") or ""
        metodo = str(metodo or "").upper()
        return metodo if metodo in {"PUE", "PPD"} else "PUE"

    for row in rows:
        factura_id = row.get("id")
        try:
            ws.append([
                _excel_text(_gas_lp_factura_date_key(row)),
                _excel_text(_gas_lp_factura_folio_label(row)),
                _excel_text(row.get("uuid_sat") or ""),
                _excel_text(_gas_lp_factura_razon_social(row)),
                _excel_number(_safe_total(row)),
                _excel_liters(row.get("volumen_litros")),
                _excel_text(_safe_metodo_pago(row)),
                _excel_text(_gas_lp_factura_estado_excel(row)),
            ])
        except Exception as exc:
            logger.error(
                "conciliacion_export_excel_error profile_id=%s period=%s factura_id=%s exception=%s traceback=%s",
                selected_perfil_id or user.get("perfil_id"),
                month,
                factura_id,
                exc,
                traceback.format_exc(),
            )
            ws.append(["", _excel_text(factura_id), "", "", 0.0, 0.0, "", ""])
    for width, column in zip([14, 18, 40, 34, 16, 12, 12, 26], "ABCDEFGH"):
        ws.column_dimensions[column].width = width
    for cell in ws["F"][1:]:
        cell.number_format = "#,##0.000"
    for cell in ws["E"][1:]:
        cell.number_format = '$#,##0.00'

    stream = BytesIO()
    try:
        wb.save(stream)
    except Exception as exc:
        logger.error(
            "conciliacion_export_excel_error profile_id=%s period=%s factura_id=%s exception=%s traceback=%s",
            selected_perfil_id or user.get("perfil_id"),
            month,
            None,
            exc,
            traceback.format_exc(),
        )
        raise _safe_internal_error("gas_lp_conciliacion_export_excel", exc)
    stream.seek(0)
    suffix = day or month
    filename = f"conciliacion_gas_lp_{suffix}.xlsx"
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/internal-auth/gas-lp/facturas/{factura_id}/complemento-pago")
async def gas_lp_generar_complemento_pago(factura_id: int, payload: GasLpComplementoPagoPayload, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_conciliacion_context(token, write=True, perfil_id=perfil_id)
    user = ctx["user"]
    profile = _gas_lp_profile(user, require_module_marker=True)
    settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
    issuer = _require_gas_lp_issuer(profile, settings)
    serie_factura = _gas_lp_internal_series(user, settings)
    folio_factura = datetime.now().strftime("GLP%Y%m%d%H%M%S")
    sb = get_supabase_admin()
    requested: dict[int, Decimal | None] = {}
    for item in payload.facturas or []:
        fid = int(item.get("factura_id") or item.get("id") or 0)
        if fid:
            requested[fid] = _money(item.get("monto")) if item.get("monto") not in {None, ""} else None
    for fid in payload.factura_ids or []:
        if int(fid or 0):
            requested.setdefault(int(fid), None)
    requested.setdefault(int(factura_id), None)
    factura_ids = list(dict.fromkeys(requested.keys()))
    rows = (
        sb.table("gas_lp_facturas")
        .select("*")
        .in_("id", factura_ids)
        .eq("tenant_id", user.get("tenant_id"))
        .execute()
        .data
        or []
    )
    match_profile = {**profile, "rfc": _gas_lp_company_rfc(user, profile)}
    rows = [row for row in rows if _gas_lp_factura_matches_company(row, user, match_profile)]
    if len(rows) != len(factura_ids):
        raise HTTPException(404, "Una factura seleccionada no existe para esta empresa.")
    facturas_by_id = {int(r["id"]): r for r in rows}
    facturas = [facturas_by_id[fid] for fid in factura_ids]
    rfc = ""
    saldos: dict[int, Decimal] = {}
    for factura in facturas:
        md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
        info = _factura_payment_info(factura)
        if info["metodo_pago"] != "PPD" and str(md.get("metodo_pago") or "").upper() != "PPD":
            raise HTTPException(400, "Solo puedes generar complemento para facturas PPD.")
        if str(factura.get("status") or "").lower().startswith("cancel"):
            raise HTTPException(400, "No se puede generar complemento sobre una factura cancelada.")
        if not factura.get("xml_content"):
            raise HTTPException(400, "Cada factura debe tener XML timbrado.")
        frfc = str(factura.get("rfc_receptor") or "").upper()
        if rfc and frfc and rfc != frfc:
            raise HTTPException(400, "Selecciona facturas del mismo cliente/RFC.")
        rfc = rfc or frfc
        saldo = _money(info["saldo_insoluto"])
        if saldo <= 0:
            raise HTTPException(400, "Una factura seleccionada ya no tiene saldo pendiente.")
        saldos[int(factura["id"])] = saldo
    total_saldo = sum(saldos.values(), Decimal("0.00"))
    total_recibido = _money(payload.monto) if payload.monto not in {None, ""} else total_saldo
    if total_recibido <= 0 or total_recibido > total_saldo:
        raise HTTPException(400, "El monto recibido debe ser mayor a cero y no exceder el saldo seleccionado.")
    remaining = total_recibido
    pagos: dict[int, Decimal] = {}
    for fid in factura_ids:
        explicit = requested.get(fid)
        amount = _money(explicit if explicit is not None else min(saldos[fid], remaining))
        if amount <= 0 or amount > saldos[fid]:
            raise HTTPException(400, "El importe asignado a una factura no es válido.")
        pagos[fid] = amount
        remaining = _money(remaining - amount)
    if remaining != Decimal("0.00"):
        raise HTTPException(400, "El monto recibido no coincide con los importes asignados.")
    xml_pago, totals = _build_gas_lp_pago20_multi_xml(facturas=facturas, issuer=issuer, fecha_pago=payload.fecha_pago, forma_pago=payload.forma_pago, pagos=pagos)
    resultado = timbrar_cfdi(xml_pago)
    if resultado.get("error"):
        raise HTTPException(400, f"PAC rechazó el complemento de pago: {resultado['error']}")
    xml_timbrado = resultado.get("xml_timbrado") or xml_pago
    now = _now_iso()
    comp_row = {
        "factura_id": factura_ids[0],
        "user_id": user.get("owner_user_id"),
        "tenant_id": user.get("tenant_id"),
        "perfil_id": user.get("perfil_id"),
        "uuid_sat": resultado.get("uuid") or "",
        "xml_content": xml_timbrado,
        "status": "timbrado",
        "fecha_pago": totals["fecha_pago"],
        "forma_pago": totals["forma_pago"],
        "monto": totals["monto"],
        "saldo_insoluto": totals["saldo_insoluto"],
        "metadata": {
            "factura_ids": factura_ids,
            "referencia": payload.referencia,
            "banco": payload.banco,
            "notas": payload.notas,
            "facturas": totals["facturas"],
            "created_by_area": "conciliacion",
            "created_by_internal": user.get("id"),
            "created_by": user.get("display_name") or "",
            "empresa_rfc": issuer.get("rfc") or profile.get("rfc") or "",
        },
        "created_at": now,
        "updated_at": now,
    }
    try:
        comp = (sb.table("gas_lp_complementos_pago").insert(comp_row).execute().data or [comp_row])[0]
    except Exception as exc:
        raise _safe_internal_error("gas_lp_complemento_pago_insert", exc)
    rels = []
    for doc in totals["facturas"]:
        rels.append({
            "complemento_id": comp.get("id"),
            "factura_id": doc["factura_id"],
            "user_id": user.get("owner_user_id"),
            "tenant_id": user.get("tenant_id"),
            "perfil_id": user.get("perfil_id"),
            "uuid_relacionado": doc["uuid_relacionado"],
            "monto": doc["monto"],
            "saldo_anterior": doc["saldo_anterior"],
            "saldo_insoluto": doc["saldo_insoluto"],
            "status": "timbrado",
            "created_at": now,
            "updated_at": now,
        })
    try:
        if comp.get("id"):
            sb.table("gas_lp_complementos_pago_facturas").insert(rels).execute()
    except Exception as exc:
        logger.info("No se pudieron guardar relaciones de complemento pago: %s", exc)
    for doc in totals["facturas"]:
        factura = facturas_by_id[int(doc["factura_id"])]
        md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
        status = "pagado_con_complemento" if _money(doc["saldo_insoluto"]) <= 0 else "pago_parcial"
        md = {**md, "payment_status": status, "saldo_insoluto": doc["saldo_insoluto"], "ultimo_complemento_pago_id": comp.get("id"), "ultimo_complemento_pago_uuid": comp.get("uuid_sat") or ""}
        sb.table("gas_lp_facturas").update({"metadata": md, "updated_at": now}).eq("id", doc["factura_id"]).execute()
    return JSONResponse({"ok": True, "complemento": comp, "facturas": totals["facturas"]})


@router.get("/internal-auth/gas-lp/complementos-pago/{complemento_id}/xml")
async def gas_lp_complemento_pago_xml(complemento_id: int, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_conciliacion_context(token, perfil_id=perfil_id)
    user = ctx["user"]
    _gas_lp_profile(user, require_module_marker=True)
    rows = (
        get_supabase_admin()
        .table("gas_lp_complementos_pago")
        .select("*")
        .eq("id", complemento_id)
        .eq("tenant_id", user.get("tenant_id"))
        .eq("perfil_id", user.get("perfil_id"))
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows or not rows[0].get("xml_content"):
        raise HTTPException(404, "Complemento de pago no encontrado.")
    return Response(content=rows[0]["xml_content"], media_type="application/xml")


@router.get("/internal-auth/gas-lp/complementos-pago/{complemento_id}/pdf")
async def gas_lp_complemento_pago_pdf(complemento_id: int, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_conciliacion_context(token, perfil_id=perfil_id)
    user = ctx["user"]
    _gas_lp_profile(user, require_module_marker=True)
    rows = (
        get_supabase_admin()
        .table("gas_lp_complementos_pago")
        .select("*")
        .eq("id", complemento_id)
        .eq("tenant_id", user.get("tenant_id"))
        .eq("perfil_id", user.get("perfil_id"))
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows or not rows[0].get("xml_content"):
        raise HTTPException(404, "Complemento de pago no encontrado.")
    xml_content = rows[0]["xml_content"]
    settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
    info = fiscal_pdf_info(xml_content, "complemento_pago_gas_lp")
    pdf_bytes = generar_pdf_gas_lp_desde_xml(xml_content, logo_data_url=settings.get("PdfLogoDataUrl", ""))
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{info.filename}"'},
    )


@router.post("/internal-auth/gas-lp/conciliacion/facturas/{factura_id}/cancelar")
async def gas_lp_conciliacion_cancelar(factura_id: int, payload: GasLpCancelacionPayload, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_conciliacion_context(token, write=True, perfil_id=perfil_id)
    user = ctx["user"]
    profile = _gas_lp_profile(user, require_module_marker=True)
    motivo = str(payload.motivo or "").strip()
    uuid_sustitucion = str(payload.uuid_sustitucion or "").strip()
    if motivo not in {"01", "02", "03", "04"}:
        raise HTTPException(400, "Motivo SAT inválido. Usa 01, 02, 03 o 04.")
    if motivo == "01" and not uuid_sustitucion:
        raise HTTPException(400, "El motivo SAT 01 requiere UUID sustituto.")
    sw_config = sw_runtime_config()
    if _sw_config_looks_like_sandbox(sw_config):
        raise HTTPException(400, "Cancelación fiscal bloqueada: SW no está en producción.")
    if not sw_config.get("real_cancelacion_flag"):
        raise HTTPException(400, "Cancelación real bloqueada: falta SW_ALLOW_REAL_CANCELACION=true.")
    now = _now_iso()
    sb = get_supabase_admin()
    rows = (
        sb.table("gas_lp_facturas")
        .select("*")
        .eq("id", factura_id)
        .eq("tenant_id", user.get("tenant_id"))
        .eq("perfil_id", user.get("perfil_id"))
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows:
        raise HTTPException(404, "Factura no encontrada.")
    factura = rows[0]
    if str(factura.get("status") or "").lower().startswith("cancel"):
        raise HTTPException(400, "La factura ya tiene estado de cancelación.")
    uuid_sat = str(factura.get("uuid_sat") or "").strip()
    if not uuid_sat:
        raise HTTPException(400, "No se puede cancelar fiscalmente: la factura no tiene UUID SAT.")
    factura_rfc_emisor = _gas_lp_factura_emisor_rfc(factura)
    factura_nombre_emisor = _gas_lp_factura_emisor_nombre(factura)
    if not factura_rfc_emisor:
        raise HTTPException(400, "No se puede cancelar fiscalmente: la factura no tiene RFC emisor guardado.")
    profile_rfc = _clean_rfc(profile.get("rfc") or "")
    if profile_rfc and profile_rfc != factura_rfc_emisor:
        logger.warning(
            "gas_lp_cancelacion_profile_invoice_rfc_mismatch factura_id=%s profile_rfc=%s factura_rfc=%s uuid=%s",
            factura_id,
            profile_rfc,
            factura_rfc_emisor,
            uuid_sat,
        )
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    base_cancel_md = {
        "cancelacion_tipo": "fiscal",
        "motivo_cancelacion": motivo,
        "uuid_sustitucion": uuid_sustitucion,
        "notas_cancelacion": str(payload.notas or "").strip(),
        "cancelacion_solicitada_por": user.get("display_name") or user.get("id"),
        "cancelacion_solicitada_at": now,
        "cancelacion_uuid_cancelado": uuid_sat,
        "cancelacion_rfc_emisor": factura_rfc_emisor,
        "cancelacion_nombre_emisor": factura_nombre_emisor,
        "cancelacion_profile_rfc": profile_rfc,
    }
    try:
        resultado = cancel_cfdi_universal(
            sb=sb,
            module="gas_lp",
            invoice_table="gas_lp_facturas",
            invoice_id=factura_id,
            uuid_sat=uuid_sat,
            rfc_emisor=factura_rfc_emisor,
            motivo=motivo,
            uuid_sustitucion=uuid_sustitucion,
            user_id=user.get("owner_user_id") or user.get("id") or "",
            perfil_id=user.get("perfil_id"),
            tenant_id=user.get("tenant_id"),
            requested_by=user.get("display_name") or user.get("id") or "",
        )
    except HTTPException as exc:
        detail = exc.detail if isinstance(exc.detail, dict) else {"message": str(exc.detail)}
        error_message = str(detail.get("message") or "SW Sapien rechazó la cancelación.")
        error_diagnostic = detail.get("diagnostic") if isinstance(detail.get("diagnostic"), dict) else {}
        err_md = {
            **md,
            **base_cancel_md,
            "estado_fiscal": "cancelacion_error",
            "cancelacion_error": error_message,
            "cancelacion_error_tecnico": error_diagnostic,
            "cancelacion_endpoint_final": error_diagnostic.get("endpoint_final"),
            "cancelacion_pac_request_id": detail.get("pac_request_id"),
            "cancelacion_pac_response_id": detail.get("pac_response_id"),
            "cancelacion_error_at": _now_iso(),
        }
        try:
            sb.table("gas_lp_facturas").update({"metadata": err_md, "updated_at": _now_iso()}).eq("id", factura_id).execute()
        except Exception as update_exc:
            logger.warning("gas_lp_cancelacion_error_metadata_update_failed factura_id=%s err=%s", factura_id, update_exc)
        raise HTTPException(exc.status_code, error_message)
    acuse = str(resultado.get("acuse") or "")
    diagnostic = resultado.get("diagnostic") if isinstance(resultado.get("diagnostic"), dict) else {}
    estado_fiscal = "cancelada_fiscalmente" if acuse else "cancelacion_solicitada"
    status_label = "Cancelada fiscalmente" if acuse else "Cancelación solicitada"
    cancel_md = {
        **md,
        **base_cancel_md,
        "estado_fiscal": estado_fiscal,
        "cancelacion_estado_fiscal_label": status_label,
        "cancelacion_pac_request_id": resultado.get("pac_request_id"),
        "cancelacion_pac_response_id": resultado.get("pac_response_id"),
        "cancelacion_acuse": acuse,
        "cancelacion_respuesta_sw": resultado.get("raw") or {},
        "cancelacion_diagnostico_http": diagnostic,
        "cancelacion_endpoint_final": diagnostic.get("endpoint_final"),
        "cancelacion_confirmada_at": _now_iso(),
    }
    data = (
        sb.table("gas_lp_facturas")
        .update({"status": status_label, "metadata": cancel_md, "updated_at": _now_iso()})
        .eq("id", factura_id)
        .execute()
        .data
        or []
    )
    return JSONResponse({
        "ok": True,
        "factura": data[0] if data else {**factura, "status": status_label, "metadata": cancel_md},
        "cancelacion": {
            "estado_fiscal": estado_fiscal,
            "status": status_label,
            "acuse": acuse,
            "pac_request_id": resultado.get("pac_request_id"),
            "pac_response_id": resultado.get("pac_response_id"),
            "respuesta_sw": resultado.get("raw") or {},
        },
    })


@router.post("/internal-auth/gas-lp/facturas")
async def gas_lp_internal_crear_factura(payload: GasLpInternalFacturaPayload, token: str):
    from routes.transporte import _normalizar_receptor_cfdi, _validar_datos_cfdi_receptor

    ctx = _gas_lp_internal_context(token, write=True)
    user = ctx["user"]
    profile = _gas_lp_profile(user)
    settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
    issuer = _require_gas_lp_issuer(profile, settings)
    is_transfer = str(payload.tipo_operacion or "").strip().lower() == "traspaso"
    receptor = {
        "rfc": issuer["rfc"],
        "nombre": issuer["nombre"],
        "cp": issuer["cp"],
        "regimen_fiscal": issuer["regimen"],
        "uso_cfdi": "S01",
    } if is_transfer else (_public_general_receptor(issuer["cp"]) if payload.publico_general else None)
    cliente_row = None
    sb = get_supabase_admin()
    if payload.cliente_id and not receptor:
        rows = (
            sb.table("gas_lp_clientes_facturacion")
            .select("*")
            .eq("id", payload.cliente_id)
            .eq("user_id", user.get("owner_user_id"))
            .eq("tenant_id", user.get("tenant_id"))
            .eq("perfil_id", user.get("perfil_id"))
            .eq("activo", True)
            .limit(1)
            .execute()
            .data
            or []
        )
        if not rows:
            raise HTTPException(404, "Cliente no encontrado para esta empresa.")
        cliente_row = rows[0]
        receptor = {
            "rfc": _clean_rfc(cliente_row.get("rfc")),
            "nombre": str(cliente_row.get("nombre") or "").strip(),
            "cp": _clean_cp(cliente_row.get("cp")),
            "regimen_fiscal": str(cliente_row.get("regimen_fiscal") or "616").strip(),
            "uso_cfdi": str(cliente_row.get("uso_cfdi") or "S01").strip(),
        }
    if not receptor:
        receptor = {
            "rfc": _clean_rfc(payload.rfc),
            "nombre": payload.nombre.strip(),
            "cp": _clean_cp(payload.cp),
            "regimen_fiscal": (payload.regimen_fiscal or "616").strip(),
            "uso_cfdi": (payload.uso_cfdi or "S01").strip(),
        }
    if receptor["rfc"] == "XAXX010101000":
        receptor = {**_public_general_receptor(issuer["cp"]), **{"uso_cfdi": receptor.get("uso_cfdi") or "S01"}}
    if not receptor.get("rfc") or not receptor.get("nombre") or not receptor.get("cp"):
        raise HTTPException(400, "Receptor incompleto: RFC, nombre y CP son obligatorios.")
    if receptor["rfc"] != "XAXX010101000":
        receptor = {
            **receptor,
            **_normalizar_receptor_cfdi(
                receptor["rfc"],
                receptor["nombre"],
                receptor["cp"],
                receptor["regimen_fiscal"],
            ),
        }
    _validar_datos_cfdi_receptor(
        receptor["rfc"],
        receptor["regimen_fiscal"],
        receptor["cp"],
        receptor["uso_cfdi"],
    )
    serie_factura = _gas_lp_internal_series(user, settings)
    folio_factura = _gas_lp_next_invoice_folio(sb, user, serie_factura)
    facilities_by_id = {
        int(f["id"]): f
        for f in get_facilities(user.get("owner_user_id"), "gas_lp", perfil_id=user.get("perfil_id"))
        if f.get("id") is not None
    }
    origen = facilities_by_id.get(int(payload.facility_id or 0), {})
    if not origen:
        raise HTTPException(400, "Selecciona la instalación origen para timbrar Gas LP.")
    destino = facilities_by_id.get(int(payload.destino_facility_id or 0), {})
    if is_transfer:
        if not destino:
            raise HTTPException(400, "Selecciona la estación destino para el traspaso.")
        if int(payload.facility_id or 0) == int(payload.destino_facility_id or 0):
            raise HTTPException(400, "Origen y destino deben ser distintos para el traspaso.")
    hyp_mode = _gas_lp_hyp_mode()
    if hyp_mode == "diagnostic" and not payload.hyp_experimental_diagnostics:
        raise HTTPException(400, "El modo HyP diagnóstico sólo permite pruebas persisted=false.")
    clave_prod_serv_original = _clean_clave_prod_serv(payload.clave_prod_serv)
    clave_prod_serv = GAS_LP_CLAVE_PROD_SERV
    clave_hyp_diagnostic_override = ""
    if payload.hyp_experimental_diagnostics and payload.hyp_clave_hyp_override:
        clave_hyp_diagnostic_override = _clean_clave_prod_serv(payload.hyp_clave_hyp_override)
        if clave_hyp_diagnostic_override not in GAS_LP_HYP_DIAGNOSTIC_CLAVES:
            raise HTTPException(400, "La clave HyP experimental sólo permite 15111510 o 15101515.")
        clave_prod_serv = clave_hyp_diagnostic_override
    hyp = {}
    if (hyp_mode == "required" or payload.hyp_experimental_diagnostics) and not is_transfer:
        hyp = _gas_lp_hyp_from_facility(origen, GAS_LP_CLAVE_PROD_SERV)
        if clave_hyp_diagnostic_override:
            hyp = {**hyp, "clave_hyp": clave_hyp_diagnostic_override}
    hyp_original = dict(hyp)
    hyp_override_aplicado = False
    if payload.hyp_experimental_diagnostics:
        override_numero = str(payload.hyp_numero_permiso_override or "").strip().upper()
        override_tipo = str(payload.hyp_tipo_permiso_override or "").strip().upper()
        if not override_numero:
            raise HTTPException(400, "La prueba experimental HyP requiere hyp_numero_permiso_override.")
        if override_tipo and override_tipo not in HYP_TIPO_PERMISOS_VALIDOS:
            raise HTTPException(400, "El TipoPermiso experimental debe usar PER01-PER11.")
        hyp = {
            **hyp,
            "numero_permiso": override_numero,
            "tipo_permiso": override_tipo or hyp.get("tipo_permiso") or "",
        }
        hyp_override_aplicado = True
    informacion_global = None
    if receptor["rfc"] == "XAXX010101000" and payload.factura_global:
        informacion_global = {
            "periodicidad": payload.informacion_global_periodicidad,
            "meses": payload.informacion_global_meses,
            "anio": payload.informacion_global_anio,
        }
    concepto_cfdi = "LITRO DE GAS LP" if (is_transfer or (hyp_mode == "disabled" and not payload.hyp_experimental_diagnostics)) else payload.concepto
    metodo_pago = "PUE" if is_transfer else payload.metodo_pago
    forma_pago = (payload.forma_pago or "01") if is_transfer else payload.forma_pago
    transfer_email_source = payload.transfer_email if payload.transfer_email_provided else (payload.transfer_email or _transfer_email_from_settings(settings))
    transfer_recipients = _clean_billing_emails(transfer_email_source) if is_transfer else []
    transfer_recipient_text = ", ".join(transfer_recipients)
    customer_recipients = [] if is_transfer else _customer_invoice_recipients(cliente_row)

    xml, totals = _build_gas_lp_consumo_xml(
        issuer=issuer,
        receptor=receptor,
        litros=payload.litros,
        precio_unitario=payload.precio_unitario,
        concepto=concepto_cfdi,
        forma_pago=forma_pago,
        metodo_pago=metodo_pago,
        descuento=payload.descuento,
        iva_rate=payload.iva_rate,
        serie=serie_factura,
        folio=folio_factura,
        comentarios=payload.comentarios,
        fecha=payload.fecha,
        clave_prod_serv=clave_prod_serv,
        no_identificacion=payload.no_identificacion,
        unidad=payload.unidad,
        hyp=hyp,
        informacion_global=informacion_global,
        allow_zero_total=is_transfer,
    )
    hyp_node_xml = _gas_lp_hyp_xml_fragment(hyp)
    sw_config = sw_runtime_config()
    sw_sandbox = _sw_config_looks_like_sandbox(sw_config)
    debug_payload = {
        "event": "gas_lp_hyp_pre_timbrado",
        "created_at": _now_iso(),
        "perfil_id": user.get("perfil_id"),
        "tenant_id": user.get("tenant_id"),
        "rfc_emisor": issuer.get("rfc") or "",
        "usuario_que_timbra_id": user.get("id"),
        "usuario_que_timbra": user.get("display_name") or "",
        "empresa_asignada_id": user.get("perfil_id"),
        "empresa_asignada": profile.get("nombre") or "",
        "empresa_asignada_rfc": profile.get("rfc") or "",
        "ambiente_sw_actual": sw_config.get("sw_env") or "",
        "app_env": sw_config.get("app_env") or "",
        "endpoint_sw_usado": sw_config.get("xml_issue_url") or "",
        "base_url_sw": sw_config.get("base_url") or "",
        "modo_sandbox": sw_sandbox,
        "timbrado_real_o_prueba": "prueba" if sw_sandbox else "real",
        "credenciales_sw_configuradas": bool(sw_config.get("has_credentials")),
        "timbrado_real_habilitado": bool(sw_config.get("real_stamping_allowed")),
        "gas_lp_hyp_mode": hyp_mode,
        "gas_lp_hyp_disabled_warning": "",
        "facility_id": payload.facility_id,
        "instalacion": origen.get("nombre") or origen.get("clave_instalacion") or "",
        "numero_permiso_instalacion": origen.get("num_permiso") or "",
        "tipo_permiso_generado": hyp.get("tipo_permiso") or "",
        "numero_permiso_hyp": hyp.get("numero_permiso") or "",
        "hyp_experimental_diagnostics": bool(payload.hyp_experimental_diagnostics),
        "hyp_override_aplicado": hyp_override_aplicado,
        "numero_permiso_original_hyp": hyp_original.get("numero_permiso") or "",
        "tipo_permiso_original_hyp": hyp_original.get("tipo_permiso") or "",
        "numero_permiso_transformado_hyp": hyp.get("numero_permiso") or "",
        "tipo_permiso_transformado_hyp": hyp.get("tipo_permiso") or "",
        "clave_prod_serv_recibida": clave_prod_serv_original,
        "clave_prod_serv": clave_prod_serv,
        "clave_hyp_diagnostic_override": clave_hyp_diagnostic_override,
        "clave_hyp": hyp.get("clave_hyp") or "",
        "subproducto_hyp": hyp.get("subproducto_hyp") or "",
        "incluye_complemento_hyp": bool(hyp_node_xml and "HidroYPetro" in xml),
        "hidroypetro_xml": hyp_node_xml,
        "cfdi_xml_enviado": xml,
    }
    _write_gas_lp_hyp_debug_log(debug_payload)
    logger.info(
        "gas_lp_hyp_pre_timbrado usuario=%s empresa=%s empresa_rfc=%s sw_env=%s app_env=%s endpoint=%s rfc_emisor=%s sandbox=%s timbrado=%s hyp_mode=%s experimental=%s facility_id=%s instalacion=%s numero_permiso_instalacion=%s numero_permiso_original=%s numero_permiso_final=%s tipo_permiso_original=%s tipo_permiso_final=%s clave_prod_serv_recibida=%s clave_prod_serv_final=%s incluye_hyp=%s clave_hyp=%s subproducto_hyp=%s hyp_xml=%s xml_enviado=%s",
        user.get("display_name") or user.get("id") or "",
        profile.get("nombre") or "",
        profile.get("rfc") or "",
        sw_config.get("sw_env") or "",
        sw_config.get("app_env") or "",
        sw_config.get("xml_issue_url") or "",
        issuer.get("rfc") or "",
        sw_sandbox,
        "prueba" if sw_sandbox else "real",
        hyp_mode,
        bool(payload.hyp_experimental_diagnostics),
        payload.facility_id,
        origen.get("nombre") or origen.get("clave_instalacion") or "",
        origen.get("num_permiso") or "",
        hyp_original.get("numero_permiso") or "",
        hyp.get("numero_permiso") or "",
        hyp_original.get("tipo_permiso") or "",
        hyp.get("tipo_permiso") or "",
        clave_prod_serv_original,
        clave_prod_serv,
        bool(hyp_node_xml and "HidroYPetro" in xml),
        hyp.get("clave_hyp") or "",
        hyp.get("subproducto_hyp") or "",
        hyp_node_xml,
        xml,
    )
    if sw_sandbox:
        logger.warning(
            "gas_lp_timbrado_bloqueado_por_ambiente sw_env=%s endpoint=%s rfc_emisor=%s",
            sw_config.get("sw_env") or "",
            sw_config.get("xml_issue_url") or "",
            issuer.get("rfc") or "",
        )
        raise HTTPException(
            400,
            "Este emisor está configurado en modo pruebas. Cambia a producción antes de timbrar CFDI real.",
        )
    resultado = timbrar_cfdi(xml)
    if payload.hyp_experimental_diagnostics:
        diagnostic_response = {
            "ok": not bool(resultado.get("error")),
            "diagnostic": True,
            "persisted": False,
            "facility_id_recibido": payload.facility_id,
            "instalacion": origen.get("nombre") or origen.get("clave_instalacion") or "",
            "rfc_emisor": issuer.get("rfc") or "",
            "numero_permiso_original": hyp_original.get("numero_permiso") or "",
            "numero_permiso_transformado": hyp.get("numero_permiso") or "",
            "tipo_permiso_original": hyp_original.get("tipo_permiso") or "",
            "tipo_permiso_final": hyp.get("tipo_permiso") or "",
            "clave_prod_serv_final": clave_prod_serv,
            "clave_hyp_final": hyp.get("clave_hyp") or "",
            "subproducto_hyp": hyp.get("subproducto_hyp") or "",
            "gas_lp_hyp_mode": hyp_mode,
            "hidroypetro_xml": hyp_node_xml,
            "xml_enviado": xml,
            "pac_sw_response": resultado,
        }
        status_code = 400 if resultado.get("error") else 200
        return JSONResponse(diagnostic_response, status_code=status_code)
    if resultado.get("error"):
        pac_error = str(resultado["error"])
        if "CCHYP107" in pac_error or "NumeroPermiso" in pac_error:
            pac_error = (
                f"{pac_error} "
                "El PAC no acepta el permiso LP/... con el tipo de permiso seleccionado. "
                "Validar con SW/SAT el TipoPermiso exacto para esa instalación Gas LP y que el permiso esté cargado en L_CNE."
            )
        raise HTTPException(400, f"PAC rechazó la factura: {pac_error}")
    now = _now_iso()
    row = {
        **_gas_lp_invoice_scope(user, profile),
        "facility_id": payload.facility_id,
        "record_uuid": totals["folio"],
        "uuid_sat": resultado.get("uuid") or "",
        "xml_content": resultado.get("xml_timbrado") or xml,
        "pdf_url": resultado.get("pdf_url") or "",
        "status": "Vigente",
        "fecha_timbrado": now,
        "rfc_receptor": receptor["rfc"],
        "volumen_litros": float(payload.litros),
        "importe": totals["subtotal"],
        "tipo_comprobante": "I",
        "distancia_km": 1,
        "metadata": {
            "portal": "asistente_gas_lp",
            "internal_user_id": user.get("id"),
            "created_by_internal_name": user.get("display_name") or "",
            "created_by": user.get("display_name") or "",
            "empresa_asignada_id": user.get("perfil_id"),
            "empresa_asignada_nombre": profile.get("nombre") or "",
            "empresa_rfc": profile.get("rfc") or "",
            "rfc_emisor": issuer.get("rfc") or "",
            "cliente_id": None if is_transfer else payload.cliente_id,
            "cliente_nombre": receptor["nombre"],
            **_invoice_email_metadata(customer_recipients),
            "concepto": payload.concepto,
            "precio_unitario": payload.precio_unitario,
            "descuento_por_litro": payload.descuento,
            "descuento": totals["descuento"],
            "iva_rate": payload.iva_rate,
            "serie": serie_factura,
            "folio_usuario": folio_factura,
            "comentarios": payload.comentarios,
            "fecha_emision": totals["fecha"],
            "clave_prod_serv": clave_prod_serv,
            "gas_lp_hyp_mode": hyp_mode,
            "gas_lp_hyp_warning": "",
            "hidrocarburos_petroliferos": hyp,
            "no_identificacion": payload.no_identificacion,
            "unidad": payload.unidad,
            "metodo_pago": metodo_pago,
            "forma_pago": forma_pago,
            "tipo_operacion": payload.tipo_operacion,
            "operation_type": "transfer" if is_transfer else payload.tipo_operacion,
            "facility_id": payload.facility_id,
            "origen_facility_id": payload.facility_id,
            "origen_facility_name": origen.get("nombre") or "",
            "origen_nombre": origen.get("nombre") or "",
            "destino_facility_id": payload.destino_facility_id,
            "destino_facility_name": destino.get("nombre") or "",
            "destino_nombre": destino.get("nombre") or "",
            "transfer_email": transfer_recipient_text,
            "created_from": "assistant_transfer" if is_transfer else "assistant_sale",
            "observaciones": payload.comentarios,
            "generar_carta_porte": payload.generar_carta_porte,
            "vehiculo_id": payload.vehiculo_id,
            "chofer_id": payload.chofer_id,
            "ruta_id": payload.ruta_id,
            "payment_status": "pendiente_complemento" if metodo_pago.upper() == "PPD" else "pagado_pue",
            "saldo_insoluto": totals["total"] if metodo_pago.upper() == "PPD" else 0,
            "iva": totals["iva"],
            "total": totals["total"],
        },
        "created_at": now,
    }
    try:
        data = sb.table("gas_lp_facturas").insert(row).execute().data or [row]
    except Exception as exc:
        raise _safe_internal_error("gas_lp_crear_factura", exc)
    factura_row = data[0]
    email_result = None
    email_results = []
    recipients = transfer_recipients if is_transfer else customer_recipients
    recipient = transfer_recipient_text if is_transfer else ", ".join(recipients)
    if payload.enviar_correo and recipients:
        try:
            xml_timbrado = factura_row.get("xml_content") or resultado.get("xml_timbrado") or xml
            info = fiscal_pdf_info(xml_timbrado, "factura_gas_lp")
            pdf_bytes = generar_pdf_gas_lp_desde_xml(
                xml_timbrado,
                logo_data_url=settings.get("PdfLogoDataUrl", ""),
                observaciones=_gas_lp_factura_observaciones(factura_row),
            )
            for email_to in recipients:
                email_result = send_gas_lp_invoice_email(
                    to_email=email_to,
                    issuer_name=issuer["nombre"],
                    customer_name=receptor["nombre"],
                    uuid_sat=factura_row.get("uuid_sat") or resultado.get("uuid") or "",
                    total=totals["total"],
                    xml_content=xml_timbrado,
                    pdf_bytes=pdf_bytes,
                    pdf_filename=info.filename,
                    serie_folio=_gas_lp_factura_folio_label(factura_row),
                )
                email_results.append({"to": email_to, **email_result.as_metadata()})
            now_email = _now_iso()
            md = factura_row.get("metadata") if isinstance(factura_row.get("metadata"), dict) else {}
            all_ok = bool(email_results) and all(item.get("ok") for item in email_results)
            first_error = next((str(item.get("error") or "") for item in email_results if not item.get("ok")), "")
            message_ids = ", ".join(str(item.get("message_id") or "") for item in email_results if item.get("message_id"))
            md = {**md, "email_delivery": email_result.as_metadata() if email_result else {}}
            if is_transfer:
                md = {
                    **md,
                    "transfer_email_delivery": email_results,
                    "transfer_email_sent_at": now_email if all_ok else md.get("transfer_email_sent_at"),
                    "transfer_email_sent_to": recipient if all_ok else md.get("transfer_email_sent_to", recipient),
                    "transfer_email_message_id": message_ids if all_ok else md.get("transfer_email_message_id", ""),
                    "transfer_email_error": "" if all_ok else first_error,
                }
            update_payload = {
                "metadata": md,
                "email_enviado": all_ok,
                "email_enviado_at": now_email if all_ok else None,
                "email_destinatario": recipient,
                "email_error": "" if all_ok else first_error,
                "updated_at": now_email,
            }
            updated = sb.table("gas_lp_facturas").update(update_payload).eq("id", factura_row.get("id")).execute().data or []
            factura_row = updated[0] if updated else {**factura_row, **update_payload}
        except Exception as exc:
            logger.exception("gas_lp_invoice_email failed: factura=%s err=%s", factura_row.get("id"), exc)
            email_result = None
    warnings = []
    if is_transfer and recipients and (not email_results or any(not item.get("ok") for item in email_results)):
        warnings.append("CFDI timbrado correctamente, pero no se pudo enviar el correo.")
    return JSONResponse({"ok": True, "factura": factura_row, "totals": totals, "email": email_result.as_metadata() if email_result else None, "warnings": warnings})


@router.post("/internal-auth/gas-lp/transfer-email-default")
async def gas_lp_transfer_email_default(payload: GasLpTransferEmailDefaultPayload, token: str):
    ctx = _gas_lp_internal_context(token, write=True)
    user = ctx["user"]
    try:
        saved = _save_transfer_email_default(
            str(user.get("owner_user_id") or ""),
            int(user.get("perfil_id") or 0),
            payload.email,
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise _safe_internal_error("gas_lp_transfer_email_default", exc)
    return JSONResponse({"ok": True, "transfer_email_default": _transfer_email_from_settings(saved)})


@router.post("/internal-auth/gas-lp/hyp-l-cne-diagnostics")
async def gas_lp_hyp_l_cne_diagnostics(payload: GasLpHypLCNEDiagnosticPayload, token: str):
    ctx = _gas_lp_internal_context(token, write=True)
    user = ctx["user"]
    facilities_by_id = {
        int(f["id"]): f
        for f in get_facilities(user.get("owner_user_id"), "gas_lp", perfil_id=user.get("perfil_id"))
        if f.get("id") is not None
    }
    requested_ids = [int(fid) for fid in (payload.facility_ids or []) if fid]
    if payload.facility_id:
        requested_ids.append(int(payload.facility_id))
    requested_ids = list(dict.fromkeys(requested_ids))
    if not requested_ids:
        raise HTTPException(400, "Selecciona al menos una instalación para diagnosticar L_CNE.")

    results: list[dict] = []
    stopped_on_success = False
    for facility_id in requested_ids:
        facility = facilities_by_id.get(facility_id)
        if not facility:
            results.append(
                {
                    "facility_id": facility_id,
                    "ok": False,
                    "error": "Instalación no encontrada para esta empresa Gas LP.",
                }
            )
            continue
        attempts = _gas_lp_lcne_diagnostic_matrix(facility, payload.probar_claves_producto)
        for attempt in attempts:
            attempt_payload = GasLpInternalFacturaPayload(
                cliente_id=payload.cliente_id,
                publico_general=False,
                litros=payload.litros,
                precio_unitario=payload.precio_unitario,
                concepto="Gas licuado de petróleo",
                forma_pago=payload.forma_pago,
                metodo_pago=payload.metodo_pago,
                descuento=payload.descuento,
                iva_rate=payload.iva_rate,
                comentarios=f"Diagnóstico L_CNE HyP {facility.get('nombre') or facility_id} {attempt['label']} {attempt['clave_hyp']}",
                clave_prod_serv=attempt["clave_hyp"],
                no_identificacion="GLP-LTR",
                unidad="Litro",
                facility_id=facility_id,
                enviar_correo=False,
                hyp_experimental_diagnostics=True,
                hyp_numero_permiso_override=attempt["numero_permiso"],
                hyp_tipo_permiso_override=attempt["tipo_permiso"],
                hyp_clave_hyp_override=attempt["clave_hyp"],
            )
            try:
                response = await gas_lp_internal_crear_factura(attempt_payload, token)
                try:
                    body = json.loads(response.body.decode("utf-8"))
                except Exception:
                    body = {"raw_response": response.body.decode("utf-8", errors="replace")}
                status_code = getattr(response, "status_code", 200)
            except HTTPException as exc:
                body = {"ok": False, "detail": exc.detail}
                status_code = exc.status_code
            pac_response = body.get("pac_sw_response") if isinstance(body, dict) else None
            pac_error = ""
            if isinstance(pac_response, dict):
                pac_error = str(pac_response.get("error") or pac_response.get("message") or pac_response.get("detail") or "")
            result = {
                "facility_id": facility_id,
                "instalacion": facility.get("nombre") or facility.get("clave_instalacion") or "",
                "permiso_real_instalacion": facility.get("num_permiso") or "",
                "attempt_label": attempt["label"],
                "permiso_xml": attempt["numero_permiso"],
                "tipo_permiso": attempt["tipo_permiso"],
                "clave_hyp": attempt["clave_hyp"],
                "clave_prod_serv": body.get("clave_prod_serv_final") if isinstance(body, dict) else attempt["clave_hyp"],
                "subproducto_hyp": body.get("subproducto_hyp") if isinstance(body, dict) else GAS_LP_HYP_SUBPRODUCTO,
                "http_status": status_code,
                "ok": bool(body.get("ok")) if isinstance(body, dict) else False,
                "diagnostic": bool(body.get("diagnostic")) if isinstance(body, dict) else False,
                "persisted": body.get("persisted") if isinstance(body, dict) else None,
                "hidroypetro_xml": body.get("hidroypetro_xml") if isinstance(body, dict) else "",
                "xml_enviado": body.get("xml_enviado") if isinstance(body, dict) else "",
                "pac_sw_response": pac_response or body,
                "error_resumen": pac_error or (str(body.get("detail") or "") if isinstance(body, dict) else ""),
            }
            results.append(result)
            _write_gas_lp_hyp_debug_log(
                {
                    "event": "gas_lp_hyp_l_cne_matrix_attempt",
                    "created_at": _now_iso(),
                    **{k: result.get(k) for k in (
                        "facility_id",
                        "instalacion",
                        "permiso_real_instalacion",
                        "attempt_label",
                        "permiso_xml",
                        "tipo_permiso",
                        "clave_hyp",
                        "http_status",
                        "ok",
                        "persisted",
                        "error_resumen",
                    )},
                }
            )
            if result["ok"] and payload.stop_on_success:
                stopped_on_success = True
                break
        if stopped_on_success:
            break

    return JSONResponse(
        {
            "ok": True,
            "diagnostic": True,
            "persisted": False,
            "stopped_on_success": stopped_on_success,
            "attempts": results,
        }
    )


@router.get("/internal-auth/gas-lp/detected-loads")
async def gas_lp_detected_loads(token: str, search: str | None = None, status: str | None = None):
    ctx = _internal_session(token, "gas_lp")
    user = ctx["user"]
    tenant_id = user.get("tenant_id")
    perfil_id = user.get("perfil_id")
    sb = get_supabase_admin()
    try:
        q = sb.table("detected_loads").select("*, cfdi_sat_inbox(uuid,rfc_emisor,nombre_emisor,fecha,total)")
        q = q.eq("tenant_id", tenant_id)
        if perfil_id is not None:
            q = q.eq("perfil_id", perfil_id)
        if status:
            q = q.eq("status", status)
        rows = q.order("created_at", desc=True).limit(50).execute().data or []
    except Exception as exc:
        raise _safe_internal_error("detected_loads", exc)

    needle = (search or "").strip().lower()
    loads = []
    for row in rows:
        cfdi = row.get("cfdi_sat_inbox") or {}
        item = {
            "id": row.get("id"),
            "source": "detected_loads",
            "status": row.get("status"),
            "status_label": _status_label(row.get("status")),
            "proveedor": cfdi.get("nombre_emisor") or row.get("proveedor_id") or "Proveedor por confirmar",
            "rfc_proveedor": cfdi.get("rfc_emisor") or "",
            "empresa": f"Perfil {perfil_id or '—'}",
            "destino_detectado": row.get("destino_detectado") or "Por confirmar",
            "producto_detectado": row.get("producto_detectado") or "Por confirmar",
            "litros_detectados": row.get("litros_detectados"),
            "unidad_detectada": row.get("unidad_detectada") or "L",
            "uuid": cfdi.get("uuid") or "",
            "fecha_detectada": row.get("fecha_detectada") or cfdi.get("fecha"),
            "confidence_score": row.get("confidence_score") or 0,
        }
        haystack = " ".join(str(item.get(k) or "") for k in (
            "proveedor", "rfc_proveedor", "uuid", "producto_detectado", "litros_detectados", "fecha_detectada"
        )).lower()
        if not needle or needle in haystack:
            loads.append(item)

    source = "real" if loads else "empty"
    states = [
        {"key": "sin_sincronizar", "label": "Sin sincronizar"},
        {"key": "buscando_cfdi", "label": "Buscando CFDI"},
        {"key": "new", "label": "Nueva carga detectada"},
        {"key": "pending_confirmation", "label": "Pendiente de confirmar"},
        {"key": "carta_porte_created", "label": "Carta Porte borrador"},
    ]
    return JSONResponse({"ok": True, "source": source, "loads": loads, "states": states})


@router.post("/internal-auth/gas-lp/detected-loads/{load_id}/action")
async def gas_lp_detected_load_action(load_id: str, payload: DetectedLoadAction, token: str):
    ctx = _internal_session(token, "gas_lp")
    user = ctx["user"]
    action = (payload.action or "").strip().lower()
    if action not in {"confirm", "ignore", "edit"}:
        raise HTTPException(400, "Acción inválida.")
    status_by_action = {
        "confirm": "carta_porte_created",
        "ignore": "rejected",
        "edit": "pending_confirmation",
    }
    update = {
        "status": status_by_action[action],
        "updated_at": _now_iso(),
    }
    if action == "confirm":
        update["confirmed_by"] = user.get("owner_user_id")
        update["confirmed_at"] = _now_iso()
    if payload.updates:
        for key in ("producto_detectado", "litros_detectados", "unidad_detectada", "origen_detectado", "destino_detectado", "assigned_operator_id"):
            if key in payload.updates:
                update[key] = payload.updates[key]
    try:
        q = get_supabase_admin().table("detected_loads").update(update).eq("id", load_id).eq("tenant_id", user.get("tenant_id"))
        if user.get("perfil_id") is not None:
            q = q.eq("perfil_id", user.get("perfil_id"))
        q.execute()
    except Exception as exc:
        raise _safe_internal_error("detected_load_action", exc)
    return JSONResponse({"ok": True, "status": update["status"], "message": _status_label(update["status"])})
