# routes/transporte.py
# ─────────────────────────────────────────────────────────────────────────────
# Endpoints del módulo TRANSPORTE DE HIDROCARBUROS
# Completamente aislado de Gas LP — no importa ni modifica nada de Gas LP.
#
# Todas las tablas usan el prefijo tr_ para no colisionar con Gas LP.
# La sección 'transporte' se valida via user_sections (Supabase).
#
# Endpoints:
#   GET  /api/tr/catalogo/productos      — ClaveProducto + SubProducto
#   POST /api/tr/viajes                  — Registrar viaje
#   GET  /api/tr/viajes                  — Listar viajes del usuario
#   GET  /api/tr/viajes/{id}             — Detalle de un viaje
#   PUT  /api/tr/viajes/{id}             — Editar viaje no timbrado
#   DELETE /api/tr/viajes/{id}           — Eliminar viaje no timbrado
#   POST /api/tr/viajes/{id}/timbrar     — Timbrar CFDI del viaje
#   POST /api/tr/viajes/{id}/cancelar    — Cancelar CFDI
#   GET  /api/tr/facturas                — Listar CFDIs timbrados
#   GET  /api/tr/facturas/{id}/xml       — Descargar XML
#   POST /api/tr/covol/generar           — Generar JSON covol mensual
#   GET  /api/tr/choferes                — CRUD choferes
#   POST /api/tr/choferes
#   PUT  /api/tr/choferes/{id}
#   DELETE /api/tr/choferes/{id}
#   GET  /api/tr/vehiculos               — CRUD vehículos
#   POST /api/tr/vehiculos
#   PUT  /api/tr/vehiculos/{id}
#   DELETE /api/tr/vehiculos/{id}
#   GET  /api/tr/rutas                   — CRUD rutas
#   POST /api/tr/rutas
#   PUT  /api/tr/rutas/{id}
#   DELETE /api/tr/rutas/{id}
#   GET  /api/tr/clientes                — CRUD clientes transporte
#   POST /api/tr/clientes
#   PUT  /api/tr/clientes/{id}
#   DELETE /api/tr/clientes/{id}
#   GET  /api/tr/settings                — Config del módulo transporte
#   PUT  /api/tr/settings
# ─────────────────────────────────────────────────────────────────────────────

from __future__ import annotations
import json
import logging
import os
import re
import hashlib
import secrets
from io import BytesIO
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from typing import Optional

from fastapi import APIRouter, File, Form, Header, HTTPException, Query, UploadFile
from fastapi.responses import JSONResponse, Response

from routes.auth import obtener_acceso_modulo, verify_token
from supabase_config import get_supabase, get_supabase_admin, get_supabase_for_user
from services.product_catalog import get_all_productos, validar_producto_completo
from services.cne_validator import validar_num_permiso
from services.transport_builder import build_cfdi_transporte, build_cfdi_cancelacion_transporte
from services.service_invoice_builder import build_cfdi_servicio_transporte, money
from services.transport_transformer import (
    build_transport_covol, save_transport_covol, transport_covol_to_json
)
from services.carta_porte_pdf import extraer_info_pdf, generar_pdf_carta_porte_desde_xml
from services.fiscal_pdf import (
    audit_fiscal_pdf_event,
    fiscal_pdf_info,
    generar_pdf_ingreso_desde_xml,
    save_fiscal_artifacts,
)
from services.carta_porte_validation import requiere_complemento_hidrocarburos, validar_xml_carta_porte_transporte
from services.sat_xml_extractor import extraer_factura_timbrada_sat
from models.transport_schemas import (
    ViajeCreate, TimbradoViajeRequest, CancelacionViajeRequest,
    FacturaServicioCreate,
    GenerarCovolRequest, ChoferTransporteCreate, VehiculoTransporteCreate,
    RutaTransporteCreate, ClienteTransporteCreate,
)
from services.sw_sapien import timbrar_cfdi, cancelar_cfdi, emitir_timbrar_json

logger = logging.getLogger(__name__)
router = APIRouter()

# ── Prefijo de todas las tablas de transporte ─────────────────────────────────
# NUNCA modificar tablas sin prefijo tr_ (esas son de Gas LP)
_TBL_VIAJES    = "tr_viajes"
_TBL_CFDI      = "tr_cfdi"
_TBL_FACT_SERV = "tr_facturas_servicio"
_TBL_FACT_SERV_CARTAS = "tr_facturas_servicio_cartas"
_TBL_CHOFERES  = "tr_choferes"
_TBL_VEHICULOS = "tr_vehiculos"
_TBL_RUTAS     = "tr_rutas"
_TBL_CLIENTES  = "tr_clientes"
_TBL_SETTINGS  = "tr_settings"
_TBL_COVOL     = "tr_covol_reports"
_TBL_EVENTOS   = "tr_viaje_eventos"
_TBL_DOCS      = "tr_viaje_documentos"
_TBL_TARIFAS   = "tr_tarifas"
_TBL_GASTOS    = "tr_gastos_viaje"
_TBL_LIQS      = "tr_liquidaciones"
_TBL_LIQ_ITEMS = "tr_liquidacion_items"
_TBL_NOTIFS    = "tr_notificaciones"
_TBL_OPER_ACC  = "tr_operador_accesos"
_TBL_IMPORTS   = "tr_importaciones"

MODULO = "transporte"
_RFC_RE = re.compile(r"^[A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3}$", re.IGNORECASE)
_CP_RE = re.compile(r"^\d{5}$")
_REGIMENES_PERSONA_MORAL = {"601", "603", "610", "620", "622", "623", "624", "626"}
_REGIMENES_PERSONA_FISICA = {"605", "606", "607", "608", "611", "612", "614", "615", "616", "621", "625", "626"}
_RFC_PRUEBAS_SAT = {
    # CSD/RFC de pruebas publicado por SAT/SW. El nombre debe ir exactamente así para CFDI 4.0.
    "EKU9003173C9": {
        "nombre": "ESCUELA KEMPER URGATE",
        "cp": "42501",
        "regimen_fiscal": "601",
    },
}


# ── Helpers de autenticación ──────────────────────────────────────────────────

def _auth(authorization: str) -> tuple[str, str]:
    """
    Valida Bearer token y devuelve (user_id, access_token).
    Verifica que el usuario tenga sección 'transporte' en user_sections.
    """
    if not authorization.startswith("Bearer "):
        raise HTTPException(401, "No autenticado.")
    token = authorization[7:]
    uid = verify_token(token)
    if not uid:
        raise HTTPException(401, "Token inválido o expirado.")
    # Verificar sección transporte
    try:
        sb = get_supabase_for_user(token)
        res = sb.table("user_sections").select("section").eq("user_id", uid).execute()
        secciones = {(r.get("section") or "").strip().lower() for r in (res.data or [])}
        if MODULO not in secciones:
            raise HTTPException(403, "Este usuario no tiene acceso al módulo de transporte.")
    except HTTPException:
        raise
    except Exception as e:
        logger.warning("No se pudo verificar sección para %s: %s", uid, e)
        raise HTTPException(403, "No se pudo verificar acceso al módulo de transporte.")
    return uid, token


def _sb(token: str):
    return get_supabase_for_user(token)


def _parse_perfil_id(raw: str | None) -> Optional[int]:
    try:
        v = int(str(raw or "").strip())
        return v if v > 0 else None
    except (TypeError, ValueError):
        return None


def _perfil(perfil_id: Optional[int] = None, x_perfil_id: str = "") -> Optional[int]:
    pid = perfil_id or _parse_perfil_id(x_perfil_id)
    if not pid:
        raise HTTPException(400, "Selecciona un perfil/empresa activo para operar Transporte.")
    return pid


def _settings_transporte(uid: str, token: str, perfil_id: Optional[int] = None) -> dict:
    """Obtiene la configuración del módulo transporte para el usuario/perfil."""
    try:
        sb  = _sb(token)
        q   = sb.table(_TBL_SETTINGS).select("data").eq("user_id", uid)
        if perfil_id:
            q = q.eq("perfil_id", perfil_id)
        else:
            q = q.is_("perfil_id", "null")
        res = q.limit(1).execute()
        rows = res.data or []
        return rows[0].get("data", {}) if rows else {}
    except Exception as e:
        logger.warning("No se pudo obtener settings transporte para %s: %s", uid, e)
        return {}


def _role_transporte(uid: str, token: str) -> str:
    acceso = obtener_acceso_modulo(uid, MODULO, access_token=token)
    return (acceso.get("role") or "user").lower()


def _require_admin_transporte(uid: str, token: str) -> None:
    if _role_transporte(uid, token) != "admin":
        raise HTTPException(403, "Acceso restringido a administradores de Transporte.")


def _get_chofer(uid: str, token: str, chofer_id: int) -> dict:
    sb  = _sb(token)
    res = sb.table(_TBL_CHOFERES).select("*").eq("id", chofer_id).eq("user_id", uid).limit(1).execute()
    rows = res.data or []
    if not rows:
        raise HTTPException(404, f"Chofer {chofer_id} no encontrado.")
    return rows[0]


def _get_vehiculo(uid: str, token: str, vehiculo_id: int) -> dict:
    sb  = _sb(token)
    res = sb.table(_TBL_VEHICULOS).select("*").eq("id", vehiculo_id).eq("user_id", uid).limit(1).execute()
    rows = res.data or []
    if not rows:
        raise HTTPException(404, f"Vehículo {vehiculo_id} no encontrado.")
    return rows[0]


def _editable_viaje(status: str) -> bool:
    """Permite cambios solo antes de timbrar Carta Porte."""
    return (status or "").lower() in {"borrador", "programado", "error"}


def _validar_rfc_cp_config(data: dict) -> None:
    for campo in ("RfcContribuyente", "RfcProveedor"):
        valor = re.sub(r"[^A-Z0-9Ñ&]", "", str(data.get(campo, "") or "").upper())
        if valor:
            data[campo] = valor
        if valor and not _RFC_RE.match(valor):
            raise HTTPException(400, f"{campo} tiene formato inválido para SAT: {valor}.")
    cp = str(data.get("CodigoPostal", "") or "").strip()
    if cp and not _CP_RE.match(cp):
        raise HTTPException(400, "CodigoPostal debe tener 5 dígitos.")
    if data.get("RfcContribuyente") and data.get("RegimenFiscal"):
        _validar_regimen_para_rfc(data.get("RfcContribuyente", ""), data.get("RegimenFiscal", ""), "emisor")


def _tipo_persona_rfc(rfc: str) -> str:
    limpio = re.sub(r"[^A-Z0-9Ñ&]", "", str(rfc or "").upper())
    if len(limpio) == 12:
        return "moral"
    if len(limpio) == 13:
        return "fisica"
    raise HTTPException(400, f"RFC emisor inválido para SAT: {limpio or '(vacío)'}.")


def _validar_regimen_para_rfc(rfc: str, regimen: str, contexto: str = "emisor") -> None:
    regimen = str(regimen or "").strip()
    tipo = _tipo_persona_rfc(rfc)
    permitidos = _REGIMENES_PERSONA_MORAL if tipo == "moral" else _REGIMENES_PERSONA_FISICA
    if regimen not in permitidos:
        etiqueta = "persona moral" if tipo == "moral" else "persona física"
        raise HTTPException(
            400,
            f"Régimen fiscal {contexto} {regimen or '(vacío)'} no corresponde al RFC {rfc} ({etiqueta}). "
            f"Corrige Configuración antes de timbrar."
        )


def _normalizar_nombre_fiscal(nombre: str) -> str:
    return re.sub(r"\s+", " ", str(nombre or "").strip().upper())


def _normalizar_receptor_cfdi(rfc: str, nombre: str, cp: str = "", regimen: str = "") -> dict:
    rfc_limpio = re.sub(r"[^A-Z0-9Ñ&]", "", str(rfc or "").upper())
    normalizado = {
        "rfc": rfc_limpio,
        "nombre": _normalizar_nombre_fiscal(nombre),
        "cp": str(cp or "").strip(),
        "regimen_fiscal": str(regimen or "").strip(),
    }
    prueba = _RFC_PRUEBAS_SAT.get(rfc_limpio)
    if prueba:
        normalizado.update(prueba)
    return normalizado


def _validar_datos_cfdi_receptor(rfc: str, regimen: str, cp: str, uso_cfdi: str) -> None:
    if not _RFC_RE.match((rfc or "").strip().upper()):
        raise HTTPException(400, "RFC receptor inválido para CFDI 4.0.")
    if not _CP_RE.match((cp or "").strip()):
        raise HTTPException(400, "Código postal receptor inválido para CFDI 4.0.")
    if not str(regimen or "").strip():
        raise HTTPException(400, "Régimen fiscal receptor requerido para CFDI 4.0.")
    _validar_regimen_para_rfc(rfc, regimen, "receptor")
    if not str(uso_cfdi or "").strip():
        raise HTTPException(400, "Uso CFDI requerido para CFDI 4.0.")


def _validar_totales_servicio(
    payload: FacturaServicioCreate,
    esperado: dict,
) -> None:
    """Evita facturar con impuestos manipulados en frontend; el servidor recalcula desde tarifas."""
    checks = {
        "subtotal": (payload.subtotal, esperado.get("subtotal")),
        "iva": (payload.iva, esperado.get("iva")),
        "retención": (payload.retencion, esperado.get("retencion")),
        "total": (payload.total, esperado.get("total")),
    }
    for label, (recibido, calc) in checks.items():
        if abs(float(recibido or 0) - float(calc or 0)) > 0.01:
            raise HTTPException(
                400,
                f"{label.title()} inválido. El sistema calculó {float(calc or 0):.2f} con las tarifas/impuestos configurados.",
            )


def _periodo_bounds(periodo: str) -> tuple[str, str]:
    """Convierte YYYY-MM a rango ISO para columnas timestamptz."""
    anio = int(periodo[:4])
    mes = int(periodo[5:7])
    inicio = datetime(anio, mes, 1, tzinfo=timezone.utc)
    if mes == 12:
        fin = datetime(anio + 1, 1, 1, tzinfo=timezone.utc)
    else:
        fin = datetime(anio, mes + 1, 1, tzinfo=timezone.utc)
    return inicio.isoformat(), fin.isoformat()


def _periodo_liquidacion_bounds(periodo: str, periodo_tipo: str = "") -> tuple[str, str]:
    """Soporta periodos mensuales YYYY-MM y quincenales YYYY-MM-Q1/Q2."""
    raw = str(periodo or datetime.now(timezone.utc).strftime("%Y-%m")).strip()
    tipo = str(periodo_tipo or "").strip().lower()
    if raw.endswith("-Q1") or raw.endswith("-Q2"):
        tipo = raw[-2:].lower()
        raw = raw[:7]
    anio = int(raw[:4])
    mes = int(raw[5:7])
    if tipo in {"q1", "quincena1", "primera"}:
        inicio = datetime(anio, mes, 1, tzinfo=timezone.utc)
        fin = datetime(anio, mes, 16, tzinfo=timezone.utc)
    elif tipo in {"q2", "quincena2", "segunda"}:
        inicio = datetime(anio, mes, 16, tzinfo=timezone.utc)
        if mes == 12:
            fin = datetime(anio + 1, 1, 1, tzinfo=timezone.utc)
        else:
            fin = datetime(anio, mes + 1, 1, tzinfo=timezone.utc)
    else:
        inicio_s, fin_s = _periodo_bounds(raw)
        return inicio_s, fin_s
    return inicio.isoformat(), fin.isoformat()


def _periodo_liquidacion_label(periodo: str, periodo_tipo: str = "") -> str:
    raw = str(periodo or datetime.now(timezone.utc).strftime("%Y-%m")).strip()
    tipo = str(periodo_tipo or "").strip().lower()
    if raw.endswith("-Q1") or raw.endswith("-Q2"):
        return raw
    if tipo in {"q1", "quincena1", "primera"}:
        return f"{raw}-Q1"
    if tipo in {"q2", "quincena2", "segunda"}:
        return f"{raw}-Q2"
    return raw


def _safe_float(v, default: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def _productos_from_row(viaje: dict) -> list[dict]:
    try:
        productos = json.loads(viaje.get("productos_json") or "[]")
        return productos if isinstance(productos, list) else []
    except Exception:
        return []


def _registrar_evento(
    sb,
    uid: str,
    perfil_id: Optional[int],
    viaje_id: int,
    event_type: str,
    title: str,
    description: str = "",
    actor_type: str = "system",
    actor_id: str = "",
    metadata: Optional[dict] = None,
) -> None:
    """Bitacora operativa no critica: nunca debe romper timbrado/facturacion."""
    try:
        sb.table(_TBL_EVENTOS).insert({
            "user_id": uid,
            "perfil_id": perfil_id,
            "viaje_id": viaje_id,
            "event_type": event_type,
            "title": title,
            "description": description,
            "actor_type": actor_type,
            "actor_id": actor_id,
            "metadata": metadata or {},
            "created_at": datetime.now(timezone.utc).isoformat(),
        }).execute()
    except Exception as e:
        logger.info("Evento operativo omitido (%s/%s): %s", viaje_id, event_type, e)


def _build_document_path(uid: str, perfil_id: Optional[int], viaje_id: int, tipo: str, filename: str) -> str:
    clean_name = re.sub(r"[^A-Za-z0-9._-]+", "_", filename or "documento")
    perfil = str(perfil_id or "default")
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    return f"{uid}/{perfil}/viajes/{viaje_id}/{tipo}/{stamp}_{clean_name}"


def _build_cfdi_document_path(uid: str, perfil_id: Optional[int], viaje_id: int, tipo: str, filename: str) -> str:
    clean_name = re.sub(r"[^A-Za-z0-9._-]+", "_", filename or "documento")
    perfil = str(perfil_id or "default")
    return f"{uid}/{perfil}/viajes/{viaje_id}/{tipo}/{clean_name}"


def _guardar_cfdi_pdf_en_expediente(sb, uid: str, cfdi_row: dict, pdf_bytes: bytes, filename: str, metadata: dict) -> None:
    """Guarda el PDF generado en Storage/documentos sin bloquear la descarga si Storage falla."""
    viaje_id = cfdi_row.get("viaje_id")
    if not viaje_id:
        return
    perfil_id = cfdi_row.get("perfil_id")
    bucket = "transport-documents"
    path = _build_cfdi_document_path(uid, perfil_id, int(viaje_id), "carta_porte_pdf", filename)
    try:
        sb.storage.from_(bucket).upload(path, pdf_bytes, {"content-type": "application/pdf", "upsert": "true"})
    except Exception as e:
        logger.info("PDF Carta Porte generado pero no guardado en Storage (%s): %s", viaje_id, e)
        return
    try:
        existentes = (
            sb.table(_TBL_DOCS)
            .select("id")
            .eq("user_id", uid)
            .eq("viaje_id", viaje_id)
            .eq("storage_path", path)
            .limit(1)
            .execute()
            .data
            or []
        )
        if not existentes:
            sb.table(_TBL_DOCS).insert({
                "user_id": uid,
                "perfil_id": perfil_id,
                "viaje_id": viaje_id,
                "tipo": "carta_porte_pdf",
                "nombre": filename,
                "storage_bucket": bucket,
                "storage_path": path,
                "mime_type": "application/pdf",
                "size_bytes": len(pdf_bytes),
                "uuid_sat": str(cfdi_row.get("uuid_sat") or ""),
                "metadata": metadata,
                "created_by": uid,
            }).execute()
            _registrar_evento(
                sb, uid, perfil_id, int(viaje_id), "documento_generado",
                "PDF Carta Porte generado", filename, "system", "ge_control",
                {"tipo": "carta_porte_pdf", "storage_path": path, **metadata},
            )
    except Exception as e:
        logger.info("PDF Carta Porte guardado en Storage pero no registrado en documentos (%s): %s", viaje_id, e)


def _guardar_cfdi_xml_en_expediente(sb, uid: str, cfdi_row: dict, xml_content: str, filename: str, metadata: dict) -> None:
    """Guarda XML timbrado en Storage/documentos por UUID sin bloquear el timbrado si Storage falla."""
    viaje_id = cfdi_row.get("viaje_id")
    if not viaje_id or not xml_content:
        return
    perfil_id = cfdi_row.get("perfil_id")
    bucket = "transport-documents"
    path = _build_cfdi_document_path(uid, perfil_id, int(viaje_id), "carta_porte_xml", filename)
    content = xml_content.encode("utf-8")
    try:
        sb.storage.from_(bucket).upload(path, content, {"content-type": "application/xml", "upsert": "true"})
    except Exception as e:
        logger.info("XML Carta Porte no guardado en Storage (%s): %s", viaje_id, e)
        return
    try:
        existentes = (
            sb.table(_TBL_DOCS)
            .select("id")
            .eq("user_id", uid)
            .eq("viaje_id", viaje_id)
            .eq("storage_path", path)
            .limit(1)
            .execute()
            .data
            or []
        )
        if not existentes:
            sb.table(_TBL_DOCS).insert({
                "user_id": uid,
                "perfil_id": perfil_id,
                "viaje_id": viaje_id,
                "tipo": "carta_porte_xml",
                "nombre": filename,
                "storage_bucket": bucket,
                "storage_path": path,
                "mime_type": "application/xml",
                "size_bytes": len(content),
                "uuid_sat": str(cfdi_row.get("uuid_sat") or ""),
                "metadata": metadata,
                "created_by": uid,
            }).execute()
    except Exception as e:
        logger.info("XML Carta Porte guardado en Storage pero no registrado en documentos (%s): %s", viaje_id, e)


def _hash_operator_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _crear_notificacion_manual(sb, uid: str, pid, viaje_id: int | None, mensaje: str, destinatario: str = "", metadata: Optional[dict] = None):
    """Registra una notificación pendiente; WhatsApp/email quedan como canal futuro, no base de datos."""
    try:
        sb.table(_TBL_NOTIFS).insert({
            "user_id": uid,
            "perfil_id": pid,
            "viaje_id": viaje_id,
            "canal": "manual",
            "destinatario": destinatario,
            "mensaje": mensaje,
            "status": "pendiente",
            "metadata": metadata or {},
        }).execute()
    except Exception as e:
        logger.info("No se pudo registrar notificación manual: %s", e)


def _normalizar(texto: str) -> str:
    return re.sub(r"\s+", " ", (texto or "").strip().lower())


def _calcular_tarifa_operativa(viaje: dict, tarifas: list[dict]) -> dict:
    """Selecciona la mejor tarifa configurable por prioridad y calcula totales."""
    productos = _productos_from_row(viaje)
    primer_producto = productos[0] if productos else {}
    litros = sum(_safe_float(p.get("volumen_litros")) for p in productos)
    kilos = sum(_safe_float(p.get("kilos") or p.get("peso_kg")) for p in productos)
    if kilos <= 0:
        kilos = litros * 0.75
    ruta_id = viaje.get("ruta_id")
    cliente_id = viaje.get("cliente_id")
    cliente_rfc = _normalizar(viaje.get("rfc_receptor"))
    origen = _normalizar(viaje.get("nombre_origen") or viaje.get("cp_origen"))
    destino = _normalizar(viaje.get("nombre_destino") or viaje.get("cp_destino"))
    producto = _normalizar(primer_producto.get("descripcion") or primer_producto.get("clave_producto") or "")

    def score(t: dict) -> int:
        s = 0
        if t.get("ruta_id") and ruta_id and int(t.get("ruta_id")) == int(ruta_id):
            s += 80
        if t.get("cliente_id"):
            if cliente_id and int(t.get("cliente_id")) == int(cliente_id):
                s += 35
            else:
                s -= 60
        if _normalizar(t.get("origen")) and _normalizar(t.get("origen")) in origen:
            s += 20
        if _normalizar(t.get("destino")) and _normalizar(t.get("destino")) in destino:
            s += 20
        tp = _normalizar(t.get("producto"))
        if tp and (tp in producto or tp in {"magna/diesel/premium", "todos", "*"}):
            s += 15
        return s - int(t.get("prioridad") or 100)

    activas = [t for t in tarifas if t.get("activo", True)]
    tarifa = sorted(activas, key=score, reverse=True)[0] if activas else {}
    regla = tarifa.get("regla_calculo") or "manual"
    rate = _safe_float(tarifa.get("tarifa"))
    if regla == "litros":
        subtotal = litros * rate
    elif regla == "kilos":
        subtotal = kilos * rate
    elif regla == "distancia":
        subtotal = _safe_float(viaje.get("distancia_km")) * rate
    elif regla == "viaje":
        subtotal = rate
    else:
        subtotal = sum(_safe_float(p.get("importe")) for p in productos)
    subtotal = round(subtotal, 2)
    aplica_iva = bool(tarifa.get("aplica_iva", True)) if tarifa else True
    aplica_retencion = bool(tarifa.get("aplica_retencion", True)) if tarifa else False
    iva_tasa = _safe_float(tarifa.get("iva_tasa"), 0.16) if tarifa else 0.16
    retencion_tasa = _safe_float(tarifa.get("retencion_tasa"), 0.04) if tarifa else 0.0
    iva = round(subtotal * iva_tasa, 2) if aplica_iva else 0.0
    ret = round(subtotal * retencion_tasa, 2) if aplica_retencion else 0.0
    total = round(subtotal + iva - ret, 2)
    return {
        "tarifa_id": tarifa.get("id"),
        "regla_calculo": regla,
        "tarifa": rate,
        "litros": round(litros, 3),
        "kilos": round(kilos, 3),
        "subtotal": subtotal,
        "iva": iva,
        "retencion": ret,
        "total": total,
        "iva_tasa": iva_tasa,
        "retencion_tasa": retencion_tasa,
        "aplica_iva": aplica_iva,
        "aplica_retencion": aplica_retencion,
        "match_score": score(tarifa) if tarifa else 0,
    }


def _sumar_calculos_servicio(viajes: list[dict], tarifas: list[dict]) -> dict:
    """Suma subtotal, IVA, retención y total de una factura de servicio desde tarifas configurables."""
    items = []
    subtotal = iva = retencion = total = 0.0
    tasas_iva: set[float] = set()
    tasas_ret: set[float] = set()
    aplica_iva = False
    aplica_retencion = False
    for viaje in viajes:
        calc = _calcular_tarifa_operativa(viaje, tarifas)
        items.append({"viaje_id": viaje.get("id"), **calc})
        subtotal += calc["subtotal"]
        iva += calc["iva"]
        retencion += calc["retencion"]
        total += calc["total"]
        if calc.get("aplica_iva"):
            aplica_iva = True
            tasas_iva.add(float(calc.get("iva_tasa") or 0))
        if calc.get("aplica_retencion"):
            aplica_retencion = True
            tasas_ret.add(float(calc.get("retencion_tasa") or 0))
    return {
        "subtotal": round(subtotal, 2),
        "iva": round(iva, 2),
        "retencion": round(retencion, 2),
        "total": round(total, 2),
        "iva_tasa": sorted(tasas_iva)[0] if len(tasas_iva) == 1 else (0.16 if aplica_iva else 0.0),
        "retencion_tasa": sorted(tasas_ret)[0] if len(tasas_ret) == 1 else (0.04 if aplica_retencion else 0.0),
        "aplica_iva": aplica_iva,
        "aplica_retencion": aplica_retencion,
        "items": items,
        "tasas_mixtas": len(tasas_iva) > 1 or len(tasas_ret) > 1,
    }


def _ruta_payload(payload: RutaTransporteCreate) -> dict:
    return {
        "nombre":        payload.nombre.strip(),
        "cp_origen":     payload.cp_origen,
        "nombre_origen": payload.nombre_origen.strip(),
        "cp_destino":    payload.cp_destino,
        "nombre_destino": payload.nombre_destino.strip(),
        "distancia_km":  payload.distancia_km,
        "duracion_estimada_min": max(int(payload.duracion_estimada_min or 0), 0),
    }


def _viaje_row(uid: str, payload: ViajeCreate, productos_json: str, volumen_total: float, status: str = "programado") -> dict:
    receptor = _normalizar_receptor_cfdi(
        payload.rfc_receptor,
        payload.nombre_receptor,
        payload.cp_receptor,
        getattr(payload, "regimen_fiscal_receptor", "601"),
    )
    return {
        "user_id":              uid,
        "perfil_id":            payload.perfil_id,
        "facility_id":          payload.facility_id,
        "chofer_id":            payload.chofer_id,
        "vehiculo_id":          payload.vehiculo_id,
        "ruta_id":              payload.ruta_id,
        "cp_origen":            payload.cp_origen,
        "nombre_origen":        payload.nombre_origen,
        "cp_destino":           payload.cp_destino,
        "nombre_destino":       payload.nombre_destino,
        "fecha_hora_salida":    payload.fecha_hora_salida,
        "fecha_hora_llegada":   payload.fecha_hora_llegada,
        "productos_json":       productos_json,
        "tipo_cfdi":            payload.tipo_cfdi,
        "rfc_receptor":         receptor["rfc"],
        "nombre_receptor":      receptor["nombre"],
        "cp_receptor":          receptor["cp"],
        "regimen_fiscal_receptor": receptor["regimen_fiscal"] or getattr(payload, "regimen_fiscal_receptor", "601"),
        "uso_cfdi":             payload.uso_cfdi,
        "num_permiso_cne":      payload.num_permiso_cne,
        "distancia_km":         payload.distancia_km,
        "duracion_estimada_min": max(int(payload.duracion_estimada_min or 0), 0),
        "volumen_total_litros": volumen_total,
        "status":               status,
        "observaciones":        payload.observaciones,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 1. CATÁLOGO DE PRODUCTOS
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/tr/catalogo/productos")
async def get_catalogo_productos(authorization: str = Header(default="")):
    """Lista todos los productos del catálogo SAT para transporte."""
    uid, _ = _auth(authorization)
    return JSONResponse({"ok": True, "productos": get_all_productos()})


@router.get("/tr/catalogo/validar-clave")
async def validar_clave_producto(
    clave_producto:    str = Query(...),
    clave_subproducto: str = Query(...),
    authorization:     str = Header(default=""),
):
    """Valida una combinación ClaveProducto + ClaveSubProducto."""
    uid, _ = _auth(authorization)
    ok, msg = validar_producto_completo(clave_producto, clave_subproducto)
    return JSONResponse({"ok": ok, "mensaje": msg})


# ══════════════════════════════════════════════════════════════════════════════
# 2. VIAJES
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/tr/viajes")
async def crear_viaje(payload: ViajeCreate, authorization: str = Header(default="")):
    """Registra un nuevo viaje de transporte de hidrocarburos."""
    uid, token = _auth(authorization)

    # Validar existencia de chofer y vehículo
    chofer   = _get_chofer(uid, token, payload.chofer_id)
    vehiculo = _get_vehiculo(uid, token, payload.vehiculo_id)

    # Validar todos los productos del viaje
    for prod in payload.productos:
        ok, msg = validar_producto_completo(prod.clave_producto, prod.clave_subproducto)
        if not ok:
            raise HTTPException(400, f"Producto inválido: {msg}")

    # Serializar productos a JSON para almacenar
    productos_json = json.dumps(
        [p.model_dump() for p in payload.productos],
        ensure_ascii=False,
    )

    # Obtener ruta si se especificó
    cp_origen  = payload.cp_origen
    cp_destino = payload.cp_destino
    nom_origen  = payload.nombre_origen
    nom_destino = payload.nombre_destino

    if payload.ruta_id:
        try:
            sb = _sb(token)
            res = sb.table(_TBL_RUTAS).select("*").eq("id", payload.ruta_id).eq("user_id", uid).limit(1).execute()
            ruta_rows = res.data or []
            if ruta_rows:
                r = ruta_rows[0]
                cp_origen   = cp_origen   or r.get("cp_origen", "")
                cp_destino  = cp_destino  or r.get("cp_destino", "")
                nom_origen  = nom_origen  or r.get("nombre_origen", "")
                nom_destino = nom_destino or r.get("nombre_destino", "")
                payload.duracion_estimada_min = payload.duracion_estimada_min or int(r.get("duracion_estimada_min") or 0)
        except Exception as e:
            logger.warning("No se pudo obtener ruta %s: %s", payload.ruta_id, e)

    volumen_total = round(sum(p.volumen_litros for p in payload.productos), 3)
    payload.cp_origen = cp_origen
    payload.cp_destino = cp_destino
    payload.nombre_origen = nom_origen
    payload.nombre_destino = nom_destino

    now = datetime.now(timezone.utc).isoformat()
    row = _viaje_row(uid, payload, productos_json, volumen_total, status="programado")
    row.update({"uuid_cfdi": "", "id_ccp": "", "created_at": now})

    try:
        sb  = _sb(token)
        res = sb.table(_TBL_VIAJES).insert(row).execute()
        viaje_id = res.data[0]["id"] if res.data else None
        if viaje_id:
            _registrar_evento(
                sb, uid, payload.perfil_id, int(viaje_id), "viaje_creado",
                "Viaje creado", "Registro inicial del viaje.", "oficina", uid,
                {"status": "programado", "volumen_total_litros": volumen_total},
            )
    except Exception as e:
        logger.error("Error al crear viaje: %s", e)
        raise HTTPException(500, f"Error al registrar viaje: {e}")

    logger.info("Viaje creado: user=%s id=%s volumen=%.2f L", uid, viaje_id, volumen_total)
    return JSONResponse({
        "ok":       True,
        "viaje_id": viaje_id,
        "volumen_total_litros": volumen_total,
        "status":   "programado",
    })


@router.put("/tr/viajes/{viaje_id}")
async def actualizar_viaje(viaje_id: int, payload: ViajeCreate, authorization: str = Header(default="")):
    """Edita un viaje mientras no tenga Carta Porte timbrada."""
    uid, token = _auth(authorization)
    sb = _sb(token)
    res = sb.table(_TBL_VIAJES).select("*").eq("id", viaje_id).eq("user_id", uid).limit(1).execute()
    rows = res.data or []
    if not rows:
        raise HTTPException(404, f"Viaje {viaje_id} no encontrado.")
    if not _editable_viaje(rows[0].get("status", "")):
        raise HTTPException(400, "Solo se pueden editar viajes en Borrador, Programado o Error.")

    _get_chofer(uid, token, payload.chofer_id)
    _get_vehiculo(uid, token, payload.vehiculo_id)
    for prod in payload.productos:
        ok, msg = validar_producto_completo(prod.clave_producto, prod.clave_subproducto)
        if not ok:
            raise HTTPException(400, f"Producto inválido: {msg}")

    if payload.ruta_id:
        ruta_res = sb.table(_TBL_RUTAS).select("*").eq("id", payload.ruta_id).eq("user_id", uid).limit(1).execute()
        ruta_rows = ruta_res.data or []
        if ruta_rows:
            r = ruta_rows[0]
            payload.cp_origen = payload.cp_origen or r.get("cp_origen", "")
            payload.cp_destino = payload.cp_destino or r.get("cp_destino", "")
            payload.nombre_origen = payload.nombre_origen or r.get("nombre_origen", "")
            payload.nombre_destino = payload.nombre_destino or r.get("nombre_destino", "")
            payload.duracion_estimada_min = payload.duracion_estimada_min or int(r.get("duracion_estimada_min") or 0)

    productos_json = json.dumps([p.model_dump() for p in payload.productos], ensure_ascii=False)
    volumen_total = round(sum(p.volumen_litros for p in payload.productos), 3)
    row = _viaje_row(uid, payload, productos_json, volumen_total, status=rows[0].get("status", "programado"))
    row.pop("user_id", None)
    try:
        sb.table(_TBL_VIAJES).update(row).eq("id", viaje_id).eq("user_id", uid).execute()
        _registrar_evento(
            sb, uid, rows[0].get("perfil_id"), viaje_id, "viaje_actualizado",
            "Viaje actualizado", "La oficina modifico datos operativos del viaje.",
            "oficina", uid, {"volumen_total_litros": volumen_total},
        )
    except Exception as e:
        raise HTTPException(500, f"Error al actualizar viaje: {e}")

    return JSONResponse({"ok": True, "viaje_id": viaje_id, "volumen_total_litros": volumen_total})


@router.delete("/tr/viajes/{viaje_id}")
async def eliminar_viaje(viaje_id: int, authorization: str = Header(default="")):
    """Elimina un viaje si todavía no tiene Carta Porte timbrada."""
    uid, token = _auth(authorization)
    sb = _sb(token)
    res = sb.table(_TBL_VIAJES).select("id,status,uuid_cfdi").eq("id", viaje_id).eq("user_id", uid).limit(1).execute()
    rows = res.data or []
    if not rows:
        raise HTTPException(404, f"Viaje {viaje_id} no encontrado.")
    row = rows[0]
    if row.get("uuid_cfdi") or not _editable_viaje(row.get("status", "")):
        raise HTTPException(400, "No se puede eliminar un viaje con Carta Porte timbrada.")
    try:
        sb.table(_TBL_VIAJES).delete().eq("id", viaje_id).eq("user_id", uid).execute()
    except Exception as e:
        raise HTTPException(500, f"Error al eliminar viaje: {e}")
    return JSONResponse({"ok": True})


@router.get("/tr/viajes")
async def listar_viajes(
    periodo:        Optional[str] = Query(None),
    status:         Optional[str] = Query(None),
    perfil_id:      Optional[int] = Query(None),
    clave_producto: Optional[str] = Query(None),
    page:           int           = Query(1, ge=1),
    page_size:      int           = Query(50, ge=1, le=200),
    authorization:  str           = Header(default=""),
):
    """Lista los viajes del usuario con filtros."""
    uid, token = _auth(authorization)
    sb = _sb(token)

    try:
        q = sb.table(_TBL_VIAJES).select("*").eq("user_id", uid).order("fecha_hora_salida", desc=True)
        if periodo:
            q = q.like("fecha_hora_salida", f"{periodo}%")
        if status:
            q = q.eq("status", status)
        if perfil_id:
            q = q.eq("perfil_id", perfil_id)

        offset = (page - 1) * page_size
        q = q.range(offset, offset + page_size - 1)

        res  = q.execute()
        rows = res.data or []

        # Si hay filtro por clave_producto, filtrar en Python (JSON en BD)
        if clave_producto:
            clave_prod_up = clave_producto.upper()
            rows = [
                r for r in rows
                if clave_prod_up in (r.get("productos_json") or "")
            ]

    except Exception as e:
        logger.error("Error al listar viajes: %s", e)
        raise HTTPException(500, f"Error al listar viajes: {e}")

    return JSONResponse({"ok": True, "viajes": rows, "total": len(rows)})


@router.get("/tr/viajes/{viaje_id}")
async def detalle_viaje(viaje_id: int, authorization: str = Header(default="")):
    """Detalle de un viaje específico."""
    uid, token = _auth(authorization)
    sb = _sb(token)
    try:
        res = sb.table(_TBL_VIAJES).select("*").eq("id", viaje_id).eq("user_id", uid).limit(1).execute()
        rows = res.data or []
        if not rows:
            raise HTTPException(404, f"Viaje {viaje_id} no encontrado.")
        viaje = rows[0]
        # Deserializar productos
        try:
            viaje["productos"] = json.loads(viaje.get("productos_json") or "[]")
        except Exception:
            viaje["productos"] = []
        return JSONResponse({"ok": True, "viaje": viaje})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error al obtener viaje: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# 3. TIMBRADO
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/tr/viajes/{viaje_id}/timbrar")
async def timbrar_viaje(
    viaje_id:      int,
    payload:       TimbradoViajeRequest,
    authorization: str = Header(default=""),
):
    """
    Timbra el CFDI de un viaje via SW Sapien.
    Genera automáticamente:
      · Complemento Carta Porte 3.1
      · Complemento Hidrocarburos y Petrolíferos 1.0
    """
    uid, token = _auth(authorization)
    sb = _sb(token)

    # Obtener viaje
    try:
        res = sb.table(_TBL_VIAJES).select("*").eq("id", viaje_id).eq("user_id", uid).limit(1).execute()
        rows = res.data or []
        if not rows:
            raise HTTPException(404, f"Viaje {viaje_id} no encontrado.")
        viaje_row = rows[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error al obtener viaje: {e}")

    if viaje_row.get("status") == "timbrado":
        raise HTTPException(400, "Este viaje ya tiene un CFDI timbrado.")

    # Obtener datos relacionados
    chofer   = _get_chofer(uid, token, viaje_row["chofer_id"])
    vehiculo = _get_vehiculo(uid, token, viaje_row["vehiculo_id"])

    # Obtener settings del módulo transporte
    settings = _settings_transporte(uid, token, viaje_row.get("perfil_id"))

    if not settings.get("RfcContribuyente"):
        raise HTTPException(400, "Configura el RFC del contribuyente en Ajustes del módulo Transporte.")

    regimen_emisor = (payload.regimen_fiscal_emisor or settings.get("RegimenFiscal") or "").strip()
    _validar_regimen_para_rfc(settings.get("RfcContribuyente", ""), regimen_emisor, "emisor")

    emisor = {
        "rfc":              settings.get("RfcContribuyente", ""),
        "nombre":           settings.get("DescripcionInstalacion", ""),
        "regimen_fiscal":   regimen_emisor,
        "domicilio_fiscal": settings.get("CodigoPostal", "20000"),
        "num_permiso_cne":  viaje_row.get("num_permiso_cne") or settings.get("NumPermiso", ""),
    }

    # Reconstruir ViajeCreate desde el row de BD
    try:
        productos_raw = json.loads(viaje_row.get("productos_json") or "[]")
        from models.transport_schemas import ProductoTransporte
        productos = [ProductoTransporte(**p) for p in productos_raw]
    except Exception as e:
        raise HTTPException(400, f"Productos del viaje inválidos: {e}")
    enforce_hidro = bool(settings.get("ValidarComplementoHidrocarburos", True))
    if enforce_hidro and requiere_complemento_hidrocarburos([p.model_dump() for p in productos]):
        raise HTTPException(
            400,
            "Timbrado bloqueado para no gastar timbres: Magna/Premium/Diésel requieren validar e incorporar "
            "el complemento Hidrocarburos y Petrolíferos junto con Carta Porte. Falta cerrar el payload exacto con SW Sapien."
        )

    from models.transport_schemas import ViajeCreate
    receptor_cfdi = _normalizar_receptor_cfdi(
        viaje_row.get("rfc_receptor", ""),
        viaje_row.get("nombre_receptor", ""),
        viaje_row.get("cp_receptor", "20000"),
        viaje_row.get("regimen_fiscal_receptor", "601"),
    )
    viaje_obj = ViajeCreate(
        chofer_id=          viaje_row["chofer_id"],
        vehiculo_id=        viaje_row["vehiculo_id"],
        ruta_id=            viaje_row.get("ruta_id"),
        cp_origen=          viaje_row.get("cp_origen", ""),
        nombre_origen=      viaje_row.get("nombre_origen", ""),
        cp_destino=         viaje_row.get("cp_destino", ""),
        nombre_destino=     viaje_row.get("nombre_destino", ""),
        fecha_hora_salida=  viaje_row["fecha_hora_salida"],
        fecha_hora_llegada= viaje_row.get("fecha_hora_llegada"),
        productos=          productos,
        tipo_cfdi=          payload.tipo_cfdi or viaje_row.get("tipo_cfdi", "T"),
        rfc_receptor=       receptor_cfdi["rfc"],
        nombre_receptor=    receptor_cfdi["nombre"],
        cp_receptor=        receptor_cfdi["cp"] or "20000",
        regimen_fiscal_receptor= receptor_cfdi["regimen_fiscal"] or "601",
        uso_cfdi=           viaje_row.get("uso_cfdi", "S01"),
        num_permiso_cne=    viaje_row.get("num_permiso_cne", ""),
        distancia_km=       float(viaje_row.get("distancia_km") or 1.0),
    )

    # Construir CFDI
    try:
        cfdi_dict, id_ccp = build_cfdi_transporte(viaje_obj, emisor, chofer, vehiculo)
    except ValueError as e:
        raise HTTPException(400, f"Error al construir CFDI: {e}")
    except Exception as e:
        logger.error("Error inesperado construyendo CFDI viaje %s: %s", viaje_id, e)
        raise HTTPException(500, f"Error interno al construir CFDI: {e}")

    # Timbrar via SW Sapien con Emision Timbrado JSON oficial.
    resultado_sw = emitir_timbrar_json(cfdi_dict)
    if not resultado_sw.get("ok"):
        err_msg = resultado_sw.get("error") or "Error desconocido"
        raise HTTPException(400, f"SW Sapien rechazó la Carta Porte: {err_msg}")

    result_data  = resultado_sw.get("data", {}) or {}
    uuid_sat     = result_data.get("uuid", "")
    xml_timbrado = result_data.get("cfdi", "")
    pdf_url      = result_data.get("pdfUrl", "")
    now_iso      = datetime.now(timezone.utc).isoformat()
    validacion_cp = validar_xml_carta_porte_transporte(
        xml_timbrado,
        [p.model_dump() for p in productos],
        enforce_hidrocarburos=enforce_hidro,
    ) if xml_timbrado else None
    carta_porte_valida = bool(validacion_cp and validacion_cp.ok)

    # Guardar CFDI en tr_cfdi
    cfdi_row = {
        "user_id":        uid,
        "viaje_id":       viaje_id,
        "perfil_id":      viaje_row.get("perfil_id"),
        "tipo_cfdi":      viaje_obj.tipo_cfdi,
        "uuid_sat":       uuid_sat,
        "id_ccp":         id_ccp,
        "xml_content":    xml_timbrado,
        "pdf_url":        pdf_url,
        "status":         "Vigente" if carta_porte_valida else "ErrorValidacion",
        "fecha_timbrado": now_iso,
        "rfc_receptor":   viaje_obj.rfc_receptor,
        "volumen_total":  float(viaje_row.get("volumen_total_litros") or 0),
        "importe_total":  round(sum(p.importe for p in productos), 2),
        "num_permiso_cne": viaje_obj.num_permiso_cne,
        "created_at":     now_iso,
    }

    try:
        inserted = sb.table(_TBL_CFDI).insert(cfdi_row).execute()
        cfdi_saved = (inserted.data or [{}])[0]
        if xml_timbrado:
            xml_filename = f"cfdi_tr_{uuid_sat or id_ccp or viaje_id}.xml"
            _guardar_cfdi_xml_en_expediente(
                sb, uid, {**cfdi_row, "id": cfdi_saved.get("id")}, xml_timbrado, xml_filename,
                {"cfdi_id": cfdi_saved.get("id"), "uuid_sat": uuid_sat, "id_ccp": id_ccp, "validacion": (validacion_cp.metadata if validacion_cp else {})},
            )
        # Actualizar status del viaje solo como timbrado cuando el XML contiene Carta Porte válida.
        sb.table(_TBL_VIAJES).update({
            "status":   "timbrado" if carta_porte_valida else "error",
            "uuid_cfdi": uuid_sat,
            "id_ccp":    id_ccp if carta_porte_valida else "",
        }).eq("id", viaje_id).execute()
        _registrar_evento(
            sb, uid, viaje_row.get("perfil_id"), viaje_id,
            "carta_porte_timbrada" if carta_porte_valida else "cfdi_timbrado_invalido_carta_porte",
            "Carta Porte timbrada" if carta_porte_valida else "CFDI timbrado sin Carta Porte válida",
            f"UUID SAT {uuid_sat}" if uuid_sat else "CFDI recibido de SW Sapien.",
            "system", "sw_sapien", {"uuid_sat": uuid_sat, "id_ccp": id_ccp},
        )
        if not carta_porte_valida:
            _registrar_evento(
                sb, uid, viaje_row.get("perfil_id"), viaje_id, "validacion_carta_porte",
                "XML no válido como Carta Porte de carretera",
                "; ".join((validacion_cp.errors if validacion_cp else ["XML vacío o inválido"])[:6]),
                "system", "ge_control", {"uuid_sat": uuid_sat, "id_ccp_generado": id_ccp, "validacion": (validacion_cp.metadata if validacion_cp else {})},
            )
    except Exception as e:
        logger.error("Error al guardar CFDI timbrado en BD: %s", e)
        # El CFDI ya fue timbrado — retornar el UUID aunque falle la BD
        return JSONResponse({
            "ok":          True,
            "viaje_id":    viaje_id,
            "uuid_sat":    uuid_sat,
            "id_ccp":      id_ccp,
            "pdf_url":     pdf_url,
            "status":      "Vigente",
            "fecha_timbrado": now_iso,
            "advertencia": f"CFDI timbrado pero error al guardar en BD: {e}",
        })

    logger.info("Viaje %s timbrado: uuid_sat=%s id_ccp=%s", viaje_id, uuid_sat, id_ccp)
    return JSONResponse({
        "ok":             True,
        "viaje_id":       viaje_id,
        "uuid_sat":       uuid_sat,
        "id_ccp":         id_ccp,
        "pdf_url":        pdf_url,
        "status":         "Vigente" if carta_porte_valida else "ErrorValidacion",
        "fecha_timbrado": now_iso,
        "advertencia": None if carta_porte_valida else "CFDI timbrado, pero XML no válido como Carta Porte de carretera. Revisa detalle fiscal.",
        "validacion_carta_porte": {
            "ok": carta_porte_valida,
            "errors": validacion_cp.errors if validacion_cp else ["XML vacío o inválido"],
            "warnings": validacion_cp.warnings if validacion_cp else [],
            "metadata": validacion_cp.metadata if validacion_cp else {},
        },
    })


@router.post("/tr/viajes/{viaje_id}/cancelar")
async def cancelar_viaje(
    viaje_id:      int,
    payload:       CancelacionViajeRequest,
    authorization: str = Header(default=""),
):
    """Cancela el CFDI de un viaje."""
    uid, token = _auth(authorization)
    sb = _sb(token)

    # Obtener CFDI del viaje
    try:
        res = sb.table(_TBL_CFDI).select("*").eq("viaje_id", viaje_id).eq("user_id", uid).limit(1).execute()
        rows = res.data or []
        if not rows:
            raise HTTPException(404, "No se encontró CFDI para este viaje.")
        cfdi_row = rows[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error al obtener CFDI: {e}")

    if cfdi_row.get("status") == "Cancelada":
        raise HTTPException(400, "Este CFDI ya está cancelado.")

    settings   = _settings_transporte(uid, token)
    rfc_emisor = settings.get("RfcContribuyente", "")

    resultado = cancelar_cfdi(cfdi_row["uuid_sat"], rfc_emisor, payload.motivo)
    if resultado["ok"]:
        try:
            sb.table(_TBL_CFDI).update({"status": "Cancelada"}).eq("id", cfdi_row["id"]).execute()
            sb.table(_TBL_VIAJES).update({"status": "cancelado"}).eq("id", viaje_id).execute()
            _registrar_evento(
                sb, uid, cfdi_row.get("perfil_id"), viaje_id, "carta_porte_cancelada",
                "CFDI/Carta Porte cancelado", f"Motivo SAT {payload.motivo}.",
                "oficina", uid, {"uuid_sat": cfdi_row.get("uuid_sat"), "motivo": payload.motivo},
            )
        except Exception as e:
            logger.error("Error al actualizar status cancelación: %s", e)

    return JSONResponse({
        "ok":     resultado["ok"],
        "status": resultado["status"],
        "error":  resultado.get("error"),
    })


# ══════════════════════════════════════════════════════════════════════════════
# 4. FACTURAS (listado y descarga)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/tr/facturas-servicio")
async def listar_facturas_servicio(
    periodo:       Optional[str] = Query(None),
    perfil_id:     Optional[int] = Query(None),
    authorization: str           = Header(default=""),
    x_perfil_id:   str           = Header(default=""),
):
    """Lista facturas del servicio de transporte emitidas o preparadas."""
    uid, token = _auth(authorization)
    try:
        pid = _perfil(perfil_id, x_perfil_id)
        q = _sb(token).table(_TBL_FACT_SERV).select("*").eq("user_id", uid).order("created_at", desc=True)
        if periodo:
            ini, fin = _periodo_bounds(periodo)
            q = q.gte("created_at", ini).lt("created_at", fin)
        if pid:
            q = q.eq("perfil_id", pid)
        res = q.execute()
        return JSONResponse({"ok": True, "facturas_servicio": res.data or []})
    except Exception as e:
        raise HTTPException(500, f"Error al listar facturas de servicio: {e}")


@router.get("/tr/facturas-servicio/{factura_id}/pdf")
async def ver_pdf_factura_servicio_transporte(
    factura_id: int,
    download: bool = Query(False),
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    sb = _sb(token)
    q = sb.table(_TBL_FACT_SERV).select("*").eq("id", factura_id).eq("user_id", uid).limit(1)
    if pid:
        q = q.eq("perfil_id", pid)
    rows = q.execute().data or []
    if not rows:
        raise HTTPException(404, "Factura de servicio no encontrada.")
    row = rows[0]
    xml_content = row.get("xml_content") or ""
    if not xml_content:
        raise HTTPException(404, "Factura de servicio sin XML timbrado para generar PDF.")
    settings = _settings_transporte(uid, token, row.get("perfil_id") or pid)
    info = fiscal_pdf_info(xml_content, "factura_servicio_transporte")
    pdf_bytes = generar_pdf_ingreso_desde_xml(
        xml_content,
        logo_data_url=settings.get("PdfLogoDataUrl", ""),
    )
    storage = save_fiscal_artifacts(
        get_supabase_admin(),
        bucket="transport-documents",
        base_path=f"{uid}/{row.get('perfil_id') or 'default'}/facturas_servicio/{factura_id}",
        xml_content=xml_content,
        pdf_bytes=pdf_bytes,
        pdf_filename=info.filename,
        metadata={"module": "transporte", "entity_type": "factura_servicio", "uuid_sat": row.get("uuid_sat") or ""},
    )
    audit_fiscal_pdf_event(
        get_supabase_admin(),
        user_id=uid,
        module="transporte",
        entity_type="factura_servicio",
        entity_id=factura_id,
        uuid_sat=row.get("uuid_sat") or "",
        action="pdf_generated_internal" if not download else "pdf_download_internal",
        metadata={**storage, "sw_pdf_url_ignored": bool(row.get("pdf_url"))},
    )
    disposition = "attachment" if download else "inline"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'{disposition}; filename="{info.filename}"'},
    )


@router.get("/tr/facturas-servicio/{factura_id}/xml")
async def descargar_xml_factura_servicio_transporte(
    factura_id: int,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_FACT_SERV).select("*").eq("id", factura_id).eq("user_id", uid).limit(1)
    if pid:
        q = q.eq("perfil_id", pid)
    rows = q.execute().data or []
    if not rows:
        raise HTTPException(404, "Factura de servicio no encontrada.")
    row = rows[0]
    if not row.get("xml_content"):
        raise HTTPException(404, "Factura de servicio sin XML timbrado.")
    info = fiscal_pdf_info(row["xml_content"], "factura_servicio_transporte")
    audit_fiscal_pdf_event(
        get_supabase_admin(),
        user_id=uid,
        module="transporte",
        entity_type="factura_servicio",
        entity_id=factura_id,
        uuid_sat=row.get("uuid_sat") or "",
        action="xml_download",
    )
    return Response(
        content=row["xml_content"],
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{info.filename.replace(".pdf", ".xml")}"'},
    )


@router.get("/tr/dashboard")
async def dashboard_transporte(
    periodo: Optional[str] = Query(None),
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    sb = _sb(token)
    pid = _perfil(perfil_id, x_perfil_id)
    periodo = periodo or datetime.now(timezone.utc).strftime("%Y-%m")
    qv = sb.table(_TBL_VIAJES).select("*").eq("user_id", uid).like("fecha_hora_salida", f"{periodo}%")
    if pid:
        qv = qv.eq("perfil_id", pid)
    viajes = qv.execute().data or []
    ini, fin = _periodo_bounds(periodo)
    qf = sb.table(_TBL_FACT_SERV).select("*").eq("user_id", uid).gte("created_at", ini).lt("created_at", fin)
    if pid:
        qf = qf.eq("perfil_id", pid)
    facturas = qf.execute().data or []
    return JSONResponse({
        "ok": True,
        "periodo": periodo,
        "total_viajes": len(viajes),
        "cartas_timbradas": len([v for v in viajes if v.get("uuid_cfdi")]),
        "pendientes": len([v for v in viajes if not v.get("uuid_cfdi")]),
        "volumen_total": round(sum(float(v.get("volumen_total_litros") or 0) for v in viajes), 2),
        "facturacion_servicio": round(sum(float(f.get("total") or 0) for f in facturas), 2),
    })


@router.get("/tr/analytics")
async def analytics_transporte(
    authorization: str = Header(default=""),
    perfil_id: Optional[int] = Query(None),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_VIAJES).select("*").eq("user_id", uid)
    if pid:
        q = q.eq("perfil_id", pid)
    rows = q.execute().data or []
    por_ruta = {}
    por_producto = {}
    for v in rows:
        ruta = f"{v.get('cp_origen') or '?'}-{v.get('cp_destino') or '?'}"
        por_ruta.setdefault(ruta, {"ruta": ruta, "viajes": 0, "volumen": 0.0})
        por_ruta[ruta]["viajes"] += 1
        por_ruta[ruta]["volumen"] += float(v.get("volumen_total_litros") or 0)
        try:
            productos = json.loads(v.get("productos_json") or "[]")
        except Exception:
            productos = []
        for p in productos:
            nombre = p.get("descripcion") or p.get("clave_producto") or "Producto"
            por_producto.setdefault(nombre, {"producto": nombre, "viajes": 0, "volumen": 0.0})
            por_producto[nombre]["viajes"] += 1
            por_producto[nombre]["volumen"] += float(p.get("volumen_litros") or 0)
    return JSONResponse({
        "ok": True,
        "rutas": sorted(por_ruta.values(), key=lambda x: x["volumen"], reverse=True),
        "productos": sorted(por_producto.values(), key=lambda x: x["volumen"], reverse=True),
    })


@router.get("/tr/forecast")
async def forecast_transporte(
    authorization: str = Header(default=""),
    perfil_id: Optional[int] = Query(None),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_VIAJES).select("fecha_hora_salida,volumen_total_litros").eq("user_id", uid).order("fecha_hora_salida")
    if pid:
        q = q.eq("perfil_id", pid)
    rows = q.execute().data or []
    por_mes = {}
    for r in rows:
        periodo = (r.get("fecha_hora_salida") or "")[:7]
        if len(periodo) == 7:
            por_mes[periodo] = por_mes.get(periodo, 0.0) + float(r.get("volumen_total_litros") or 0)
    series = [por_mes[k] for k in sorted(por_mes)]
    if not series:
        return JSONResponse({"ok": True, "modelo": "sin_datos", "pronostico_volumen": 0, "periodos": []})
    prom = sum(series[-3:]) / min(len(series), 3)
    if len(series) >= 2:
        tendencia = (series[-1] - series[0]) / max(len(series) - 1, 1)
    else:
        tendencia = 0.0
    pronostico = max(round(prom + tendencia, 2), 0)
    return JSONResponse({"ok": True, "modelo": "promedio_movil_3m_con_tendencia", "pronostico_volumen": pronostico, "periodos": sorted(por_mes), "volumenes": series})


# ══════════════════════════════════════════════════════════════════════════════
# 3B. OPERACION: Dashboard, Viaje 360, documentos, tarifas y operador
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/tr/dashboard-operativo")
async def dashboard_operativo_transporte(
    periodo: Optional[str] = Query(None),
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    periodo = periodo or datetime.now(timezone.utc).strftime("%Y-%m")
    q = _sb(token).table(_TBL_VIAJES).select("*").eq("user_id", uid).like("fecha_hora_salida", f"{periodo}%")
    if pid:
        q = q.eq("perfil_id", pid)
    viajes = q.execute().data or []
    resumen = {
        "programados": 0, "sin_confirmacion": 0, "en_ruta": 0, "entregados": 0,
        "cartas_pendientes": 0, "facturas_pendientes": 0, "liquidaciones_pendientes": 0,
    }
    for v in viajes:
        op = (v.get("operacion_status") or v.get("status") or "programado").lower()
        if op in {"programado", "borrador", "asignado"}:
            resumen["programados"] += 1
        if op == "en_ruta":
            resumen["en_ruta"] += 1
        if op in {"entregado", "cerrado"}:
            resumen["entregados"] += 1
        if not v.get("fecha_entrega_confirmada") and op not in {"cancelado", "cerrado"}:
            resumen["sin_confirmacion"] += 1
        if not v.get("uuid_cfdi"):
            resumen["cartas_pendientes"] += 1
        if v.get("uuid_cfdi") and (v.get("factura_status") or "pendiente") == "pendiente":
            resumen["facturas_pendientes"] += 1
        if (v.get("liquidacion_status") or "pendiente") == "pendiente":
            resumen["liquidaciones_pendientes"] += 1
    resumen["alertas"] = [
        {"tipo": "carta_porte", "label": "Viajes sin Carta Porte", "count": resumen["cartas_pendientes"]},
        {"tipo": "factura", "label": "Cartas Porte pendientes de factura", "count": resumen["facturas_pendientes"]},
        {"tipo": "operador", "label": "Viajes sin confirmacion de operador", "count": resumen["sin_confirmacion"]},
    ]
    return JSONResponse({"ok": True, "periodo": periodo, "resumen": resumen, "viajes": viajes[:100]})


@router.get("/tr/carta-aporte/tareas")
async def tareas_carta_aporte_desde_sat(
    perfil_id: Optional[int] = Query(None),
    force: bool = Query(False),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    """
    Tarea operativa diaria: a partir de las 02:00 CDMX propone generar Carta Aporte
    con las facturas timbradas en la ultima hora, leyendo datos fiscales desde XML SAT.
    Disponible para administradores y operadores con acceso a Transporte.
    """
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    now_mx = datetime.now(ZoneInfo("America/Mexico_City"))
    if not force and (now_mx.hour, now_mx.minute) < (2, 0):
        return JSONResponse({
            "ok": True,
            "tasks": [],
            "next_run_local": now_mx.replace(hour=2, minute=0, second=0, microsecond=0).isoformat(),
            "message": "La tarea automática de Carta Aporte aparece a partir de las 02:00.",
        })

    fin = datetime.now(timezone.utc)
    ini = fin - timedelta(hours=1)
    q = (
        _sb(token).table(_TBL_CFDI)
        .select("id, viaje_id, uuid_sat, fecha_timbrado, xml_content, status")
        .eq("user_id", uid)
        .eq("status", "Vigente")
        .gte("fecha_timbrado", ini.isoformat())
        .lte("fecha_timbrado", fin.isoformat())
        .order("fecha_timbrado", desc=True)
    )
    if pid:
        q = q.eq("perfil_id", pid)
    rows = q.execute().data or []
    facturas = []
    errores = []
    for row in rows:
        xml = row.get("xml_content") or ""
        if not xml:
            errores.append({"cfdi_id": row.get("id"), "error": "CFDI sin XML timbrado almacenado."})
            continue
        try:
            data = extraer_factura_timbrada_sat(xml).as_dict()
            facturas.append({
                "cfdi_id": row.get("id"),
                "viaje_id": row.get("viaje_id"),
                **data,
            })
        except Exception as exc:
            errores.append({"cfdi_id": row.get("id"), "error": f"No se pudo leer XML SAT: {exc}"})

    tasks = []
    if facturas:
        tasks.append({
            "tipo": "generar_carta_aporte",
            "titulo": "Generar Carta Aporte con las facturas timbradas en la ultima hora.",
            "perfil_id": pid,
            "window_start": ini.isoformat(),
            "window_end": fin.isoformat(),
            "facturas": facturas,
            "errores": errores,
            "manual_capture_required": False,
        })
    return JSONResponse({"ok": True, "tasks": tasks, "errores": errores})


@router.get("/tr/viajes/{viaje_id}/360")
async def viaje_360(viaje_id: int, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    viaje_res = sb.table(_TBL_VIAJES).select("*").eq("id", viaje_id).eq("user_id", uid).limit(1).execute()
    rows = viaje_res.data or []
    if not rows:
        raise HTTPException(404, "Viaje no encontrado.")
    viaje = rows[0]
    try:
        viaje["productos"] = _productos_from_row(viaje)
        chofer = sb.table(_TBL_CHOFERES).select("*").eq("id", viaje.get("chofer_id")).eq("user_id", uid).limit(1).execute().data or []
        vehiculo = sb.table(_TBL_VEHICULOS).select("*").eq("id", viaje.get("vehiculo_id")).eq("user_id", uid).limit(1).execute().data or []
        docs = sb.table(_TBL_DOCS).select("*").eq("user_id", uid).eq("viaje_id", viaje_id).order("created_at", desc=True).execute().data or []
        eventos = sb.table(_TBL_EVENTOS).select("*").eq("user_id", uid).eq("viaje_id", viaje_id).order("created_at", desc=False).execute().data or []
        facturas = sb.table(_TBL_FACT_SERV_CARTAS).select("factura_servicio_id,viaje_id,created_at").eq("user_id", uid).eq("viaje_id", viaje_id).execute().data or []
        gastos = sb.table(_TBL_GASTOS).select("*").eq("user_id", uid).eq("viaje_id", viaje_id).execute().data or []
    except Exception as e:
        logger.info("Viaje 360 parcial para %s: %s", viaje_id, e)
        docs, eventos, facturas, gastos, chofer, vehiculo = [], [], [], [], [], []
    return JSONResponse({
        "ok": True, "viaje": viaje, "chofer": chofer[0] if chofer else None,
        "vehiculo": vehiculo[0] if vehiculo else None, "documentos": docs,
        "eventos": eventos, "facturas_servicio": facturas, "gastos": gastos,
    })


@router.post("/tr/viajes/{viaje_id}/eventos")
async def crear_evento_manual(viaje_id: int, payload: dict, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    v = sb.table(_TBL_VIAJES).select("id,perfil_id").eq("id", viaje_id).eq("user_id", uid).limit(1).execute().data or []
    if not v:
        raise HTTPException(404, "Viaje no encontrado.")
    _registrar_evento(
        sb, uid, v[0].get("perfil_id"), viaje_id,
        str(payload.get("event_type") or "nota_manual"),
        str(payload.get("title") or "Nota manual"),
        str(payload.get("description") or ""),
        "oficina", uid, payload.get("metadata") if isinstance(payload.get("metadata"), dict) else {},
    )
    return JSONResponse({"ok": True})


@router.post("/tr/viajes/{viaje_id}/operacion-status")
async def actualizar_operacion_status(viaje_id: int, payload: dict, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    status = str(payload.get("operacion_status") or "").strip().lower()
    validos = {"programado", "asignado", "recibido", "en_ruta", "entregado", "problema", "cerrado", "cancelado"}
    if status not in validos:
        raise HTTPException(400, "Estatus operativo invalido.")
    rows = sb.table(_TBL_VIAJES).select("id,perfil_id").eq("id", viaje_id).eq("user_id", uid).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Viaje no encontrado.")
    update = {"operacion_status": status}
    if status == "entregado":
        update["fecha_entrega_confirmada"] = datetime.now(timezone.utc).isoformat()
    if status == "cerrado":
        update["closed_at"] = datetime.now(timezone.utc).isoformat()
    sb.table(_TBL_VIAJES).update(update).eq("id", viaje_id).eq("user_id", uid).execute()
    _registrar_evento(sb, uid, rows[0].get("perfil_id"), viaje_id, "operacion_actualizada", f"Estatus operativo: {status}", str(payload.get("nota") or ""), "oficina", uid, update)
    return JSONResponse({"ok": True, "operacion_status": status})


@router.post("/tr/viajes/{viaje_id}/gastos")
async def crear_gasto_viaje(viaje_id: int, payload: dict, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    viaje_rows = sb.table(_TBL_VIAJES).select("id,perfil_id,chofer_id").eq("id", viaje_id).eq("user_id", uid).limit(1).execute().data or []
    if not viaje_rows:
        raise HTTPException(404, "Viaje no encontrado.")
    importe = _safe_float(payload.get("importe"))
    if importe <= 0:
        raise HTTPException(400, "El gasto debe ser mayor a 0.")
    v = viaje_rows[0]
    row = {
        "user_id": uid,
        "perfil_id": v.get("perfil_id"),
        "viaje_id": viaje_id,
        "chofer_id": v.get("chofer_id"),
        "tipo": str(payload.get("tipo") or "otro"),
        "descripcion": str(payload.get("descripcion") or ""),
        "importe": importe,
        "moneda": str(payload.get("moneda") or "MXN"),
        "status": str(payload.get("status") or "aprobado"),
        "metadata": payload.get("metadata") if isinstance(payload.get("metadata"), dict) else {},
    }
    res = sb.table(_TBL_GASTOS).insert(row).execute()
    _registrar_evento(sb, uid, v.get("perfil_id"), viaje_id, "gasto_registrado", "Gasto registrado", f"{row['tipo']}: ${importe:.2f}", "oficina", uid, row)
    return JSONResponse({"ok": True, "gasto": (res.data or [None])[0]})


@router.put("/tr/gastos/{gasto_id}")
async def actualizar_gasto_viaje(gasto_id: int, payload: dict, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    allowed = {"tipo", "descripcion", "importe", "moneda", "status", "metadata"}
    row = {k: v for k, v in payload.items() if k in allowed}
    _sb(token).table(_TBL_GASTOS).update(row).eq("id", gasto_id).eq("user_id", uid).execute()
    return JSONResponse({"ok": True})


@router.get("/tr/viajes/{viaje_id}/documentos")
async def listar_documentos_viaje(viaje_id: int, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    rows = _sb(token).table(_TBL_DOCS).select("*").eq("user_id", uid).eq("viaje_id", viaje_id).order("created_at", desc=True).execute().data or []
    return JSONResponse({"ok": True, "documentos": rows})


@router.post("/tr/viajes/{viaje_id}/documentos")
async def registrar_documento_viaje(viaje_id: int, payload: dict, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    v = sb.table(_TBL_VIAJES).select("id,perfil_id").eq("id", viaje_id).eq("user_id", uid).limit(1).execute().data or []
    if not v:
        raise HTTPException(404, "Viaje no encontrado.")
    pid = v[0].get("perfil_id")
    row = {
        "user_id": uid, "perfil_id": pid, "viaje_id": viaje_id,
        "tipo": str(payload.get("tipo") or "otro"),
        "nombre": str(payload.get("nombre") or payload.get("storage_path") or "Documento"),
        "storage_bucket": str(payload.get("storage_bucket") or "transport-documents"),
        "storage_path": str(payload.get("storage_path") or ""),
        "mime_type": str(payload.get("mime_type") or ""),
        "size_bytes": int(payload.get("size_bytes") or 0),
        "uuid_sat": str(payload.get("uuid_sat") or ""),
        "metadata": payload.get("metadata") if isinstance(payload.get("metadata"), dict) else {},
        "created_by": uid,
    }
    res = sb.table(_TBL_DOCS).insert(row).execute()
    _registrar_evento(sb, uid, pid, viaje_id, "documento_registrado", "Documento registrado", row["nombre"], "oficina", uid, {"tipo": row["tipo"]})
    return JSONResponse({"ok": True, "documento": (res.data or [None])[0]})


@router.post("/tr/viajes/{viaje_id}/documentos/upload")
async def subir_documento_viaje(
    viaje_id: int,
    tipo: str = Form("otro"),
    file: UploadFile = File(...),
    authorization: str = Header(default=""),
):
    uid, token = _auth(authorization)
    sb = _sb(token)
    v = sb.table(_TBL_VIAJES).select("id,perfil_id").eq("id", viaje_id).eq("user_id", uid).limit(1).execute().data or []
    if not v:
        raise HTTPException(404, "Viaje no encontrado.")
    pid = v[0].get("perfil_id")
    content = await file.read()
    bucket = "transport-documents"
    path = _build_document_path(uid, pid, viaje_id, tipo, file.filename or "documento")
    try:
        sb.storage.from_(bucket).upload(path, content, {"content-type": file.content_type or "application/octet-stream", "upsert": "true"})
    except Exception as e:
        raise HTTPException(500, f"No se pudo subir a Supabase Storage. Verifica que exista el bucket '{bucket}': {e}")
    row = {
        "user_id": uid, "perfil_id": pid, "viaje_id": viaje_id, "tipo": tipo,
        "nombre": file.filename or "Documento", "storage_bucket": bucket,
        "storage_path": path, "mime_type": file.content_type or "",
        "size_bytes": len(content), "created_by": uid,
    }
    res = sb.table(_TBL_DOCS).insert(row).execute()
    _registrar_evento(sb, uid, pid, viaje_id, "documento_subido", "Documento subido", row["nombre"], "oficina", uid, {"tipo": tipo, "storage_path": path})
    return JSONResponse({"ok": True, "documento": (res.data or [None])[0]})


@router.get("/tr/tarifas")
async def listar_tarifas(perfil_id: Optional[int] = Query(None), authorization: str = Header(default=""), x_perfil_id: str = Header(default="")):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_TARIFAS).select("*").eq("user_id", uid).eq("activo", True).order("prioridad")
    if pid:
        q = q.eq("perfil_id", pid)
    return JSONResponse({"ok": True, "tarifas": q.execute().data or []})


@router.post("/tr/tarifas")
async def crear_tarifa(payload: dict, authorization: str = Header(default=""), x_perfil_id: str = Header(default="")):
    uid, token = _auth(authorization)
    pid = _perfil(payload.get("perfil_id"), x_perfil_id)
    row = {
        "user_id": uid, "perfil_id": pid, "cliente_id": payload.get("cliente_id"),
        "ruta_id": payload.get("ruta_id"), "origen": str(payload.get("origen") or ""),
        "destino": str(payload.get("destino") or ""), "producto": str(payload.get("producto") or ""),
        "regla_calculo": str(payload.get("regla_calculo") or "litros"),
        "tarifa": _safe_float(payload.get("tarifa")), "iva_tasa": _safe_float(payload.get("iva_tasa"), 0.16),
        "retencion_tasa": _safe_float(payload.get("retencion_tasa"), 0.04),
        "aplica_iva": bool(payload.get("aplica_iva", True)), "aplica_retencion": bool(payload.get("aplica_retencion", True)),
        "moneda": str(payload.get("moneda") or "MXN"), "prioridad": int(payload.get("prioridad") or 100),
        "vigencia_desde": payload.get("vigencia_desde") or None,
        "vigencia_hasta": payload.get("vigencia_hasta") or None,
        "observaciones": str(payload.get("observaciones") or ""),
        "metadata": payload.get("metadata") if isinstance(payload.get("metadata"), dict) else {},
    }
    res = _sb(token).table(_TBL_TARIFAS).insert(row).execute()
    return JSONResponse({"ok": True, "tarifa": (res.data or [None])[0]})


@router.put("/tr/tarifas/{tarifa_id}")
async def actualizar_tarifa(tarifa_id: int, payload: dict, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    allowed = {"cliente_id","ruta_id","origen","destino","producto","regla_calculo","tarifa","iva_tasa","retencion_tasa","aplica_iva","aplica_retencion","moneda","prioridad","activo","vigencia_desde","vigencia_hasta","observaciones","metadata"}
    row = {k: v for k, v in payload.items() if k in allowed}
    row["updated_at"] = datetime.now(timezone.utc).isoformat()
    _sb(token).table(_TBL_TARIFAS).update(row).eq("id", tarifa_id).eq("user_id", uid).execute()
    return JSONResponse({"ok": True})


@router.post("/tr/viajes/{viaje_id}/calcular-tarifa")
async def calcular_tarifa_viaje(viaje_id: int, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    viaje_rows = sb.table(_TBL_VIAJES).select("*").eq("id", viaje_id).eq("user_id", uid).limit(1).execute().data or []
    if not viaje_rows:
        raise HTTPException(404, "Viaje no encontrado.")
    viaje = viaje_rows[0]
    q = sb.table(_TBL_TARIFAS).select("*").eq("user_id", uid).eq("activo", True)
    if viaje.get("perfil_id"):
        q = q.eq("perfil_id", viaje.get("perfil_id"))
    calc = _calcular_tarifa_operativa(viaje, q.execute().data or [])
    try:
        sb.table(_TBL_VIAJES).update({"tarifa_total": calc["subtotal"], "retencion": calc["retencion"], "total_operativo": calc["total"]}).eq("id", viaje_id).eq("user_id", uid).execute()
    except Exception as e:
        logger.info("No se pudieron guardar totales operativos para viaje %s: %s", viaje_id, e)
    _registrar_evento(sb, uid, viaje.get("perfil_id"), viaje_id, "tarifa_calculada", "Tarifa calculada", "", "system", "tarifas", calc)
    return JSONResponse({"ok": True, "calculo": calc})


@router.post("/tr/operador/acceso")
async def crear_acceso_operador(payload: dict, authorization: str = Header(default=""), x_perfil_id: str = Header(default="")):
    uid, token = _auth(authorization)
    pid = _perfil(payload.get("perfil_id"), x_perfil_id)
    chofer_id = int(payload.get("chofer_id") or 0)
    if not chofer_id:
        raise HTTPException(400, "chofer_id requerido.")
    token_plain = secrets.token_urlsafe(24)
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    _sb(token).table(_TBL_OPER_ACC).insert({
        "user_id": uid,
        "perfil_id": pid,
        "chofer_id": chofer_id,
        "token_hash": _hash_operator_token(token_plain),
        "status": "activo",
        "expires_at": expires_at.isoformat(),
    }).execute()
    return JSONResponse({"ok": True, "token": token_plain, "url": f"/operador/transporte?token={token_plain}"})


def _operador_context(token_plain: str):
    sb = get_supabase_admin()
    rows = sb.table(_TBL_OPER_ACC).select("*").eq("token_hash", _hash_operator_token(token_plain)).eq("status", "activo").limit(1).execute().data or []
    if not rows:
        raise HTTPException(401, "Acceso de operador invalido.")
    acc = rows[0]
    expires_at = acc.get("expires_at")
    if expires_at:
        try:
            exp = datetime.fromisoformat(str(expires_at).replace("Z", "+00:00"))
            if exp <= datetime.now(timezone.utc):
                try:
                    sb.table(_TBL_OPER_ACC).update({"status": "expirado"}).eq("id", acc["id"]).execute()
                except Exception:
                    pass
                raise HTTPException(401, "Acceso de operador expirado.")
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(401, "Acceso de operador inválido.")
    try:
        sb.table(_TBL_OPER_ACC).update({"last_used_at": datetime.now(timezone.utc).isoformat()}).eq("id", acc["id"]).execute()
    except Exception:
        pass
    return sb, acc


def _operador_meta(sb, acc: dict) -> dict:
    chofer = {}
    empresa = {}
    notificaciones = []
    try:
        rows = sb.table(_TBL_CHOFERES).select("id,nombre,rfc,licencia,telefono").eq("id", acc.get("chofer_id")).eq("user_id", acc.get("user_id")).limit(1).execute().data or []
        chofer = rows[0] if rows else {}
    except Exception:
        chofer = {}
    try:
        if acc.get("perfil_id"):
            rows = sb.table("perfiles_empresa").select("id,nombre,rfc").eq("id", acc.get("perfil_id")).limit(1).execute().data or []
            empresa = rows[0] if rows else {}
    except Exception:
        empresa = {}
    try:
        q = sb.table(_TBL_NOTIFS).select("*").eq("user_id", acc.get("user_id")).order("created_at", desc=True).limit(5)
        if acc.get("perfil_id"):
            q = q.eq("perfil_id", acc.get("perfil_id"))
        notificaciones = q.execute().data or []
    except Exception:
        notificaciones = []
    return {"chofer": chofer, "empresa": empresa, "notificaciones": notificaciones}


@router.get("/tr/operador/viajes")
async def operador_viajes(token: str = Query(...)):
    sb, acc = _operador_context(token)
    q = sb.table(_TBL_VIAJES).select("*").eq("user_id", acc["user_id"]).eq("chofer_id", acc["chofer_id"])
    if acc.get("perfil_id"):
        q = q.eq("perfil_id", acc.get("perfil_id"))
    viajes = q.in_("operacion_status", ["programado", "asignado", "recibido", "en_ruta", "problema"]).order("fecha_hora_salida").execute().data or []
    cfdis = []
    ids = [int(v["id"]) for v in viajes if v.get("id")]
    if ids:
        cfdis = sb.table(_TBL_CFDI).select("viaje_id,uuid_sat,id_ccp,status").eq("user_id", acc["user_id"]).in_("viaje_id", ids).execute().data or []
    cfdi_map = {int(c.get("viaje_id")): c for c in cfdis if c.get("viaje_id")}
    for v in viajes:
        v["productos"] = _productos_from_row(v)
        v["cfdi"] = cfdi_map.get(int(v.get("id") or 0), {})
        v["tiene_pdf_carta_porte"] = bool(v.get("uuid_cfdi") or v["cfdi"].get("uuid_sat"))
    pending_docs = [v for v in viajes if not v.get("tiene_pdf_carta_porte")]
    return JSONResponse({
        "ok": True,
        "viajes": viajes,
        "meta": _operador_meta(sb, acc),
        "resumen": {
            "viajes_activos": len(viajes),
            "documentos_pendientes": len(pending_docs),
            "expires_at": acc.get("expires_at"),
        },
    })


@router.get("/tr/operador/carta-aporte/tareas")
async def operador_tareas_carta_aporte(token: str = Query(...), force: bool = Query(False)):
    sb, acc = _operador_context(token)
    now_mx = datetime.now(ZoneInfo("America/Mexico_City"))
    if not force and (now_mx.hour, now_mx.minute) < (2, 0):
        return JSONResponse({"ok": True, "tasks": []})
    fin = datetime.now(timezone.utc)
    ini = fin - timedelta(hours=1)
    q = (
        sb.table(_TBL_CFDI)
        .select("id, viaje_id, uuid_sat, fecha_timbrado, xml_content, status")
        .eq("user_id", acc["user_id"])
        .eq("status", "Vigente")
        .gte("fecha_timbrado", ini.isoformat())
        .lte("fecha_timbrado", fin.isoformat())
        .order("fecha_timbrado", desc=True)
    )
    if acc.get("perfil_id"):
        q = q.eq("perfil_id", acc.get("perfil_id"))
    rows = q.execute().data or []
    facturas = []
    errores = []
    for row in rows:
        try:
            data = extraer_factura_timbrada_sat(row.get("xml_content") or "").as_dict()
            facturas.append({"cfdi_id": row.get("id"), "viaje_id": row.get("viaje_id"), **data})
        except Exception as exc:
            errores.append({"cfdi_id": row.get("id"), "error": f"No se pudo leer XML SAT: {exc}"})
    tasks = [{
        "tipo": "generar_carta_aporte",
        "titulo": "Generar Carta Aporte con las facturas timbradas en la ultima hora.",
        "perfil_id": acc.get("perfil_id"),
        "window_start": ini.isoformat(),
        "window_end": fin.isoformat(),
        "facturas": facturas,
        "errores": errores,
        "manual_capture_required": False,
    }] if facturas else []
    return JSONResponse({"ok": True, "tasks": tasks, "errores": errores})


@router.post("/tr/operador/viajes/{viaje_id}/accion")
async def operador_accion(viaje_id: int, payload: dict, token: str = Query(...)):
    sb, acc = _operador_context(token)
    accion = str(payload.get("accion") or "").strip()
    mapping = {"recibido": ("recibido", "Ya lo recibio"), "en_camino": ("en_ruta", "Va en camino"), "entregado": ("entregado", "Ya entrego"), "problema": ("problema", "Reporto problema")}
    if accion not in mapping:
        raise HTTPException(400, "Accion no valida.")
    status, title = mapping[accion]
    viaje_rows = sb.table(_TBL_VIAJES).select("id,user_id,perfil_id,chofer_id,nombre_origen,nombre_destino,cp_origen,cp_destino").eq("id", viaje_id).eq("user_id", acc["user_id"]).eq("chofer_id", acc["chofer_id"]).limit(1).execute().data or []
    if not viaje_rows:
        raise HTTPException(404, "Viaje no encontrado para este operador.")
    update = {"operacion_status": status}
    if accion == "entregado":
        update["fecha_entrega_confirmada"] = datetime.now(timezone.utc).isoformat()
    uq = sb.table(_TBL_VIAJES).update(update).eq("id", viaje_id).eq("user_id", acc["user_id"]).eq("chofer_id", acc["chofer_id"])
    if acc.get("perfil_id"):
        uq = uq.eq("perfil_id", acc.get("perfil_id"))
    uq.execute()
    _registrar_evento(sb, acc["user_id"], viaje_rows[0].get("perfil_id"), viaje_id, f"operador_{accion}", title, str(payload.get("nota") or ""), "operador", str(acc["chofer_id"]), {"accion": accion})
    if accion == "problema":
        v = viaje_rows[0]
        ruta = f"{v.get('nombre_origen') or v.get('cp_origen') or '?'} → {v.get('nombre_destino') or v.get('cp_destino') or '?'}"
        _crear_notificacion_manual(
            sb,
            acc["user_id"],
            v.get("perfil_id"),
            viaje_id,
            f"Operador reportó problema en viaje #{viaje_id}: {ruta}. {str(payload.get('nota') or '').strip()}",
            metadata={"accion": accion, "chofer_id": acc["chofer_id"]},
        )
    return JSONResponse({"ok": True, "operacion_status": status})


@router.get("/tr/operador/viajes/{viaje_id}/pdf")
async def operador_pdf_carta_porte(viaje_id: int, token: str = Query(...), download: bool = Query(False)):
    """PDF imprimible de Carta Porte para el operador asignado."""
    sb, acc = _operador_context(token)
    viaje_rows = (
        sb.table(_TBL_VIAJES)
        .select("id,user_id,perfil_id,chofer_id")
        .eq("id", viaje_id)
        .eq("user_id", acc["user_id"])
        .eq("chofer_id", acc["chofer_id"])
        .limit(1)
        .execute()
        .data
        or []
    )
    if not viaje_rows:
        raise HTTPException(404, "Viaje no encontrado para este operador.")
    cfdi_rows = (
        sb.table(_TBL_CFDI)
        .select("id,user_id,perfil_id,viaje_id,uuid_sat,id_ccp,xml_content,pdf_url")
        .eq("user_id", acc["user_id"])
        .eq("viaje_id", viaje_id)
        .order("fecha_timbrado", desc=True)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not cfdi_rows:
        raise HTTPException(404, "Este viaje todavía no tiene Carta Porte timbrada.")
    row = cfdi_rows[0]
    if not row.get("xml_content"):
        raise HTTPException(404, "La Carta Porte no tiene XML guardado.")
    viaje = viaje_rows[0]
    productos = []
    viaje_full = sb.table(_TBL_VIAJES).select("productos_json").eq("id", viaje_id).eq("user_id", acc["user_id"]).limit(1).execute().data or []
    if viaje_full:
        productos = _productos_from_row(viaje_full[0])
    validacion = validar_xml_carta_porte_transporte(row["xml_content"], productos)
    if validacion.bloquea_pdf:
        raise HTTPException(409, "PDF bloqueado: XML no válido como Carta Porte de carretera. " + "; ".join(validacion.errors[:4]))
    info = extraer_info_pdf(row["xml_content"])
    settings_rows = sb.table(_TBL_SETTINGS).select("data").eq("user_id", acc["user_id"]).eq("perfil_id", viaje.get("perfil_id")).limit(1).execute().data or []
    settings = settings_rows[0].get("data", {}) if settings_rows else {}
    pdf_bytes = generar_pdf_carta_porte_desde_xml(row["xml_content"], settings.get("PdfLogoDataUrl", ""))
    disposition = "attachment" if download else "inline"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'{disposition}; filename="{info.filename}"'},
    )


@router.get("/tr/cartas-porte-facturables")
async def listar_cartas_porte_facturables(
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    """Cartas Porte timbradas que todavia no han sido usadas en factura de servicio."""
    uid, token = _auth(authorization)
    sb = _sb(token)
    pid = _perfil(perfil_id, x_perfil_id)
    try:
        fact_q = sb.table(_TBL_FACT_SERV_CARTAS).select("viaje_id").eq("user_id", uid)
        if pid:
            fact_q = fact_q.eq("perfil_id", pid)
        fact_res = fact_q.execute()
        facturados = {int(r.get("viaje_id")) for r in (fact_res.data or []) if r.get("viaje_id")}
    except Exception:
        facturados = set()
    try:
        cfdi_q = sb.table(_TBL_CFDI).select("*").eq("user_id", uid).eq("status", "Vigente")
        if pid:
            cfdi_q = cfdi_q.eq("perfil_id", pid)
        cfdi_res = cfdi_q.order("fecha_timbrado", desc=True).execute()
        cfdis = [c for c in (cfdi_res.data or []) if int(c.get("viaje_id") or 0) not in facturados]
        viajes_ids = [int(c.get("viaje_id")) for c in cfdis if c.get("viaje_id")]
        viajes_map = {}
        if viajes_ids:
            vq = sb.table(_TBL_VIAJES).select("*").eq("user_id", uid).in_("id", viajes_ids)
            if pid:
                vq = vq.eq("perfil_id", pid)
            v_res = vq.execute()
            viajes_map = {int(v["id"]): v for v in (v_res.data or [])}
        cq = sb.table(_TBL_CLIENTES).select("*").eq("user_id", uid).eq("activo", True)
        if pid:
            cq = cq.eq("perfil_id", pid)
        clientes_res = cq.execute()
        clientes = clientes_res.data or []
        clientes_by_rfc = {str(c.get("rfc") or "").upper(): c for c in clientes}
        tq = sb.table(_TBL_TARIFAS).select("*").eq("user_id", uid).eq("activo", True)
        if pid:
            tq = tq.eq("perfil_id", pid)
        tarifas = tq.execute().data or []
        items = []
        for cfdi in cfdis:
            viaje = viajes_map.get(int(cfdi.get("viaje_id") or 0), {})
            cliente = clientes_by_rfc.get(str(viaje.get("rfc_receptor") or cfdi.get("rfc_receptor") or "").upper(), {})
            if cliente.get("id"):
                viaje = {**viaje, "cliente_id": cliente.get("id")}
            calc = _calcular_tarifa_operativa(viaje, tarifas)
            items.append({
                "viaje_id": cfdi.get("viaje_id"),
                "cfdi_id": cfdi.get("id"),
                "uuid_cfdi": cfdi.get("uuid_sat"),
                "id_ccp": cfdi.get("id_ccp"),
                "folio": cfdi.get("id_ccp") or cfdi.get("uuid_sat"),
                "cliente_id": cliente.get("id"),
                "rfc_receptor": cliente.get("rfc") or viaje.get("rfc_receptor") or cfdi.get("rfc_receptor"),
                "nombre_receptor": cliente.get("nombre") or viaje.get("nombre_receptor"),
                "cp_receptor": cliente.get("cp") or viaje.get("cp_receptor"),
                "regimen_fiscal": cliente.get("regimen_fiscal") or "601",
                "uso_cfdi": cliente.get("uso_cfdi") or viaje.get("uso_cfdi") or "G03",
                "subtotal": calc["subtotal"],
                "iva": calc["iva"],
                "retencion": calc["retencion"],
                "total": calc["total"],
                "iva_tasa": calc["iva_tasa"],
                "retencion_tasa": calc["retencion_tasa"],
                "aplica_iva": calc["aplica_iva"],
                "aplica_retencion": calc["aplica_retencion"],
                "tarifa_id": calc.get("tarifa_id"),
                "regla_calculo": calc.get("regla_calculo"),
            })
        return JSONResponse({"ok": True, "cartas": items})
    except Exception as e:
        raise HTTPException(500, f"Error al listar Cartas Porte facturables: {e}")


@router.post("/tr/facturas-servicio")
async def crear_factura_servicio(payload: FacturaServicioCreate, authorization: str = Header(default="")):
    """
    Prepara una factura de ingreso por servicio de transporte y la relaciona con una o varias Cartas Porte.
    El timbrado fiscal puede conectarse al PAC reutilizando esta misma estructura.
    """
    uid, token = _auth(authorization)
    sb = _sb(token)

    _validar_datos_cfdi_receptor(payload.rfc_receptor, payload.regimen_fiscal, payload.cp_receptor, payload.uso_cfdi)
    viajes_res = sb.table(_TBL_VIAJES).select("*").eq("user_id", uid).in_("id", payload.viaje_ids).execute()
    viajes = viajes_res.data or []
    perfil_factura = payload.perfil_id or (viajes[0].get("perfil_id") if viajes else None)
    encontrados = {int(v["id"]) for v in viajes}
    faltantes = [vid for vid in payload.viaje_ids if vid not in encontrados]
    if faltantes:
        raise HTTPException(404, f"Viajes no encontrados: {faltantes}")
    no_timbrados = [v["id"] for v in viajes if not v.get("uuid_cfdi")]
    if no_timbrados:
        raise HTTPException(400, f"Para facturar el servicio, primero timbra la Carta Porte de los viajes: {no_timbrados}")
    try:
        ya_q = sb.table(_TBL_FACT_SERV_CARTAS).select("viaje_id").eq("user_id", uid).in_("viaje_id", payload.viaje_ids)
        if perfil_factura:
            ya_q = ya_q.eq("perfil_id", perfil_factura)
        ya_res = ya_q.execute()
        ya = [r.get("viaje_id") for r in (ya_res.data or [])]
        if ya:
            raise HTTPException(400, f"Estas Cartas Porte ya tienen factura de servicio: {ya}")
    except HTTPException:
        raise
    except Exception:
        # Compatibilidad con bases que aun no tienen la tabla de control.
        existentes = sb.table(_TBL_FACT_SERV).select("viaje_ids").eq("user_id", uid).execute().data or []
        usados = set()
        for f in existentes:
            vals = f.get("viaje_ids") or []
            if isinstance(vals, list):
                usados.update(int(v) for v in vals if str(v).isdigit())
        repetidos = [v for v in payload.viaje_ids if v in usados]
        if repetidos:
            raise HTTPException(400, f"Estas Cartas Porte ya tienen factura de servicio: {repetidos}")

    tq = sb.table(_TBL_TARIFAS).select("*").eq("user_id", uid).eq("activo", True)
    if perfil_factura:
        tq = tq.eq("perfil_id", perfil_factura)
    tarifas = tq.execute().data or []
    if payload.cliente_id:
        viajes_calc = [{**v, "cliente_id": payload.cliente_id} for v in viajes]
    else:
        viajes_calc = viajes
    calculo_servicio = _sumar_calculos_servicio(viajes_calc, tarifas)
    sin_tarifa = [i.get("viaje_id") for i in calculo_servicio.get("items", []) if not i.get("tarifa_id")]
    if sin_tarifa:
        raise HTTPException(400, f"Configura una tarifa de servicio antes de facturar estas Cartas Porte: {sin_tarifa}")
    if calculo_servicio.get("tasas_mixtas"):
        raise HTTPException(400, "No mezcles Cartas Porte con tasas distintas de IVA/retención en una sola factura de servicio.")
    _validar_totales_servicio(payload, calculo_servicio)

    settings = _settings_transporte(uid, token, perfil_factura)
    emisor = {
        "rfc": settings.get("RfcContribuyente", ""),
        "nombre": settings.get("DescripcionInstalacion", ""),
        "regimen_fiscal": settings.get("RegimenFiscal", "601"),
        "domicilio_fiscal": settings.get("CodigoPostal", ""),
    }
    if not emisor["rfc"] or not emisor["nombre"] or not emisor["domicilio_fiscal"]:
        raise HTTPException(400, "Configura RFC, razón social y código postal del contribuyente antes de facturar.")
    receptor = {
        "rfc": payload.rfc_receptor,
        "nombre": payload.nombre_receptor,
        "cp": payload.cp_receptor,
        "regimen_fiscal": payload.regimen_fiscal,
        "uso_cfdi": payload.uso_cfdi,
    }
    cfdi_dict = build_cfdi_servicio_transporte(
        emisor=emisor,
        receptor=receptor,
        cartas_porte=viajes,
        subtotal=calculo_servicio["subtotal"],
        iva=calculo_servicio["iva"],
        retencion=calculo_servicio["retencion"],
        iva_tasa=calculo_servicio["iva_tasa"],
        retencion_tasa=calculo_servicio["retencion_tasa"],
        aplica_iva=calculo_servicio["aplica_iva"],
        aplica_retencion=calculo_servicio["aplica_retencion"],
        forma_pago=payload.forma_pago,
        metodo_pago=payload.metodo_pago,
        uso_cfdi=payload.uso_cfdi,
    )
    sw = emitir_timbrar_json(cfdi_dict)
    if not sw.get("ok"):
        raise HTTPException(400, f"SW Sapien rechazó la factura de servicio: {sw.get('error')}")
    sw_data = sw.get("data") or {}

    now_iso = datetime.now(timezone.utc).isoformat()
    row = {
        "user_id":         uid,
        "perfil_id":       perfil_factura,
        "cliente_id":      payload.cliente_id,
        "viaje_ids":       payload.viaje_ids,
        "cfdi_relacionados": [
            {"viaje_id": v["id"], "uuid_cfdi": v.get("uuid_cfdi", ""), "id_ccp": v.get("id_ccp", "")}
            for v in viajes
        ],
        "rfc_receptor":    payload.rfc_receptor,
        "nombre_receptor": payload.nombre_receptor,
        "cp_receptor":     payload.cp_receptor,
        "regimen_fiscal":  payload.regimen_fiscal,
        "uso_cfdi":        payload.uso_cfdi,
        "concepto":        payload.concepto,
        "subtotal":        calculo_servicio["subtotal"],
        "iva":             calculo_servicio["iva"],
        "retencion":       calculo_servicio["retencion"],
        "total":           calculo_servicio["total"],
        "iva_tasa":        calculo_servicio["iva_tasa"],
        "retencion_tasa":  calculo_servicio["retencion_tasa"],
        "aplica_iva":      calculo_servicio["aplica_iva"],
        "aplica_retencion": calculo_servicio["aplica_retencion"],
        "calculo_json":    calculo_servicio,
        "forma_pago":      payload.forma_pago,
        "metodo_pago":     payload.metodo_pago,
        "moneda":          payload.moneda,
        "uuid_sat":        sw_data.get("uuid", ""),
        "xml_content":     sw_data.get("cfdi", ""),
        "pdf_url":         sw_data.get("pdfUrl", ""),
        "status":          "timbrada",
        "created_at":      now_iso,
    }
    try:
        res = sb.table(_TBL_FACT_SERV).insert(row).execute()
        factura_id = res.data[0]["id"] if res.data else None
        try:
            sb.table(_TBL_FACT_SERV_CARTAS).insert([
                {"user_id": uid, "perfil_id": perfil_factura, "factura_servicio_id": factura_id, "viaje_id": vid, "created_at": now_iso}
                for vid in payload.viaje_ids
            ]).execute()
        except Exception as e:
            logger.warning("No se pudo registrar bloqueo de doble factura: %s", e)
        for vid in payload.viaje_ids:
            _registrar_evento(
                sb, uid, perfil_factura, int(vid), "factura_servicio_timbrada",
                "Factura de servicio timbrada",
                f"UUID SAT {sw_data.get('uuid', '')}" if sw_data.get("uuid") else "Factura de servicio generada.",
                "system", "sw_sapien", {"factura_servicio_id": factura_id, "uuid_sat": sw_data.get("uuid", "")},
            )
        return JSONResponse({"ok": True, "id": factura_id, "status": "timbrada", "uuid_sat": sw_data.get("uuid", "")})
    except Exception as e:
        raise HTTPException(500, f"Error al crear factura de servicio: {e}")


@router.get("/tr/liquidaciones")
async def listar_liquidaciones(
    periodo: Optional[str] = Query(None),
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_LIQS).select("*").eq("user_id", uid).order("created_at", desc=True)
    if pid:
        q = q.eq("perfil_id", pid)
    if periodo:
        if re.match(r"^\d{4}-\d{2}$", periodo):
            q = q.in_("periodo", [periodo, f"{periodo}-Q1", f"{periodo}-Q2"])
        else:
            q = q.eq("periodo", periodo)
    return JSONResponse({"ok": True, "liquidaciones": q.execute().data or []})


@router.get("/tr/liquidaciones/{liquidacion_id}")
async def detalle_liquidacion(liquidacion_id: int, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    liq = sb.table(_TBL_LIQS).select("*").eq("id", liquidacion_id).eq("user_id", uid).limit(1).execute().data or []
    if not liq:
        raise HTTPException(404, "Liquidacion no encontrada.")
    items = sb.table(_TBL_LIQ_ITEMS).select("*").eq("liquidacion_id", liquidacion_id).eq("user_id", uid).execute().data or []
    return JSONResponse({"ok": True, "liquidacion": liq[0], "items": items})


@router.get("/tr/liquidaciones/{liquidacion_id}/export.xlsx")
async def exportar_liquidacion_xlsx(liquidacion_id: int, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    liq_rows = sb.table(_TBL_LIQS).select("*").eq("id", liquidacion_id).eq("user_id", uid).limit(1).execute().data or []
    if not liq_rows:
        raise HTTPException(404, "Liquidación no encontrada.")
    liq = liq_rows[0]
    items = sb.table(_TBL_LIQ_ITEMS).select("*").eq("liquidacion_id", liquidacion_id).eq("user_id", uid).execute().data or []
    chofer = {}
    if liq.get("chofer_id"):
        ch = sb.table(_TBL_CHOFERES).select("*").eq("id", liq.get("chofer_id")).eq("user_id", uid).limit(1).execute().data or []
        chofer = ch[0] if ch else {}
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Liquidacion"
        ws["A1"] = "GE CONTROL - Liquidación de chofer"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A3"] = "Folio"; ws["B3"] = liq.get("id")
        ws["A4"] = "Chofer"; ws["B4"] = chofer.get("nombre") or liq.get("chofer_id")
        ws["A5"] = "Periodo"; ws["B5"] = liq.get("periodo")
        ws["A6"] = "Estatus"; ws["B6"] = liq.get("status")
        headers = ["Viaje", "Concepto", "Litros", "Kilos", "Tarifa", "Subtotal", "IVA", "Retención", "Gastos", "Total"]
        ws.append([])
        ws.append(headers)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="7A1E2C")
        for it in items:
            ws.append([
                it.get("viaje_id"), it.get("concepto"), float(it.get("litros") or 0),
                float(it.get("kilos") or 0), float(it.get("tarifa") or 0),
                float(it.get("subtotal") or 0), float(it.get("iva") or 0),
                float(it.get("retencion") or 0), float(it.get("gastos") or 0),
                float(it.get("total") or 0),
            ])
        ws.append([])
        ws.append(["Subtotal", "", "", "", "", float(liq.get("subtotal") or 0)])
        ws.append(["IVA", "", "", "", "", float(liq.get("iva") or 0)])
        ws.append(["Retención", "", "", "", "", float(liq.get("retencion") or 0)])
        ws.append(["Gastos", "", "", "", "", float(liq.get("gastos") or 0)])
        ws.append(["Comisión extra", "", "", "", "", float(liq.get("comision_extra") or 0)])
        ws.append(["Descuentos", "", "", "", "", float(liq.get("descuentos") or 0)])
        ws.append(["Anticipos", "", "", "", "", float(liq.get("anticipos") or 0)])
        ws.append(["Total a pagar", "", "", "", "", float(liq.get("total") or 0)])
        for col in "ABCDEFGHIJ":
            ws.column_dimensions[col].width = 16
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
    except Exception as e:
        raise HTTPException(500, f"No se pudo generar Excel de liquidación: {e}")
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="liquidacion_{liquidacion_id}.xlsx"'},
    )


@router.post("/tr/liquidaciones/generar")
async def generar_liquidacion(payload: dict, authorization: str = Header(default=""), x_perfil_id: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    pid = _perfil(payload.get("perfil_id"), x_perfil_id)
    chofer_id = int(payload.get("chofer_id") or 0)
    periodo = _periodo_liquidacion_label(str(payload.get("periodo") or datetime.now(timezone.utc).strftime("%Y-%m")), str(payload.get("periodo_tipo") or ""))
    periodo_inicio, periodo_fin = _periodo_liquidacion_bounds(periodo, str(payload.get("periodo_tipo") or ""))
    if not chofer_id:
        raise HTTPException(400, "chofer_id requerido.")
    q = sb.table(_TBL_VIAJES).select("*").eq("user_id", uid).eq("chofer_id", chofer_id).gte("fecha_hora_salida", periodo_inicio).lt("fecha_hora_salida", periodo_fin)
    if pid:
        q = q.eq("perfil_id", pid)
    viajes = q.execute().data or []
    viajes = [v for v in viajes if (v.get("liquidacion_status") or "pendiente") in {"pendiente", "error", "borrador"}]
    if not viajes:
        raise HTTPException(404, "No hay viajes pendientes de liquidar para ese chofer/periodo.")

    tq = sb.table(_TBL_TARIFAS).select("*").eq("user_id", uid).eq("activo", True)
    if pid:
        tq = tq.eq("perfil_id", pid)
    tarifas = tq.execute().data or []

    items = []
    subtotal = iva = retencion = total = 0.0
    sin_tarifa = []
    for v in viajes:
        calc = _calcular_tarifa_operativa(v, tarifas)
        if not calc.get("tarifa_id"):
            sin_tarifa.append(v["id"])
            continue
        gastos = sb.table(_TBL_GASTOS).select("importe").eq("user_id", uid).eq("viaje_id", v["id"]).eq("status", "aprobado").execute().data or []
        gastos_total = round(sum(_safe_float(g.get("importe")) for g in gastos), 2)
        item_total = round(calc["total"] + gastos_total, 2)
        subtotal += calc["subtotal"]
        iva += calc["iva"]
        retencion += calc["retencion"]
        total += item_total
        items.append({
            "user_id": uid, "perfil_id": pid, "viaje_id": v["id"],
            "concepto": f"Flete viaje #{v['id']}",
            "litros": calc["litros"], "kilos": calc["kilos"], "tarifa": calc["tarifa"],
            "subtotal": calc["subtotal"], "iva": calc["iva"], "retencion": calc["retencion"],
            "gastos": gastos_total, "total": item_total, "metadata": calc,
        })
    if sin_tarifa:
        raise HTTPException(400, f"Configura tarifa antes de liquidar estos viajes: {sin_tarifa}")
    if not items:
        raise HTTPException(404, "No hay viajes con tarifa configurada para liquidar.")

    now_iso = datetime.now(timezone.utc).isoformat()
    anticipos = _safe_float(payload.get("anticipos"))
    comision_extra = _safe_float(payload.get("comision_extra"))
    descuentos = _safe_float(payload.get("descuentos"))
    liq_row = {
        "user_id": uid, "perfil_id": pid, "chofer_id": chofer_id, "periodo": periodo,
        "periodo_inicio": periodo_inicio, "periodo_fin": periodo_fin,
        "subtotal": round(subtotal, 2), "iva": round(iva, 2), "retencion": round(retencion, 2),
        "gastos": round(sum(i["gastos"] for i in items), 2),
        "anticipos": anticipos,
        "comision_extra": comision_extra,
        "descuentos": descuentos,
        "total": round(total + comision_extra - anticipos - descuentos, 2),
        "status": str(payload.get("status") or "emitida"),
        "notas": str(payload.get("notas") or ""),
        "metodo_pago": str(payload.get("metodo_pago") or ""),
        "referencia_pago": str(payload.get("referencia_pago") or ""),
        "metadata": {"periodo_inicio": periodo_inicio, "periodo_fin": periodo_fin, "items": len(items)},
        "created_at": now_iso,
    }
    res = sb.table(_TBL_LIQS).insert(liq_row).execute()
    liquidacion_id = res.data[0]["id"] if res.data else None
    for item in items:
        item["liquidacion_id"] = liquidacion_id
    if items:
        sb.table(_TBL_LIQ_ITEMS).insert(items).execute()
        ids = [i["viaje_id"] for i in items]
        sb.table(_TBL_VIAJES).update({"liquidacion_status": "emitida"}).eq("user_id", uid).in_("id", ids).execute()
        for vid in ids:
            _registrar_evento(sb, uid, pid, int(vid), "liquidacion_generada", "Liquidacion generada", f"Liquidacion #{liquidacion_id}", "oficina", uid, {"liquidacion_id": liquidacion_id})
    return JSONResponse({"ok": True, "liquidacion_id": liquidacion_id, "items": len(items), "total": liq_row["total"]})


@router.post("/tr/liquidaciones/{liquidacion_id}/pagar")
async def pagar_liquidacion(liquidacion_id: int, payload: dict, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    now_iso = datetime.now(timezone.utc).isoformat()
    metodo = str(payload.get("metodo_pago") or "").strip() or "efectivo"
    referencia = str(payload.get("referencia_pago") or "").strip()
    sb.table(_TBL_LIQS).update({
        "status": "pagada",
        "paid_at": now_iso,
        "metodo_pago": metodo,
        "referencia_pago": referencia,
    }).eq("id", liquidacion_id).eq("user_id", uid).execute()
    items = sb.table(_TBL_LIQ_ITEMS).select("viaje_id,perfil_id").eq("liquidacion_id", liquidacion_id).eq("user_id", uid).execute().data or []
    ids = [int(i["viaje_id"]) for i in items if i.get("viaje_id")]
    if ids:
        sb.table(_TBL_VIAJES).update({"liquidacion_status": "pagada"}).eq("user_id", uid).in_("id", ids).execute()
        for item in items:
            _registrar_evento(sb, uid, item.get("perfil_id"), int(item["viaje_id"]), "liquidacion_pagada", "Liquidacion pagada", f"Liquidacion #{liquidacion_id} · {metodo}", "oficina", uid, {"liquidacion_id": liquidacion_id, "metodo_pago": metodo, "referencia_pago": referencia})
    return JSONResponse({"ok": True})


@router.post("/tr/importar/excel-ruth")
async def importar_excel_ruth(
    file: UploadFile = File(...),
    dry_run: bool = Form(True),
    perfil_id: Optional[int] = Form(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    """Importador historico no destructivo: extrae resumen y tarifas del Excel operativo."""
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(await file.read()), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el Excel: {e}")

    resumen: dict = {"sheets": {}, "tarifas_detectadas": 0, "viajes_detectados": 0}
    for s in wb.sheetnames:
        ws = wb[s]
        nonempty = 0
        for row in ws.iter_rows():
            if any(c.value is not None and str(c.value).strip() for c in row):
                nonempty += 1
        resumen["sheets"][s] = {"rows": ws.max_row, "cols": ws.max_column, "nonempty_rows": nonempty}

    tarifas = []
    if "Precio.Tarifas" in wb.sheetnames:
        ws = wb["Precio.Tarifas"]
        for r in range(7, 13):
            origen, destino, producto, tiempos, tarifa = [ws.cell(r, c).value for c in range(2, 7)]
            if origen and destino and tarifa not in (None, ""):
                tarifas.append({
                    "user_id": uid, "perfil_id": pid, "origen": str(origen), "destino": str(destino),
                    "producto": str(producto or ""), "regla_calculo": "litros",
                    "tarifa": _safe_float(tarifa), "metadata": {"tiempos": str(tiempos or ""), "fuente": "Facturas de Ingreso Ruth.xlsx"},
                })
        for r in range(36, 40):
            destino = ws.cell(r, 2).value
            for c in range(3, 9):
                origen = ws.cell(35, c).value
                tarifa = ws.cell(r, c).value
                if origen and destino and tarifa not in (None, ""):
                    tarifas.append({
                        "user_id": uid, "perfil_id": pid, "origen": str(origen), "destino": str(destino),
                        "producto": "Gas LP", "regla_calculo": "kilos",
                        "tarifa": _safe_float(tarifa), "metadata": {"fuente": "Facturas de Ingreso Ruth.xlsx"},
                    })
    resumen["tarifas_detectadas"] = len(tarifas)
    for sheet in ("Gasolina Tabla", "Gas Tabla", "Gas LP", "Gaso Antiguo"):
        if sheet in wb.sheetnames:
            resumen["viajes_detectados"] += max((resumen["sheets"][sheet]["nonempty_rows"] - 1), 0)

    sb = _sb(token)
    inserted = 0
    if not dry_run and tarifas:
        res = sb.table(_TBL_TARIFAS).insert(tarifas).execute()
        inserted = len(res.data or [])
    sb.table(_TBL_IMPORTS).insert({
        "user_id": uid, "perfil_id": pid, "fuente": "excel_ruth",
        "filename": file.filename or "Facturas de Ingreso Ruth.xlsx",
        "resumen": resumen, "status": "preview" if dry_run else "procesada",
    }).execute()
    return JSONResponse({"ok": True, "dry_run": dry_run, "resumen": resumen, "tarifas_insertadas": inserted})


@router.get("/tr/facturas")
async def listar_facturas_transporte(
    periodo:       Optional[str] = Query(None),
    perfil_id:     Optional[int] = Query(None),
    authorization: str           = Header(default=""),
):
    """Lista los CFDIs timbrados del módulo transporte."""
    uid, token = _auth(authorization)
    sb = _sb(token)
    try:
        q = sb.table(_TBL_CFDI).select("*").eq("user_id", uid).order("fecha_timbrado", desc=True)
        if periodo:
            ini, fin = _periodo_bounds(periodo)
            q = q.gte("fecha_timbrado", ini).lt("fecha_timbrado", fin)
        if perfil_id:
            q = q.eq("perfil_id", perfil_id)
        res  = q.execute()
        rows = res.data or []
        # Omitir xml_content del listado (pesado)
        for r in rows:
            r.pop("xml_content", None)
        return JSONResponse({"ok": True, "facturas": rows})
    except Exception as e:
        raise HTTPException(500, f"Error al listar facturas: {e}")


@router.get("/tr/facturas/{cfdi_id}/xml")
async def descargar_xml_transporte(cfdi_id: int, authorization: str = Header(default="")):
    """Descarga el XML timbrado de un CFDI de transporte."""
    uid, token = _auth(authorization)
    sb = _sb(token)
    try:
        res = sb.table(_TBL_CFDI).select("uuid_sat,xml_content").eq("id", cfdi_id).eq("user_id", uid).limit(1).execute()
        rows = res.data or []
        if not rows:
            raise HTTPException(404, "CFDI no encontrado.")
        row = rows[0]
        return Response(
            content=row["xml_content"],
            media_type="application/xml",
            headers={"Content-Disposition": f'attachment; filename="cfdi_tr_{row["uuid_sat"]}.xml"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error al obtener XML: {e}")


@router.get("/tr/facturas/{cfdi_id}/pdf")
async def ver_pdf_carta_porte_transporte(
    cfdi_id: int,
    download: bool = Query(False),
    authorization: str = Header(default=""),
):
    """
    Genera y entrega la representación impresa del CFDI/Carta Porte desde el XML timbrado.
    No depende de que SW Sapien regrese pdfUrl.
    """
    uid, token = _auth(authorization)
    sb = _sb(token)
    try:
        res = (
            sb.table(_TBL_CFDI)
            .select("id,user_id,perfil_id,viaje_id,uuid_sat,id_ccp,xml_content,pdf_url")
            .eq("id", cfdi_id)
            .eq("user_id", uid)
            .limit(1)
            .execute()
        )
        rows = res.data or []
        if not rows:
            raise HTTPException(404, "CFDI no encontrado.")
        row = rows[0]
        xml_content = row.get("xml_content") or ""
        if not xml_content:
            raise HTTPException(404, "Este CFDI no tiene XML timbrado guardado.")

        viaje_rows = []
        productos = []
        if row.get("viaje_id"):
            viaje_rows = sb.table(_TBL_VIAJES).select("id,perfil_id,productos_json").eq("id", row.get("viaje_id")).eq("user_id", uid).limit(1).execute().data or []
            if viaje_rows:
                productos = _productos_from_row(viaje_rows[0])
        validacion = validar_xml_carta_porte_transporte(xml_content, productos)
        if validacion.bloquea_pdf:
            raise HTTPException(409, "PDF bloqueado: XML no válido como Carta Porte de carretera. " + "; ".join(validacion.errors[:5]))
        settings = _settings_transporte(uid, token, row.get("perfil_id"))
        info = extraer_info_pdf(xml_content)
        pdf_bytes = generar_pdf_carta_porte_desde_xml(xml_content, settings.get("PdfLogoDataUrl", ""))
        _guardar_cfdi_pdf_en_expediente(
            sb,
            uid,
            row,
            pdf_bytes,
            info.filename,
            {
                "cfdi_id": cfdi_id,
                "uuid_sat": info.uuid,
                "id_ccp": info.id_ccp,
                "has_carta_porte": info.has_carta_porte,
                "source": "xml_timbrado",
            },
        )
        disposition = "attachment" if download else "inline"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'{disposition}; filename="{info.filename}"'},
        )
    except HTTPException:
        raise
    except RuntimeError as e:
        raise HTTPException(500, str(e))
    except Exception as e:
        logger.exception("Error generando PDF Carta Porte cfdi=%s", cfdi_id)
        raise HTTPException(500, f"Error al generar PDF Carta Porte: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# 5. CONTROLES VOLUMÉTRICOS (covol)
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/tr/covol/generar")
async def generar_covol_transporte(
    payload:       GenerarCovolRequest,
    authorization: str = Header(default=""),
):
    """
    Genera el JSON de Controles Volumétricos mensual para transporte.
    Toma todos los viajes timbrados del periodo y los consolida.
    """
    uid, token = _auth(authorization)
    sb = _sb(token)

    periodo  = f"{payload.anio:04d}-{payload.mes:02d}"
    settings = _settings_transporte(uid, token, payload.perfil_id)

    if not settings.get("RfcContribuyente"):
        raise HTTPException(400, "Configura el RFC del contribuyente en Ajustes del módulo Transporte.")

    # Obtener viajes timbrados del periodo
    try:
        q = (
            sb.table(_TBL_VIAJES)
            .select("*")
            .eq("user_id", uid)
            .eq("status", "timbrado")
            .like("fecha_hora_salida", f"{periodo}%")
        )
        if payload.perfil_id:
            q = q.eq("perfil_id", payload.perfil_id)
        res   = q.execute()
        viajes_raw = res.data or []
    except Exception as e:
        raise HTTPException(500, f"Error al obtener viajes del periodo: {e}")

    if not viajes_raw:
        raise HTTPException(404, f"No hay viajes timbrados en el periodo {periodo}.")

    selected_permiso = (payload.num_permiso_cne or settings.get("NumPermiso", "") or "").strip()
    if not selected_permiso:
        raise HTTPException(
            400,
            "El número de permiso CNE es requerido para generar el JSON mensual de Transporte.",
        )

    def _permiso_viaje(row: dict) -> str:
        return (row.get("num_permiso_cne") or settings.get("NumPermiso", "") or "").strip()

    permisos_detectados = sorted({
        p for p in (_permiso_viaje(v) for v in viajes_raw) if p
    })
    viajes_raw = [v for v in viajes_raw if _permiso_viaje(v) == selected_permiso]
    if not viajes_raw:
        detalle = f" Permisos detectados en el periodo: {', '.join(permisos_detectados)}." if permisos_detectados else ""
        raise HTTPException(
            404,
            f"No hay viajes timbrados del periodo {periodo} para el permiso CNE {selected_permiso}.{detalle}",
        )

    # Convertir viajes_raw a formato esperado por transport_transformer
    viajes_para_covol: list[dict] = []
    for v in viajes_raw:
        try:
            productos_json = json.loads(v.get("productos_json") or "[]")
        except Exception:
            productos_json = []
        viajes_para_covol.append({
            "uuid_cfdi":         v.get("uuid_cfdi", ""),
            "id_ccp":            v.get("id_ccp", ""),
            "num_permiso_cne":    _permiso_viaje(v),
            "tipo_movimiento":   "descarga",   # El autotanque entrega → descarga en destino
            "fecha_hora_salida": v.get("fecha_hora_salida", ""),
            "rfc_receptor":      v.get("rfc_receptor", ""),
            "nombre_receptor":   v.get("nombre_receptor", ""),
            "productos":         productos_json,
        })

    # Preparar settings para el transformer
    covol_settings = {
        **settings,
        "NumPermiso":          selected_permiso,
        "ClaveInstalacion":    payload.clave_instalacion or settings.get("ClaveInstalacion", ""),
        "DescripcionInstalacion": payload.descripcion_instalacion or settings.get("DescripcionInstalacion", ""),
        "ModalidadPermiso":    settings.get("ModalidadPermiso", "PER51"),
    }

    try:
        sat_dict, meta = build_transport_covol(
            viajes=                  viajes_para_covol,
            settings=                covol_settings,
            anio=                    payload.anio,
            mes=                     payload.mes,
            inventario_inicial_litros= payload.inventario_inicial_litros,
        )
        archivos = save_transport_covol(sat_dict, meta, covol_settings)
    except Exception as e:
        logger.error("Error al generar covol transporte: %s", e)
        raise HTTPException(500, f"Error al generar reporte: {e}")

    # Guardar reporte en tr_covol_reports
    now_iso = datetime.now(timezone.utc).isoformat()
    try:
        sb.table(_TBL_COVOL).insert({
            "user_id":        uid,
            "perfil_id":      payload.perfil_id,
            "periodo":        periodo,
            "filename_base":  meta.get("first_uuid", "")[:8],
            "json_name":      archivos["json_name"],
            "zip_name":       archivos["zip_name"],
            "json_content":   archivos["json_content"],
            "zip_b64":        archivos["zip_b64"],
            "total_cargas":   meta.get("total_cargas", 0),
            "total_descargas": meta.get("total_descargas", 0),
            "num_productos":  meta.get("num_productos", 0),
            "created_at":     now_iso,
        }).execute()
    except Exception as e:
        logger.warning("No se pudo guardar covol en BD: %s", e)

    return JSONResponse({
        "ok":           True,
        "periodo":      periodo,
        "json_name":    archivos["json_name"],
        "zip_name":     archivos["zip_name"],
        "json_content": archivos["json_content"],
        "zip_b64":      archivos["zip_b64"],
        "num_permiso_cne": selected_permiso,
        "permisos_detectados": permisos_detectados,
        "meta":         {**meta, "num_permiso_cne": selected_permiso, "permisos_detectados": permisos_detectados},
    })


# ══════════════════════════════════════════════════════════════════════════════
# 6. CATÁLOGOS: Choferes, Vehículos, Rutas, Clientes
# ══════════════════════════════════════════════════════════════════════════════

# ── Choferes ──────────────────────────────────────────────────────────────────

@router.get("/tr/choferes")
async def listar_choferes(
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    sb = _sb(token)
    pid = _perfil(perfil_id, x_perfil_id)
    q = sb.table(_TBL_CHOFERES).select("*").eq("user_id", uid).eq("activo", True)
    if pid:
        q = q.eq("perfil_id", pid)
    res = q.order("nombre").execute()
    return JSONResponse({"choferes": res.data or []})


@router.post("/tr/choferes")
async def crear_chofer(
    payload: ChoferTransporteCreate,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    sb = _sb(token)
    pid = _perfil(perfil_id, x_perfil_id)
    try:
        res = sb.table(_TBL_CHOFERES).insert({
            "user_id":      uid,
            "perfil_id":    pid,
            "nombre":       payload.nombre.strip(),
            "rfc":          payload.rfc,
            "licencia":     payload.licencia.strip(),
            "tipo_licencia": payload.tipo_licencia,
            "telefono":     payload.telefono.strip(),
            "curp":         payload.curp.strip().upper(),
            "activo":       True,
            "created_at":   datetime.now(timezone.utc).isoformat(),
        }).execute()
        return JSONResponse({"ok": True, "id": res.data[0]["id"] if res.data else None})
    except Exception as e:
        raise HTTPException(500, f"Error al crear chofer: {e}")


@router.put("/tr/choferes/{chofer_id}")
async def actualizar_chofer(
    chofer_id: int, payload: ChoferTransporteCreate,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    sb = _sb(token)
    pid = _perfil(perfil_id, x_perfil_id)
    q = sb.table(_TBL_CHOFERES).update({
        "nombre":       payload.nombre.strip(),
        "rfc":          payload.rfc,
        "licencia":     payload.licencia.strip(),
        "tipo_licencia": payload.tipo_licencia,
        "telefono":     payload.telefono.strip(),
        "curp":         payload.curp.strip().upper(),
    }).eq("id", chofer_id).eq("user_id", uid)
    if pid:
        q = q.eq("perfil_id", pid)
    q.execute()
    return JSONResponse({"ok": True})


@router.delete("/tr/choferes/{chofer_id}")
async def eliminar_chofer(
    chofer_id: int,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_CHOFERES).update({"activo": False}).eq("id", chofer_id).eq("user_id", uid)
    if pid:
        q = q.eq("perfil_id", pid)
    q.execute()
    return JSONResponse({"ok": True})


# ── Vehículos ─────────────────────────────────────────────────────────────────

@router.get("/tr/vehiculos")
async def listar_vehiculos(
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_VEHICULOS).select("*").eq("user_id", uid).eq("activo", True)
    if pid:
        q = q.eq("perfil_id", pid)
    res = q.order("placas").execute()
    return JSONResponse({"vehiculos": res.data or []})


@router.post("/tr/vehiculos")
async def crear_vehiculo(
    payload: VehiculoTransporteCreate,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    try:
        res = _sb(token).table(_TBL_VEHICULOS).insert({
            "user_id":           uid,
            "perfil_id":         pid,
            "placas":            payload.placas,
            "modelo":            payload.modelo.strip(),
            "anio":              payload.anio,
            "config_vehicular":  payload.config_vehicular,
            "aseguradora":       payload.aseguradora.strip(),
            "poliza_seguro":     payload.poliza_seguro.strip(),
            "permiso_sct":       payload.permiso_sct.strip(),
            "num_permiso_sct":   payload.num_permiso_sct.strip(),
            "capacidad_litros":  payload.capacidad_litros,
            "num_ejes":          payload.num_ejes,
            "activo":            True,
            "created_at":        datetime.now(timezone.utc).isoformat(),
        }).execute()
        return JSONResponse({"ok": True, "id": res.data[0]["id"] if res.data else None})
    except Exception as e:
        raise HTTPException(500, f"Error al crear vehículo: {e}")


@router.put("/tr/vehiculos/{vehiculo_id}")
async def actualizar_vehiculo(
    vehiculo_id: int, payload: VehiculoTransporteCreate,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_VEHICULOS).update({
        "placas":          payload.placas,
        "modelo":          payload.modelo.strip(),
        "anio":            payload.anio,
        "config_vehicular": payload.config_vehicular,
        "aseguradora":     payload.aseguradora.strip(),
        "poliza_seguro":   payload.poliza_seguro.strip(),
        "permiso_sct":     payload.permiso_sct.strip(),
        "num_permiso_sct": payload.num_permiso_sct.strip(),
        "capacidad_litros": payload.capacidad_litros,
    }).eq("id", vehiculo_id).eq("user_id", uid)
    if pid:
        q = q.eq("perfil_id", pid)
    q.execute()
    return JSONResponse({"ok": True})


@router.delete("/tr/vehiculos/{vehiculo_id}")
async def eliminar_vehiculo(
    vehiculo_id: int,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_VEHICULOS).update({"activo": False}).eq("id", vehiculo_id).eq("user_id", uid)
    if pid:
        q = q.eq("perfil_id", pid)
    q.execute()
    return JSONResponse({"ok": True})


# ── Rutas ─────────────────────────────────────────────────────────────────────

@router.get("/tr/rutas")
async def listar_rutas(
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_RUTAS).select("*").eq("user_id", uid).eq("activo", True)
    if pid:
        q = q.eq("perfil_id", pid)
    res = q.order("nombre").execute()
    return JSONResponse({"rutas": res.data or []})


@router.post("/tr/rutas")
async def crear_ruta(
    payload: RutaTransporteCreate,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    try:
        row = _ruta_payload(payload)
        row.update({
            "user_id":       uid,
            "perfil_id":     pid,
            "activo":        True,
            "created_at":    datetime.now(timezone.utc).isoformat(),
        })
        res = _sb(token).table(_TBL_RUTAS).insert(row).execute()
        return JSONResponse({"ok": True, "id": res.data[0]["id"] if res.data else None})
    except Exception as e:
        raise HTTPException(500, f"Error al crear ruta: {e}")


@router.put("/tr/rutas/{ruta_id}")
async def actualizar_ruta(
    ruta_id: int, payload: RutaTransporteCreate,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_RUTAS).update(_ruta_payload(payload)).eq("id", ruta_id).eq("user_id", uid)
    if pid:
        q = q.eq("perfil_id", pid)
    q.execute()
    return JSONResponse({"ok": True})


@router.delete("/tr/rutas/{ruta_id}")
async def eliminar_ruta(
    ruta_id: int,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_RUTAS).update({"activo": False}).eq("id", ruta_id).eq("user_id", uid)
    if pid:
        q = q.eq("perfil_id", pid)
    q.execute()
    return JSONResponse({"ok": True})


# ── Clientes transporte ────────────────────────────────────────────────────────

@router.get("/tr/clientes")
async def listar_clientes_transporte(
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_CLIENTES).select("*").eq("user_id", uid).eq("activo", True)
    if pid:
        q = q.eq("perfil_id", pid)
    res = q.order("nombre").execute()
    return JSONResponse({"clientes": res.data or []})


@router.post("/tr/clientes")
async def crear_cliente_transporte(
    payload: ClienteTransporteCreate,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    receptor = _normalizar_receptor_cfdi(payload.rfc, payload.nombre, payload.cp, payload.regimen_fiscal)
    _validar_datos_cfdi_receptor(receptor["rfc"], receptor["regimen_fiscal"], receptor["cp"], payload.uso_cfdi)
    try:
        res = _sb(token).table(_TBL_CLIENTES).insert({
            "user_id":        uid,
            "perfil_id":      pid,
            "rfc":            receptor["rfc"],
            "nombre":         receptor["nombre"],
            "cp":             receptor["cp"],
            "regimen_fiscal": receptor["regimen_fiscal"],
            "uso_cfdi":       payload.uso_cfdi,
            "activo":         True,
            "created_at":     datetime.now(timezone.utc).isoformat(),
        }).execute()
        return JSONResponse({"ok": True, "id": res.data[0]["id"] if res.data else None})
    except Exception as e:
        raise HTTPException(500, f"Error al crear cliente: {e}")


@router.put("/tr/clientes/{cliente_id}")
async def actualizar_cliente_transporte(
    cliente_id: int, payload: ClienteTransporteCreate,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    receptor = _normalizar_receptor_cfdi(payload.rfc, payload.nombre, payload.cp, payload.regimen_fiscal)
    _validar_datos_cfdi_receptor(receptor["rfc"], receptor["regimen_fiscal"], receptor["cp"], payload.uso_cfdi)
    q = _sb(token).table(_TBL_CLIENTES).update({
        "rfc":            receptor["rfc"],
        "nombre":         receptor["nombre"],
        "cp":             receptor["cp"],
        "regimen_fiscal": receptor["regimen_fiscal"],
        "uso_cfdi":       payload.uso_cfdi,
    }).eq("id", cliente_id).eq("user_id", uid)
    if pid:
        q = q.eq("perfil_id", pid)
    q.execute()
    return JSONResponse({"ok": True})


@router.delete("/tr/clientes/{cliente_id}")
async def eliminar_cliente_transporte(
    cliente_id: int,
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil(perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_CLIENTES).update({"activo": False}).eq("id", cliente_id).eq("user_id", uid)
    if pid:
        q = q.eq("perfil_id", pid)
    q.execute()
    return JSONResponse({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
# 7. SETTINGS DEL MÓDULO TRANSPORTE
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/tr/settings")
async def get_settings_transporte(
    perfil_id:     Optional[int] = Query(None),
    authorization: str           = Header(default=""),
    x_perfil_id:   str           = Header(default=""),
):
    """Obtiene la configuración del módulo transporte."""
    uid, token = _auth(authorization)
    settings = _settings_transporte(uid, token, _perfil(perfil_id, x_perfil_id))
    return JSONResponse({"ok": True, "settings": settings})


@router.put("/tr/settings")
async def update_settings_transporte(
    data:          dict,
    perfil_id:     Optional[int] = Query(None),
    authorization: str           = Header(default=""),
    x_perfil_id:   str           = Header(default=""),
):
    """
    Guarda/actualiza la configuración del módulo transporte.
    Campos esperados en data:
      RfcContribuyente, DescripcionInstalacion, CodigoPostal, RegimenFiscal,
      NumPermiso, ClaveInstalacion, ModalidadPermiso, NumeroAutotanques,
      RfcProveedor, Caracter, display_name
    """
    uid, token = _auth(authorization)
    sb = _sb(token)
    perfil_id = _perfil(perfil_id, x_perfil_id)
    now_iso = datetime.now(timezone.utc).isoformat()

    # Limpiar campos sensibles
    data_limpia = {
        k: v for k, v in data.items()
        if k != "perfil_id" and isinstance(v, (str, int, float, bool, list, dict))
    }
    _validar_rfc_cp_config(data_limpia)

    try:
        # Verificar si ya existe un registro
        q = sb.table(_TBL_SETTINGS).select("id").eq("user_id", uid)
        if perfil_id:
            q = q.eq("perfil_id", perfil_id)
        res = q.limit(1).execute()
        rows = res.data or []

        if rows:
            sb.table(_TBL_SETTINGS).update({
                "data":       data_limpia,
                "updated_at": now_iso,
            }).eq("id", rows[0]["id"]).execute()
        else:
            sb.table(_TBL_SETTINGS).insert({
                "user_id":    uid,
                "perfil_id":  perfil_id,
                "data":       data_limpia,
                "updated_at": now_iso,
                "created_at": now_iso,
            }).execute()

        return JSONResponse({"ok": True})
    except Exception as e:
        raise HTTPException(500, f"Error al guardar configuración: {e}")
