from __future__ import annotations

from .core import *
import os
import re

from fastapi import File, Form, UploadFile
from models.transport_schemas import FacturaServicioCreate, GenerarCovolRequest, ViajeCreate
from services.service_invoice_builder import (
    CLAVE_SERVICIO_TRANSPORTE,
    DESCRIPCION_CARTA_INGRESO,
    build_cfdi_ingreso_carta_porte,
    build_cfdi_servicio_transporte,
)
from services.sw_sapien import emitir_timbrar_json, timbrar_cfdi
from services.transport_builder import build_cfdi_transporte_xml

_TBL_VIAJES = "tr_viajes"
_TBL_CFDI = "tr_cfdi"
_TBL_CLIENTES = "tr_clientes"
_TBL_CHOFERES = "tr_choferes"
_TBL_VEHICULOS = "tr_vehiculos"
_TBL_TARIFAS = "tr_tarifas"
_TBL_FACT_SERV = "tr_facturas_servicio"
_TBL_FACT_SERV_CARTAS = "tr_facturas_servicio_cartas"
_TBL_LIQS = "tr_liquidaciones"
_TBL_LIQ_ITEMS = "tr_liquidacion_items"
_TBL_GASTOS = "tr_gastos"
_TBL_IMPORTS = "tr_imports"
_TBL_COVOL = "tr_covol_reports"
_TBL_OPER_ACC = "tr_operador_accesos"
_ENABLE_SIMPLE_SERVICE_INVOICE = (os.environ.get("ENABLE_SIMPLE_SERVICE_INVOICE") or "").strip().lower() in {"1", "true", "yes", "on", "si", "sí"}


def _auth(authorization: str) -> tuple[str, str]:
    if not authorization.startswith("Bearer "):
        raise HTTPException(401, "No autenticado.")
    token = authorization[7:]
    uid = verify_token(token)
    if not uid:
        raise HTTPException(401, "Token inválido o expirado.")
    return uid, token


def _sb(token: str):
    try:
        from supabase_config import get_supabase_for_user

        return get_supabase_for_user(token)
    except Exception:
        return get_supabase_admin()


def _perfil_autorizado(uid: str, token: str, perfil_id=None, x_perfil_id: str = "") -> int | None:
    raw = perfil_id or (x_perfil_id or "").strip() or None
    if not raw:
        return None
    try:
        pid = int(raw)
    except (TypeError, ValueError):
        raise HTTPException(400, "Perfil inválido.")
    rows = (
        _sb(token)
        .table("perfiles_empresa")
        .select("id")
        .eq("id", pid)
        .eq("user_id", uid)
        .eq("activo", True)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows:
        raise HTTPException(403, "La empresa seleccionada no pertenece a tu usuario o está inactiva.")
    return pid


def _require_admin_transporte(uid: str, token: str) -> None:
    role = (obtener_acceso_modulo(uid, "transporte", access_token=token).get("role") or "user").lower()
    if role != "admin":
        raise HTTPException(403, "Solo administradores de Transporte pueden realizar esta acción.")


def _safe_float(value, default: float = 0.0) -> float:
    try:
        return float(value if value is not None and value != "" else default)
    except (TypeError, ValueError):
        return default


def _rfc_tipo_persona(rfc: str) -> str:
    clean = re.sub(r"[^A-Z0-9]", "", str(rfc or "").strip().upper())
    if len(clean) == 13:
        return "fisica"
    if len(clean) == 12:
        return "moral"
    return ""


def _fact_serv_requires_retencion(emisor_rfc: str, receptor_rfc: str) -> bool:
    return _rfc_tipo_persona(emisor_rfc) == "fisica" and _rfc_tipo_persona(receptor_rfc) == "moral"


def _validar_datos_cfdi_receptor(rfc: str, regimen: str, cp: str, uso_cfdi: str) -> None:
    from routes.core import _gas_lp_validar_datos_cfdi_receptor

    _gas_lp_validar_datos_cfdi_receptor(rfc, regimen, cp, uso_cfdi)


def _settings_transporte(uid: str, token: str, perfil_id) -> dict:
    if not perfil_id:
        return {}
    from routes.settings import _load as load_settings
    from routes.transporte_v2 import _deep_merge, _parse_json_value

    settings = load_settings(uid, int(perfil_id)) or {}
    try:
        rows = (
            _sb(token)
            .table("tr_settings")
            .select("data")
            .eq("user_id", uid)
            .eq("perfil_id", int(perfil_id))
            .limit(1)
            .execute()
            .data
            or []
        )
        if rows:
            tr_settings = _parse_json_value(rows[0].get("data"), {})
            if isinstance(tr_settings, dict):
                settings = _deep_merge(settings, tr_settings)
    except Exception as exc:
        logger.info("No se pudo mezclar tr_settings Transporte para Carta Ingreso: %s", exc)
    fiscal = settings.get("perfil_fiscal") if isinstance(settings.get("perfil_fiscal"), dict) else {}
    fiscal_rfc = str(fiscal.get("rfc_contribuyente") or "").strip()
    fiscal_nombre = str(fiscal.get("nombre_fiscal") or "").strip()
    fiscal_cp = str(fiscal.get("cp_fiscal") or "").strip()
    fiscal_regimen = str(fiscal.get("regimen_fiscal") or "").strip()
    settings["RfcContribuyente"] = fiscal_rfc or settings.get("RfcContribuyente") or settings.get("rfc") or ""
    settings["DescripcionInstalacion"] = fiscal_nombre or settings.get("DescripcionInstalacion") or settings.get("nombre") or ""
    settings["CodigoPostal"] = fiscal_cp or settings.get("CodigoPostal") or settings.get("codigo_postal") or ""
    settings["RegimenFiscal"] = fiscal_regimen or settings.get("RegimenFiscal") or settings.get("regimen_fiscal") or "601"
    return settings


def _cliente_por_receptor(sb, uid: str, perfil_id, rfc: str) -> dict:
    q = sb.table(_TBL_CLIENTES).select("*").eq("user_id", uid).eq("rfc", str(rfc or "").upper()).limit(1)
    if perfil_id:
        q = q.eq("perfil_id", perfil_id)
    rows = q.execute().data or []
    return rows[0] if rows else {}


def _service_invoice_catalog_email(cliente: dict | None) -> str:
    """Correo vigente del catálogo; nunca toma respaldos históricos del viaje/payload."""
    row = cliente or {}
    metadata = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
    return _clean_billing_email(
        row.get("email_facturacion")
        or row.get("email")
        or metadata.get("email_facturacion")
        or metadata.get("email")
        or metadata.get("correo")
    )


def _tariff_product_text(viaje: dict) -> str:
    productos = viaje.get("productos_json")
    if isinstance(productos, str):
        try:
            productos = json.loads(productos)
        except Exception:
            productos = []
    first = (productos or [{}])[0] if isinstance(productos, list) and productos else {}
    meta = _fact_serv_trip_meta(viaje)
    return " ".join(str(v or "") for v in [
        viaje.get("producto"),
        viaje.get("tipo_producto"),
        first.get("descripcion"),
        first.get("clave_prodserv_cfdi"),
        first.get("clave_producto"),
        meta.get("producto"),
    ]).strip().upper()


def _fact_serv_product_metadata(viaje: dict, *, base_carta: dict | None = None) -> dict:
    productos = viaje.get("productos_json")
    if isinstance(productos, str):
        try:
            productos = json.loads(productos)
        except Exception:
            productos = []
    first = (productos or [{}])[0] if isinstance(productos, list) and productos else {}
    meta = _fact_serv_trip_meta(viaje)
    producto_nombre = (
        viaje.get("producto")
        or viaje.get("producto_descripcion")
        or first.get("descripcion")
        or first.get("nombre")
        or meta.get("producto")
        or meta.get("producto_descripcion")
        or ""
    )
    producto_id = (
        viaje.get("producto_id")
        or viaje.get("producto_operacion_id")
        or first.get("producto_id")
        or first.get("id")
        or meta.get("producto_id")
    )
    product_text = _tariff_product_text(viaje)
    familia = ""
    if "GAS L" in product_text or "GAS LP" in product_text or "15111510" in product_text:
        familia = "gas_lp"
    elif any(token in product_text for token in ("MAGNA", "PREMIUM", "DIESEL", "DIÉSEL", "GASOLINA", "151015")):
        familia = "petroliferos"
    no_carta = "T"
    serie = str(meta.get("serie_carta_porte") or meta.get("serie") or "T").strip() or "T"
    folio = str(meta.get("folio_carta_porte") or meta.get("folio") or "").strip()
    if folio:
        no_carta = f"{serie}-{folio}"
    return {
        "producto_id": producto_id,
        "producto_nombre": producto_nombre,
        "producto_descripcion": producto_nombre,
        "producto_familia": familia,
        "familia_producto": familia,
        "litros": _safe_float(viaje.get("volumen_litros") or viaje.get("volumen_total_litros") or first.get("volumen_litros") or first.get("cantidad_litros")),
        "kilos": _safe_float(viaje.get("peso_kg") or first.get("peso_kg")),
        "origen": viaje.get("nombre_origen") or viaje.get("origen") or meta.get("nombre_origen") or meta.get("origen") or "",
        "destino": viaje.get("nombre_destino") or viaje.get("destino") or meta.get("nombre_destino") or meta.get("destino") or "",
        "no_carta_porte": no_carta,
        "serie_carta_porte": serie,
        "folio_carta_porte": folio,
        "fecha_carta_porte": viaje.get("fecha_salida") or viaje.get("fecha_hora_salida") or viaje.get("created_at") or "",
        "uuid_carta_porte_base": (base_carta or {}).get("uuid_sat") or (base_carta or {}).get("uuid_cfdi") or "",
        "id_ccp_carta_porte_base": (base_carta or {}).get("id_ccp") or "",
    }


def _tariff_match(viaje: dict, tarifa: dict) -> bool:
    if tarifa.get("cliente_id") and str(tarifa.get("cliente_id")) != str(viaje.get("cliente_id") or ""):
        return False
    productos = viaje.get("productos_json")
    if isinstance(productos, str):
        try:
            productos = json.loads(productos)
        except Exception:
            productos = []
    first = (productos or [{}])[0] if isinstance(productos, list) and productos else {}
    tarifa_producto_id = tarifa.get("producto_id")
    viaje_producto_id = (
        viaje.get("producto_id")
        or viaje.get("producto_operacion_id")
        or first.get("producto_id")
        or first.get("id")
    )
    mismo_producto_id = bool(
        tarifa_producto_id
        and viaje_producto_id
        and str(tarifa_producto_id) == str(viaje_producto_id)
    )
    producto_tarifa = str(tarifa.get("producto") or "").strip().upper()
    tarifa_meta = _fact_serv_trip_meta(tarifa)
    familia_tarifa = str(tarifa_meta.get("familia_producto") or "").strip().lower()
    texto_viaje = _tariff_product_text(viaje)
    es_petrolifero = any(value in texto_viaje for value in ("MAGNA", "PREMIUM", "DIESEL", "DIÉSEL", "GASOLINA"))
    familia_compatible = (
        (familia_tarifa == "petroliferos" and es_petrolifero)
        or (familia_tarifa == "gas_lp" and any(value in texto_viaje for value in ("GAS L.P", "GAS LP", "15111510")))
    )
    if producto_tarifa and not mismo_producto_id and not familia_compatible and producto_tarifa not in texto_viaje:
        return False
    def route_text(value):
        return " ".join("".join(ch if ch.isalnum() else " " for ch in str(value or "").upper()).split())

    origen = route_text(tarifa.get("origen"))
    viaje_origen = route_text(viaje.get("nombre_origen") or viaje.get("origen"))
    if origen and not (origen == viaje_origen or origen in viaje_origen or viaje_origen in origen):
        return False
    destino = route_text(tarifa.get("destino"))
    viaje_destino = route_text(viaje.get("nombre_destino") or viaje.get("destino"))
    if destino and not (destino == viaje_destino or destino in viaje_destino or viaje_destino in destino):
        return False
    has_route_text = bool(origen or destino)
    same_route_id = bool(tarifa.get("ruta_id") and str(tarifa.get("ruta_id")) == str(viaje.get("ruta_id") or ""))
    if tarifa.get("ruta_id") and not same_route_id and not has_route_text:
        return False
    return True


def _tariff_specificity(viaje: dict, tarifa: dict) -> int:
    if not _tariff_match(viaje, tarifa):
        return -1
    productos = viaje.get("productos_json")
    if isinstance(productos, str):
        try:
            productos = json.loads(productos)
        except Exception:
            productos = []
    first = (productos or [{}])[0] if isinstance(productos, list) and productos else {}
    viaje_producto_id = (
        viaje.get("producto_id")
        or viaje.get("producto_operacion_id")
        or first.get("producto_id")
        or first.get("id")
    )
    score = 0
    if tarifa.get("ruta_id") and str(tarifa.get("ruta_id")) == str(viaje.get("ruta_id") or ""):
        score += 100
    if str(tarifa.get("origen") or "").strip() or str(tarifa.get("destino") or "").strip():
        # Una coincidencia validada de origen/destino es suficiente aunque el
        # ruta_id histórico del viaje haya quedado desfasado.
        score += 120
    if tarifa.get("producto_id") and viaje_producto_id and str(tarifa.get("producto_id")) == str(viaje_producto_id):
        score += 80
    elif str(tarifa.get("producto") or "").strip().upper() and str(tarifa.get("producto") or "").strip().upper() in _tariff_product_text(viaje):
        score += 40
    if tarifa.get("cliente_id") and str(tarifa.get("cliente_id")) == str(viaje.get("cliente_id") or ""):
        score += 20
    if str(tarifa.get("origen") or "").strip():
        score += 5
    if str(tarifa.get("destino") or "").strip():
        score += 5
    try:
        score -= int(tarifa.get("prioridad") or 100)
    except Exception:
        score -= 100
    return score


def _calcular_tarifa_operativa(viaje: dict, tarifas: list[dict]) -> dict:
    matches = [(t, _tariff_specificity(viaje, t)) for t in (tarifas or [])]
    valid_matches = sorted((item for item in matches if item[1] >= 0), key=lambda item: item[1], reverse=True)
    tarifa = valid_matches[0][0] if valid_matches else None
    tarifa_error = ""
    if valid_matches:
        top_score = valid_matches[0][1]
        top_matches = [item[0] for item in valid_matches if item[1] == top_score]
        top_values = {
            (
                round(_safe_float(item.get("tarifa")), 6),
                str(item.get("regla_calculo") or item.get("base_calculo") or "").strip().lower(),
            )
            for item in top_matches
        }
        if len(top_values) > 1:
            tarifa = None
            tarifa_error = "Tarifa ambigua: existen configuraciones igualmente específicas con importes distintos."
    productos = viaje.get("productos_json")
    if isinstance(productos, str):
        try:
            productos = json.loads(productos)
        except Exception:
            productos = []
    first = (productos or [{}])[0] if isinstance(productos, list) and productos else {}
    regla = str(
        (tarifa or {}).get("regla_calculo")
        or (tarifa or {}).get("base_calculo")
        or ""
    ).strip().lower()
    if not regla:
        product_text = _tariff_product_text(viaje)
        regla = "kilos" if "GAS L" in product_text or "15111510" in product_text else "litros"
    litros = _safe_float(viaje.get("volumen_total_litros") or viaje.get("volumen_litros") or first.get("volumen_litros") or first.get("cantidad_litros"))
    kilos = _safe_float(viaje.get("peso_kg") or first.get("peso_kg"))
    if regla in {"kg", "kilo", "kilos"}:
        base = kilos
    elif regla == "distancia":
        base = _safe_float(viaje.get("distancia_km"), 1)
    elif regla == "viaje":
        base = 1.0
    else:
        base = litros
    tarifa_val = _safe_float((tarifa or {}).get("tarifa"))
    subtotal = round(base * tarifa_val, 2) if tarifa else round(_safe_float(viaje.get("subtotal_flete") or viaje.get("tarifa_total")), 2)
    iva_tasa = _safe_float((tarifa or {}).get("iva_tasa"), 0.16)
    retencion_tasa = _safe_float((tarifa or {}).get("retencion_tasa"), 0.04)
    aplica_iva = bool((tarifa or {}).get("aplica_iva", True))
    aplica_retencion = bool((tarifa or {}).get("aplica_retencion", True))
    iva = round(subtotal * iva_tasa, 2) if aplica_iva else 0.0
    retencion = round(subtotal * retencion_tasa, 2) if aplica_retencion else 0.0
    total = round(subtotal + iva - retencion, 2)
    return {
        "viaje_id": viaje.get("id"),
        "tarifa_id": (tarifa or {}).get("id"),
        "tarifa_error": tarifa_error,
        "regla_calculo": regla,
        "cantidad_base": base,
        "tarifa": tarifa_val,
        "subtotal": subtotal,
        "iva": iva,
        "retencion": retencion,
        "total": total,
        "iva_tasa": iva_tasa,
        "retencion_tasa": retencion_tasa,
        "aplica_iva": aplica_iva,
        "aplica_retencion": aplica_retencion,
    }


def _sumar_calculos_servicio(viajes: list[dict], tarifas: list[dict]) -> dict:
    items = [_calcular_tarifa_operativa(v, tarifas) for v in viajes]
    iva_rates = {item["iva_tasa"] for item in items}
    ret_rates = {item["retencion_tasa"] for item in items}
    return {
        "items": items,
        "subtotal": round(sum(item["subtotal"] for item in items), 2),
        "iva": round(sum(item["iva"] for item in items), 2),
        "retencion": round(sum(item["retencion"] for item in items), 2),
        "total": round(sum(item["total"] for item in items), 2),
        "iva_tasa": next(iter(iva_rates), 0.16),
        "retencion_tasa": next(iter(ret_rates), 0.04),
        "aplica_iva": any(item["aplica_iva"] for item in items),
        "aplica_retencion": any(item["aplica_retencion"] for item in items),
        "tasas_mixtas": len(iva_rates) > 1 or len(ret_rates) > 1,
    }


def _fact_serv_apply_required_retencion(calculo: dict, *, emisor_rfc: str, receptor_rfc: str) -> dict:
    if not _fact_serv_requires_retencion(emisor_rfc, receptor_rfc):
        return calculo
    items = []
    for raw in calculo.get("items") or []:
        item = dict(raw or {})
        subtotal = round(_safe_float(item.get("subtotal")), 2)
        iva_tasa = _safe_float(item.get("iva_tasa"), _safe_float(calculo.get("iva_tasa"), 0.16))
        retencion_tasa = _safe_float(item.get("retencion_tasa"), _safe_float(calculo.get("retencion_tasa"), 0.04)) or 0.04
        aplica_iva = bool(item.get("aplica_iva", calculo.get("aplica_iva", True)))
        iva = round(subtotal * iva_tasa, 2) if aplica_iva else 0.0
        retencion = round(subtotal * retencion_tasa, 2)
        item.update({
            "iva": iva,
            "retencion": retencion,
            "total": round(subtotal + iva - retencion, 2),
            "retencion_tasa": retencion_tasa,
            "aplica_retencion": True,
            "retencion_forzada_pf_pm": True,
        })
        items.append(item)
    if not items:
        return calculo
    iva_rates = {item["iva_tasa"] for item in items}
    ret_rates = {item["retencion_tasa"] for item in items}
    calculo.update({
        "items": items,
        "iva": round(sum(item["iva"] for item in items), 2),
        "retencion": round(sum(item["retencion"] for item in items), 2),
        "total": round(sum(item["total"] for item in items), 2),
        "iva_tasa": next(iter(iva_rates), _safe_float(calculo.get("iva_tasa"), 0.16)),
        "retencion_tasa": next(iter(ret_rates), _safe_float(calculo.get("retencion_tasa"), 0.04) or 0.04),
        "aplica_retencion": True,
        "retencion_forzada_pf_pm": True,
        "tasas_mixtas": len(iva_rates) > 1 or len(ret_rates) > 1,
    })
    return calculo


def _validar_totales_servicio(payload: FacturaServicioCreate, calculo: dict) -> None:
    expected = round(_safe_float(calculo.get("total")), 2)
    got = round(_safe_float(payload.total), 2)
    if abs(expected - got) > 0.05:
        raise HTTPException(400, f"El total cambió al recalcular la tarifa. Esperado ${expected:,.2f}, recibido ${got:,.2f}.")


def _periodo_bounds(periodo: str) -> tuple[str, str]:
    match = re.fullmatch(r"(\d{4})-(\d{2})", str(periodo or "").strip())
    if not match:
        raise HTTPException(400, "Periodo inválido. Usa formato YYYY-MM.")
    year, month = int(match.group(1)), int(match.group(2))
    if month < 1 or month > 12:
        raise HTTPException(400, "Mes inválido.")
    start = datetime(year, month, 1, tzinfo=timezone.utc)
    end = datetime(year + (1 if month == 12 else 0), 1 if month == 12 else month + 1, 1, tzinfo=timezone.utc)
    return start.isoformat(), end.isoformat()


def _registrar_evento(sb, uid: str, perfil_id, viaje_id: int, tipo: str, titulo: str, descripcion: str, actor_tipo: str, actor_id: str, metadata: dict) -> None:
    try:
        sb.table("tr_eventos").insert({
            "user_id": uid,
            "perfil_id": perfil_id,
            "viaje_id": viaje_id,
            "tipo": tipo,
            "titulo": titulo,
            "descripcion": descripcion,
            "actor_tipo": actor_tipo,
            "actor_id": actor_id,
            "metadata": metadata or {},
            "created_at": datetime.now(timezone.utc).isoformat(),
        }).execute()
    except Exception as exc:
        logger.info("No se pudo registrar evento transporte %s/%s: %s", viaje_id, tipo, exc)


def _base_cartas_porte_timbradas(sb, uid: str, perfil_id, viaje_ids: list[int]) -> dict[int, dict]:
    q = (
        sb.table(_TBL_CFDI)
        .select("id,viaje_id,uuid_sat,id_ccp,status,tipo_cfdi,fecha_timbrado,cancelacion_status,cancelacion_resultado")
        .eq("user_id", uid)
        .eq("tipo_cfdi", "T")
        .in_("viaje_id", viaje_ids)
    )
    if perfil_id:
        q = q.eq("perfil_id", perfil_id)
    rows = q.order("fecha_timbrado", desc=True).execute().data or []
    found: dict[int, dict] = {}
    for row in rows:
        vid = int(row.get("viaje_id") or 0)
        if not vid or vid in found or _fact_serv_cfdi_cancelada(row):
            continue
        if str(row.get("status") or "").lower() != "vigente":
            continue
        if not str(row.get("uuid_sat") or "").strip():
            continue
        found[vid] = row
    return found


def _get_row(sb, table: str, uid: str, perfil_id, row_id) -> dict:
    if not row_id:
        return {}
    q = sb.table(table).select("*").eq("id", row_id).eq("user_id", uid).limit(1)
    if perfil_id:
        q = q.eq("perfil_id", perfil_id)
    rows = q.execute().data or []
    return rows[0] if rows else {}


def _build_carta_ingreso_viaje(viaje: dict, payload: FacturaServicioCreate, calculo: dict, settings: dict, sb, uid: str, perfil_id) -> tuple[ViajeCreate, dict, dict]:
    from routes.transporte_v2 import (
        TBL_PRODUCTOS,
        _first_text,
        _meta,
        _stamp_expand_route_locations,
        _stamp_expand_vehicle_trailers,
        _stamp_make_producto,
        _stamp_transportista_permiso,
        _stamp_vehicle_payload,
    )

    viaje_expanded = dict(viaje or {})
    if viaje_expanded.get("ruta_id"):
        ruta = _get_row(sb, "tr_rutas", uid, perfil_id, viaje_expanded.get("ruta_id"))
        if ruta:
            viaje_expanded.update({k: v for k, v in _stamp_expand_route_locations(sb, uid, perfil_id, ruta).items() if v and not viaje_expanded.get(k)})
    meta = _meta(viaje_expanded)
    producto_row = _get_row(sb, TBL_PRODUCTOS, uid, perfil_id, viaje_expanded.get("producto_operacion_id") or viaje_expanded.get("producto_id"))
    producto = _stamp_make_producto(viaje_expanded, producto_row, settings)
    producto = producto.model_copy(update={"importe": _safe_float(calculo.get("subtotal"))})
    producto_text = _first_text(
        producto_row.get("tipo_producto"),
        producto_row.get("descripcion"),
        producto_row.get("nombre"),
        producto.descripcion,
        meta.get("producto"),
    )
    permiso_transportista = _stamp_transportista_permiso(sb, uid, perfil_id, producto_text)
    chofer = _get_row(sb, _TBL_CHOFERES, uid, perfil_id, viaje_expanded.get("chofer_id"))
    vehiculo_raw = _get_row(sb, _TBL_VEHICULOS, uid, perfil_id, viaje_expanded.get("vehiculo_id"))
    vehiculo = _stamp_vehicle_payload(_stamp_expand_vehicle_trailers(sb, uid, perfil_id, vehiculo_raw))
    if not chofer:
        raise HTTPException(400, "Falta operador/chofer para generar Carta Ingreso con Carta Porte.")
    if not vehiculo_raw:
        raise HTTPException(400, "Falta vehículo para generar Carta Ingreso con Carta Porte.")
    viaje_obj = ViajeCreate(
        perfil_id=perfil_id,
        facility_id=viaje_expanded.get("facility_id"),
        chofer_id=int(viaje_expanded.get("chofer_id") or 0),
        vehiculo_id=int(viaje_expanded.get("vehiculo_id") or 0),
        ruta_id=viaje_expanded.get("ruta_id"),
        proveedor_id=viaje_expanded.get("proveedor_id"),
        origen_id=viaje_expanded.get("origen_id"),
        destino_id=viaje_expanded.get("destino_id"),
        producto_operacion_id=viaje_expanded.get("producto_operacion_id") or viaje_expanded.get("producto_id"),
        programa_fecha=viaje_expanded.get("programa_fecha"),
        cp_origen=_first_text(viaje_expanded.get("cp_origen"), meta.get("cp_origen")),
        nombre_origen=_first_text(viaje_expanded.get("nombre_origen"), viaje_expanded.get("origen"), meta.get("nombre_origen")),
        rfc_origen=_first_text(viaje_expanded.get("rfc_origen"), meta.get("rfc_origen")),
        id_ubicacion_origen=_first_text(viaje_expanded.get("id_ubicacion_origen"), meta.get("id_ubicacion_origen")),
        estado_origen=_first_text(viaje_expanded.get("estado_origen"), meta.get("estado_origen")),
        municipio_origen=_first_text(viaje_expanded.get("municipio_origen"), meta.get("municipio_origen")),
        localidad_origen=_first_text(viaje_expanded.get("localidad_origen"), meta.get("localidad_origen")),
        calle_origen=_first_text(viaje_expanded.get("calle_origen"), meta.get("calle_origen")),
        cp_destino=_first_text(viaje_expanded.get("cp_destino"), meta.get("cp_destino")),
        nombre_destino=_first_text(viaje_expanded.get("nombre_destino"), viaje_expanded.get("destino"), meta.get("nombre_destino")),
        rfc_destino=_first_text(viaje_expanded.get("rfc_destino"), meta.get("rfc_destino"), payload.rfc_receptor),
        id_ubicacion_destino=_first_text(viaje_expanded.get("id_ubicacion_destino"), meta.get("id_ubicacion_destino")),
        estado_destino=_first_text(viaje_expanded.get("estado_destino"), meta.get("estado_destino")),
        municipio_destino=_first_text(viaje_expanded.get("municipio_destino"), meta.get("municipio_destino")),
        localidad_destino=_first_text(viaje_expanded.get("localidad_destino"), meta.get("localidad_destino")),
        calle_destino=_first_text(viaje_expanded.get("calle_destino"), meta.get("calle_destino")),
        fecha_hora_salida=_first_text(viaje_expanded.get("fecha_hora_salida"), viaje_expanded.get("fecha_salida")),
        fecha_hora_llegada=_first_text(viaje_expanded.get("fecha_hora_llegada"), viaje_expanded.get("fecha_llegada"), viaje_expanded.get("fecha_hora_salida")),
        productos=[producto],
        tipo_cfdi="I",
        rfc_receptor=payload.rfc_receptor,
        nombre_receptor=payload.nombre_receptor,
        cp_receptor=payload.cp_receptor,
        regimen_fiscal_receptor=payload.regimen_fiscal,
        uso_cfdi=payload.uso_cfdi,
        num_permiso_cne=_first_text(
            viaje_expanded.get("num_permiso_cne"),
            meta.get("num_permiso_cne"),
            (permiso_transportista or {}).get("permiso_cre"),
            settings.get("NumPermisoCNE"),
            settings.get("num_permiso_cne"),
        ),
        iva_tasa=_safe_float(calculo.get("iva_tasa"), 0.16),
        retencion_tasa=_safe_float(calculo.get("retencion_tasa"), 0.04),
        aplica_iva=bool(calculo.get("aplica_iva", True)),
        aplica_retencion=bool(calculo.get("aplica_retencion", False)),
        distancia_km=_safe_float(viaje_expanded.get("distancia_km") or meta.get("distancia_km"), 1.0),
        observaciones=_first_text(viaje_expanded.get("observaciones"), meta.get("observaciones")),
    )
    return viaje_obj, chofer, vehiculo


def _insert_factura_servicio_tolerant(sb, row: dict):
    try:
        return sb.table(_TBL_FACT_SERV).insert(row).execute()
    except Exception as exc:
        fallback = dict(row)
        metadata = dict(fallback.get("metadata") or {})
        for key in ("tipo", "uuid_carta_ingreso", "uuid_carta_porte_base", "status_timbrado", "xml_path", "pdf_path"):
            if key in fallback:
                metadata[key] = fallback.pop(key)
        fallback["metadata"] = metadata
        logger.info("Insert Carta Ingreso sin columnas nuevas, usando metadata: %s", exc)
        return sb.table(_TBL_FACT_SERV).insert(fallback).execute()


def _service_invoice_payment_defaults(cliente_cfg: dict, settings: dict) -> dict:
    meta = cliente_cfg.get("metadata") if isinstance(cliente_cfg.get("metadata"), dict) else {}
    settings_meta = settings.get("metadata") if isinstance(settings.get("metadata"), dict) else {}
    metodo = str(
        cliente_cfg.get("metodo_pago_default")
        or meta.get("metodo_pago_default")
        or meta.get("metodo_pago")
        or settings.get("MetodoPagoServicio")
        or settings_meta.get("metodo_pago_servicio")
        or "PUE"
    ).strip().upper()
    forma = str(
        cliente_cfg.get("forma_pago_default")
        or meta.get("forma_pago_default")
        or meta.get("forma_pago")
        or settings.get("FormaPagoServicio")
        or settings_meta.get("forma_pago_servicio")
        or "03"
    ).strip()
    if metodo not in {"PUE", "PPD"}:
        metodo = "PUE"
    if not re.fullmatch(r"\d{2}", forma):
        forma = "03" if metodo == "PUE" else "99"
    if metodo == "PPD" and forma == "03" and not (
        cliente_cfg.get("forma_pago_default") or meta.get("forma_pago_default") or meta.get("forma_pago")
    ):
        forma = "99"
    return {"metodo_pago": metodo, "forma_pago": forma}


def _fact_serv_trip_meta(row: dict) -> dict:
    meta = row.get("metadata") or row.get("defaults_json") or {}
    if isinstance(meta, str):
        try:
            parsed = json.loads(meta)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}
    return meta if isinstance(meta, dict) else {}


def _fact_serv_trip_period(row: dict) -> str:
    return str(row.get("fecha_hora_salida") or row.get("fecha_salida") or row.get("created_at") or "")[:7]


def _fact_serv_trip_permit(row: dict) -> str:
    meta = _fact_serv_trip_meta(row)
    return str(
        row.get("num_permiso_cne")
        or meta.get("num_permiso_cne")
        or meta.get("permiso_transportista")
        or meta.get("permiso_cre_transportista")
        or meta.get("permiso_cre")
        or ""
    ).strip()


def _fact_serv_cfdi_cancelada(row: dict) -> bool:
    if not row:
        return False
    result = row.get("cancelacion_resultado") or {}
    if isinstance(result, str):
        try:
            parsed = json.loads(result)
            result = parsed if isinstance(parsed, dict) else {}
        except Exception:
            result = {}
    status = str(row.get("cancelacion_status") or result.get("status") or "").strip().lower()
    if isinstance(result, dict) and (result.get("operativa") is True or result.get("warning")):
        return False
    if isinstance(result, dict) and result.get("ok") is True:
        return True
    if status in {"cancelled", "cancelado", "cancelada", "ok"}:
        return True
    return "cancel" in str(row.get("status") or "").lower() and status not in {"error", "rechazada", "rejected"}


def _fact_serv_amount(value, default: float = 0.0) -> float:
    try:
        return float(value if value is not None else default)
    except (TypeError, ValueError):
        return default


def _fact_serv_metadata(row: dict) -> dict:
    meta = row.get("metadata") or {}
    if isinstance(meta, str):
        try:
            meta = json.loads(meta)
        except Exception:
            meta = {}
    return meta if isinstance(meta, dict) else {}


def _fact_serv_invoice_cancelada(row: dict) -> bool:
    meta = _fact_serv_metadata(row)
    cancel_result = meta.get("cancelacion_resultado") or row.get("cancelacion_resultado") or {}
    text = " ".join(str(v or "") for v in (
        row.get("status"),
        row.get("estatus"),
        row.get("cancelacion_status"),
        meta.get("status"),
        meta.get("estatus"),
        meta.get("cancelacion_status"),
        "cancelada" if meta.get("canceled_at") else "",
        cancel_result.get("status") if isinstance(cancel_result, dict) else "",
        "cancelada" if isinstance(cancel_result, dict) and cancel_result.get("manual") else "",
    )).lower()
    return "cancel" in text


def _fact_serv_calc_base(item: dict, calculo: dict) -> float:
    rule = str(
        item.get("regla_calculo")
        or item.get("base_calculo")
        or calculo.get("regla_calculo")
        or calculo.get("base_calculo")
        or ""
    ).strip().lower()
    if rule in {"litro", "litros"}:
        return _fact_serv_amount(item.get("litros") or item.get("volumen_litros") or item.get("cantidad_base"))
    if rule in {"kg", "kilo", "kilos"}:
        return _fact_serv_amount(item.get("kilos") or item.get("peso_kg") or item.get("cantidad_base"))
    if rule == "distancia":
        return _fact_serv_amount(item.get("distancia_km") or item.get("distancia") or item.get("cantidad_base"))
    if rule == "viaje":
        return 1.0
    base = _fact_serv_amount(item.get("cantidad_base"))
    if base > 0:
        return base
    litros = _fact_serv_amount(item.get("litros") or item.get("volumen_litros"))
    kilos = _fact_serv_amount(item.get("kilos") or item.get("peso_kg"))
    return litros or kilos or 1.0


def _fact_serv_apply_tariff_override(calculo: dict, payload: FacturaServicioCreate) -> dict:
    override = _fact_serv_amount(getattr(payload, "override_tarifa", None), -1)
    if override < 0:
        return calculo
    items = calculo.get("items") if isinstance(calculo.get("items"), list) else []
    if len(items) != 1:
        return calculo
    item = dict(items[0] or {})
    base = _fact_serv_calc_base(item, calculo)
    subtotal = round(base * override, 2)
    iva_tasa = _fact_serv_amount(calculo.get("iva_tasa", item.get("iva_tasa")), 0.16)
    retencion_tasa = _fact_serv_amount(calculo.get("retencion_tasa", item.get("retencion_tasa")), 0.04)
    aplica_iva = bool(calculo.get("aplica_iva", item.get("aplica_iva", iva_tasa > 0)))
    aplica_retencion = bool(calculo.get("aplica_retencion", item.get("aplica_retencion", retencion_tasa > 0)))
    iva = round(subtotal * iva_tasa, 2) if aplica_iva else 0.0
    retencion = round(subtotal * retencion_tasa, 2) if aplica_retencion else 0.0
    total = round(subtotal + iva - retencion, 2)
    motivo = getattr(payload, "override_tarifa_motivo", "") or "Tarifa editada en revisión"
    item.update({
        "cantidad_base": base,
        "tarifa": override,
        "tarifa_override": True,
        "tarifa_override_motivo": motivo,
        "subtotal": subtotal,
        "iva": iva,
        "retencion": retencion,
        "total": total,
    })
    calculo.update({
        "items": [item],
        "subtotal": subtotal,
        "iva": iva,
        "retencion": retencion,
        "total": total,
        "iva_tasa": iva_tasa,
        "retencion_tasa": retencion_tasa,
        "aplica_iva": aplica_iva,
        "aplica_retencion": aplica_retencion,
        "tarifa_override": True,
        "tarifa_override_motivo": motivo,
    })
    return calculo


def _next_carta_ingreso_folio(sb, uid: str, perfil_id, *, prefix: str = "F") -> str:
    """Siguiente folio interno CI-F-N sin reutilizar folios cancelados."""
    max_num = 0
    max_row_id = 0
    try:
        # Las columnas serie_folio/folio_cfdi no existen aun en todos los
        # esquemas. Si se solicitan explicitamente, Supabase rechaza toda la
        # consulta y el consecutivo vuelve silenciosamente a F-1.
        q = sb.table(_TBL_FACT_SERV).select("id,metadata,xml_content").eq("user_id", uid)
        if perfil_id:
            q = q.eq("perfil_id", perfil_id)
        rows = q.order("created_at", desc=True).limit(1000).execute().data or []
    except Exception:
        rows = []
    pattern = re.compile(r"(?:CI-)?F-?(\d+)$", re.I)
    for row in rows:
        max_row_id = max(max_row_id, int(row.get("id") or 0))
        meta = _fact_serv_metadata(row)
        values = [
            meta.get("serie_folio"),
            meta.get("folio_cfdi"),
            meta.get("folio_solicitado"),
        ]
        xml = row.get("xml_content") or ""
        if xml:
            try:
                info = fiscal_pdf_info(xml, "carta_ingreso_transporte")
                values.append(info.serie_folio)
            except Exception:
                pass
        for value in values:
            match = pattern.search(str(value or "").strip())
            if match:
                try:
                    max_num = max(max_num, int(match.group(1)))
                except ValueError:
                    pass
    # Red de seguridad para registros legacy sin folio en metadata ni XML.
    return f"{prefix}-{(max_num + 1) if max_num else (max_row_id + 1)}"


@router.get("/tr/cartas-porte-facturables")
async def listar_cartas_porte_facturables(
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    """Cartas Porte timbradas que todavia no han sido usadas en Carta Ingreso."""
    uid, token = _auth(authorization)
    sb = _sb(token)
    pid = _perfil_autorizado(uid, token, perfil_id, x_perfil_id)
    try:
        fact_q = sb.table(_TBL_FACT_SERV_CARTAS).select("factura_servicio_id,viaje_id").eq("user_id", uid)
        if pid:
            fact_q = fact_q.eq("perfil_id", pid)
        fact_res = fact_q.execute()
        links = fact_res.data or []
        factura_ids = [int(r.get("factura_servicio_id") or 0) for r in links if r.get("factura_servicio_id")]
        facturas_activas = set()
        if factura_ids:
            fq = sb.table(_TBL_FACT_SERV).select("id,status").eq("user_id", uid).in_("id", factura_ids)
            if pid:
                fq = fq.eq("perfil_id", pid)
            facturas_activas = {
                int(r.get("id") or 0)
                for r in (fq.execute().data or [])
                if "cancel" not in str(r.get("status") or "").lower()
            }
        facturados = {
            int(r.get("viaje_id"))
            for r in links
            if r.get("viaje_id") and int(r.get("factura_servicio_id") or 0) in facturas_activas
        }
    except Exception:
        facturados = set()
    try:
        cfdi_q = (
            sb.table(_TBL_CFDI)
            .select("id,user_id,perfil_id,viaje_id,uuid_sat,id_ccp,rfc_receptor,status,tipo_cfdi,fecha_timbrado")
            .eq("user_id", uid)
            .eq("status", "Vigente")
            .eq("tipo_cfdi", "T")
        )
        if pid:
            cfdi_q = cfdi_q.eq("perfil_id", pid)
        cfdi_res = cfdi_q.order("fecha_timbrado", desc=True).limit(1000).execute()
        cfdis = [c for c in (cfdi_res.data or []) if int(c.get("viaje_id") or 0) not in facturados]
        viajes_ids = [int(c.get("viaje_id")) for c in cfdis if c.get("viaje_id")]
        viajes_map = {}
        if viajes_ids:
            vq = sb.table(_TBL_VIAJES).select("*").eq("user_id", uid).in_("id", viajes_ids)
            if pid:
                vq = vq.eq("perfil_id", pid)
            v_res = vq.execute()
            viajes_map = {int(v["id"]): v for v in (v_res.data or [])}
        cq = sb.table(_TBL_CLIENTES).select("*").eq("user_id", uid).eq("activo", True)
        if pid:
            cq = cq.eq("perfil_id", pid)
        clientes_res = cq.execute()
        clientes = clientes_res.data or []
        clientes_by_rfc = {str(c.get("rfc") or "").upper(): c for c in clientes}
        tq = sb.table(_TBL_TARIFAS).select("*").eq("user_id", uid).eq("activo", True)
        if pid:
            tq = tq.eq("perfil_id", pid)
        tarifas = tq.execute().data or []
        items = []
        for cfdi in cfdis:
            viaje = viajes_map.get(int(cfdi.get("viaje_id") or 0), {})
            cliente = clientes_by_rfc.get(str(viaje.get("rfc_receptor") or cfdi.get("rfc_receptor") or "").upper(), {})
            if cliente.get("id"):
                viaje = {**viaje, "cliente_id": cliente.get("id")}
            calc = _calcular_tarifa_operativa(viaje, tarifas)
            items.append({
                "viaje_id": cfdi.get("viaje_id"),
                "cfdi_id": cfdi.get("id"),
                "uuid_cfdi": cfdi.get("uuid_sat"),
                "id_ccp": cfdi.get("id_ccp"),
                "folio": cfdi.get("id_ccp") or cfdi.get("uuid_sat"),
                "cliente_id": cliente.get("id"),
                "rfc_receptor": cliente.get("rfc") or viaje.get("rfc_receptor") or cfdi.get("rfc_receptor"),
                "nombre_receptor": cliente.get("nombre") or viaje.get("nombre_receptor"),
                "cp_receptor": cliente.get("cp") or viaje.get("cp_receptor"),
                "regimen_fiscal": cliente.get("regimen_fiscal") or "601",
                "uso_cfdi": cliente.get("uso_cfdi") or viaje.get("uso_cfdi") or "G03",
                "subtotal": calc["subtotal"],
                "iva": calc["iva"],
                "retencion": calc["retencion"],
                "total": calc["total"],
                "iva_tasa": calc["iva_tasa"],
                "retencion_tasa": calc["retencion_tasa"],
                "aplica_iva": calc["aplica_iva"],
                "aplica_retencion": calc["aplica_retencion"],
                "tarifa_id": calc.get("tarifa_id"),
                "regla_calculo": calc.get("regla_calculo"),
            })
        return JSONResponse({"ok": True, "cartas": items})
    except Exception as e:
        raise HTTPException(500, f"Error al listar Cartas Porte facturables: {e}")


@router.post("/tr/facturas-servicio")
async def crear_factura_servicio(payload: FacturaServicioCreate, authorization: str = Header(default="")):
    """
    Timbrado visible de Carta Ingreso: CFDI I con Complemento Carta Porte 3.1.
    El CFDI de ingreso simple queda solo como compatibilidad legacy por feature flag.
    """
    uid, token = _auth(authorization)
    sb = _sb(token)

    _validar_datos_cfdi_receptor(payload.rfc_receptor, payload.regimen_fiscal, payload.cp_receptor, payload.uso_cfdi)
    viajes_res = sb.table(_TBL_VIAJES).select("*").eq("user_id", uid).in_("id", payload.viaje_ids).execute()
    viajes = viajes_res.data or []
    perfil_factura = payload.perfil_id or (viajes[0].get("perfil_id") if viajes else None)
    encontrados = {int(v["id"]) for v in viajes}
    faltantes = [vid for vid in payload.viaje_ids if vid not in encontrados]
    if faltantes:
        raise HTTPException(404, f"Viajes no encontrados: {faltantes}")
    base_cartas = _base_cartas_porte_timbradas(sb, uid, perfil_factura, payload.viaje_ids)
    no_timbrados = [v["id"] for v in viajes if int(v["id"]) not in base_cartas]
    if no_timbrados:
        raise HTTPException(400, f"Para timbrar Carta Ingreso, primero timbra la Carta Porte Traslado de los viajes: {no_timbrados}")
    cancelados = []
    for v in viajes:
        meta = _fact_serv_trip_meta(v)
        text = " ".join(str(value or "") for value in [
            v.get("status"), v.get("estatus"), v.get("carta_porte_status"),
            meta.get("status"), meta.get("estatus"), meta.get("carta_porte_status"),
        ]).lower()
        if "cancel" in text:
            cancelados.append(v.get("id"))
    if cancelados:
        raise HTTPException(400, f"Estas Cartas Porte están canceladas y no se pueden facturar: {cancelados}")
    try:
        cfdi_q = (
            sb.table(_TBL_CFDI)
            .select("viaje_id,status,cancelacion_status,cancelacion_resultado")
            .eq("user_id", uid)
            .eq("tipo_cfdi", "T")
            .in_("viaje_id", payload.viaje_ids)
        )
        if perfil_factura:
            cfdi_q = cfdi_q.eq("perfil_id", perfil_factura)
        cfdi_cancelados = [
            int(row.get("viaje_id") or 0)
            for row in (cfdi_q.execute().data or [])
            if row.get("viaje_id") and _fact_serv_cfdi_cancelada(row)
        ]
    except Exception:
        cfdi_cancelados = []
    if cfdi_cancelados:
        raise HTTPException(400, f"Estas Cartas Porte están canceladas fiscalmente y no se pueden facturar: {sorted(set(cfdi_cancelados))}")
    if perfil_factura:
        cerrados = []
        for v in viajes:
            periodo_viaje = _fact_serv_trip_period(v)
            permiso_viaje = _fact_serv_trip_permit(v)
            if not periodo_viaje or not permiso_viaje:
                continue
            try:
                cierre = (
                    sb.table("tr_covol_month_closures")
                    .select("id")
                    .eq("user_id", uid)
                    .eq("perfil_id", perfil_factura)
                    .eq("periodo", periodo_viaje)
                    .eq("num_permiso_cne", permiso_viaje)
                    .eq("status", "cerrado")
                    .limit(1)
                    .execute()
                    .data
                    or []
                )
                if cierre:
                    cerrados.append({"viaje": v.get("id"), "periodo": periodo_viaje, "permiso": permiso_viaje})
            except Exception:
                pass
        if cerrados:
            raise HTTPException(409, f"El mes ya está cerrado para estas Cartas Porte: {cerrados}")
    try:
        ya_q = sb.table(_TBL_FACT_SERV_CARTAS).select("factura_servicio_id,viaje_id").eq("user_id", uid).in_("viaje_id", payload.viaje_ids)
        if perfil_factura:
            ya_q = ya_q.eq("perfil_id", perfil_factura)
        ya_res = ya_q.execute()
        links = ya_res.data or []
        factura_ids = [int(r.get("factura_servicio_id") or 0) for r in links if r.get("factura_servicio_id")]
        facturas_activas = set()
        facturas_canceladas = set()
        if factura_ids:
            fq = sb.table(_TBL_FACT_SERV).select("id,status,estatus,cancelacion_status,cancelacion_resultado,metadata").eq("user_id", uid).in_("id", factura_ids)
            if perfil_factura:
                fq = fq.eq("perfil_id", perfil_factura)
            for r in (fq.execute().data or []):
                fid = int(r.get("id") or 0)
                if not fid:
                    continue
                if _fact_serv_invoice_cancelada(r):
                    facturas_canceladas.add(fid)
                else:
                    facturas_activas.add(fid)
        if facturas_canceladas:
            try:
                clean_q = sb.table(_TBL_FACT_SERV_CARTAS).delete().eq("user_id", uid).in_("factura_servicio_id", sorted(facturas_canceladas))
                if perfil_factura:
                    clean_q = clean_q.eq("perfil_id", perfil_factura)
                clean_q.execute()
                links = [r for r in links if int(r.get("factura_servicio_id") or 0) not in facturas_canceladas]
            except Exception as exc:
                logger.info("No se pudieron limpiar relaciones de Carta Ingreso cancelada: %s", exc)
        ya = [r.get("viaje_id") for r in links if int(r.get("factura_servicio_id") or 0) in facturas_activas]
        if ya:
            raise HTTPException(400, f"Estas Cartas Porte ya tienen Carta Ingreso: {ya}")
    except HTTPException:
        raise
    except Exception:
        # Compatibilidad con bases que aun no tienen la tabla de control.
        existentes = sb.table(_TBL_FACT_SERV).select("viaje_ids,status,estatus,cancelacion_status,cancelacion_resultado,metadata").eq("user_id", uid).execute().data or []
        usados = set()
        for f in existentes:
            if _fact_serv_invoice_cancelada(f):
                continue
            vals = f.get("viaje_ids") or []
            if isinstance(vals, list):
                usados.update(int(v) for v in vals if str(v).isdigit())
        repetidos = [v for v in payload.viaje_ids if v in usados]
        if repetidos:
            raise HTTPException(400, f"Estas Cartas Porte ya tienen Carta Ingreso: {repetidos}")

    settings = _settings_transporte(uid, token, perfil_factura)
    emisor = {
        "rfc": settings.get("RfcContribuyente", ""),
        "nombre": settings.get("DescripcionInstalacion", ""),
        "regimen_fiscal": settings.get("RegimenFiscal", "601"),
        "domicilio_fiscal": settings.get("CodigoPostal", ""),
    }
    if not emisor["rfc"] or not emisor["nombre"] or not emisor["domicilio_fiscal"]:
        raise HTTPException(400, "Configura RFC, razón social y código postal del contribuyente antes de facturar.")

    tq = sb.table(_TBL_TARIFAS).select("*").eq("user_id", uid).eq("activo", True)
    if perfil_factura:
        tq = tq.eq("perfil_id", perfil_factura)
    tarifas = tq.execute().data or []
    if payload.cliente_id:
        viajes_calc = [{**v, "cliente_id": payload.cliente_id} for v in viajes]
    else:
        viajes_calc = viajes
    calculo_servicio = _sumar_calculos_servicio(viajes_calc, tarifas)
    calculo_servicio = _fact_serv_apply_tariff_override(calculo_servicio, payload)
    calculo_servicio = _fact_serv_apply_required_retencion(
        calculo_servicio,
        emisor_rfc=emisor["rfc"],
        receptor_rfc=payload.rfc_receptor,
    )
    sin_tarifa = [
        i.get("viaje_id")
        for i in calculo_servicio.get("items", [])
        if not i.get("tarifa_id") and not i.get("tarifa_override")
    ]
    if sin_tarifa:
        errores_tarifa = [
            i.get("tarifa_error")
            for i in calculo_servicio.get("items", [])
            if i.get("tarifa_error")
        ]
        if errores_tarifa:
            raise HTTPException(400, f"No se puede timbrar: {errores_tarifa[0]} Viajes: {sin_tarifa}")
        raise HTTPException(400, f"Configura una tarifa de servicio antes de timbrar Carta Ingreso para estos viajes: {sin_tarifa}")
    if calculo_servicio.get("tasas_mixtas"):
        raise HTTPException(400, "No mezcles Cartas Porte con tasas distintas de IVA/retención en una sola Carta Ingreso.")
    _validar_totales_servicio(payload, calculo_servicio)

    receptor = {
        "rfc": payload.rfc_receptor,
        "nombre": payload.nombre_receptor,
        "cp": payload.cp_receptor,
        "regimen_fiscal": payload.regimen_fiscal,
        "uso_cfdi": payload.uso_cfdi,
    }
    cliente_cfg = {}
    if payload.cliente_id:
        cliente_rows = sb.table(_TBL_CLIENTES).select("*").eq("user_id", uid).eq("id", payload.cliente_id).limit(1).execute().data or []
        cliente_cfg = cliente_rows[0] if cliente_rows else {}
    if not cliente_cfg:
        cliente_cfg = _cliente_por_receptor(sb, uid, perfil_factura, payload.rfc_receptor)
    email_receptor = _service_invoice_catalog_email(cliente_cfg)
    fiscal_defaults = _service_invoice_payment_defaults(cliente_cfg, settings)
    forma_pago = payload.forma_pago or fiscal_defaults["forma_pago"]
    metodo_pago = payload.metodo_pago or fiscal_defaults["metodo_pago"]
    if fiscal_defaults["metodo_pago"] != "PUE" and metodo_pago == "PUE":
        metodo_pago = fiscal_defaults["metodo_pago"]
    if fiscal_defaults["forma_pago"] != "03" and forma_pago == "03":
        forma_pago = fiscal_defaults["forma_pago"]
    clave_carta_ingreso = str(
        settings.get("ClaveProdServCartaIngreso")
        or settings.get("clave_prodserv_carta_ingreso")
        or (settings.get("facturacion") or {}).get("clave_prodserv_carta_ingreso")
        or (settings.get("metadata") or {}).get("clave_prodserv_carta_ingreso")
        or CLAVE_SERVICIO_TRANSPORTE
    ).strip() or CLAVE_SERVICIO_TRANSPORTE
    descripcion_carta_ingreso = (payload.concepto or DESCRIPCION_CARTA_INGRESO).strip() or DESCRIPCION_CARTA_INGRESO
    id_ccp_carta_ingreso = ""
    if _ENABLE_SIMPLE_SERVICE_INVOICE:
        cfdi_dict = build_cfdi_servicio_transporte(
            emisor=emisor,
            receptor=receptor,
            cartas_porte=[base_cartas[int(v["id"])] for v in viajes],
            subtotal=calculo_servicio["subtotal"],
            iva=calculo_servicio["iva"],
            retencion=calculo_servicio["retencion"],
            iva_tasa=calculo_servicio["iva_tasa"],
            retencion_tasa=calculo_servicio["retencion_tasa"],
            aplica_iva=calculo_servicio["aplica_iva"],
            aplica_retencion=calculo_servicio["aplica_retencion"],
            forma_pago=forma_pago,
            metodo_pago=metodo_pago,
            uso_cfdi=payload.uso_cfdi,
            clave_prod_serv=clave_carta_ingreso,
            descripcion=descripcion_carta_ingreso,
        )
        tipo_registro = "factura_servicio"
    else:
        if len(viajes) != 1:
            raise HTTPException(400, "Por ahora timbra una Carta Ingreso por viaje/Carta Porte base para conservar el complemento Carta Porte 3.1 completo.")
        viaje_obj, chofer_row, vehiculo_row = _build_carta_ingreso_viaje(viajes[0], payload, calculo_servicio, settings, sb, uid, perfil_factura)
        folio_carta_ingreso = _next_carta_ingreso_folio(sb, uid, perfil_factura)
        try:
            cfdi_dict, id_ccp_carta_ingreso = build_cfdi_ingreso_carta_porte(
                viaje=viaje_obj,
                emisor=emisor,
                receptor=receptor,
                chofer=chofer_row,
                vehiculo=vehiculo_row,
                cartas_porte_base=[base_cartas[int(viajes[0]["id"])]],
                subtotal=calculo_servicio["subtotal"],
                iva=calculo_servicio["iva"],
                retencion=calculo_servicio["retencion"],
                iva_tasa=calculo_servicio["iva_tasa"],
                retencion_tasa=calculo_servicio["retencion_tasa"],
                aplica_iva=calculo_servicio["aplica_iva"],
                aplica_retencion=calculo_servicio["aplica_retencion"],
                forma_pago=forma_pago,
                metodo_pago=metodo_pago,
                uso_cfdi=payload.uso_cfdi,
                folio=folio_carta_ingreso,
                clave_prod_serv=clave_carta_ingreso,
                descripcion=descripcion_carta_ingreso,
            )
        except ValueError as exc:
            raise HTTPException(400, f"Faltan datos obligatorios Carta Porte 3.1 para Carta Ingreso: {exc}") from exc
        tipo_registro = "carta_ingreso"
    if tipo_registro == "carta_ingreso":
        xml_pre_timbrado = build_cfdi_transporte_xml(cfdi_dict)
        sw_xml = timbrar_cfdi(xml_pre_timbrado)
        if sw_xml.get("error"):
            raise HTTPException(400, f"SW Sapien rechazó la Carta Ingreso: {sw_xml.get('error')}")
        sw_data = {
            "uuid": sw_xml.get("uuid", ""),
            "cfdi": sw_xml.get("xml_timbrado", ""),
            "pdfUrl": sw_xml.get("pdf_url", ""),
        }
    else:
        sw = emitir_timbrar_json(cfdi_dict)
        if not sw.get("ok"):
            raise HTTPException(400, f"SW Sapien rechazó la Carta Ingreso: {sw.get('error')}")
        sw_data = sw.get("data") or {}

    cfdi_info = None
    if sw_data.get("cfdi"):
        try:
            cfdi_info = fiscal_pdf_info(
                sw_data.get("cfdi") or "",
                "carta_ingreso_transporte" if tipo_registro == "carta_ingreso" else "factura_servicio_transporte",
            )
        except Exception:
            cfdi_info = None

    now_iso = datetime.now(timezone.utc).isoformat()
    fecha_cfdi = str(cfdi_dict.get("Fecha") or "")
    product_metadata = {}
    if viajes:
        first_viaje = viajes[0]
        product_metadata = _fact_serv_product_metadata(
            first_viaje,
            base_carta=base_cartas.get(int(first_viaje["id"])),
        )
    row = {
        "user_id":         uid,
        "perfil_id":       perfil_factura,
        "cliente_id":      payload.cliente_id,
        "viaje_ids":       payload.viaje_ids,
        "cfdi_relacionados": [
            {
                "viaje_id": v["id"],
                "uuid_cfdi": base_cartas[int(v["id"])].get("uuid_sat", ""),
                "uuid_sat": base_cartas[int(v["id"])].get("uuid_sat", ""),
                "id_ccp": base_cartas[int(v["id"])].get("id_ccp", ""),
            }
            for v in viajes
        ],
        "rfc_receptor":    payload.rfc_receptor,
        "nombre_receptor": payload.nombre_receptor,
        "cp_receptor":     payload.cp_receptor,
        "regimen_fiscal":  payload.regimen_fiscal,
        "uso_cfdi":        payload.uso_cfdi,
        "concepto":        descripcion_carta_ingreso,
        "subtotal":        calculo_servicio["subtotal"],
        "iva":             calculo_servicio["iva"],
        "retencion":       calculo_servicio["retencion"],
        "total":           calculo_servicio["total"],
        "iva_tasa":        calculo_servicio["iva_tasa"],
        "retencion_tasa":  calculo_servicio["retencion_tasa"],
        "aplica_iva":      calculo_servicio["aplica_iva"],
        "aplica_retencion": calculo_servicio["aplica_retencion"],
        "calculo_json":    calculo_servicio,
        "forma_pago":      forma_pago,
        "metodo_pago":     metodo_pago,
        "moneda":          payload.moneda,
        "uuid_sat":        sw_data.get("uuid", ""),
        "xml_content":     sw_data.get("cfdi", ""),
        "pdf_url":         sw_data.get("pdfUrl", ""),
        "status":          "timbrada",
        "tipo":            tipo_registro,
        "uuid_carta_ingreso": sw_data.get("uuid", "") if tipo_registro == "carta_ingreso" else "",
        "uuid_carta_porte_base": ",".join(base_cartas[int(v["id"])].get("uuid_sat", "") for v in viajes),
        "status_timbrado": "timbrada",
        "xml_path":        "",
        "pdf_path":        "",
        "metadata":        {
            "tipo": tipo_registro,
            "cfdi_tipo": cfdi_dict.get("TipoDeComprobante"),
            "folio_solicitado": folio_carta_ingreso if tipo_registro == "carta_ingreso" else "",
            "fecha_emision": fecha_cfdi,
            "fecha_cfdi": fecha_cfdi,
            "id_ccp_carta_ingreso": id_ccp_carta_ingreso,
            "uuid_carta_ingreso": sw_data.get("uuid", "") if tipo_registro == "carta_ingreso" else "",
            "uuid_carta_porte_base": [base_cartas[int(v["id"])].get("uuid_sat", "") for v in viajes],
            "serie_folio": cfdi_info.serie_folio if cfdi_info else "",
            "folio_cfdi": cfdi_info.serie_folio if cfdi_info else "",
            "concepto_clave_prod_serv": clave_carta_ingreso,
            "legacy_simple_service_invoice": _ENABLE_SIMPLE_SERVICE_INVOICE,
            "email_receptor": email_receptor,
            "email_delivery": {
                "status": "pendiente" if email_receptor else "omitido",
                "provider": "resend",
                **({} if email_receptor else {"reason": "cliente_sin_correo"}),
            },
            **product_metadata,
        },
        "created_at":      now_iso,
    }
    try:
        res = _insert_factura_servicio_tolerant(sb, row)
        factura_id = res.data[0]["id"] if res.data else None
        try:
            sb.table(_TBL_FACT_SERV).update({
                "idempotency_key": f"{'-'.join(str(v) for v in sorted(payload.viaje_ids))}:{tipo_registro}",
            }).eq("id", factura_id).eq("user_id", uid).execute()
        except Exception as exc:
            logger.info("Columnas idempotency factura servicio aun no disponibles factura=%s: %s", factura_id, exc)
        try:
            sb.table(_TBL_FACT_SERV_CARTAS).insert([
                {"user_id": uid, "perfil_id": perfil_factura, "factura_servicio_id": factura_id, "viaje_id": vid, "created_at": now_iso}
                for vid in payload.viaje_ids
            ]).execute()
        except Exception as e:
            logger.warning("No se pudo registrar bloqueo de doble factura: %s", e)
        for vid in payload.viaje_ids:
            _registrar_evento(
                sb, uid, perfil_factura, int(vid), "carta_ingreso_timbrada",
                "Carta Ingreso timbrada",
                f"UUID SAT {sw_data.get('uuid', '')}" if sw_data.get("uuid") else "Carta Ingreso generada.",
                "system", "sw_sapien", {"factura_servicio_id": factura_id, "uuid_sat": sw_data.get("uuid", ""), "tipo": tipo_registro},
            )
            try:
                viaje_meta = _fact_serv_trip_meta(next((v for v in viajes if int(v.get("id") or 0) == int(vid)), {}))
                viaje_meta.update({
                    "factura_servicio_calculo": calculo_servicio,
                    "factura_servicio_tarifa": (calculo_servicio.get("items") or [{}])[0].get("tarifa"),
                    "factura_servicio_fecha_cfdi": fecha_cfdi,
                })
                sb.table(_TBL_VIAJES).update({
                    "factura_servicio_status": "timbrada",
                    "factura_servicio_uuid": sw_data.get("uuid", ""),
                    "factura_servicio_pdf_url": f"/api/tr-v2/facturas-servicio/{factura_id}/pdf?download=true",
                    "factura_servicio_xml_url": f"/api/tr-v2/facturas-servicio/{factura_id}/xml",
                    "metadata": viaje_meta,
                }).eq("id", int(vid)).eq("user_id", uid).execute()
            except Exception as exc:
                logger.info("Columnas separadas factura servicio aun no disponibles viaje=%s: %s", vid, exc)
                try:
                    sb.table(_TBL_VIAJES).update({"metadata": viaje_meta}).eq("id", int(vid)).eq("user_id", uid).execute()
                except Exception as meta_exc:
                    logger.info("No se pudo guardar metadata de calculo Carta Ingreso viaje=%s: %s", vid, meta_exc)
        version_xml(
            module="transporte",
            entity_type=tipo_registro,
            entity_id=factura_id,
            uuid_sat=sw_data.get("uuid", ""),
            xml_content=sw_data.get("cfdi", ""),
            user_id=uid,
            perfil_id=perfil_factura,
            source="sw_sapien",
        )
        email_delivery = {
            "ok": False,
            "skipped": True,
            "error": "Carta Ingreso sin XML para adjuntar." if email_receptor else "Cliente sin correo fiscal.",
            "provider": "resend",
        }
        xml_content = sw_data.get("cfdi", "") or ""
        if xml_content and email_receptor:
            try:
                info = fiscal_pdf_info(
                    xml_content,
                    "carta_ingreso_transporte" if tipo_registro == "carta_ingreso" else "factura_servicio_transporte",
                )
                logo_data_url = settings.get("PdfLogoDataUrl", "") or (settings.get("perfil_fiscal") or {}).get("logo_data_url", "")
                pdf_theme = settings.get("perfil_fiscal") if isinstance(settings.get("perfil_fiscal"), dict) else {}
                pdf_bytes = (
                    generar_pdf_ingreso_carta_porte_desde_xml(xml_content, logo_data_url=logo_data_url, pdf_theme=pdf_theme)
                    if tipo_registro == "carta_ingreso"
                    else generar_pdf_ingreso_desde_xml(xml_content, logo_data_url=logo_data_url)
                )
                email_result = send_gas_lp_invoice_email(
                    to_email=email_receptor,
                    issuer_name=emisor.get("nombre", ""),
                    customer_name=payload.nombre_receptor,
                    uuid_sat=sw_data.get("uuid", ""),
                    total=calculo_servicio["total"],
                    xml_content=xml_content,
                    pdf_bytes=pdf_bytes,
                    pdf_filename=info.filename,
                    serie_folio=info.serie_folio or (f"FS-{factura_id or ''}" if tipo_registro != "carta_ingreso" else "CI"),
                )
                email_delivery = email_result.as_metadata()
            except Exception as exc:
                email_delivery = {"ok": False, "skipped": False, "error": str(exc)[:500], "provider": "resend"}
        try:
            merged_metadata = dict(row.get("metadata") or {})
            merged_metadata.update({"email_receptor": email_receptor, "email_delivery": email_delivery})
            update_payload = {"metadata": merged_metadata, "email_receptor": email_receptor}
            try:
                sb.table(_TBL_FACT_SERV).update(update_payload).eq("id", factura_id).eq("user_id", uid).execute()
            except Exception:
                sb.table(_TBL_FACT_SERV).update({"metadata": update_payload["metadata"]}).eq("id", factura_id).eq("user_id", uid).execute()
        except Exception as exc:
            logger.warning("No se pudo guardar auditoría email Carta Ingreso: %s", exc)
        return JSONResponse({"ok": True, "id": factura_id, "status": "timbrada", "uuid_sat": sw_data.get("uuid", ""), "email_delivery": email_delivery})
    except Exception as e:
        raise HTTPException(500, f"Error al crear Carta Ingreso: {e}")


@router.post("/tr/sat-sync/manual-xml")
async def sat_sync_manual_xml_transporte(
    files: list[UploadFile] = File(...),
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    _require_admin_transporte(uid, token)
    scope = _perfil_sat_scope(uid, token, perfil_id, x_perfil_id)
    if not files:
        raise HTTPException(400, "Sube al menos un XML SAT.")
    if len(files) > 50:
        raise HTTPException(400, "Máximo 50 XML por carga manual.")

    xml_items = []
    for file in files:
        filename = file.filename or "cfdi.xml"
        if not filename.lower().endswith(".xml"):
            raise HTTPException(400, f"Solo se aceptan XML SAT. Archivo inválido: {filename}")
        content = await file.read()
        if len(content) > 2_000_000:
            raise HTTPException(400, f"XML demasiado grande: {filename}")
        xml_items.append({"filename": filename, "content": content})

    result = ingest_manual_sat_xmls(
        sb=get_supabase_admin(),
        window=SatSyncWindow(
            tenant_id=scope["tenant_id"],
            company_id=scope["company_id"],
            perfil_id=scope["perfil_id"],
            sync_type="both",
            provider="manual",
        ),
        xml_items=xml_items,
        created_by=uid,
    )
    return JSONResponse(result, status_code=200 if result.get("ok") else 207)


@router.get("/tr/liquidaciones")
async def listar_liquidaciones(
    periodo: Optional[str] = Query(None),
    perfil_id: Optional[int] = Query(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    uid, token = _auth(authorization)
    pid = _perfil_autorizado(uid, token, perfil_id, x_perfil_id)
    q = _sb(token).table(_TBL_LIQS).select("*").eq("user_id", uid).order("created_at", desc=True)
    if pid:
        q = q.eq("perfil_id", pid)
    if periodo:
        if re.match(r"^\d{4}-\d{2}$", periodo):
            q = q.in_("periodo", [periodo, f"{periodo}-Q1", f"{periodo}-Q2"])
        else:
            q = q.eq("periodo", periodo)
    return JSONResponse({"ok": True, "liquidaciones": q.execute().data or []})


@router.get("/tr/liquidaciones/{liquidacion_id}")
async def detalle_liquidacion(liquidacion_id: int, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    liq = sb.table(_TBL_LIQS).select("*").eq("id", liquidacion_id).eq("user_id", uid).limit(1).execute().data or []
    if not liq:
        raise HTTPException(404, "Liquidacion no encontrada.")
    items = sb.table(_TBL_LIQ_ITEMS).select("*").eq("liquidacion_id", liquidacion_id).eq("user_id", uid).execute().data or []
    return JSONResponse({"ok": True, "liquidacion": liq[0], "items": items})


@router.get("/tr/liquidaciones/{liquidacion_id}/export.xlsx")
async def exportar_liquidacion_xlsx(liquidacion_id: int, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    liq_rows = sb.table(_TBL_LIQS).select("*").eq("id", liquidacion_id).eq("user_id", uid).limit(1).execute().data or []
    if not liq_rows:
        raise HTTPException(404, "Liquidación no encontrada.")
    liq = liq_rows[0]
    items = sb.table(_TBL_LIQ_ITEMS).select("*").eq("liquidacion_id", liquidacion_id).eq("user_id", uid).execute().data or []
    chofer = {}
    if liq.get("chofer_id"):
        ch = sb.table(_TBL_CHOFERES).select("*").eq("id", liq.get("chofer_id")).eq("user_id", uid).limit(1).execute().data or []
        chofer = ch[0] if ch else {}
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Liquidacion"
        ws["A1"] = "GE CONTROL - Liquidación de chofer"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A3"] = "Folio"; ws["B3"] = liq.get("id")
        ws["A4"] = "Chofer"; ws["B4"] = chofer.get("nombre") or liq.get("chofer_id")
        ws["A5"] = "Periodo"; ws["B5"] = liq.get("periodo")
        ws["A6"] = "Estatus"; ws["B6"] = liq.get("status")
        headers = ["Viaje", "Concepto", "Litros", "Kilos", "Tarifa", "Subtotal", "IVA", "Retención", "Gastos", "Total"]
        ws.append([])
        ws.append(headers)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="7A1E2C")
        for it in items:
            ws.append([
                it.get("viaje_id"), it.get("concepto"), float(it.get("litros") or 0),
                float(it.get("kilos") or 0), float(it.get("tarifa") or 0),
                float(it.get("subtotal") or 0), float(it.get("iva") or 0),
                float(it.get("retencion") or 0), float(it.get("gastos") or 0),
                float(it.get("total") or 0),
            ])
        ws.append([])
        ws.append(["Subtotal", "", "", "", "", float(liq.get("subtotal") or 0)])
        ws.append(["IVA", "", "", "", "", float(liq.get("iva") or 0)])
        ws.append(["Retención", "", "", "", "", float(liq.get("retencion") or 0)])
        ws.append(["Gastos", "", "", "", "", float(liq.get("gastos") or 0)])
        ws.append(["Comisión extra", "", "", "", "", float(liq.get("comision_extra") or 0)])
        ws.append(["Descuentos", "", "", "", "", float(liq.get("descuentos") or 0)])
        ws.append(["Anticipos", "", "", "", "", float(liq.get("anticipos") or 0)])
        ws.append(["Total a pagar", "", "", "", "", float(liq.get("total") or 0)])
        for col in "ABCDEFGHIJ":
            ws.column_dimensions[col].width = 16
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
    except Exception as e:
        raise HTTPException(500, f"No se pudo generar Excel de liquidación: {e}")
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="liquidacion_{liquidacion_id}.xlsx"'},
    )


@router.post("/tr/liquidaciones/generar")
async def generar_liquidacion(payload: dict, authorization: str = Header(default=""), x_perfil_id: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    pid = _perfil_autorizado(uid, token, payload.get("perfil_id"), x_perfil_id)
    chofer_id = int(payload.get("chofer_id") or 0)
    periodo = _periodo_liquidacion_label(str(payload.get("periodo") or datetime.now(timezone.utc).strftime("%Y-%m")), str(payload.get("periodo_tipo") or ""))
    periodo_inicio, periodo_fin = _periodo_liquidacion_bounds(periodo, str(payload.get("periodo_tipo") or ""))
    if not chofer_id:
        raise HTTPException(400, "chofer_id requerido.")
    q = sb.table(_TBL_VIAJES).select("*").eq("user_id", uid).eq("chofer_id", chofer_id).gte("fecha_hora_salida", periodo_inicio).lt("fecha_hora_salida", periodo_fin)
    if pid:
        q = q.eq("perfil_id", pid)
    viajes = q.execute().data or []
    viajes = [v for v in viajes if (v.get("liquidacion_status") or "pendiente") in {"pendiente", "error", "borrador"}]
    if not viajes:
        raise HTTPException(404, "No hay viajes pendientes de liquidar para ese chofer/periodo.")

    tq = sb.table(_TBL_TARIFAS).select("*").eq("user_id", uid).eq("activo", True)
    if pid:
        tq = tq.eq("perfil_id", pid)
    tarifas = tq.execute().data or []

    items = []
    subtotal = iva = retencion = total = 0.0
    sin_tarifa = []
    for v in viajes:
        calc = _calcular_tarifa_operativa(v, tarifas)
        if not calc.get("tarifa_id"):
            sin_tarifa.append(v["id"])
            continue
        gastos = sb.table(_TBL_GASTOS).select("importe").eq("user_id", uid).eq("viaje_id", v["id"]).eq("status", "aprobado").execute().data or []
        gastos_total = round(sum(_safe_float(g.get("importe")) for g in gastos), 2)
        item_total = round(calc["total"] + gastos_total, 2)
        subtotal += calc["subtotal"]
        iva += calc["iva"]
        retencion += calc["retencion"]
        total += item_total
        items.append({
            "user_id": uid, "perfil_id": pid, "viaje_id": v["id"],
            "concepto": f"Flete viaje #{v['id']}",
            "litros": calc["litros"], "kilos": calc["kilos"], "tarifa": calc["tarifa"],
            "subtotal": calc["subtotal"], "iva": calc["iva"], "retencion": calc["retencion"],
            "gastos": gastos_total, "total": item_total, "metadata": calc,
        })
    if sin_tarifa:
        raise HTTPException(400, f"Configura tarifa antes de liquidar estos viajes: {sin_tarifa}")
    if not items:
        raise HTTPException(404, "No hay viajes con tarifa configurada para liquidar.")

    now_iso = datetime.now(timezone.utc).isoformat()
    anticipos = _safe_float(payload.get("anticipos"))
    comision_extra = _safe_float(payload.get("comision_extra"))
    descuentos = _safe_float(payload.get("descuentos"))
    liq_row = {
        "user_id": uid, "perfil_id": pid, "chofer_id": chofer_id, "periodo": periodo,
        "periodo_inicio": periodo_inicio, "periodo_fin": periodo_fin,
        "subtotal": round(subtotal, 2), "iva": round(iva, 2), "retencion": round(retencion, 2),
        "gastos": round(sum(i["gastos"] for i in items), 2),
        "anticipos": anticipos,
        "comision_extra": comision_extra,
        "descuentos": descuentos,
        "pago_nomina": _safe_float(payload.get("pago_nomina")),
        "pago_banco": _safe_float(payload.get("pago_banco")),
        "diferencia_efectivo": _safe_float(payload.get("diferencia_efectivo")),
        "total": round(total + comision_extra - anticipos - descuentos, 2),
        "status": str(payload.get("status") or "emitida"),
        "notas": str(payload.get("notas") or ""),
        "metodo_pago": str(payload.get("metodo_pago") or ""),
        "referencia_pago": str(payload.get("referencia_pago") or ""),
        "metadata": {"periodo_inicio": periodo_inicio, "periodo_fin": periodo_fin, "items": len(items)},
        "created_at": now_iso,
    }
    res = sb.table(_TBL_LIQS).insert(liq_row).execute()
    liquidacion_id = res.data[0]["id"] if res.data else None
    for item in items:
        item["liquidacion_id"] = liquidacion_id
    if items:
        sb.table(_TBL_LIQ_ITEMS).insert(items).execute()
        ids = [i["viaje_id"] for i in items]
        sb.table(_TBL_VIAJES).update({"liquidacion_status": "emitida"}).eq("user_id", uid).in_("id", ids).execute()
        for vid in ids:
            _registrar_evento(sb, uid, pid, int(vid), "liquidacion_generada", "Liquidacion generada", f"Liquidacion #{liquidacion_id}", "oficina", uid, {"liquidacion_id": liquidacion_id})
    return JSONResponse({"ok": True, "liquidacion_id": liquidacion_id, "items": len(items), "total": liq_row["total"]})


@router.post("/tr/liquidaciones/{liquidacion_id}/pagar")
async def pagar_liquidacion(liquidacion_id: int, payload: dict, authorization: str = Header(default="")):
    uid, token = _auth(authorization)
    sb = _sb(token)
    now_iso = datetime.now(timezone.utc).isoformat()
    metodo = str(payload.get("metodo_pago") or "").strip() or "efectivo"
    referencia = str(payload.get("referencia_pago") or "").strip()
    sb.table(_TBL_LIQS).update({
        "status": "pagada",
        "paid_at": now_iso,
        "metodo_pago": metodo,
        "referencia_pago": referencia,
        "pago_nomina": _safe_float(payload.get("pago_nomina")),
        "pago_banco": _safe_float(payload.get("pago_banco")),
        "diferencia_efectivo": _safe_float(payload.get("diferencia_efectivo")),
    }).eq("id", liquidacion_id).eq("user_id", uid).execute()
    items = sb.table(_TBL_LIQ_ITEMS).select("viaje_id,perfil_id").eq("liquidacion_id", liquidacion_id).eq("user_id", uid).execute().data or []
    ids = [int(i["viaje_id"]) for i in items if i.get("viaje_id")]
    if ids:
        sb.table(_TBL_VIAJES).update({"liquidacion_status": "pagada"}).eq("user_id", uid).in_("id", ids).execute()
        for item in items:
            _registrar_evento(sb, uid, item.get("perfil_id"), int(item["viaje_id"]), "liquidacion_pagada", "Liquidacion pagada", f"Liquidacion #{liquidacion_id} · {metodo}", "oficina", uid, {"liquidacion_id": liquidacion_id, "metodo_pago": metodo, "referencia_pago": referencia})
    return JSONResponse({"ok": True})


@router.post("/tr/importar/excel-ruth")
async def importar_excel_ruth(
    file: UploadFile = File(...),
    dry_run: bool = Form(True),
    perfil_id: Optional[int] = Form(None),
    authorization: str = Header(default=""),
    x_perfil_id: str = Header(default=""),
):
    """Importador historico no destructivo: extrae resumen y tarifas del Excel operativo."""
    uid, token = _auth(authorization)
    pid = _perfil_autorizado(uid, token, perfil_id, x_perfil_id)
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(await file.read()), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el Excel: {e}")

    resumen: dict = {"sheets": {}, "tarifas_detectadas": 0, "viajes_detectados": 0}
    for s in wb.sheetnames:
        ws = wb[s]
        nonempty = 0
        for row in ws.iter_rows():
            if any(c.value is not None and str(c.value).strip() for c in row):
                nonempty += 1
        resumen["sheets"][s] = {"rows": ws.max_row, "cols": ws.max_column, "nonempty_rows": nonempty}

    tarifas = []
    if "Precio.Tarifas" in wb.sheetnames:
        ws = wb["Precio.Tarifas"]
        for r in range(7, 13):
            origen, destino, producto, tiempos, tarifa = [ws.cell(r, c).value for c in range(2, 7)]
            if origen and destino and tarifa not in (None, ""):
                tarifas.append({
                    "user_id": uid, "perfil_id": pid, "origen": str(origen), "destino": str(destino),
                    "producto": str(producto or ""), "regla_calculo": "litros",
                    "tarifa": _safe_float(tarifa), "metadata": {"tiempos": str(tiempos or ""), "fuente": "Facturas de Ingreso Ruth.xlsx"},
                })
        for r in range(36, 40):
            destino = ws.cell(r, 2).value
            for c in range(3, 9):
                origen = ws.cell(35, c).value
                tarifa = ws.cell(r, c).value
                if origen and destino and tarifa not in (None, ""):
                    tarifas.append({
                        "user_id": uid, "perfil_id": pid, "origen": str(origen), "destino": str(destino),
                        "producto": "Gas LP", "regla_calculo": "kilos",
                        "tarifa": _safe_float(tarifa), "metadata": {"fuente": "Facturas de Ingreso Ruth.xlsx"},
                    })
    resumen["tarifas_detectadas"] = len(tarifas)
    for sheet in ("Gasolina Tabla", "Gas Tabla", "Gas LP", "Gaso Antiguo"):
        if sheet in wb.sheetnames:
            resumen["viajes_detectados"] += max((resumen["sheets"][sheet]["nonempty_rows"] - 1), 0)

    sb = _sb(token)
    inserted = 0
    if not dry_run and tarifas:
        res = sb.table(_TBL_TARIFAS).insert(tarifas).execute()
        inserted = len(res.data or [])
    sb.table(_TBL_IMPORTS).insert({
        "user_id": uid, "perfil_id": pid, "fuente": "excel_ruth",
        "filename": file.filename or "Facturas de Ingreso Ruth.xlsx",
        "resumen": resumen, "status": "preview" if dry_run else "procesada",
    }).execute()
    return JSONResponse({"ok": True, "dry_run": dry_run, "resumen": resumen, "tarifas_insertadas": inserted})


@router.get("/tr/facturas")
async def listar_facturas_transporte(
    periodo:       Optional[str] = Query(None),
    perfil_id:     Optional[int] = Query(None),
    authorization: str           = Header(default=""),
):
    """Lista los CFDIs timbrados del módulo transporte."""
    uid, token = _auth(authorization)
    sb = _sb(token)
    try:
        q = sb.table(_TBL_CFDI).select("id,user_id,perfil_id,viaje_id,tipo_cfdi,uuid_sat,id_ccp,pdf_url,status,fecha_timbrado,rfc_receptor,created_at,updated_at").eq("user_id", uid).order("fecha_timbrado", desc=True)
        if periodo:
            ini, fin = _periodo_bounds(periodo)
            q = q.gte("fecha_timbrado", ini).lt("fecha_timbrado", fin)
        if perfil_id:
            q = q.eq("perfil_id", perfil_id)
        res  = q.execute()
        rows = res.data or []
        return JSONResponse({"ok": True, "facturas": rows})
    except Exception as e:
        raise HTTPException(500, f"Error al listar facturas: {e}")


@router.get("/tr/facturas/{cfdi_id}/xml")
async def descargar_xml_transporte(cfdi_id: int, authorization: str = Header(default="")):
    """Descarga el XML timbrado de un CFDI de transporte."""
    uid, token = _auth(authorization)
    sb = _sb(token)
    try:
        res = sb.table(_TBL_CFDI).select("uuid_sat,xml_content").eq("id", cfdi_id).eq("user_id", uid).limit(1).execute()
        rows = res.data or []
        if not rows:
            raise HTTPException(404, "CFDI no encontrado.")
        row = rows[0]
        audit_fiscal_pdf_event(
            get_supabase_admin(),
            user_id=uid,
            module="transporte",
            entity_type="carta_porte",
            entity_id=cfdi_id,
            uuid_sat=row.get("uuid_sat") or "",
            action="xml_download",
        )
        return Response(
            content=row["xml_content"],
            media_type="application/xml",
            headers={"Content-Disposition": f'attachment; filename="cfdi_tr_{row["uuid_sat"]}.xml"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error al obtener XML: {e}")


@router.get("/tr/facturas/{cfdi_id}/pdf")
async def ver_pdf_carta_porte_transporte(
    cfdi_id: int,
    download: bool = Query(False),
    authorization: str = Header(default=""),
):
    """
    Genera y entrega la representación impresa del CFDI/Carta Porte desde el XML timbrado.
    No depende de que SW Sapien regrese pdfUrl.
    """
    uid, token = _auth(authorization)
    sb = _sb(token)
    try:
        res = (
            sb.table(_TBL_CFDI)
            .select("id,user_id,perfil_id,viaje_id,uuid_sat,id_ccp,xml_content,pdf_url")
            .eq("id", cfdi_id)
            .eq("user_id", uid)
            .limit(1)
            .execute()
        )
        rows = res.data or []
        if not rows:
            raise HTTPException(404, "CFDI no encontrado.")
        row = rows[0]
        xml_content = row.get("xml_content") or ""
        if not xml_content:
            raise HTTPException(404, "Este CFDI no tiene XML timbrado guardado.")

        viaje_rows = []
        productos = []
        if row.get("viaje_id"):
            viaje_rows = sb.table(_TBL_VIAJES).select("id,perfil_id,productos_json").eq("id", row.get("viaje_id")).eq("user_id", uid).limit(1).execute().data or []
            if viaje_rows:
                productos = _productos_from_row(viaje_rows[0])
        validacion = validar_xml_carta_porte_transporte(xml_content, productos)
        if validacion.bloquea_pdf:
            raise HTTPException(409, "PDF bloqueado: XML no válido como Carta Porte de carretera. " + "; ".join(validacion.errors[:5]))
        settings = _settings_transporte(uid, token, row.get("perfil_id"))
        info = extraer_info_pdf(xml_content)
        pdf_bytes = generar_pdf_carta_porte_desde_xml(xml_content, settings.get("PdfLogoDataUrl", ""))
        _guardar_cfdi_pdf_en_expediente(
            sb,
            uid,
            row,
            pdf_bytes,
            info.filename,
            {
                "cfdi_id": cfdi_id,
                "uuid_sat": info.uuid,
                "id_ccp": info.id_ccp,
                "has_carta_porte": info.has_carta_porte,
                "source": "xml_timbrado",
            },
        )
        disposition = "attachment" if download else "inline"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'{disposition}; filename="{info.filename}"'},
        )
    except HTTPException:
        raise
    except RuntimeError as e:
        raise HTTPException(500, str(e))
    except Exception as e:
        logger.exception("Error generando PDF Carta Porte cfdi=%s", cfdi_id)
        raise HTTPException(500, f"Error al generar PDF Carta Porte: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# 5. CONTROLES VOLUMÉTRICOS (covol)
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/tr/covol/generar")
async def generar_covol_transporte(
    payload:       GenerarCovolRequest,
    authorization: str = Header(default=""),
):
    """
    Genera el JSON de Controles Volumétricos mensual para transporte.
    Toma todos los viajes timbrados del periodo y los consolida.
    """
    uid, token = _auth(authorization)
    sb = _sb(token)

    periodo  = f"{payload.anio:04d}-{payload.mes:02d}"
    settings = _settings_transporte(uid, token, payload.perfil_id)

    if not settings.get("RfcContribuyente"):
        raise HTTPException(400, "Configura el RFC del contribuyente en Ajustes del módulo Transporte.")

    # Obtener viajes timbrados del periodo
    try:
        q = (
            sb.table(_TBL_VIAJES)
            .select("*")
            .eq("user_id", uid)
            .eq("status", "timbrado")
            .like("fecha_hora_salida", f"{periodo}%")
        )
        if payload.perfil_id:
            q = q.eq("perfil_id", payload.perfil_id)
        res   = q.execute()
        viajes_raw = res.data or []
    except Exception as e:
        raise HTTPException(500, f"Error al obtener viajes del periodo: {e}")

    if not viajes_raw:
        raise HTTPException(404, f"No hay viajes timbrados en el periodo {periodo}.")

    selected_permiso = (payload.num_permiso_cne or settings.get("NumPermiso", "") or "").strip()
    if not selected_permiso:
        raise HTTPException(
            400,
            "El número de permiso CNE es requerido para generar el JSON mensual de Transporte.",
        )

    def _permiso_viaje(row: dict) -> str:
        return (row.get("num_permiso_cne") or settings.get("NumPermiso", "") or "").strip()

    permisos_detectados = sorted({
        p for p in (_permiso_viaje(v) for v in viajes_raw) if p
    })
    viajes_raw = [v for v in viajes_raw if _permiso_viaje(v) == selected_permiso]
    if not viajes_raw:
        detalle = f" Permisos detectados en el periodo: {', '.join(permisos_detectados)}." if permisos_detectados else ""
        raise HTTPException(
            404,
            f"No hay viajes timbrados del periodo {periodo} para el permiso CNE {selected_permiso}.{detalle}",
        )

    # Convertir viajes_raw a formato esperado por transport_transformer
    viajes_para_covol: list[dict] = []
    for v in viajes_raw:
        try:
            productos_json = json.loads(v.get("productos_json") or "[]")
        except Exception:
            productos_json = []
        viajes_para_covol.append({
            "uuid_cfdi":         v.get("uuid_cfdi", ""),
            "id_ccp":            v.get("id_ccp", ""),
            "num_permiso_cne":    _permiso_viaje(v),
            "tipo_movimiento":   "descarga",   # El autotanque entrega → descarga en destino
            "fecha_hora_salida": v.get("fecha_hora_salida", ""),
            "rfc_receptor":      v.get("rfc_receptor", ""),
            "nombre_receptor":   v.get("nombre_receptor", ""),
            "productos":         productos_json,
        })

    # Preparar settings para el transformer
    covol_settings = {
        **settings,
        "NumPermiso":          selected_permiso,
        "ClaveInstalacion":    payload.clave_instalacion or settings.get("ClaveInstalacion", ""),
        "DescripcionInstalacion": payload.descripcion_instalacion or settings.get("DescripcionInstalacion", ""),
        "ModalidadPermiso":    settings.get("ModalidadPermiso", "PER51"),
    }

    try:
        sat_dict, meta = build_transport_covol(
            viajes=                  viajes_para_covol,
            settings=                covol_settings,
            anio=                    payload.anio,
            mes=                     payload.mes,
            inventario_inicial_litros= payload.inventario_inicial_litros,
        )
        archivos = save_transport_covol(sat_dict, meta, covol_settings)
    except Exception as e:
        logger.error("Error al generar covol transporte: %s", e)
        raise HTTPException(500, f"Error al generar reporte: {e}")

    # Guardar reporte en tr_covol_reports
    now_iso = datetime.now(timezone.utc).isoformat()
    try:
        sb.table(_TBL_COVOL).insert({
            "user_id":        uid,
            "perfil_id":      payload.perfil_id,
            "periodo":        periodo,
            "filename_base":  meta.get("first_uuid", "")[:8],
            "json_name":      archivos["json_name"],
            "zip_name":       archivos["zip_name"],
            "json_content":   archivos["json_content"],
            "zip_b64":        archivos["zip_b64"],
            "total_cargas":   meta.get("total_cargas", 0),
            "total_descargas": meta.get("total_descargas", 0),
            "num_productos":  meta.get("num_productos", 0),
            "created_at":     now_iso,
        }).execute()
    except Exception as e:
        logger.warning("No se pudo guardar covol en BD: %s", e)

    return JSONResponse({
        "ok":           True,
        "periodo":      periodo,
        "json_name":    archivos["json_name"],
        "zip_name":     archivos["zip_name"],
        "json_content": archivos["json_content"],
        "zip_b64":      archivos["zip_b64"],
        "num_permiso_cne": selected_permiso,
        "permisos_detectados": permisos_detectados,
        "meta":         {**meta, "num_permiso_cne": selected_permiso, "permisos_detectados": permisos_detectados},
    })


# ══════════════════════════════════════════════════════════════════════════════
# 6. CATÁLOGOS: Choferes, Vehículos, Rutas, Clientes
# ══════════════════════════════════════════════════════════════════════════════

# ── Choferes ──────────────────────────────────────────────────────────────────
