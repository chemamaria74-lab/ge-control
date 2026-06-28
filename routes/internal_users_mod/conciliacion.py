import re

from .core import *

_CONCILIACION_PUBLICO_RFC = "XAXX010101000"
_CONCILIACION_COMPACT_METADATA_KEYS = {
    "asistente_nombre",
    "cliente_id",
    "cliente_nombre",
    "cliente_real_nombre",
    "cliente_observado",
    "comentarios",
    "created_by",
    "created_by_area",
    "created_by_internal",
    "created_by_internal_name",
    "descuento",
    "descuento_capturado",
    "descuento_confirmado",
    "descuento_por_litro",
    "descuento_preview",
    "destino_facility_name",
    "destino_nombre",
    "empresa_asignada_nombre",
    "empresa_nombre",
    "fecha_cfdi",
    "fecha_emision",
    "folio",
    "folio_usuario",
    "forma_pago",
    "internal_user_id",
    "is_transfer",
    "iva",
    "litros",
    "metodo_pago",
    "observaciones",
    "operation_type",
    "origen_facility_name",
    "origen_nombre",
    "payment_status",
    "portal",
    "precio_unitario",
    "receptor_nombre",
    "saldo_insoluto",
    "subtotal",
    "tipo_descuento",
    "tipo_operacion",
    "total",
    "usuario_nombre",
}


def _conciliacion_clean_text(value) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text[:120]


def _conciliacion_public_name_key(value) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", str(value or "").upper()).strip()


def _conciliacion_extract_observed_client(text: str) -> tuple[str, str]:
    raw = _conciliacion_clean_text(text)
    if not raw:
        return "", ""
    patterns = (
        r"(?:cliente|clienta|nombre|para|referencia|observaci[oó]n|comentario)\s*[:=\-]\s*([^|;\n\r]{3,80})",
        r"(?:cliente|clienta)\s+([^|;\n\r]{3,80})",
    )
    for pattern in patterns:
        match = re.search(pattern, raw, flags=re.IGNORECASE)
        if match:
            candidate = _conciliacion_clean_text(match.group(1))
            if _conciliacion_public_name_key(candidate) not in {"", "PUBLICO EN GENERAL", "PUBLICO GENERAL"}:
                return candidate, "observaciones"
    if _conciliacion_public_name_key(raw) not in {"PUBLICO EN GENERAL", "PUBLICO GENERAL"} and len(raw.split()) >= 2:
        return raw, "observaciones"
    return "", ""


def _conciliacion_real_cliente(row: dict, metadata: dict) -> dict:
    is_public = str(row.get("rfc_receptor") or metadata.get("cliente_rfc") or "").upper() == _CONCILIACION_PUBLICO_RFC
    observed_fields = (
        "cliente_real_nombre",
        "cliente_observado",
        "nombre_cliente_real",
        "cliente_referencia",
        "nota",
        "referencia",
    )
    observed = ""
    source = ""
    if is_public:
        for key in observed_fields:
            candidate = _conciliacion_clean_text(metadata.get(key))
            if candidate and _conciliacion_public_name_key(candidate) not in {"PUBLICO EN GENERAL", "PUBLICO GENERAL"}:
                observed = candidate
                source = key
                break
        if not observed:
            for key in ("comentarios", "observaciones", "nota", "referencia"):
                observed, source = _conciliacion_extract_observed_client(metadata.get(key) or "")
                if observed:
                    break
    base = _conciliacion_clean_text(metadata.get("cliente_nombre") or row.get("nombre_receptor") or row.get("rfc_receptor") or "—")
    return {
        "nombre": observed or base,
        "fuente": source or ("metadata" if base else ""),
        "es_publico_general": is_public,
        "observaciones": _conciliacion_clean_text(metadata.get("observaciones") or metadata.get("comentarios")),
    }


def _conciliacion_discount_info(row: dict, metadata: dict, info: dict) -> dict:
    tipo = str(metadata.get("tipo_descuento") or "").strip()
    litros = float(_money(info.get("litros") or row.get("volumen_litros") or metadata.get("litros") or 0))
    por_litro = float(_money(metadata.get("descuento_por_litro") or 0))
    candidates = []
    for key in ("descuento_confirmado", "descuento_preview", "descuento", "descuento_capturado"):
        value = metadata.get(key)
        if isinstance(value, dict):
            value = value.get("total") or value.get("monto") or value.get("descuento_total_aplicado")
        candidates.append(float(_money(value or 0)))
    if por_litro > 0 and litros > 0:
        candidates.append(round(por_litro * litros, 2))
    total_desc = round(max(candidates or [0.0]), 2)
    return {
        "tipo": tipo or ("por_litro" if por_litro > 0 else ""),
        "capturado": float(_money(metadata.get("descuento_capturado") or 0)),
        "por_litro": round(por_litro, 6),
        "total": total_desc,
        "tiene_descuento": total_desc > 0,
    }


def _conciliacion_compact_metadata(metadata: dict) -> dict:
    compact = {}
    for key in _CONCILIACION_COMPACT_METADATA_KEYS:
        if key in metadata:
            compact[key] = metadata.get(key)
    return compact


def _conciliacion_compact_cliente(cliente: dict) -> dict:
    credit = clienteCredit = cliente.get("credito_ppd") if isinstance(cliente.get("credito_ppd"), dict) else {}
    metadata = cliente.get("metadata") if isinstance(cliente.get("metadata"), dict) else {}
    if not credit:
        raw_credit = metadata.get("credito_ppd") or metadata.get("credito") or {}
        credit = raw_credit if isinstance(raw_credit, dict) else {}
    return {
        "id": cliente.get("id"),
        "rfc": cliente.get("rfc"),
        "nombre": cliente.get("nombre"),
        "credito_habilitado": cliente.get("credito_habilitado") or credit.get("credito_habilitado") or credit.get("habilitado") or False,
        "dias_credito": cliente.get("dias_credito") or credit.get("dias_credito") or credit.get("dias") or 0,
        "limite_credito": cliente.get("limite_credito") or credit.get("limite_credito") or credit.get("limite"),
        "credito_notas": cliente.get("credito_notas") or credit.get("credito_notas") or credit.get("notas") or "",
        "metadata": {"credito_ppd": clienteCredit or credit},
    }


def _conciliacion_compact_factura(row: dict) -> dict:
    keys = (
        "id", "user_id", "tenant_id", "perfil_id", "facility_id", "record_uuid", "uuid_sat", "status",
        "fecha_timbrado", "rfc_receptor", "tipo_comprobante", "created_at", "updated_at",
        "email_enviado", "email_destinatario", "email_error", "fecha_factura_key", "realizado_por",
    )
    compact = {key: row.get(key) for key in keys if key in row}
    compact["volumen_litros"] = float(_money(row.get("volumen_litros") or 0))
    compact["importe"] = float(_money(row.get("importe") or 0))
    compact["metadata"] = _conciliacion_compact_metadata(row.get("metadata") if isinstance(row.get("metadata"), dict) else {})
    for key in (
        "payment_info",
        "fiscal_status",
        "bank_reconciliation",
        "issuer_info",
        "complementos_pago",
        "latest_complemento_pago",
        "cliente_real",
        "cliente_display",
        "observaciones",
        "discount_info",
    ):
        if key in row:
            compact[key] = row.get(key)
    return compact

@router.get("/internal-auth/gas-lp/conciliacion/perfiles")
async def gas_lp_conciliacion_perfiles(token: str):
    if str(token or "").count(".") == 2:
        uid = verify_token(token)
        if not uid:
            raise HTTPException(401, "Sesión inválida o expirada.")
        access = _resolve_active_module_access(uid, "gas_lp", access_token=token)
        role = (access.get("role") or "").lower()
        if role not in {"admin", "conciliacion", "asistente_facturacion"}:
            raise HTTPException(403, "Tu usuario no tiene acceso a conciliación Gas LP.")
        perfiles = _gas_lp_conciliacion_visible_profiles(uid, access, token)
        perfil_id = None if role == "admin" else access.get("perfil_id")
        return JSONResponse({"ok": True, "perfil_id": perfil_id, "perfiles": perfiles})

    ctx = _gas_lp_conciliacion_context(token)
    user = ctx["user"]
    profile = _gas_lp_profile(user, require_module_marker=True)
    perfiles = [{"id": profile.get("id"), "nombre": profile.get("nombre"), "rfc": profile.get("rfc"), "descripcion": ""}]
    return JSONResponse({"ok": True, "perfil_id": user.get("perfil_id"), "perfiles": perfiles})


@router.get("/internal-auth/gas-lp/conciliacion/facilities")
async def gas_lp_conciliacion_facilities(token: str, perfil_id: int | None = None):
    ctx = _gas_lp_conciliacion_context(token, perfil_id=perfil_id)
    user = ctx["user"]
    rows = _gas_lp_admin_facilities(user)
    settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
    precio_venta_litro, precio_venta_litro_configurado = _configured_setting(
        settings,
        ("precio_venta_litro", "PrecioVentaLitro", "precio_default_litro", "precio_litro"),
    )
    return JSONResponse({
        "ok": True,
        "facilities": rows,
        "precio_venta_litro": precio_venta_litro,
        "precio_venta_litro_configurado": precio_venta_litro_configurado,
    })


@router.get("/internal-auth/gas-lp/conciliacion/summary")
async def gas_lp_conciliacion_summary(token: str, periodo: str | None = None, perfil_id: int | None = None):
    ctx = _gas_lp_conciliacion_context(token, perfil_id=perfil_id)
    user = ctx["user"]
    profile = _gas_lp_profile(user, require_module_marker=True)
    month = (periodo or datetime.now().strftime("%Y-%m"))[:7]
    sb = get_supabase_admin()
    try:
        rows = _gas_lp_company_facturas_rows(
            sb,
            user,
            profile,
            month=month,
            limit=10000,
            include_carta_porte=False,
            select="*",
            company_fallback=True,
            visibility_log=False,
        )
    except Exception as exc:
        raise _safe_internal_error("gas_lp_conciliacion_summary", exc)
    _gas_lp_attach_internal_creators(sb, rows)
    comp_by_factura = _gas_lp_complementos_por_factura(sb, [_safe_int_id(r.get("id")) for r in rows if _safe_int_id(r.get("id"))])
    try:
        clientes_query = (
            sb.table("gas_lp_clientes_facturacion")
            .select("*")
            .eq("user_id", user.get("owner_user_id"))
            .eq("perfil_id", user.get("perfil_id"))
            .eq("activo", True)
        )
        if user.get("tenant_id"):
            clientes_query = clientes_query.eq("tenant_id", user.get("tenant_id"))
        else:
            clientes_query = clientes_query.is_("tenant_id", "null")
        clientes = clientes_query.order("nombre", desc=False).execute().data or []
        clientes = [_normalize_gas_lp_cliente_credit(row) for row in clientes]
    except Exception as exc:
        logger.warning(
            "gas_lp_conciliacion_clientes_credito_skipped tenant=%s perfil=%s err=%s",
            user.get("tenant_id"),
            user.get("perfil_id"),
            exc,
        )
        clientes = []
    factura_ids = [_safe_int_id(r.get("id")) for r in rows if _safe_int_id(r.get("id"))]
    try:
        bank_rows = (
            sb.table("gas_lp_invoice_bank_reconciliations")
            .select("id,factura_id,amount,difference,status,payment_detected_at,confirmed_by,confirmed_by_name,confirmed_at,reference_note,comment,updated_at")
            .in_("factura_id", factura_ids)
            .execute()
            .data
            or []
        ) if factura_ids else []
    except Exception as exc:
        logger.warning("gas_lp_conciliacion_bank_reconciliations_skipped perfil=%s err=%s", user.get("perfil_id"), exc)
        bank_rows = []
    bank_by_factura = {_safe_int_id(row.get("factura_id")): row for row in bank_rows}
    clientes_compactos = [_conciliacion_compact_cliente(row) for row in clientes]
    clientes_by_id = {_safe_int_id(c.get("id")): c for c in clientes_compactos if _safe_int_id(c.get("id"))}
    clientes_by_rfc = {str(c.get("rfc") or "").upper(): c for c in clientes_compactos if c.get("rfc")}
    total = credito = publico = complementos_pendientes = 0.0
    descuentos_periodo = facturas_con_descuento = publico_observado = 0
    credito_pagado = saldo_vencido = 0.0
    facturas_vencidas = facturas_vigentes = 0
    clientes_con_saldo = set()
    today_key = datetime.now().strftime("%Y-%m-%d")
    for row in rows:
        md = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
        info = _factura_payment_info(row)
        row["fecha_factura_key"] = _gas_lp_factura_date_key(row)
        row["payment_info"] = _payment_info_json(info)
        row["fiscal_status"] = _gas_lp_factura_fiscal_status_info(row)
        bank_row = bank_by_factura.get(_safe_int_id(row.get("id"))) or {}
        row["bank_reconciliation"] = {
            "id": bank_row.get("id"),
            "factura_id": bank_row.get("factura_id") or row.get("id"),
            "amount": float(_money(bank_row.get("amount") or 0)),
            "difference": float(_money(bank_row.get("difference") or 0)),
            "status": str(bank_row.get("status") or "pendiente"),
            "payment_detected_at": bank_row.get("payment_detected_at") or "",
            "confirmed_by": bank_row.get("confirmed_by") or "",
            "confirmed_by_name": bank_row.get("confirmed_by_name") or "",
            "confirmed_at": bank_row.get("confirmed_at") or "",
            "reference_note": bank_row.get("reference_note") or "",
            "comment": bank_row.get("comment") or "",
            "updated_at": bank_row.get("updated_at") or "",
        }
        row["issuer_info"] = {
            "rfc": _gas_lp_factura_emisor_rfc(row),
            "nombre": _gas_lp_factura_emisor_nombre(row),
        }
        row["realizado_por"] = _gas_lp_factura_realizado_por(row)
        comps = comp_by_factura.get(_safe_int_id(row.get("id")), [])
        row["complementos_pago"] = comps
        if comps:
            row["latest_complemento_pago"] = comps[0]
        cliente_real = _conciliacion_real_cliente(row, md)
        row["cliente_real"] = cliente_real
        row["cliente_display"] = cliente_real["nombre"]
        row["observaciones"] = cliente_real.get("observaciones") or ""
        discount_info = _conciliacion_discount_info(row, md, info)
        row["discount_info"] = discount_info
        if str(row.get("status") or "").lower().startswith("cancel"):
            continue
        amount = float(info["total"])
        total += amount
        if discount_info["tiene_descuento"]:
            descuentos_periodo += discount_info["total"]
            facturas_con_descuento += 1
        if cliente_real["es_publico_general"]:
            publico += amount
            if cliente_real.get("fuente") and cliente_real.get("fuente") != "metadata":
                publico_observado += 1
        if info["metodo_pago"] == "PPD" or str(md.get("metodo_pago") or "").upper() == "PPD":
            saldo = float(info["saldo_insoluto"])
            if saldo > 0:
                credito += saldo
                complementos_pendientes += 1
                cliente = clientes_by_id.get(_safe_int_id(md.get("cliente_id"))) or clientes_by_rfc.get(str(row.get("rfc_receptor") or "").upper()) or {}
                cliente_key = str(cliente.get("id") or row.get("rfc_receptor") or cliente_real.get("nombre") or row.get("id"))
                clientes_con_saldo.add(cliente_key)
                dias_credito = int(float(cliente.get("dias_credito") or 0))
                vencimiento = ""
                if dias_credito > 0:
                    emision_key = row.get("fecha_factura_key") or ""
                    try:
                        venc = datetime.strptime(emision_key[:10], "%Y-%m-%d")
                        vencimiento = (venc + timedelta(days=dias_credito)).strftime("%Y-%m-%d")
                    except Exception:
                        vencimiento = ""
                if vencimiento and vencimiento < today_key:
                    facturas_vencidas += 1
                    saldo_vencido += saldo
                else:
                    facturas_vigentes += 1
            credito_pagado += max(0.0, amount - saldo)
    facturas_compactas = [_conciliacion_compact_factura(row) for row in rows]
    return JSONResponse({
        "ok": True,
        "periodo": month,
        "company": {"id": profile.get("id"), "name": profile.get("nombre"), "rfc": profile.get("rfc")},
        "kpis": {
            "facturas": len(rows),
            "total_facturado": round(total, 2),
            "credito_estimado": round(credito, 2),
            "publico_general": round(publico, 2),
            "complementos_pendientes": int(complementos_pendientes),
            "credito_pendiente": round(credito, 2),
            "credito_pagado": round(credito_pagado, 2),
            "clientes_con_saldo": len(clientes_con_saldo),
            "facturas_ppd_pendientes": int(complementos_pendientes),
            "facturas_vencidas": int(facturas_vencidas),
            "facturas_vigentes": int(facturas_vigentes),
            "saldo_vencido": round(saldo_vencido, 2),
            "descuentos_periodo": round(descuentos_periodo, 2),
            "facturas_con_descuento": int(facturas_con_descuento),
            "publico_general_con_cliente_observado": int(publico_observado),
        },
        "credito_summary": {
            "pendiente": round(credito, 2),
            "pagado": round(credito_pagado, 2),
            "clientes_con_saldo": len(clientes_con_saldo),
            "facturas_vencidas": int(facturas_vencidas),
            "facturas_vigentes": int(facturas_vigentes),
            "saldo_vencido": round(saldo_vencido, 2),
        },
        "descuentos_summary": {
            "total": round(descuentos_periodo, 2),
            "facturas": int(facturas_con_descuento),
            "publico_general_observado": int(publico_observado),
        },
        "facturas": facturas_compactas,
        "clientes": clientes_compactos,
    })


@router.post("/internal-auth/gas-lp/conciliacion/facturar-publico-general")
async def gas_lp_conciliacion_facturar_publico_general(payload: GasLpConciliacionPublicoGeneralPayload, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_conciliacion_context(token, write=True, perfil_id=perfil_id)
    user = ctx["user"]
    profile = _gas_lp_profile(user, require_module_marker=True)
    settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
    issuer = _require_gas_lp_issuer(profile, settings)
    receptor = _public_general_receptor(issuer["cp"])
    sb = get_supabase_admin()
    serie_factura = _gas_lp_internal_series(user, settings)
    folio_factura = _gas_lp_next_invoice_folio(sb, user, serie_factura)
    facilities_by_id = {
        int(f["id"]): f
        for f in _gas_lp_admin_facilities(user)
        if f.get("id") is not None
    }
    origen = facilities_by_id.get(int(payload.facility_id or 0), {})
    if not origen:
        raise HTTPException(400, "Selecciona la instalación origen para timbrar Público en General.")
    hyp_mode = _gas_lp_hyp_mode()
    hyp = {}
    if hyp_mode == "required":
        hyp = _gas_lp_hyp_from_facility(origen, GAS_LP_CLAVE_PROD_SERV)
    informacion_global = None
    if payload.factura_global:
        informacion_global = {
            "periodicidad": payload.informacion_global_periodicidad,
            "meses": payload.informacion_global_meses,
            "anio": payload.informacion_global_anio,
        }
    concepto_cfdi = "LITRO DE GAS LP" if hyp_mode == "disabled" else "Gas licuado de petróleo"
    xml, totals = _build_gas_lp_consumo_xml(
        issuer=issuer,
        receptor=receptor,
        litros=payload.litros,
        precio_unitario=payload.precio_unitario,
        concepto=concepto_cfdi,
        forma_pago=payload.forma_pago,
        metodo_pago=payload.metodo_pago,
        descuento=payload.descuento,
        descuento_total_base=payload.descuento_capturado if str(payload.tipo_descuento or "").strip().lower() == "total_pesos" else None,
        iva_rate=payload.iva_rate,
        serie=serie_factura,
        folio=folio_factura,
        comentarios=payload.comentarios,
        fecha=payload.fecha,
        clave_prod_serv=GAS_LP_CLAVE_PROD_SERV,
        no_identificacion="GLP-LTR",
        unidad="Litro",
        hyp=hyp,
        informacion_global=informacion_global,
    )
    _gas_lp_validate_invoice_preview_totals(
        payload,
        totals,
        context="conciliacion_publico_general",
        cliente_tipo="publico_general",
        cliente=receptor.get("nombre") or "",
        rfc=receptor.get("rfc") or "",
        instalacion=origen.get("nombre") or origen.get("clave_instalacion") or str(payload.facility_id),
    )
    sw_config = sw_runtime_config()
    if _sw_config_looks_like_sandbox(sw_config):
        raise HTTPException(400, "Este emisor está configurado en modo pruebas. Cambia a producción antes de timbrar CFDI real.")
    resultado = timbrar_cfdi(xml)
    if resultado.get("error"):
        raise HTTPException(400, f"PAC rechazó la factura: {resultado['error']}")
    now = _now_iso()
    created_by_name = str(user.get("display_name") or "").strip() or "Conciliación"
    row = {
        **_gas_lp_invoice_scope(user, profile),
        "facility_id": payload.facility_id,
        "record_uuid": totals["folio"],
        "uuid_sat": resultado.get("uuid") or "",
        "xml_content": resultado.get("xml_timbrado") or xml,
        "pdf_url": resultado.get("pdf_url") or "",
        "status": "Vigente",
        "fecha_timbrado": now,
        "rfc_receptor": receptor["rfc"],
        "volumen_litros": float(payload.litros),
        "importe": totals["subtotal"],
        "tipo_comprobante": "I",
        "distancia_km": 1,
        "metadata": {
            "portal": "conciliacion_gas_lp",
            "created_by_area": "conciliacion",
            "internal_user_id": user.get("id"),
            "created_by_internal_name": created_by_name,
            "created_by": created_by_name,
            "empresa_asignada_id": user.get("perfil_id"),
            "empresa_asignada_nombre": profile.get("nombre") or "",
            "empresa_rfc": profile.get("rfc") or "",
            "rfc_emisor": issuer.get("rfc") or "",
            "cliente_id": None,
            "cliente_nombre": receptor["nombre"],
            "cliente_email": "",
            "concepto": "LITRO DE GAS LP",
            "precio_unitario": payload.precio_unitario,
            "descuento_por_litro": payload.descuento,
            "tipo_descuento": payload.tipo_descuento,
            "descuento_capturado": payload.descuento_capturado,
            "subtotal_confirmado": payload.subtotal_preview,
            "descuento_confirmado": payload.descuento_preview,
            "iva_confirmado": payload.iva_preview,
            "total_confirmado": payload.total_preview,
            "precio_confirmado": payload.precio_unitario,
            "litros_confirmados": payload.litros,
            "tipo_descuento_confirmado": payload.tipo_descuento,
            "descuento_preview": payload.descuento_preview,
            "total_preview": payload.total_preview,
            "descuento": totals["descuento"],
            "iva_rate": payload.iva_rate,
            "serie": serie_factura,
            "folio_usuario": folio_factura,
            "comentarios": payload.comentarios,
            "fecha_emision": totals["fecha"],
            "clave_prod_serv": GAS_LP_CLAVE_PROD_SERV,
            "gas_lp_hyp_mode": hyp_mode,
            "hidrocarburos_petroliferos": hyp,
            "no_identificacion": "GLP-LTR",
            "unidad": "Litro",
            "metodo_pago": payload.metodo_pago,
            "forma_pago": payload.forma_pago,
            "tipo_operacion": "venta_publico_general",
            "facility_id": payload.facility_id,
            "origen_nombre": origen.get("nombre") or "",
            "payment_status": "pendiente_complemento" if payload.metodo_pago.upper() == "PPD" else "pagado_pue",
            "saldo_insoluto": totals["total"] if payload.metodo_pago.upper() == "PPD" else 0,
            "iva": totals["iva"],
            "total": totals["total"],
        },
        "created_at": now,
    }
    try:
        factura = (sb.table("gas_lp_facturas").insert(row).execute().data or [row])[0]
    except Exception as exc:
        raise _safe_internal_error("gas_lp_conciliacion_facturar_publico_general", exc)
    return JSONResponse({"ok": True, "factura": factura, "totals": totals})


@router.get("/internal-auth/gas-lp/conciliacion/export-excel")
async def gas_lp_conciliacion_export_excel(
    token: str,
    period: str | None = None,
    periodo: str | None = None,
    fecha: str | None = None,
    tipo: str | None = None,
    profile_id: int | None = None,
    perfil_id: int | None = None,
):
    selected_perfil_id = perfil_id if perfil_id is not None else profile_id
    ctx = _gas_lp_conciliacion_context(token, perfil_id=selected_perfil_id)
    user = ctx["user"]
    day = str(fecha or "").strip()[:10]
    month = str(periodo or period or "").strip()[:7]
    if day:
        try:
            datetime.strptime(day, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(400, "Selecciona una fecha válida para exportar.")
        month = day[:7]
    elif len(month) == 7 and month[4] == "-":
        try:
            datetime.strptime(f"{month}-01", "%Y-%m-%d")
        except ValueError:
            month = datetime.now().strftime("%Y-%m")
    else:
        month = datetime.now().strftime("%Y-%m")

    start = datetime.strptime(day or f"{month}-01", "%Y-%m-%d")
    if day:
        end = start + timedelta(days=1)
    elif start.month == 12:
        end = start.replace(year=start.year + 1, month=1)
    else:
        end = start.replace(month=start.month + 1)

    sb = get_supabase_admin()
    try:
        profile = _gas_lp_profile(user, require_module_marker=True)
        rows = _gas_lp_company_facturas_rows(sb, user, profile, month=month, limit=10000, include_carta_porte=False)
    except Exception as exc:
        logger.error(
            "conciliacion_export_excel_error profile_id=%s period=%s factura_id=%s exception=%s traceback=%s",
            selected_perfil_id or user.get("perfil_id"),
            month,
            None,
            exc,
            traceback.format_exc(),
        )
        raise _safe_internal_error("gas_lp_conciliacion_export_excel", exc)
    try:
        _gas_lp_attach_internal_creators(sb, rows)
    except Exception as exc:
        logger.error(
            "conciliacion_export_excel_error profile_id=%s period=%s factura_id=%s exception=%s traceback=%s",
            selected_perfil_id or user.get("perfil_id"),
            month,
            None,
            exc,
            traceback.format_exc(),
        )

    if day:
        rows = [row for row in rows if _gas_lp_factura_date_key(row) == day]
    else:
        rows = [row for row in rows if _gas_lp_factura_date_key(row).startswith(month)]
    tipo_filter = str(tipo or "").strip().lower()
    if tipo_filter == "factura":
        rows = [row for row in rows if not ((row.get("metadata") or {}).get("tipo_operacion") == "traspaso" or (row.get("metadata") or {}).get("is_transfer"))]
    elif tipo_filter == "traspaso":
        rows = [row for row in rows if (row.get("metadata") or {}).get("tipo_operacion") == "traspaso" or (row.get("metadata") or {}).get("is_transfer")]
    elif tipo_filter == "complemento":
        rows = []

    complementos = []
    if tipo_filter in {"", "complemento"}:
        try:
            comp_q = (
                sb.table("gas_lp_complementos_pago")
                .select("*")
                .eq("tenant_id", user.get("tenant_id"))
                .eq("perfil_id", user.get("perfil_id"))
                .gte("created_at", start.strftime("%Y-%m-%dT00:00:00"))
                .lt("created_at", end.strftime("%Y-%m-%dT00:00:00"))
                .order("created_at", desc=True)
                .limit(10000)
            )
            complementos = comp_q.execute().data or []
            _gas_lp_attach_complemento_creators(sb, complementos)
        except Exception as exc:
            logger.warning("conciliacion_export_complementos_failed profile_id=%s period=%s err=%s", user.get("perfil_id"), month, exc)
            complementos = []
    comp_ids = [_safe_int_id(c.get("id")) for c in complementos if _safe_int_id(c.get("id"))]
    comp_rels = []
    if comp_ids:
        try:
            comp_rels = (
                sb.table("gas_lp_complementos_pago_facturas")
                .select("*")
                .in_("complemento_id", comp_ids)
                .execute()
                .data
                or []
            )
        except Exception:
            comp_rels = []
    comp_rels_by_id: dict[int, list[dict]] = {}
    comp_factura_ids: list[int] = []
    for rel in comp_rels:
        comp_id = _safe_int_id(rel.get("complemento_id"))
        comp_rels_by_id.setdefault(comp_id, []).append(rel)
        fid = _safe_int_id(rel.get("factura_id"))
        if fid:
            comp_factura_ids.append(fid)
    for comp in complementos:
        comp_factura_ids.extend(_gas_lp_complemento_factura_ids(comp))
    comp_facturas_by_id = _gas_lp_facturas_by_ids_for_company(sb, user, profile, list(dict.fromkeys(comp_factura_ids)))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Documentos"
    headers = [
        "Fecha",
        "Folio de fact",
        "UUID",
        "Cliente",
        "Monto",
        "Litros",
        "Método",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7A1E2C")

    def _excel_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return str(value)

    def _excel_number(value) -> float:
        try:
            return float(_money(value))
        except Exception:
            return 0.0

    def _excel_liters(value) -> float:
        try:
            return float(Decimal(str(value or 0)).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))
        except Exception:
            return 0.0

    def _safe_total(row: dict):
        try:
            return _factura_payment_info(row).get("total")
        except Exception:
            try:
                return _gas_lp_factura_total_con_iva(row)
            except Exception:
                return 0

    def _safe_payment_info(row: dict) -> dict:
        try:
            return _factura_payment_info(row)
        except Exception:
            return {}

    def _safe_metodo_pago(row: dict) -> str:
        try:
            metodo = _gas_lp_factura_metodo_pago(row)
        except Exception:
            md = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
            metodo = md.get("metodo_pago") or ""
        metodo = str(metodo or "").upper()
        return metodo if metodo in {"PUE", "PPD"} else "PUE"

    for row in rows:
        factura_id = row.get("id")
        try:
            info = _safe_payment_info(row)
            ws.append([
                _excel_text(_gas_lp_factura_date_key(row)),
                _excel_text(_gas_lp_factura_folio_label(row)),
                _excel_text(row.get("uuid_sat") or ""),
                _excel_text(_gas_lp_factura_razon_social(row)),
                _excel_number(_safe_total(row)),
                _excel_liters(info.get("litros") or row.get("volumen_litros")),
                _excel_text(info.get("metodo_pago") or _safe_metodo_pago(row)),
            ])
        except Exception as exc:
            logger.error(
                "conciliacion_export_excel_error profile_id=%s period=%s factura_id=%s exception=%s traceback=%s",
                selected_perfil_id or user.get("perfil_id"),
                month,
                factura_id,
                exc,
                traceback.format_exc(),
            )
            ws.append(["", _excel_text(factura_id), "", "", 0.0, 0.0, ""])
    for comp in complementos:
        comp_id = _safe_int_id(comp.get("id"))
        rels = comp_rels_by_id.get(comp_id, [])
        ids = [*_gas_lp_complemento_factura_ids(comp), *[_safe_int_id(rel.get("factura_id")) for rel in rels]]
        ids = list(dict.fromkeys(fid for fid in ids if fid))
        facturas_comp = [comp_facturas_by_id[fid] for fid in ids if fid in comp_facturas_by_id]
        receptor = _gas_lp_complemento_receptor_info(comp, facturas_comp)
        refs = []
        for rel in rels:
            fid = _safe_int_id(rel.get("factura_id"))
            factura = comp_facturas_by_id.get(fid, {})
            refs.append(_gas_lp_factura_folio_label(factura) or rel.get("uuid_relacionado") or "")
        if not refs:
            refs = [_gas_lp_factura_folio_label(comp_facturas_by_id.get(fid, {})) for fid in ids]
        email_error = str(comp.get("email_error") or "")
        ws.append([
            _excel_text(str(comp.get("created_at") or "")[:10]),
            "",
            _excel_text(comp.get("uuid_sat") or ""),
            _excel_text(receptor.get("nombre") or "Cliente"),
            _excel_number(comp.get("monto") or 0),
            0.0,
            "Complemento",
        ])
    for width, column in zip([14, 20, 40, 34, 16, 14, 12], "ABCDEFG"):
        ws.column_dimensions[column].width = width
    for column in ("E",):
        for cell in ws[column][1:]:
            cell.number_format = '$#,##0.00'
    for cell in ws["F"][1:]:
        cell.number_format = "#,##0.0000"

    stream = BytesIO()
    try:
        wb.save(stream)
    except Exception as exc:
        logger.error(
            "conciliacion_export_excel_error profile_id=%s period=%s factura_id=%s exception=%s traceback=%s",
            selected_perfil_id or user.get("perfil_id"),
            month,
            None,
            exc,
            traceback.format_exc(),
        )
        raise _safe_internal_error("gas_lp_conciliacion_export_excel", exc)
    stream.seek(0)
    suffix = day or month
    filename = f"conciliacion_gas_lp_{suffix}.xlsx"
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
