from .core import *
from .catalogos_clientes import _gas_lp_clientes_scope_query
from services.carta_porte_pdf import (
    es_carta_porte_traslado,
    extraer_info_pdf as carta_porte_pdf_info,
    generar_pdf_carta_porte_desde_xml,
)


def _gas_lp_credit_reminder_parse_days(value: str | int | None) -> list[int]:
    raw = str(value if value not in {None, ""} else "2,1")
    days: list[int] = []
    for part in raw.split(","):
        text = part.strip()
        if not text:
            continue
        try:
            day = int(text)
        except ValueError:
            raise HTTPException(400, "El parámetro dias debe contener enteros separados por coma.")
        if day < 0 or day > 365:
            raise HTTPException(400, "Los días de recordatorio deben estar entre 0 y 365.")
        if day not in days:
            days.append(day)
    return days or [2, 1]


def _gas_lp_credit_reminder_today_key(today_key: str | None = None) -> str:
    if today_key:
        try:
            datetime.strptime(str(today_key)[:10], "%Y-%m-%d")
            return str(today_key)[:10]
        except ValueError:
            raise HTTPException(400, "Fecha de referencia inválida.")
    return datetime.now(_gas_lp_cfdi_timezone()).strftime("%Y-%m-%d")


def _gas_lp_credit_reminder_add_days(key: str, days: int) -> str:
    try:
        base = datetime.strptime(str(key or "")[:10], "%Y-%m-%d")
    except ValueError:
        return ""
    return (base + timedelta(days=int(days or 0))).strftime("%Y-%m-%d")


def _gas_lp_credit_reminder_day_diff(target_key: str, today_key: str) -> int | None:
    try:
        target = datetime.strptime(str(target_key or "")[:10], "%Y-%m-%d")
        today = datetime.strptime(str(today_key or "")[:10], "%Y-%m-%d")
    except ValueError:
        return None
    return (target.date() - today.date()).days


def _gas_lp_credit_reminder_client_maps(clientes: list[dict]) -> tuple[dict[int, dict], dict[str, dict]]:
    by_id: dict[int, dict] = {}
    by_rfc: dict[str, dict] = {}
    for raw in clientes or []:
        cliente = _normalize_gas_lp_cliente_credit(raw)
        cid = _safe_int_id(cliente.get("id"))
        if cid:
            by_id[cid] = cliente
        rfc = str(cliente.get("rfc") or "").strip().upper()
        if rfc and rfc not in by_rfc:
            by_rfc[rfc] = cliente
    return by_id, by_rfc


def _gas_lp_credit_reminder_cliente_for_factura(factura: dict, clientes_by_id: dict[int, dict], clientes_by_rfc: dict[str, dict]) -> dict | None:
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    cliente_id = _safe_int_id(md.get("cliente_id"))
    if cliente_id and cliente_id in clientes_by_id:
        return clientes_by_id[cliente_id]
    rfc = str(factura.get("rfc_receptor") or "").strip().upper()
    return clientes_by_rfc.get(rfc)


def _gas_lp_credit_reminder_exclusion(factura: dict, reason: str, details: str = "") -> dict:
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    return {
        "factura_id": _safe_int_id(factura.get("id")),
        "uuid": factura.get("uuid_sat") or "",
        "folio": _gas_lp_factura_folio_label(factura),
        "cliente_id": _safe_int_id(md.get("cliente_id")),
        "cliente_nombre": md.get("cliente_nombre") or factura.get("rfc_receptor") or "",
        "rfc_receptor": factura.get("rfc_receptor") or "",
        "razon_exclusion": reason,
        "detalle": details,
    }


def _gas_lp_credit_reminder_evaluate_factura(factura: dict, cliente: dict | None, target_days: list[int], today_key: str) -> tuple[dict | None, dict | None]:
    md = factura.get("metadata") if isinstance(factura.get("metadata"), dict) else {}
    if _gas_lp_factura_cancelada(factura):
        return None, _gas_lp_credit_reminder_exclusion(factura, "cancelada")
    if md.get("tipo_operacion") == "traspaso" or md.get("is_transfer") is True or md.get("operation_type") == "transfer":
        return None, _gas_lp_credit_reminder_exclusion(factura, "traspaso")

    info = _factura_payment_info(factura)
    metodo_pago = str(info.get("metodo_pago") or md.get("metodo_pago") or "").upper()
    if metodo_pago != "PPD":
        return None, _gas_lp_credit_reminder_exclusion(factura, "no_ppd")

    payment_status = str(info.get("payment_status") or md.get("payment_status") or "").lower()
    if payment_status in {"pagado_con_complemento", "pagado_manual", "pagado_pue"}:
        return None, _gas_lp_credit_reminder_exclusion(factura, "pagada")

    saldo = _money(info.get("saldo_insoluto"))
    if saldo <= 0:
        return None, _gas_lp_credit_reminder_exclusion(factura, "saldo_cero")

    if not cliente:
        return None, _gas_lp_credit_reminder_exclusion(factura, "cliente_sin_credito", "No se encontró cliente por cliente_id ni RFC.")

    credit = _normalize_gas_lp_cliente_credit(cliente)
    if not bool(credit.get("credito_habilitado")):
        return None, _gas_lp_credit_reminder_exclusion(factura, "cliente_sin_credito")
    dias_credito = int(credit.get("dias_credito") or 0)
    if dias_credito <= 0:
        return None, _gas_lp_credit_reminder_exclusion(factura, "sin_dias_credito")

    try:
        emails = _customer_invoice_recipients(credit)
    except HTTPException as exc:
        return None, _gas_lp_credit_reminder_exclusion(factura, "sin_email", str(exc.detail or "Correo inválido."))
    if not emails:
        return None, _gas_lp_credit_reminder_exclusion(factura, "sin_email")

    fecha_emision = _gas_lp_factura_date_key(factura)
    if not fecha_emision:
        return None, _gas_lp_credit_reminder_exclusion(factura, "fecha_invalida")
    fecha_vencimiento = _gas_lp_credit_reminder_add_days(fecha_emision, dias_credito)
    dias_restantes = _gas_lp_credit_reminder_day_diff(fecha_vencimiento, today_key)
    if dias_restantes is None:
        return None, _gas_lp_credit_reminder_exclusion(factura, "fecha_invalida")
    if dias_restantes not in target_days:
        return None, _gas_lp_credit_reminder_exclusion(factura, "fuera_de_ventana", f"Faltan {dias_restantes} días.")

    total = _money(info.get("total"))
    reminder_type = f"before_{dias_restantes}" if dias_restantes > 0 else "due_today"
    candidate = {
        "factura_id": _safe_int_id(factura.get("id")),
        "uuid": factura.get("uuid_sat") or "",
        "folio": _gas_lp_factura_folio_label(factura),
        "cliente_id": _safe_int_id(credit.get("id") or md.get("cliente_id")),
        "cliente_nombre": credit.get("nombre") or md.get("cliente_nombre") or factura.get("rfc_receptor") or "",
        "rfc_receptor": factura.get("rfc_receptor") or credit.get("rfc") or "",
        "emails": emails,
        "fecha_emision": fecha_emision,
        "dias_credito": dias_credito,
        "fecha_vencimiento": fecha_vencimiento,
        "dias_restantes": dias_restantes,
        "total": float(total),
        "saldo_pendiente": float(saldo),
        "payment_status": info.get("payment_status") or md.get("payment_status") or "",
        "metodo_pago": metodo_pago,
        "tipo_recordatorio": reminder_type,
        "razon_elegibilidad": f"Factura PPD pendiente con vencimiento en {dias_restantes} día{'s' if dias_restantes != 1 else ''}.",
    }
    return candidate, None


@router.get("/internal-auth/gas-lp/credito/recordatorios/candidatos")
async def gas_lp_credito_recordatorios_candidatos(
    token: str,
    dias: str = "2,1",
    dry_run: bool = True,
    include_exclusions: bool = False,
):
    ctx = _gas_lp_factura_access_context(token, write=False)
    user = ctx["user"]
    profile = _gas_lp_profile(user)
    target_days = _gas_lp_credit_reminder_parse_days(dias)
    today_key = _gas_lp_credit_reminder_today_key()
    sb = get_supabase_admin()
    try:
        facturas = _gas_lp_company_facturas_rows(sb, user, profile, month="", limit=10000, include_carta_porte=False)
    except Exception as exc:
        raise _safe_internal_error("gas_lp_credito_recordatorios_facturas", exc)
    try:
        clientes = (
            _gas_lp_clientes_scope_query(sb.table("gas_lp_clientes_facturacion").select("*"), user)
            .eq("activo", True)
            .execute()
            .data
            or []
        )
    except Exception as exc:
        raise _safe_internal_error("gas_lp_credito_recordatorios_clientes", exc)

    clientes_by_id, clientes_by_rfc = _gas_lp_credit_reminder_client_maps(clientes)
    candidatos: list[dict] = []
    exclusiones: list[dict] = []
    for factura in facturas:
        cliente = _gas_lp_credit_reminder_cliente_for_factura(factura, clientes_by_id, clientes_by_rfc)
        candidato, exclusion = _gas_lp_credit_reminder_evaluate_factura(factura, cliente, target_days, today_key)
        if candidato:
            candidatos.append(candidato)
        elif exclusion and include_exclusions:
            exclusiones.append(exclusion)

    response = {
        "ok": True,
        "dry_run": True,
        "requested_dry_run": bool(dry_run),
        "send_enabled": False,
        "today": today_key,
        "dias": target_days,
        "candidatos": candidatos,
        "summary": {
            "facturas_revisadas": len(facturas),
            "candidatos": len(candidatos),
            "exclusiones": len(exclusiones) if include_exclusions else 0,
        },
    }
    if include_exclusions:
        response["exclusiones"] = exclusiones
    return JSONResponse(response)


@router.get("/internal-auth/gas-lp/facturas")
async def gas_lp_internal_facturas(token: str, mes: str | None = None):
    ctx = _gas_lp_internal_context(token)
    user = ctx["user"]
    profile = _gas_lp_profile(user)
    sb = get_supabase_admin()
    month = str(mes or "").strip()[:7]
    if len(month) == 7 and month[4] == "-":
        try:
            datetime.strptime(f"{month}-01", "%Y-%m-%d")
        except ValueError:
            month = ""
    else:
        month = ""
    try:
        rows = _gas_lp_company_facturas_rows(sb, user, profile, month=month, limit=10000 if month else 1000)
    except Exception as exc:
        raise _safe_internal_error("gas_lp_facturas", exc)
    try:
        _gas_lp_attach_internal_creators(sb, rows)
    except Exception as exc:
        logger.warning("gas_lp_facturas_attach_creators_failed perfil=%s err=%s", user.get("perfil_id"), exc)
    try:
        _gas_lp_attach_cliente_email_recipients(sb, user, rows)
    except Exception as exc:
        logger.warning("gas_lp_facturas_attach_client_emails_failed perfil=%s err=%s", user.get("perfil_id"), exc)
    try:
        comp_by_factura = _gas_lp_complementos_por_factura(sb, [_safe_int_id(r.get("id")) for r in rows if _safe_int_id(r.get("id"))])
    except Exception as exc:
        logger.warning("gas_lp_facturas_attach_complementos_failed perfil=%s err=%s", user.get("perfil_id"), exc)
        comp_by_factura = {}
    for row in rows:
        try:
            row["fecha_factura_key"] = _gas_lp_factura_date_key(row)
            row["payment_info"] = _payment_info_json(_factura_payment_info(row))
            row["fiscal_status"] = _gas_lp_factura_fiscal_status_info(row)
            row["realizado_por"] = _gas_lp_factura_realizado_por(row)
            row["carta_porte_summary"] = _gas_lp_factura_carta_porte_summary(row.get("xml_content") or "")
            comps = comp_by_factura.get(_safe_int_id(row.get("id")), [])
            row["complementos_pago"] = comps
            if comps:
                row["latest_complemento_pago"] = comps[0]
        except Exception as exc:
            logger.warning("gas_lp_factura_row_normalize_failed id=%s perfil=%s err=%s", row.get("id"), user.get("perfil_id"), exc)
            row["fecha_factura_key"] = _gas_lp_factura_date_key(row)
            row["payment_info"] = _payment_info_json({"metodo_pago": "", "forma_pago": "", "total": 0, "saldo_insoluto": 0, "payment_status": ""})
            row["fiscal_status"] = _gas_lp_factura_fiscal_status_info(row)
            row["realizado_por"] = _gas_lp_factura_realizado_por(row)
            row["carta_porte_summary"] = _gas_lp_factura_carta_porte_summary(row.get("xml_content") or "")
            row["complementos_pago"] = []
    return JSONResponse({"ok": True, "facturas": rows})


@router.get("/internal-auth/gas-lp/facturas/export-dia")
async def gas_lp_internal_facturas_export_dia(token: str, fecha: str):
    ctx = _gas_lp_internal_context(token)
    user = ctx["user"]
    profile = _gas_lp_profile(user)
    day = str(fecha or "").strip()[:10]
    try:
        datetime.strptime(day, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(400, "Selecciona una fecha válida para exportar.")
    sb = get_supabase_admin()
    try:
        rows = _gas_lp_company_facturas_rows(sb, user, profile, month=day[:7], limit=10000)
    except Exception as exc:
        raise _safe_internal_error("gas_lp_facturas_export_dia", exc)
    rows = [row for row in rows if _gas_lp_factura_date_key(row) == day]
    try:
        _gas_lp_attach_internal_creators(sb, rows)
    except Exception:
        pass
    try:
        comp_q = (
            sb.table("gas_lp_complementos_pago")
            .select("*")
            .eq("tenant_id", user.get("tenant_id"))
            .eq("perfil_id", user.get("perfil_id"))
            .gte("created_at", f"{day}T00:00:00")
            .lt("created_at", f"{day}T23:59:59")
            .order("created_at", desc=True)
            .limit(10000)
        )
        complementos = comp_q.execute().data or []
        _gas_lp_attach_complemento_creators(sb, complementos)
    except Exception as exc:
        logger.warning("gas_lp_facturas_export_dia_complementos_failed perfil=%s day=%s err=%s", user.get("perfil_id"), day, exc)
        complementos = []
    comp_ids = [_safe_int_id(c.get("id")) for c in complementos if _safe_int_id(c.get("id"))]
    comp_rels = []
    if comp_ids:
        try:
            comp_rels = (
                sb.table("gas_lp_complementos_pago_facturas")
                .select("*")
                .in_("complemento_id", comp_ids)
                .execute()
                .data
                or []
            )
        except Exception:
            comp_rels = []
    comp_rels_by_id: dict[int, list[dict]] = {}
    comp_factura_ids: list[int] = []
    for rel in comp_rels:
        comp_id = _safe_int_id(rel.get("complemento_id"))
        comp_rels_by_id.setdefault(comp_id, []).append(rel)
        fid = _safe_int_id(rel.get("factura_id"))
        if fid:
            comp_factura_ids.append(fid)
    for comp in complementos:
        comp_factura_ids.extend(_gas_lp_complemento_factura_ids(comp))
    comp_facturas_by_id = _gas_lp_facturas_by_ids_for_company(sb, user, profile, list(dict.fromkeys(comp_factura_ids)))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Documentos"
    headers = [
        "Fecha",
        "Folio de fact",
        "UUID",
        "Cliente",
        "Monto",
        "Litros",
        "Método",
        "Estado fiscal",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7A1E2C")
    for row in rows:
        md = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
        info = _factura_payment_info(row)
        ws.append([
            _gas_lp_factura_date_key(row),
            _gas_lp_factura_folio_label(row),
            row.get("uuid_sat") or "",
            _gas_lp_factura_razon_social(row),
            float(_money(info.get("total"))),
            float(Decimal(str(info.get("litros") or row.get("volumen_litros") or 0)).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)),
            info.get("metodo_pago") or _gas_lp_factura_metodo_pago(row),
            _gas_lp_factura_fiscal_status_info(row).get("label") or "",
        ])
    for comp in complementos:
        comp_id = _safe_int_id(comp.get("id"))
        rels = comp_rels_by_id.get(comp_id, [])
        ids = [*_gas_lp_complemento_factura_ids(comp), *[_safe_int_id(rel.get("factura_id")) for rel in rels]]
        ids = list(dict.fromkeys(fid for fid in ids if fid))
        facturas_comp = [comp_facturas_by_id[fid] for fid in ids if fid in comp_facturas_by_id]
        receptor = _gas_lp_complemento_receptor_info(comp, facturas_comp)
        refs = []
        for rel in rels:
            fid = _safe_int_id(rel.get("factura_id"))
            factura = comp_facturas_by_id.get(fid, {})
            refs.append(_gas_lp_factura_folio_label(factura) or rel.get("uuid_relacionado") or "")
        if not refs:
            refs = [_gas_lp_factura_folio_label(comp_facturas_by_id.get(fid, {})) for fid in ids]
        email_error = str(comp.get("email_error") or "")
        ws.append([
            str(comp.get("created_at") or "")[:10],
            "",
            comp.get("uuid_sat") or "",
            receptor.get("nombre") or "Cliente",
            float(comp.get("monto") or 0),
            0,
            "Complemento",
            "Vigente",
        ])
    for width, column in zip([14, 20, 40, 34, 16, 14, 12, 16], "ABCDEFGH"):
        ws.column_dimensions[column].width = width
    for column in ("E",):
        for cell in ws[column][1:]:
            cell.number_format = '$#,##0.00'
    for cell in ws["F"][1:]:
        cell.number_format = "#,##0.0000"
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"facturas_gas_lp_{day}.xlsx"
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/internal-auth/gas-lp/facturas/{factura_id}/xml")
async def gas_lp_internal_factura_xml(factura_id: int, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_factura_access_context(token, perfil_id=perfil_id)
    row = _gas_lp_internal_factura(ctx["user"], factura_id)
    xml_content = row.get("xml_content") or ""
    if not xml_content:
        raise HTTPException(404, "Factura sin XML timbrado.")
    info = fiscal_pdf_info(xml_content, "factura_gas_lp")
    filename = info.filename.replace(".pdf", ".xml")
    return Response(
        content=xml_content,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/internal-auth/gas-lp/facturas/{factura_id}/pac-audit")
async def gas_lp_internal_factura_pac_audit(factura_id: int, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_factura_access_context(token, perfil_id=perfil_id)
    row = _gas_lp_internal_factura(ctx["user"], factura_id)
    sb = get_supabase_admin()
    uuid_sat = str(row.get("uuid_sat") or "").strip()
    xml_content = str(row.get("xml_content") or "")
    xml_summary = _gas_lp_factura_pac_xml_summary(xml_content)
    responses = []
    try:
        query = sb.table("pac_responses").select("*").order("id", desc=True).limit(20)
        if uuid_sat:
            query = query.eq("uuid_sat", uuid_sat)
        responses = query.execute().data or []
    except Exception as exc:
        logger.warning("gas_lp_pac_audit_responses_lookup_failed factura_id=%s uuid=%s err=%s", factura_id, uuid_sat, exc)
        responses = []
    request_ids = sorted({
        int(resp.get("request_id") or 0)
        for resp in responses
        if resp.get("request_id")
    })
    requests_by_id = {}
    if request_ids:
        try:
            reqs = sb.table("pac_requests").select("*").in_("id", request_ids).execute().data or []
            requests_by_id = {int(req.get("id") or 0): req for req in reqs}
        except Exception as exc:
            logger.warning("gas_lp_pac_audit_requests_lookup_failed factura_id=%s uuid=%s err=%s", factura_id, uuid_sat, exc)
    audit = []
    for resp in responses:
        req = requests_by_id.get(int(resp.get("request_id") or 0), {})
        audit.append({
            "pac_response_id": resp.get("id"),
            "pac_request_id": resp.get("request_id"),
            "provider": resp.get("provider") or req.get("provider") or "",
            "environment": req.get("environment") or "",
            "operation": req.get("operation") or "",
            "request_created_at": req.get("created_at") or "",
            "response_created_at": resp.get("created_at") or "",
            "request_payload": req.get("request_payload") or {},
            "response_payload": resp.get("response_payload") or {},
            "status": resp.get("status") or "",
            "error_message": resp.get("error_message") or "",
            "uuid_sat": resp.get("uuid_sat") or "",
            "pdf_url": resp.get("pdf_url") or "",
            "xml_original": resp.get("xml_original") or "",
            "xml_timbrado": resp.get("xml_timbrado") or "",
        })
    return JSONResponse({
        "ok": True,
        "factura": {
            "id": row.get("id"),
            "uuid_sat": uuid_sat,
            "record_uuid": row.get("record_uuid") or "",
            "fecha_timbrado": row.get("fecha_timbrado") or "",
            "created_at": row.get("created_at") or "",
            "source": row.get("source") or "",
            "metadata": row.get("metadata") or {},
        },
        "xml_summary": xml_summary,
        "audit_count": len(audit),
        "audit": audit,
    })


@router.get("/internal-auth/gas-lp/facturas/{factura_id}/pdf")
async def gas_lp_internal_factura_pdf(factura_id: int, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_factura_access_context(token, perfil_id=perfil_id)
    user = ctx["user"]
    row = _gas_lp_internal_factura(user, factura_id)
    xml_content = row.get("xml_content") or ""
    if not xml_content:
        raise HTTPException(404, "Factura sin XML timbrado para generar PDF.")
    settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
    if es_carta_porte_traslado(xml_content):
        info = carta_porte_pdf_info(xml_content)
        pdf_bytes = generar_pdf_carta_porte_desde_xml(xml_content, logo_data_url=settings.get("PdfLogoDataUrl", ""))
    else:
        pac_pdf_url = str(row.get("pdf_url") or "").strip()
        if pac_pdf_url:
            return RedirectResponse(pac_pdf_url, status_code=302)
        info = fiscal_pdf_info(xml_content, "factura_gas_lp")
        pdf_bytes = generar_pdf_gas_lp_desde_xml(
            xml_content,
            logo_data_url=settings.get("PdfLogoDataUrl", ""),
            observaciones=_gas_lp_factura_observaciones(row),
        )
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{info.filename}"'},
    )


@router.post("/internal-auth/gas-lp/facturas/{factura_id}/send-email")
async def gas_lp_internal_factura_send_email(factura_id: int, payload: GasLpSendEmailPayload, token: str, perfil_id: int | None = None):
    ctx = _gas_lp_factura_access_context(token, write=True, perfil_id=perfil_id)
    user = ctx["user"]
    profile = _gas_lp_profile(user)
    settings = _gas_lp_settings(user.get("owner_user_id"), int(user.get("perfil_id")))
    issuer = _require_gas_lp_issuer(profile, settings)
    row = _gas_lp_internal_factura(user, factura_id)
    xml_content = str(row.get("xml_content") or "")
    if not xml_content:
        raise HTTPException(400, "La factura no tiene XML timbrado para enviar.")
    uuid_sat = str(row.get("uuid_sat") or "").strip()
    if not uuid_sat:
        raise HTTPException(400, "La factura no tiene UUID timbrado para enviar.")
    md = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
    if md.get("tipo_operacion") == "traspaso":
        fallback_email = md.get("transfer_email") or md.get("transfer_email_sent_to") or row.get("email_destinatario")
    else:
        cliente_rows = []
        cliente_id = int(md.get("cliente_id") or 0)
        if cliente_id:
            try:
                cliente_rows = (
                    get_supabase_admin()
                    .table("gas_lp_clientes_facturacion")
                    .select("*")
                    .eq("id", cliente_id)
                    .eq("user_id", user.get("owner_user_id"))
                    .eq("tenant_id", user.get("tenant_id"))
                    .eq("perfil_id", user.get("perfil_id"))
                    .eq("activo", True)
                    .limit(1)
                    .execute()
                    .data
                    or []
                )
            except Exception as exc:
                logger.warning("gas_lp_factura_send_email_cliente_lookup_failed factura=%s cliente=%s err=%s", factura_id, cliente_id, exc)
        fallback_recipients = _customer_invoice_recipients(cliente_rows[0]) if cliente_rows else _invoice_email_recipients(
            md.get("cliente_email") or row.get("email_destinatario"),
            md.get("email_adicional_1") or "",
            md.get("email_adicional_2") or "",
            fallback=md.get("email_sent_to") or md.get("email_last_attempt_to") or "",
        )
        fallback_email = ", ".join(fallback_recipients)
    recipients = _invoice_email_recipients(payload.email, payload.email_adicional_1, payload.email_adicional_2, fallback=fallback_email)
    recipient = ", ".join(recipients)
    if not recipients:
        raise HTTPException(400, "Captura un correo destino para enviar XML/PDF.")
    try:
        if es_carta_porte_traslado(xml_content):
            info = carta_porte_pdf_info(xml_content)
            pdf_bytes = generar_pdf_carta_porte_desde_xml(xml_content, logo_data_url=settings.get("PdfLogoDataUrl", ""))
        else:
            info = fiscal_pdf_info(xml_content, "factura_gas_lp")
            pdf_bytes = generar_pdf_gas_lp_desde_xml(
                xml_content,
                logo_data_url=settings.get("PdfLogoDataUrl", ""),
                observaciones=_gas_lp_factura_observaciones(row),
            )
    except Exception as exc:
        raise _safe_internal_error("gas_lp_factura_send_email_pdf", exc)
    email_results = []
    email_result = None
    for email_to in recipients:
        email_result = send_gas_lp_invoice_email(
            to_email=email_to,
            issuer_name=issuer["nombre"],
            customer_name=str(md.get("cliente_nombre") or row.get("rfc_receptor") or "Cliente"),
            uuid_sat=uuid_sat,
            total=md.get("total") or _gas_lp_factura_total_con_iva(row),
            xml_content=xml_content,
            pdf_bytes=pdf_bytes,
            pdf_filename=info.filename,
            serie_folio=_gas_lp_factura_folio_label(row),
        )
        email_results.append({"to": email_to, **email_result.as_metadata()})
    now_email = _now_iso()
    all_ok = bool(email_results) and all(item.get("ok") for item in email_results)
    first_error = next((str(item.get("error") or "") for item in email_results if not item.get("ok")), "")
    message_ids = ", ".join(str(item.get("message_id") or "") for item in email_results if item.get("message_id"))
    updated_md = {
        **md,
        "email_delivery": email_result.as_metadata() if email_result else {},
        **_invoice_email_metadata(recipients),
        "email_sent_at": now_email if all_ok else md.get("email_sent_at"),
        "email_sent_to": recipient if all_ok else md.get("email_sent_to", recipient),
        "resend_message_id": message_ids if all_ok else md.get("resend_message_id", ""),
        "email_error": "" if all_ok else first_error,
        "email_last_attempt_at": now_email,
        "email_last_attempt_to": recipient,
    }
    if md.get("tipo_operacion") == "traspaso":
        updated_md = {
            **updated_md,
            "transfer_email_delivery": email_results,
            "transfer_email_sent_at": now_email if all_ok else md.get("transfer_email_sent_at"),
            "transfer_email_sent_to": recipient if all_ok else md.get("transfer_email_sent_to", recipient),
            "transfer_email_message_id": message_ids if all_ok else md.get("transfer_email_message_id", ""),
            "transfer_email_error": "" if all_ok else first_error,
        }
    update_payload = {
        "metadata": updated_md,
        "email_enviado": all_ok,
        "email_enviado_at": now_email if all_ok else row.get("email_enviado_at"),
        "email_destinatario": recipient,
        "email_error": "" if all_ok else first_error,
        "updated_at": now_email,
    }
    try:
        updated = (
            get_supabase_admin()
            .table("gas_lp_facturas")
            .update(update_payload)
            .eq("id", factura_id)
            .execute()
            .data
            or []
        )
    except Exception as exc:
        raise _safe_internal_error("gas_lp_factura_send_email_update", exc)
    factura = updated[0] if updated else {**row, **update_payload}
    response = {"ok": all_ok, "factura": factura, "email": email_result.as_metadata() if email_result else {}, "email_results": email_results}
    if not all_ok:
        response["message"] = first_error or "No se pudo enviar el correo."
    return JSONResponse(response, status_code=200 if all_ok else 400)
