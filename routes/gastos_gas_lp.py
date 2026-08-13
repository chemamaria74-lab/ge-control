from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
import base64
import os
import re
import unicodedata
from io import BytesIO
from typing import Any, Literal

from fastapi import APIRouter, Header, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

from routes.auth import verify_token
from routes.flotilla import _internal_fleet_context
from routes.internal_users_mod.core import _gas_lp_conciliacion_context
from services.email_delivery import send_gas_lp_expense_payment_email
from services.database import get_facilities
from supabase_config import get_supabase_admin


router = APIRouter()
MONEY_TOLERANCE = 0.005
PAYMENT_BALANCE_TOLERANCE = 1.00


class ConceptCreate(BaseModel):
    name: str = Field(min_length=2, max_length=120)


class ConceptUpdate(ConceptCreate):
    status: Literal["active", "inactive"] = "active"


class DriverCreate(BaseModel):
    name: str = Field(min_length=2, max_length=180)
    license_number: str = Field(default="", max_length=80)
    license_type: str = Field(default="", max_length=40)
    issued_on: date | None = None
    expires_on: date


class SupplierCreate(BaseModel):
    commercial_name: str = Field(min_length=2, max_length=180)
    legal_name: str = Field(default="", max_length=220)
    rfc: str = Field(default="", max_length=20)
    bank_name: str = Field(default="", max_length=120)
    account_number: str = Field(default="", max_length=34)
    payment_email: str = Field(default="", max_length=180)


class SupplierReview(BaseModel):
    action: Literal["validate", "reject"]
    reason: str = Field(default="", max_length=500)


class SupplierUpdate(BaseModel):
    commercial_name: str = Field(min_length=2, max_length=180)
    legal_name: str = Field(default="", max_length=220)
    rfc: str = Field(default="", max_length=20)
    bank_name: str = Field(default="", max_length=120)
    account_number: str = Field(default="", max_length=34)
    payment_email: str = Field(default="", max_length=180)
    status: Literal["active", "inactive"] = "active"


class VoucherCreate(BaseModel):
    group_id: int
    expense_zone_id: int
    vehicle_id: int
    supplier_id: int
    concept_id: int
    issued_on: date = Field(default_factory=date.today)
    description: str = Field(default="", max_length=500)
    driver_name: str = Field(default="", max_length=180)


class VoucherAmount(BaseModel):
    amount_mxn: float = Field(gt=0, le=100_000_000)


class CancelPayload(BaseModel):
    reason: str = Field(min_length=3, max_length=500)


class VoucherInvoiceCreate(BaseModel):
    voucher_ids: list[int] = Field(min_length=1, max_length=100)
    invoice_number: str = Field(min_length=1, max_length=100)
    invoice_date: date
    total_mxn: float = Field(gt=0, le=100_000_000)


class DirectInvoiceCreate(BaseModel):
    supplier_id: int
    concept_id: int
    invoice_number: str = Field(min_length=1, max_length=100)
    invoice_date: date
    total_mxn: float = Field(gt=0, le=100_000_000)
    period_key: str = Field(default="", pattern=r"^$|^\d{4}-\d{2}$")
    description: str = Field(default="", max_length=500)
    group_id: int | None = None
    facility_id: int | None = None
    expense_zone_id: int | None = None
    payment_target: Literal["supplier", "reimbursement"] = "supplier"
    reimbursement_recipient_id: int | None = None
    reimbursement_account_id: int | None = None
    expense_type: Literal["direct", "credit_note"] = "direct"


class DirectInvoiceBatchLine(BaseModel):
    invoice_number: str = Field(min_length=1, max_length=100)
    invoice_date: date
    total_mxn: float = Field(gt=0, le=100_000_000)


class DirectInvoiceBatchCreate(BaseModel):
    supplier_id: int
    concept_id: int
    invoices: list[DirectInvoiceBatchLine] = Field(min_length=2, max_length=100)
    period_key: str = Field(default="", pattern=r"^$|^\d{4}-\d{2}$")
    description: str = Field(default="", max_length=500)
    group_id: int | None = None
    facility_id: int | None = None
    expense_zone_id: int | None = None
    payment_target: Literal["supplier", "reimbursement"] = "supplier"
    reimbursement_recipient_id: int | None = None
    reimbursement_account_id: int | None = None


class SupplierAdvanceCreate(BaseModel):
    supplier_id: int
    concept_id: int | None = None
    expense_zone_id: int | None = None
    paid_on: date
    amount_mxn: float = Field(gt=0, le=100_000_000)
    reference: str = Field(default="", max_length=160)
    description: str = Field(default="", max_length=500)


class AdvanceAllocationInput(BaseModel):
    advance_id: int
    amount_mxn: float = Field(gt=0, le=100_000_000)


class SupplierAdvanceApply(BaseModel):
    supplier_id: int
    concept_id: int | None = None
    expense_zone_id: int | None = None
    invoice_number: str = Field(min_length=1, max_length=100)
    invoice_date: date
    total_mxn: float = Field(gt=0, le=100_000_000)
    description: str = Field(default="", max_length=500)
    allocations: list[AdvanceAllocationInput] = Field(min_length=1, max_length=100)


class DirectInvoiceUpdate(BaseModel):
    invoice_number: str = Field(min_length=1, max_length=100)
    invoice_date: date
    total_mxn: float = Field(gt=0, le=100_000_000)
    description: str = Field(default="", max_length=500)
    observation: str = Field(default="", max_length=500)


class PaidInvoiceDateUpdate(BaseModel):
    invoice_date: date


class ReimbursementAccountInput(BaseModel):
    account_type: Literal["payroll", "credit_card"]
    label: str = Field(default="", max_length=80)
    bank_name: str = Field(default="", max_length=120)
    account_number: str = Field(default="", max_length=34)
    card_last_four: str = Field(default="", max_length=4)


class ReimbursementRecipientCreate(BaseModel):
    name: str = Field(min_length=2, max_length=180)
    email: str = Field(min_length=5, max_length=180)
    accounts: list[ReimbursementAccountInput] = Field(default_factory=list, max_length=8)


class ReimbursementRecipientUpdate(ReimbursementRecipientCreate):
    status: Literal["active", "inactive"] = "active"


class PaymentAllocation(BaseModel):
    invoice_id: int
    amount_mxn: float = Field(gt=0, le=100_000_000)


class ExpensePaymentCreate(BaseModel):
    invoice_allocations: list[PaymentAllocation] = Field(min_length=1, max_length=200)
    credit_note_ids: list[int] = Field(default_factory=list, max_length=100)
    paid_on: date
    amount_mxn: float = Field(gt=0, le=100_000_000)
    method: str = Field(default="", max_length=80)
    reference: str = Field(default="", max_length=160)
    notes: str = Field(default="", max_length=500)


class ExpensePaymentDateUpdate(BaseModel):
    paid_on: date


class ExpenseZoneCreate(BaseModel):
    name: str = Field(min_length=2, max_length=120)


class ExpenseZoneUpdate(ExpenseZoneCreate):
    status: Literal["active", "inactive"] = "active"


class InvoiceTransition(BaseModel):
    action: Literal["accept", "observe", "reject", "send_to_accountant", "mark_paid", "cancel", "withdraw_from_payments"]
    observation: str = Field(default="", max_length=500)
    paid_on: date | None = None
    paid_amount_mxn: float | None = Field(default=None, gt=0, le=100_000_000)


class InvoiceCorrection(BaseModel):
    invoice_number: str = Field(min_length=1, max_length=100)
    invoice_date: date
    total_mxn: float = Field(gt=0, le=100_000_000)


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _normalize(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").strip())
    return re.sub(r"\s+", " ", "".join(c for c in text if not unicodedata.combining(c))).upper()


def _insert_expense_invoices(ctx: dict[str, Any], rows: dict[str, Any] | list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Insert invoices while translating the database idempotency guard for the UI."""
    try:
        return ctx["sb"].table("gas_lp_expense_invoices").insert(rows).execute().data or []
    except Exception as exc:
        detail = str(exc).lower()
        if "gas_lp_expense_invoices_active_identity_uidx" in detail or "23505" in detail:
            raise HTTPException(
                409,
                "Una o más facturas ya estaban registradas. No se guardó ninguna para evitar duplicados.",
            ) from exc
        raise


def _code(value: str, length: int) -> str:
    cleaned = re.sub(r"[^A-Z0-9]", "", _normalize(value))
    return (cleaned or "X")[:length]


def _validate_supplier_fields(rfc: str, email: str) -> tuple[str, str]:
    clean_rfc = _normalize(rfc)
    clean_email = str(email or "").strip().lower()
    if clean_rfc and not re.fullmatch(r"[A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3}", clean_rfc):
        raise HTTPException(400, "El RFC del proveedor no tiene un formato válido.")
    if clean_email and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", clean_email):
        raise HTTPException(400, "El correo de pagos no tiene un formato válido.")
    return clean_rfc, clean_email


def _clean_account_number(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", str(value or "").strip()).upper()


def _recipient_account_rows(ctx: dict[str, Any], recipient_id: int,
                            accounts: list[ReimbursementAccountInput]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for account in accounts:
        account_number = _clean_account_number(account.account_number)
        last_four = re.sub(r"\D", "", account.card_last_four)
        if account.account_type == "payroll" and not account_number:
            raise HTTPException(400, "Captura la cuenta o CLABE para el destino de nómina.")
        if account.account_type == "credit_card" and len(last_four) != 4:
            raise HTTPException(400, "Captura los últimos 4 dígitos de la tarjeta de crédito.")
        rows.append({
            "tenant_id": ctx["tenant_id"], "profile_id": ctx["perfil_id"],
            "recipient_id": recipient_id, "account_type": account.account_type,
            "label": account.label.strip() or ("Nómina" if account.account_type == "payroll" else "Tarjeta de crédito"),
            "bank_name": account.bank_name.strip(), "account_number": account_number,
            "card_last_four": last_four, "status": "active", "updated_at": _now(),
        })
    return rows


def _ctx(authorization: str, fleet_access: str, token: str, profile_header: str = "") -> dict[str, Any]:
    sb = get_supabase_admin()
    raw_profile_header = str(profile_header or "").strip()
    raw_profile_id, _, requested_expense_module = raw_profile_header.partition("|")
    requested_expense_module = requested_expense_module.strip().lower()
    if requested_expense_module not in {"", "gas_lp", "transporte", "control_administrativo"}:
        raise HTTPException(400, "Módulo de gastos inválido.")
    # The query token belongs to the Gas LP supervision/conciliation portal,
    # even when it happens to be a Supabase JWT.  Standalone expense portals
    # authenticate with the Authorization header instead.  Token shape is not
    # an authorization contract: treating every three-part token as a generic
    # module session made valid supervisors lose their write permission.
    if token and requested_expense_module in {"transporte", "control_administrativo"}:
        authorization = f"Bearer {token}"
        token = ""
    if fleet_access:
        ctx = _internal_fleet_context(fleet_access)
        ctx["is_manager"] = ctx.get("fleet_access_level") == "zone_manager"
        ctx["is_admin"] = not ctx["is_manager"]
        ctx["actor_id"] = str(ctx.get("internal_user_id") or "")
        ctx["actor_name"] = ctx.get("display_name") or "Gerente"
        return ctx
    if token:
        requested_profile_id = int(raw_profile_id) if raw_profile_id.isdigit() else None
        if "~" in token:
            token, raw_profile_id = token.rsplit("~", 1)
            # Compatibility for sessions opened by an older portal build. The
            # explicit profile header is authoritative in the current contract.
            if requested_profile_id is None:
                requested_profile_id = int(raw_profile_id) if raw_profile_id.isdigit() else None
            token = re.sub(r"(?:~\d+)+$", "", token)
        session_ctx = _gas_lp_conciliacion_context(token, write=True, perfil_id=requested_profile_id)
        user = session_ctx["user"]
        if str(user.get("role") or "").lower() not in {"conciliacion", "admin"}:
            raise HTTPException(403, "Este usuario no tiene acceso a Gastos y pagos.")
        return {
            "sb": sb, "tenant_id": str(user["tenant_id"]), "perfil_id": int(user["perfil_id"]),
            "allowed_group_ids": None, "is_manager": False, "is_admin": True,
            "actor_id": str(user["id"]), "actor_name": user.get("display_name") or "Asistente de gastos",
        }
    if authorization.startswith("Bearer "):
        access_token = authorization[7:].strip()
        uid = verify_token(access_token)
        if not uid:
            # Un JWT vencido o inválido es una sesión terminada, no una
            # degradación de permisos. Responder 401 permite renovar el token
            # o sacar al usuario del portal con un mensaje claro.
            raise HTTPException(401, "La sesión expiró; inicia sesión nuevamente.")
        requested_profile_id = int(raw_profile_id) if raw_profile_id.isdigit() else 0
        module = requested_expense_module or "gas_lp"
        if requested_profile_id:
            memberships = (sb.table("company_module_memberships").select("module")
                           .eq("profile_id", requested_profile_id).eq("status", "active")
                           .eq("expense_enabled", True).execute().data or [])
            modules = [str(row.get("module") or "") for row in memberships]
            if requested_expense_module:
                if requested_expense_module not in modules:
                    raise HTTPException(403, "Esta empresa no tiene habilitado Gastos y pagos en este módulo.")
            elif "control_administrativo" in modules and "gas_lp" not in modules:
                module = "control_administrativo"
            elif "transporte" in modules and "gas_lp" not in modules:
                module = "transporte"
        access = _server_module_access(sb, str(uid or ""), module, requested_profile_id)
        access_role = str(access.get("role") or "").lower()
        requested_perfil = requested_profile_id or int(access.get("perfil_id") or 0)
        owner_write = (
            module == "control_administrativo" and access_role == "user"
            and _control_admin_owner_can_manage(sb, str(uid), requested_perfil)
        )
        if access_role != "admin" and not owner_write:
            raise HTTPException(403, "Se requiere administración del módulo de gastos.")
        perfil_id = requested_perfil
        if not perfil_id:
            raise HTTPException(400, "Selecciona una empresa activa.")
        return {
            "sb": sb, "tenant_id": str(access["tenant_id"]), "perfil_id": perfil_id,
            "allowed_group_ids": None, "is_manager": False, "is_admin": True,
            "actor_id": str(uid), "actor_name": "Administración", "expense_module": module,
        }
    raise HTTPException(401, "Sesión requerida.")


def _server_module_access(sb: Any, user_id: str, module: str, perfil_id: int = 0) -> dict[str, Any]:
    """Resolve expense administration from server-side authorization rows.

    The JWT has already been verified before this function runs. Reading the
    role again through a user-scoped PostgREST client made valid admins appear
    as ordinary users whenever that secondary RLS read returned no rows. Use
    the service client already required by the expenses backend, while keeping
    the lookup constrained to the verified user, active module, tenant and
    selected company.
    """
    if not user_id or module not in {"gas_lp", "transporte", "control_administrativo"}:
        return {}
    rows = (sb.table("user_sections")
            .select("role,status,tenant_id,perfil_id,display_name")
            .eq("user_id", user_id).eq("section", module).eq("status", "active")
            .execute().data or [])
    if not rows:
        return {}

    def priority(row: dict[str, Any]) -> tuple[int, int]:
        assigned = int(row.get("perfil_id") or 0)
        # A tenant-wide administrator must win over a profile-specific ordinary
        # user row. Some accounts legitimately keep both rows after a module or
        # company activation; preferring the exact profile first downgraded the
        # request and made every expense write fail with the admin-only error.
        return (1 if str(row.get("role") or "").lower() == "admin" else 0,
                2 if perfil_id and assigned == perfil_id else 1 if assigned == 0 else 0)

    access = max(rows, key=priority)
    tenant_id = str(access.get("tenant_id") or "")
    assigned_profile = int(access.get("perfil_id") or 0)
    role = str(access.get("role") or "user").lower()
    if perfil_id:
        profiles = (sb.table("perfiles_empresa").select("id,user_id,tenant_id,activo")
                    .eq("id", perfil_id).eq("tenant_id", tenant_id).eq("activo", True)
                    .limit(1).execute().data or [])
        if not profiles:
            return {}
        if assigned_profile and assigned_profile != perfil_id:
            return {}
        if role != "admin" and str(profiles[0].get("user_id") or "") != user_id:
            return {}
    return access


def _profile(ctx: dict[str, Any]) -> dict[str, Any]:
    rows = (ctx["sb"].table("perfiles_empresa").select("id,nombre,rfc,user_id")
            .eq("tenant_id", ctx["tenant_id"]).eq("id", ctx["perfil_id"]).limit(1).execute().data or [])
    if not rows:
        raise HTTPException(403, "Empresa no disponible.")
    return rows[0]


def _control_admin_owner_can_manage(sb: Any, user_id: str, perfil_id: int) -> bool:
    """Company owners in Control administrativo may manage their own expenses.

    Some legacy Control administrativo memberships were created with role
    ``user`` even though the user owns the selected company.  Keep the write
    permission scoped to that exact profile instead of treating every module
    user as an administrator.
    """
    if not user_id or not perfil_id:
        return False
    rows = (sb.table("perfiles_empresa").select("id")
            .eq("id", perfil_id).eq("user_id", user_id).eq("activo", True)
            .limit(1).execute().data or [])
    return bool(rows)


def _profile_facilities(ctx: dict[str, Any], profile: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    """Read Mowry/Supabase company facilities; expense zones do not depend on fleet scopes."""
    profile = profile or _profile(ctx)
    rows = get_facilities(str(profile.get("user_id") or ""), "gas_lp", perfil_id=ctx["perfil_id"])
    if not rows:
        rows = (ctx["sb"].table("user_facilities").select("*")
                .eq("perfil_id", ctx["perfil_id"]).eq("modulo_propietario", "gas_lp")
                .order("nombre").execute().data or [])
    return [row for row in rows if row.get("activo", True) is not False and row.get("status", "active") != "inactive"]


def _expense_zones(ctx: dict[str, Any]) -> list[dict[str, Any]]:
    return (_base_query(ctx, "gas_lp_expense_zones").eq("status", "active").order("name").execute().data or [])


def _profile_expense_groups(ctx: dict[str, Any]) -> list[dict[str, Any]]:
    """Return only the Motive zones assigned to the company, never its root company group."""
    scopes = (ctx["sb"].table("fleet_profile_group_scopes").select("group_id,scope_type")
              .eq("tenant_id", ctx["tenant_id"]).eq("profile_id", ctx["perfil_id"])
              .eq("status", "active").execute().data or [])
    group_ids = [int(row["group_id"]) for row in scopes if row.get("scope_type") == "zone"]
    if not group_ids:
        root_ids = {int(row["group_id"]) for row in scopes if row.get("scope_type") == "company_root"}
        group_ids = [int(row["group_id"]) for row in scopes if int(row["group_id"]) not in root_ids]
    if not group_ids:
        return []
    groups = (ctx["sb"].table("fleet_groups").select("id,name,path")
              .eq("tenant_id", ctx["tenant_id"]).in_("id", group_ids).order("name").execute().data or [])
    if ctx.get("allowed_group_ids") is not None:
        allowed = {int(value) for value in ctx["allowed_group_ids"]}
        groups = [row for row in groups if int(row["id"]) in allowed]
    return groups


def _allowed_group(ctx: dict[str, Any], group_id: int | None) -> None:
    allowed = ctx.get("allowed_group_ids")
    if group_id is not None and allowed is not None and int(group_id) not in {int(x) for x in allowed}:
        raise HTTPException(403, "La zona no está asignada a este gerente.")


def _audit(ctx: dict[str, Any], entity_type: str, entity_id: int, action: str,
           before: dict | None = None, after: dict | None = None) -> None:
    ctx["sb"].table("gas_lp_expense_audit_log").insert({
        "tenant_id": ctx["tenant_id"], "profile_id": ctx["perfil_id"],
        "entity_type": entity_type, "entity_id": entity_id, "action": action,
        "actor_type": "manager" if ctx["is_manager"] else "admin",
        "actor_id": ctx["actor_id"], "before_data": before or {}, "after_data": after or {},
    }).execute()


def _base_query(ctx: dict[str, Any], table: str):
    return ctx["sb"].table(table).select("*").eq("tenant_id", ctx["tenant_id"]).eq("profile_id", ctx["perfil_id"])


def _is_without_folio(value: Any) -> bool:
    """Return true for the conventional labels used when a receipt has no folio."""
    compact = re.sub(r"[^A-Z0-9]", "", _normalize(str(value or "")))
    return compact in {"SF", "SINFOLIO", "SINNUMERO", "NA"}


def _invoice_alerts(ctx: dict[str, Any], *, supplier_id: int, invoice_number: str,
                    invoice_date: date, total_mxn: float, exclude_invoice_id: int | None = None) -> list[str]:
    rows = _base_query(ctx, "gas_lp_expense_invoices").execute().data or []
    rows = [
        row for row in rows
        if int(row.get("id") or 0) != int(exclude_invoice_id or 0)
        and row.get("status") not in {"rejected", "cancelled"}
    ]
    alerts: list[str] = []
    if not _is_without_folio(invoice_number) and any(
        int(row.get("supplier_id") or 0) == int(supplier_id)
        and not _is_without_folio(row.get("invoice_number"))
        and _normalize(row.get("invoice_number")) == _normalize(invoice_number)
        for row in rows
    ):
        alerts.append("Número de factura repetido para este proveedor.")
    if any(
        int(row.get("supplier_id") or 0) == int(supplier_id)
        and str(row.get("invoice_date") or "")[:10] == invoice_date.isoformat()
        and abs(float(row.get("total_mxn") or 0) - round(total_mxn, 2)) <= MONEY_TOLERANCE
        for row in rows
    ):
        alerts.append("Ya existe una factura del mismo proveedor, fecha e importe.")
    return alerts


def _send_payment_notification(ctx: dict[str, Any], invoice: dict[str, Any]) -> dict[str, Any]:
    suppliers = (_base_query(ctx, "gas_lp_expense_suppliers").eq("id", invoice["supplier_id"])
                 .limit(1).execute().data or [{}])
    supplier = suppliers[0]
    profile = _profile(ctx)
    paid_on = str(invoice.get("paid_on") or "")[:10]
    delivery = send_gas_lp_expense_payment_email(
        to_email=supplier.get("payment_email"), supplier_name=supplier.get("commercial_name") or "",
        company_name=profile.get("nombre") or "", invoice_number=invoice.get("invoice_number") or "",
        paid_on=paid_on, amount=invoice.get("paid_amount_mxn") or invoice.get("total_mxn") or 0,
        invoices=[{
            "invoice_number": invoice.get("invoice_number") or "",
            "invoice_date": invoice.get("invoice_date") or "",
            "total_mxn": invoice.get("total_mxn") or 0,
            "amount_paid_mxn": invoice.get("paid_amount_mxn") or invoice.get("total_mxn") or 0,
        }],
        idempotency_key=f"gas-lp-expense-{ctx['tenant_id']}-{invoice['id']}-{paid_on}",
    )
    return {
        "payment_email_status": "sent" if delivery.ok else ("skipped" if delivery.skipped else "failed"),
        "payment_email_metadata": delivery.as_metadata(), "updated_at": _now(),
    }


@router.get("/gastos/bootstrap")
def bootstrap(
    token: str = Query(default=""), authorization: str = Header(default=""),
    x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
    x_perfil_id: str = Header(default="", alias="X-Perfil-ID"),
):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    profile = _profile(ctx)
    facilities = _profile_facilities(ctx, profile)
    groups = _profile_expense_groups(ctx)
    if not groups:
        return {
            "identity": {"is_manager": ctx["is_manager"], "is_admin": ctx["is_admin"],
                         "name": ctx["actor_name"], "profile_id": ctx["perfil_id"]},
            "company": profile, "groups": [], "facilities": facilities,
            "expense_zones": _expense_zones(ctx), "vehicles": [], "mobile_drivers": [],
        }
    vehicles = (ctx["sb"].table("fleet_vehicles")
                .select("id,vehicle_number,current_driver_name,status")
                .eq("tenant_id", ctx["tenant_id"]).order("vehicle_number").execute().data or [])
    memberships_query = (ctx["sb"].table("fleet_vehicle_groups").select("vehicle_id,group_id")
                         .eq("tenant_id", ctx["tenant_id"]).in_("group_id", [int(row["id"]) for row in groups]))
    if ctx.get("allowed_group_ids") is not None:
        memberships_query = memberships_query.in_("group_id", list(ctx["allowed_group_ids"]))
    memberships = memberships_query.execute().data or []
    groups_by_vehicle: defaultdict[int, list[int]] = defaultdict(list)
    for membership in memberships:
        groups_by_vehicle[int(membership["vehicle_id"])].append(int(membership["group_id"]))
    for vehicle in vehicles:
        vehicle["group_ids"] = groups_by_vehicle.get(int(vehicle["id"]), [])
    vehicle_ids = set(groups_by_vehicle)
    vehicles = [row for row in vehicles if int(row["id"]) in vehicle_ids]
    mobile_drivers_by_name: dict[str, dict[str, Any]] = {}

    def add_mobile_driver(name: Any, vehicle_id: Any) -> None:
        clean_name = str(name or "").strip()
        clean_vehicle_id = int(vehicle_id or 0)
        if not clean_name or clean_vehicle_id not in vehicle_ids:
            return
        key = _normalize(clean_name)
        driver = mobile_drivers_by_name.setdefault(key, {"name": clean_name, "group_ids": set()})
        driver["group_ids"].update(groups_by_vehicle.get(clean_vehicle_id, []))

    for vehicle in vehicles:
        add_mobile_driver(vehicle.get("current_driver_name"), vehicle.get("id"))
    try:
        periods = (ctx["sb"].table("fleet_driving_periods")
                   .select("vehicle_id,driver_name")
                   .eq("tenant_id", ctx["tenant_id"])
                   .in_("vehicle_id", list(vehicle_ids))
                   .order("started_at", desc=True).limit(5000).execute().data or [])
        latest_driver_by_vehicle: dict[int, str] = {}
        for period in periods:
            add_mobile_driver(period.get("driver_name"), period.get("vehicle_id"))
            period_vehicle_id = int(period.get("vehicle_id") or 0)
            period_driver_name = str(period.get("driver_name") or "").strip()
            if period_vehicle_id and period_driver_name and period_vehicle_id not in latest_driver_by_vehicle:
                latest_driver_by_vehicle[period_vehicle_id] = period_driver_name
        for vehicle in vehicles:
            if not str(vehicle.get("current_driver_name") or "").strip():
                vehicle["current_driver_name"] = latest_driver_by_vehicle.get(int(vehicle["id"]), "")
    except Exception:
        # La lista sigue mostrando los choferes actuales aunque un despliegue
        # anterior todavía no tenga el histórico de conducción de Móvil.
        pass
    mobile_drivers = [
        {"name": row["name"], "group_ids": sorted(row["group_ids"]), "source": "mobile"}
        for row in mobile_drivers_by_name.values()
    ]
    mobile_drivers.sort(key=lambda row: _normalize(row["name"]))
    return {
        "identity": {"is_manager": ctx["is_manager"], "is_admin": ctx["is_admin"],
                     "name": ctx["actor_name"], "profile_id": ctx["perfil_id"]},
        "company": profile, "groups": groups, "facilities": facilities,
        "expense_zones": _expense_zones(ctx), "vehicles": vehicles,
        "mobile_drivers": mobile_drivers,
    }


@router.get("/gastos/company")
def company_fiscal_information(
    token: str = Query(default=""), authorization: str = Header(default=""),
    x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
    x_perfil_id: str = Header(default="", alias="X-Perfil-ID"),
):
    """Fiscal record for the selected Control administrativo company."""
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    membership = (ctx["sb"].table("company_module_memberships").select("profile_id")
                  .eq("tenant_id", ctx["tenant_id"]).eq("profile_id", ctx["perfil_id"])
                  .eq("module", "control_administrativo").eq("status", "active")
                  .limit(1).execute().data or [])
    if not membership:
        raise HTTPException(403, "La información fiscal está disponible en Control administrativo.")
    profile = _profile(ctx)
    details = (ctx["sb"].table("company_fiscal_details").select("*")
               .eq("tenant_id", ctx["tenant_id"]).eq("profile_id", ctx["perfil_id"])
               .limit(1).execute().data or [])
    return {"company": profile, "fiscal": details[0] if details else None}


@router.get("/gastos/concepts")
def list_concepts(limit: int = Query(default=300, ge=1, le=500),
                  token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    return {"items": _base_query(ctx, "gas_lp_expense_concepts").order("name").limit(limit).execute().data or []}


@router.get("/gastos/drivers")
def list_drivers(limit: int = Query(default=300, ge=1, le=500),
                 token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    return {"items": _base_query(ctx, "gas_lp_expense_drivers").order("name").limit(limit).execute().data or []}


@router.post("/gastos/drivers", status_code=201)
def create_driver(payload: DriverCreate, token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if payload.issued_on and payload.expires_on < payload.issued_on:
        raise HTTPException(400, "La vigencia no puede ser anterior a la fecha de expedición.")
    normalized = _normalize(payload.name)
    existing = (_base_query(ctx, "gas_lp_expense_drivers").eq("normalized_name", normalized)
                .eq("status", "active").limit(1).execute().data or [])
    if existing:
        raise HTTPException(409, "Este chofer ya existe en el catálogo.")
    row = {"tenant_id": ctx["tenant_id"], "profile_id": ctx["perfil_id"],
           "name": payload.name.strip(), "normalized_name": normalized,
           "license_number": payload.license_number.strip(), "license_type": payload.license_type.strip(),
           "issued_on": payload.issued_on.isoformat() if payload.issued_on else None,
           "expires_on": payload.expires_on.isoformat(),
           "created_by": ctx["actor_id"]}
    created = ctx["sb"].table("gas_lp_expense_drivers").insert(row).execute().data[0]
    _audit(ctx, "driver", int(created["id"]), "created", after=created)
    return {"item": created}


@router.post("/gastos/concepts", status_code=201)
def create_concept(payload: ConceptCreate, token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    normalized = _normalize(payload.name)
    existing = (_base_query(ctx, "gas_lp_expense_concepts").eq("normalized_name", normalized)
                .limit(1).execute().data or [])
    if existing:
        raise HTTPException(409, "Este concepto ya existe en el catálogo.")
    row = {
        "tenant_id": ctx["tenant_id"], "profile_id": ctx["perfil_id"], "name": payload.name.strip(),
        "normalized_name": normalized, "created_by": ctx["actor_id"],
    }
    created = ctx["sb"].table("gas_lp_expense_concepts").insert(row).execute().data[0]
    _audit(ctx, "concept", int(created["id"]), "created", after=created)
    return {"item": created}


@router.put("/gastos/concepts/{concept_id}")
def update_concept(concept_id: int, payload: ConceptUpdate, token: str = Query(default=""),
                   authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    rows = _base_query(ctx, "gas_lp_expense_concepts").eq("id", concept_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Concepto no encontrado.")
    normalized = _normalize(payload.name)
    duplicate = (_base_query(ctx, "gas_lp_expense_concepts").eq("normalized_name", normalized)
                 .neq("id", concept_id).limit(1).execute().data or [])
    if duplicate:
        raise HTTPException(409, "Este concepto ya existe en el catálogo.")
    update = {"name": payload.name.strip(), "normalized_name": normalized,
              "status": payload.status, "updated_at": _now()}
    ctx["sb"].table("gas_lp_expense_concepts").update(update).eq("id", concept_id).execute()
    _audit(ctx, "concept", concept_id, "updated", before=rows[0], after=update)
    return {"ok": True, "item": {**rows[0], **update}}


@router.delete("/gastos/concepts/{concept_id}")
def delete_concept(concept_id: int, token: str = Query(default=""), authorization: str = Header(default=""),
                   x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                   x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    """Delete an unused concept; historical expense relations remain protected."""
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Gastos y pagos puede eliminar conceptos.")
    rows = _base_query(ctx, "gas_lp_expense_concepts").eq("id", concept_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Concepto no encontrado.")
    usages = []
    for table, label in (
        ("gas_lp_expense_invoices", "gastos o facturas"),
        ("gas_lp_expense_vouchers", "vales"),
        ("gas_lp_expense_advances", "anticipos"),
    ):
        found = _base_query(ctx, table).select("id").eq("concept_id", concept_id).limit(1).execute().data or []
        if found:
            usages.append(label)
    if usages:
        raise HTTPException(409, "No se puede eliminar porque este concepto está usado en " + ", ".join(usages) + ".")
    ctx["sb"].table("gas_lp_expense_concepts").delete().eq("tenant_id", ctx["tenant_id"]).eq(
        "profile_id", ctx["perfil_id"]
    ).eq("id", concept_id).execute()
    _audit(ctx, "concept", concept_id, "deleted", before=rows[0])
    return {"ok": True, "deleted": True, "id": concept_id}


@router.get("/gastos/suppliers")
def list_suppliers(limit: int = Query(default=300, ge=1, le=500),
                   token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    return {"items": _base_query(ctx, "gas_lp_expense_suppliers").order("commercial_name").limit(limit).execute().data or []}


@router.post("/gastos/suppliers", status_code=201)
def create_supplier(payload: SupplierCreate, token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    clean_rfc, clean_email = _validate_supplier_fields(payload.rfc, payload.payment_email)
    row = {
        "tenant_id": ctx["tenant_id"], "profile_id": ctx["perfil_id"],
        "commercial_name": payload.commercial_name.strip(), "normalized_name": _normalize(payload.commercial_name),
        "legal_name": payload.legal_name.strip(),
        "rfc": clean_rfc, "bank_name": payload.bank_name.strip(),
        "account_number": _clean_account_number(payload.account_number), "payment_email": clean_email,
        "validation_status": "pending" if ctx["is_manager"] else "validated",
        "created_by_type": "manager" if ctx["is_manager"] else "admin", "created_by": ctx["actor_id"],
        "validated_by": None if ctx["is_manager"] else ctx["actor_id"],
        "validated_at": None if ctx["is_manager"] else _now(),
    }
    created = ctx["sb"].table("gas_lp_expense_suppliers").insert(row).execute().data[0]
    _audit(ctx, "supplier", int(created["id"]), "created", after=created)
    return {"item": created}


@router.post("/gastos/suppliers/{supplier_id}/review")
def review_supplier(supplier_id: int, payload: SupplierReview, token: str = Query(default=""),
                    authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo la asistente de gastos puede validar proveedores.")
    rows = _base_query(ctx, "gas_lp_expense_suppliers").eq("id", supplier_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Proveedor no encontrado.")
    before = rows[0]
    update = {
        "validation_status": "validated" if payload.action == "validate" else "rejected",
        "validated_by": ctx["actor_id"], "validated_at": _now(),
        "rejection_reason": payload.reason.strip() if payload.action == "reject" else "", "updated_at": _now(),
    }
    ctx["sb"].table("gas_lp_expense_suppliers").update(update).eq("id", supplier_id).execute()
    _audit(ctx, "supplier", supplier_id, payload.action, before=before, after=update)
    return {"ok": True}


@router.put("/gastos/suppliers/{supplier_id}")
def update_supplier(supplier_id: int, payload: SupplierUpdate, token: str = Query(default=""),
                    authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    rows = _base_query(ctx, "gas_lp_expense_suppliers").eq("id", supplier_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Proveedor no encontrado.")
    before = rows[0]
    if ctx["is_manager"] and (
        before.get("created_by_type") != "manager" or str(before.get("created_by")) != ctx["actor_id"]
        or before.get("validation_status") not in {"pending", "rejected"}
    ):
        raise HTTPException(403, "El gerente solo puede corregir proveedores propios pendientes o rechazados.")
    clean_rfc, clean_email = _validate_supplier_fields(payload.rfc, payload.payment_email)
    update = {
        "commercial_name": payload.commercial_name.strip(),
        "normalized_name": _normalize(payload.commercial_name), "legal_name": payload.legal_name.strip(),
        "rfc": clean_rfc, "bank_name": payload.bank_name.strip(),
        "account_number": _clean_account_number(payload.account_number),
        "payment_email": clean_email, "status": payload.status,
        "updated_at": _now(),
    }
    if ctx["is_manager"]:
        update.update({"validation_status": "pending", "rejection_reason": ""})
    ctx["sb"].table("gas_lp_expense_suppliers").update(update).eq("id", supplier_id).execute()
    _audit(ctx, "supplier", supplier_id, "updated", before=before, after=update)
    return {"ok": True, "item": {**before, **update}}


@router.get("/gastos/reimbursement-recipients")
def list_reimbursement_recipients(limit: int = Query(default=300, ge=1, le=500),
                                  token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    items = _base_query(ctx, "gas_lp_expense_recipients").order("name").limit(limit).execute().data or []
    recipient_ids = [int(row["id"]) for row in items]
    accounts = (_base_query(ctx, "gas_lp_expense_recipient_accounts")
                .in_("recipient_id", recipient_ids).order("account_type").execute().data or []) if recipient_ids else []
    by_recipient: defaultdict[int, list[dict[str, Any]]] = defaultdict(list)
    for account in accounts:
        by_recipient[int(account["recipient_id"])].append(account)
    for item in items:
        item["accounts"] = by_recipient.get(int(item["id"]), [])
    return {"items": items}


@router.post("/gastos/reimbursement-recipients", status_code=201)
def create_reimbursement_recipient(payload: ReimbursementRecipientCreate, token: str = Query(default=""),
                                   authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    _, clean_email = _validate_supplier_fields("", payload.email)
    row = {
        "tenant_id": ctx["tenant_id"], "profile_id": ctx["perfil_id"],
        "name": payload.name.strip(), "normalized_name": _normalize(payload.name),
        "email": clean_email, "created_by": ctx["actor_id"],
    }
    created = ctx["sb"].table("gas_lp_expense_recipients").insert(row).execute().data[0]
    account_rows = _recipient_account_rows(ctx, int(created["id"]), payload.accounts)
    if account_rows:
        created["accounts"] = ctx["sb"].table("gas_lp_expense_recipient_accounts").insert(account_rows).execute().data or []
    _audit(ctx, "reimbursement_recipient", int(created["id"]), "created", after=created)
    return {"item": created}


@router.put("/gastos/reimbursement-recipients/{recipient_id}")
def update_reimbursement_recipient(recipient_id: int, payload: ReimbursementRecipientUpdate,
                                   token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Supervisión de gastos administra personas a reembolsar.")
    rows = _base_query(ctx, "gas_lp_expense_recipients").eq("id", recipient_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Persona no encontrada.")
    _, clean_email = _validate_supplier_fields("", payload.email)
    update = {
        "name": payload.name.strip(), "normalized_name": _normalize(payload.name), "email": clean_email,
        "account_holder": payload.name.strip(), "status": payload.status, "updated_at": _now(),
    }
    ctx["sb"].table("gas_lp_expense_recipients").update(update).eq("tenant_id", ctx["tenant_id"]).eq(
        "profile_id", ctx["perfil_id"]
    ).eq("id", recipient_id).execute()
    ctx["sb"].table("gas_lp_expense_recipient_accounts").update({"status": "inactive", "updated_at": _now()}).eq(
        "tenant_id", ctx["tenant_id"]
    ).eq("profile_id", ctx["perfil_id"]).eq("recipient_id", recipient_id).execute()
    account_rows = _recipient_account_rows(ctx, recipient_id, payload.accounts)
    if account_rows:
        ctx["sb"].table("gas_lp_expense_recipient_accounts").upsert(
            account_rows, on_conflict="tenant_id,profile_id,recipient_id,account_type,label"
        ).execute()
    _audit(ctx, "reimbursement_recipient", recipient_id, "updated", before=rows[0], after=update)
    return {"ok": True, "item": {**rows[0], **update}}


@router.get("/gastos/expense-zones")
def list_expense_zones(token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    return {"items": _base_query(ctx, "gas_lp_expense_zones").order("name").execute().data or []}


@router.post("/gastos/expense-zones", status_code=201)
def create_expense_zone(payload: ExpenseZoneCreate, token: str = Query(default=""),
                        authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Supervisión de gastos puede agregar zonas internas.")
    normalized = _normalize(payload.name)
    duplicate = (_base_query(ctx, "gas_lp_expense_zones").eq("normalized_name", normalized)
                 .eq("status", "active").limit(1).execute().data or [])
    if duplicate:
        raise HTTPException(409, "Esta zona de gastos ya existe.")
    row = {"tenant_id": ctx["tenant_id"], "profile_id": ctx["perfil_id"],
           "name": payload.name.strip(), "normalized_name": normalized, "created_by": ctx["actor_id"]}
    created = ctx["sb"].table("gas_lp_expense_zones").insert(row).execute().data[0]
    _audit(ctx, "expense_zone", int(created["id"]), "created", after=created)
    return {"item": created}


@router.put("/gastos/expense-zones/{zone_id}")
def update_expense_zone(zone_id: int, payload: ExpenseZoneUpdate, token: str = Query(default=""),
                        authorization: str = Header(default=""),
                        x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                        x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Supervisión de gastos administra las zonas.")
    rows = _base_query(ctx, "gas_lp_expense_zones").eq("id", zone_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Zona no encontrada.")
    normalized = _normalize(payload.name)
    duplicate = (_base_query(ctx, "gas_lp_expense_zones").eq("normalized_name", normalized)
                 .neq("id", zone_id).eq("status", "active").limit(1).execute().data or [])
    if duplicate:
        raise HTTPException(409, "Esta zona de gastos ya existe.")
    update = {"name": payload.name.strip(), "normalized_name": normalized,
              "status": payload.status, "updated_at": _now()}
    ctx["sb"].table("gas_lp_expense_zones").update(update).eq("tenant_id", ctx["tenant_id"]).eq(
        "profile_id", ctx["perfil_id"]
    ).eq("id", zone_id).execute()
    _audit(ctx, "expense_zone", zone_id, "updated", before=rows[0], after=update)
    return {"ok": True, "item": {**rows[0], **update}}


@router.get("/gastos/vouchers")
def list_vouchers(status: str = Query(default=""), search: str = Query(default="", max_length=80),
                  limit: int = Query(default=200, ge=1, le=500), token: str = Query(default=""),
                  authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    query = _base_query(ctx, "gas_lp_expense_vouchers")
    if ctx["is_manager"]:
        query = query.eq("created_by_internal_user_id", int(ctx["actor_id"]))
    if status:
        query = query.eq("status", status)
    if search.strip():
        query = query.ilike("folio", f"%{search.strip()}%")
    return {"items": query.order("created_at", desc=True).limit(limit).execute().data or []}


@router.post("/gastos/vouchers", status_code=201)
def create_voucher(payload: VoucherCreate, token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_manager"]:
        raise HTTPException(403, "Los vales los genera un gerente de zona.")
    _allowed_group(ctx, payload.group_id)
    expense_zone = (_base_query(ctx, "gas_lp_expense_zones").eq("id", payload.expense_zone_id)
                    .eq("status", "active").limit(1).execute().data or [])
    if not expense_zone:
        raise HTTPException(400, "Selecciona una zona activa de Gastos.")
    supplier = (_base_query(ctx, "gas_lp_expense_suppliers").eq("id", payload.supplier_id)
                .eq("validation_status", "validated").eq("status", "active").limit(1).execute().data or [])
    concept = (_base_query(ctx, "gas_lp_expense_concepts").eq("id", payload.concept_id)
               .eq("status", "active").limit(1).execute().data or [])
    if not supplier:
        raise HTTPException(409, "Selecciona un proveedor validado.")
    if not concept:
        raise HTTPException(409, "Selecciona un concepto activo del catálogo.")
    groups = (ctx["sb"].table("fleet_groups").select("id,name").eq("tenant_id", ctx["tenant_id"])
              .eq("id", payload.group_id).limit(1).execute().data or [])
    vehicle = (ctx["sb"].table("fleet_vehicles").select("id,current_driver_name").eq("tenant_id", ctx["tenant_id"])
               .eq("id", payload.vehicle_id).limit(1).execute().data or [])
    if not groups or not vehicle:
        raise HTTPException(400, "Zona o unidad no disponible.")
    membership = (ctx["sb"].table("fleet_vehicle_groups").select("vehicle_id")
                  .eq("tenant_id", ctx["tenant_id"]).eq("group_id", payload.group_id)
                  .eq("vehicle_id", payload.vehicle_id).limit(1).execute().data or [])
    if not membership:
        raise HTTPException(400, "La unidad no pertenece a la zona seleccionada.")
    profile = _profile(ctx)
    folio_result = ctx["sb"].rpc("next_gas_lp_expense_voucher_folio", {
        "p_tenant_id": ctx["tenant_id"], "p_profile_id": ctx["perfil_id"],
        "p_group_id": payload.group_id, "p_year": payload.issued_on.year,
        "p_company_code": _code(profile.get("nombre") or profile.get("rfc"), 1),
        "p_zone_code": _code(groups[0]["name"], 3),
    }).execute().data
    folio = str(folio_result)
    row = {
        "tenant_id": ctx["tenant_id"], "profile_id": ctx["perfil_id"], "group_id": payload.group_id,
        "expense_zone_id": payload.expense_zone_id,
        "vehicle_id": payload.vehicle_id, "supplier_id": payload.supplier_id, "concept_id": payload.concept_id,
        "folio": folio, "issued_on": payload.issued_on.isoformat(), "description": payload.description.strip(),
        "driver_name": payload.driver_name.strip() or vehicle[0].get("current_driver_name") or "",
        "created_by_internal_user_id": int(ctx["actor_id"]), "created_by_name": ctx["actor_name"],
    }
    created = ctx["sb"].table("gas_lp_expense_vouchers").insert(row).execute().data[0]
    _audit(ctx, "voucher", int(created["id"]), "confirmed", after=created)
    return {"item": created}


def _manager_voucher(ctx: dict[str, Any], voucher_id: int) -> dict[str, Any]:
    rows = _base_query(ctx, "gas_lp_expense_vouchers").eq("id", voucher_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Vale no encontrado.")
    row = rows[0]
    if ctx["is_manager"] and int(row["created_by_internal_user_id"]) != int(ctx["actor_id"]):
        raise HTTPException(403, "Este vale pertenece a otro gerente.")
    return row


@router.put("/gastos/vouchers/{voucher_id}/amount")
def set_voucher_amount(voucher_id: int, payload: VoucherAmount, token: str = Query(default=""),
                       authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    row = _manager_voucher(ctx, voucher_id)
    if row["status"] not in {"amount_pending", "ready_to_invoice"}:
        raise HTTPException(409, "El monto ya no puede modificarse.")
    update = {"amount_mxn": round(payload.amount_mxn, 2), "status": "ready_to_invoice", "updated_at": _now()}
    ctx["sb"].table("gas_lp_expense_vouchers").update(update).eq("id", voucher_id).execute()
    _audit(ctx, "voucher", voucher_id, "amount_updated", before={"amount_mxn": row.get("amount_mxn")}, after=update)
    return {"ok": True, "item": {**row, **update}}


@router.post("/gastos/vouchers/{voucher_id}/cancel")
def cancel_voucher(voucher_id: int, payload: CancelPayload, token: str = Query(default=""),
                   authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    row = _manager_voucher(ctx, voucher_id)
    if row["status"] == "invoiced":
        raise HTTPException(409, "Un vale facturado no puede cancelarse.")
    if row["status"] == "cancelled":
        raise HTTPException(409, "El vale ya está cancelado.")
    update = {"status": "cancelled", "cancelled_at": _now(),
              "cancellation_reason": payload.reason.strip(), "updated_at": _now()}
    ctx["sb"].table("gas_lp_expense_vouchers").update(update).eq("id", voucher_id).execute()
    _audit(ctx, "voucher", voucher_id, "cancelled", before=row, after=update)
    return {"ok": True}


@router.get("/gastos/vouchers/{voucher_id}/print")
def print_voucher(voucher_id: int, token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    row = _manager_voucher(ctx, voucher_id)
    profile = _profile(ctx)
    suppliers = (_base_query(ctx, "gas_lp_expense_suppliers").eq("id", row["supplier_id"]).limit(1).execute().data or [{}])
    concepts = (_base_query(ctx, "gas_lp_expense_concepts").eq("id", row["concept_id"]).limit(1).execute().data or [{}])
    vehicles = (ctx["sb"].table("fleet_vehicles").select("vehicle_number").eq("tenant_id", ctx["tenant_id"])
                .eq("id", row["vehicle_id"]).limit(1).execute().data or [{}])
    expense_zones = (_base_query(ctx, "gas_lp_expense_zones")
                     .eq("id", row.get("expense_zone_id") or 0).limit(1).execute().data or [{}])
    output = BytesIO()
    width, height = letter[0] / 2, letter[1] / 2
    pdf = canvas.Canvas(output, pagesize=(width, height))
    pdf.setFillColor(HexColor("#7A1E2C")); pdf.rect(0, height - 58, width, 58, fill=1, stroke=0)
    logo = None
    try:
        from routes.settings import _load as load_settings
        data_url = str(load_settings(str(profile.get("user_id") or ""), int(profile["id"])).get("PdfLogoDataUrl") or "")
        if data_url.startswith("data:image/") and "," in data_url:
            logo = ImageReader(BytesIO(base64.b64decode(data_url.split(",", 1)[1])))
    except Exception:
        logo = None
    if logo is None:
        fallback = os.path.join(os.path.dirname(__file__), "..", "static", "img", "ge-control-logo-light.png")
        if os.path.exists(fallback):
            logo = ImageReader(fallback)
    if logo is not None:
        pdf.drawImage(logo, 18, height - 48, width=82, height=32, preserveAspectRatio=True, mask="auto")
    pdf.setFillColorRGB(1, 1, 1); pdf.setFont("Helvetica-Bold", 12)
    pdf.drawRightString(width - 18, height - 35, str(profile.get("nombre") or "GE CONTROL")[:30])
    pdf.setFillColorRGB(0.1, 0.1, 0.1); pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(18, height - 84, f"VALE {row['folio']}")
    pdf.setFont("Helvetica", 9)
    lines = [
        ("Fecha", str(row["issued_on"])), ("Zona", expense_zones[0].get("name") or "General de la empresa"),
        ("Proveedor", suppliers[0].get("commercial_name") or "—"),
        ("Concepto", concepts[0].get("name") or "—"), ("Unidad", vehicles[0].get("vehicle_number") or "—"),
        ("Chofer", row.get("driver_name") or "—"), ("Gerente", row.get("created_by_name") or "—"),
        ("Monto", "PENDIENTE" if row.get("amount_mxn") is None else f"${float(row['amount_mxn']):,.2f} MXN"),
    ]
    y = height - 110
    for label, value in lines:
        pdf.setFont("Helvetica-Bold", 8); pdf.drawString(18, y, f"{label}:")
        pdf.setFont("Helvetica", 8); pdf.drawString(78, y, str(value)[:42]); y -= 20
    pdf.setFont("Helvetica", 7); pdf.setFillColor(HexColor("#666666"))
    pdf.drawString(18, 18, "Original digital en GE Control · El folio no puede reutilizarse.")
    pdf.save(); output.seek(0)
    return StreamingResponse(output, media_type="application/pdf",
                             headers={"Content-Disposition": f'inline; filename="vale-{row["folio"]}.pdf"'})


@router.post("/gastos/advances", status_code=201)
def create_supplier_advance(payload: SupplierAdvanceCreate, token: str = Query(default=""),
                            authorization: str = Header(default=""),
                            x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                            x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Gastos y pagos puede registrar anticipos.")
    suppliers = (_base_query(ctx, "gas_lp_expense_suppliers").eq("id", payload.supplier_id)
                 .eq("validation_status", "validated").eq("status", "active").limit(1).execute().data or [])
    if not suppliers:
        raise HTTPException(400, "Proveedor no disponible en esta empresa.")
    if payload.concept_id:
        concepts = (_base_query(ctx, "gas_lp_expense_concepts").eq("id", payload.concept_id)
                    .eq("status", "active").limit(1).execute().data or [])
        if not concepts:
            raise HTTPException(400, "Concepto no disponible en esta empresa.")
    if payload.expense_zone_id and not any(
        int(row.get("id") or 0) == payload.expense_zone_id for row in _expense_zones(ctx)
    ):
        raise HTTPException(400, "La zona interna seleccionada no pertenece a esta empresa.")
    row = {
        "tenant_id": ctx["tenant_id"], "profile_id": ctx["perfil_id"],
        "supplier_id": payload.supplier_id, "concept_id": payload.concept_id,
        "expense_zone_id": payload.expense_zone_id, "paid_on": payload.paid_on.isoformat(),
        "amount_mxn": round(payload.amount_mxn, 2), "reference": payload.reference.strip(),
        "description": payload.description.strip(), "status": "pending", "created_by": ctx["actor_id"],
    }
    created = ctx["sb"].table("gas_lp_expense_advances").insert(row).execute().data[0]
    _audit(ctx, "advance", int(created["id"]), "created", after=created)
    return {"item": created}


@router.get("/gastos/advances")
def list_supplier_advances(status: str = Query(default="open"), limit: int = Query(default=300, ge=1, le=500),
                           token: str = Query(default=""), authorization: str = Header(default=""),
                           x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                           x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    query = _base_query(ctx, "gas_lp_expense_advances")
    if status == "open":
        query = query.in_("status", ["pending", "partial"])
    elif status:
        query = query.eq("status", status)
    items = query.order("paid_on", desc=True).order("created_at", desc=True).limit(limit).execute().data or []
    advance_ids = [int(row["id"]) for row in items]
    applications = (ctx["sb"].table("gas_lp_expense_advance_applications")
                    .select("advance_id,invoice_id,amount_mxn").in_("advance_id", advance_ids)
                    .execute().data or []) if advance_ids else []
    applied: defaultdict[int, float] = defaultdict(float)
    for application in applications:
        applied[int(application["advance_id"])] += float(application.get("amount_mxn") or 0)
    for item in items:
        used = round(applied[int(item["id"])], 2)
        item["applied_amount_mxn"] = used
        item["available_amount_mxn"] = round(max(0, float(item.get("amount_mxn") or 0) - used), 2)
    return {"items": items}


@router.post("/gastos/advances/apply", status_code=201)
def apply_supplier_advances(payload: SupplierAdvanceApply, token: str = Query(default=""),
                            authorization: str = Header(default=""),
                            x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                            x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Gastos y pagos puede relacionar anticipos.")
    allocation_ids = [int(row.advance_id) for row in payload.allocations]
    if len(set(allocation_ids)) != len(allocation_ids):
        raise HTTPException(400, "Un anticipo no puede repetirse en la misma factura.")
    suppliers = (_base_query(ctx, "gas_lp_expense_suppliers").eq("id", payload.supplier_id)
                 .eq("validation_status", "validated").eq("status", "active").limit(1).execute().data or [])
    if not suppliers:
        raise HTTPException(400, "Proveedor no disponible en esta empresa.")
    alerts = _invoice_alerts(
        ctx, supplier_id=payload.supplier_id, invoice_number=payload.invoice_number,
        invoice_date=payload.invoice_date, total_mxn=payload.total_mxn,
    )
    if alerts:
        raise HTTPException(409, " ".join(alerts))
    result = ctx["sb"].rpc("apply_gas_lp_expense_advances", {
        "p_tenant_id": ctx["tenant_id"], "p_profile_id": ctx["perfil_id"],
        "p_supplier_id": payload.supplier_id, "p_concept_id": payload.concept_id,
        "p_expense_zone_id": payload.expense_zone_id, "p_invoice_number": payload.invoice_number.strip(),
        "p_invoice_date": payload.invoice_date.isoformat(), "p_total_mxn": round(payload.total_mxn, 2),
        "p_description": payload.description.strip(), "p_created_by": ctx["actor_id"],
        "p_advance_ids": allocation_ids,
        "p_amounts": [round(float(row.amount_mxn), 2) for row in payload.allocations],
    }).execute().data
    invoice_id = int(result)
    invoice = (_base_query(ctx, "gas_lp_expense_invoices").eq("id", invoice_id).limit(1).execute().data or [{}])[0]
    _audit(ctx, "invoice", invoice_id, "created_from_advances", after={
        **invoice, "advance_allocations": [row.model_dump() for row in payload.allocations],
    })
    return {"item": invoice}


@router.delete("/gastos/advances/{advance_id}")
def delete_supplier_advance(advance_id: int, token: str = Query(default=""),
                            authorization: str = Header(default=""),
                            x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                            x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    rows = _base_query(ctx, "gas_lp_expense_advances").eq("id", advance_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Anticipo no encontrado.")
    row = rows[0]
    links = (ctx["sb"].table("gas_lp_expense_advance_applications").select("invoice_id")
             .eq("advance_id", advance_id).limit(1).execute().data or [])
    if links:
        raise HTTPException(409, "Este anticipo ya está relacionado con una factura y no se puede eliminar.")
    update = {"status": "cancelled", "cancelled_by": ctx["actor_id"], "cancelled_at": _now(),
              "cancellation_reason": "Anticipo eliminado por error.", "updated_at": _now()}
    ctx["sb"].table("gas_lp_expense_advances").update(update).eq("tenant_id", ctx["tenant_id"]).eq(
        "profile_id", ctx["perfil_id"]
    ).eq("id", advance_id).execute()
    _audit(ctx, "advance", advance_id, "cancelled", before=row, after=update)
    return {"ok": True, "deleted": True}


@router.post("/gastos/invoices/from-vouchers", status_code=201)
def create_invoice_from_vouchers(payload: VoucherInvoiceCreate, token: str = Query(default=""),
                                 authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    rows_query = _base_query(ctx, "gas_lp_expense_vouchers").in_("id", payload.voucher_ids)
    if ctx["is_manager"]:
        rows_query = rows_query.eq("created_by_internal_user_id", int(ctx["actor_id"]))
    rows = rows_query.execute().data or []
    if len(rows) != len(set(payload.voucher_ids)):
        raise HTTPException(400, "Uno o más vales no están disponibles.")
    if any(row["status"] != "ready_to_invoice" for row in rows):
        raise HTTPException(409, "Todos los vales deben estar listos para facturar.")
    if len({int(row["supplier_id"]) for row in rows}) != 1:
        raise HTTPException(400, "Una factura solo puede agrupar vales del mismo proveedor.")
    if len({int(row["group_id"]) for row in rows}) != 1:
        raise HTTPException(400, "Una factura solo puede agrupar vales de la misma zona.")
    if len({int(row["created_by_internal_user_id"]) for row in rows}) != 1:
        raise HTTPException(400, "Una factura solo puede agrupar vales capturados por el mismo gerente.")
    total = round(sum(float(row["amount_mxn"]) for row in rows), 2)
    if abs(total - round(payload.total_mxn, 2)) > MONEY_TOLERANCE:
        raise HTTPException(400, f"El total debe coincidir con la suma de los vales: ${total:,.2f}.")
    alerts = _invoice_alerts(
        ctx, supplier_id=int(rows[0]["supplier_id"]), invoice_number=payload.invoice_number,
        invoice_date=payload.invoice_date, total_mxn=total,
    )
    invoice_id = ctx["sb"].rpc("create_gas_lp_voucher_invoice", {
        "p_tenant_id": ctx["tenant_id"], "p_profile_id": ctx["perfil_id"],
        "p_manager_id": int(rows[0]["created_by_internal_user_id"]), "p_voucher_ids": payload.voucher_ids,
        "p_invoice_number": payload.invoice_number.strip(),
        "p_invoice_date": payload.invoice_date.isoformat(), "p_total_mxn": total,
    }).execute().data
    invoice_rows = _base_query(ctx, "gas_lp_expense_invoices").eq("id", invoice_id).limit(1).execute().data or []
    if not invoice_rows:
        raise HTTPException(500, "La factura se registró, pero no pudo recuperarse.")
    invoice = invoice_rows[0]
    if alerts:
        ctx["sb"].table("gas_lp_expense_invoices").update({
            "observation": "Alerta: " + " ".join(alerts), "updated_at": _now(),
        }).eq("id", invoice_id).execute()
        invoice["observation"] = "Alerta: " + " ".join(alerts)
    _audit(ctx, "invoice", int(invoice["id"]), "submitted", after=invoice)
    return {"item": invoice, "alerts": alerts}


@router.post("/gastos/invoices/direct", status_code=201)
def create_direct_invoice(payload: DirectInvoiceCreate, token: str = Query(default=""),
                          authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Gastos y pagos puede registrar gastos directos.")
    if payload.expense_type == "credit_note" and payload.payment_target != "supplier":
        raise HTTPException(400, "Una nota de crédito debe quedar asociada al proveedor.")
    if payload.payment_target == "reimbursement" and not payload.reimbursement_recipient_id:
        raise HTTPException(400, "Selecciona a la persona que recibirá el reembolso.")
    if payload.payment_target == "reimbursement" and not payload.reimbursement_account_id:
        raise HTTPException(400, "Selecciona si el reembolso será a nómina o tarjeta de crédito.")
    if payload.payment_target == "supplier" and (payload.reimbursement_recipient_id or payload.reimbursement_account_id):
        raise HTTPException(400, "Un pago al proveedor no debe incluir una persona a reembolsar.")
    selected_zone_count = sum(value is not None for value in (payload.group_id, payload.facility_id, payload.expense_zone_id))
    if selected_zone_count > 1:
        raise HTTPException(400, "Selecciona una sola zona o centro de costo.")
    if payload.group_id and not any(int(row["id"]) == payload.group_id for row in _profile_expense_groups(ctx)):
        raise HTTPException(400, "La zona de Motive seleccionada no pertenece a esta empresa.")
    if payload.facility_id and not any(int(row.get("id") or 0) == payload.facility_id for row in _profile_facilities(ctx)):
        raise HTTPException(400, "La zona seleccionada no pertenece a esta empresa.")
    if payload.expense_zone_id and not any(int(row.get("id") or 0) == payload.expense_zone_id for row in _expense_zones(ctx)):
        raise HTTPException(400, "La zona interna seleccionada no pertenece a esta empresa.")
    if payload.reimbursement_recipient_id:
        recipients = (_base_query(ctx, "gas_lp_expense_recipients")
                      .eq("id", payload.reimbursement_recipient_id).eq("status", "active").limit(1).execute().data or [])
        if not recipients:
            raise HTTPException(400, "La persona a reembolsar no está disponible.")
        accounts = (_base_query(ctx, "gas_lp_expense_recipient_accounts")
                    .eq("id", payload.reimbursement_account_id).eq("recipient_id", payload.reimbursement_recipient_id)
                    .eq("status", "active").limit(1).execute().data or [])
        if not accounts:
            raise HTTPException(400, "El destino de reembolso no pertenece a esta persona.")
    supplier = (_base_query(ctx, "gas_lp_expense_suppliers").eq("id", payload.supplier_id)
                .eq("validation_status", "validated").eq("status", "active").limit(1).execute().data or [])
    if not supplier:
        raise HTTPException(400, "Proveedor no disponible en esta empresa.")
    concept = (_base_query(ctx, "gas_lp_expense_concepts").eq("id", payload.concept_id)
               .eq("status", "active").limit(1).execute().data or [])
    if not concept:
        raise HTTPException(400, "Concepto no disponible en esta empresa.")
    alerts = _invoice_alerts(
        ctx, supplier_id=payload.supplier_id, invoice_number=payload.invoice_number,
        invoice_date=payload.invoice_date, total_mxn=payload.total_mxn,
    )
    row = {
        "tenant_id": ctx["tenant_id"], "profile_id": ctx["perfil_id"], "supplier_id": payload.supplier_id,
        "expense_type": payload.expense_type, "invoice_number": payload.invoice_number.strip(),
        "invoice_date": payload.invoice_date.isoformat(), "total_mxn": round(payload.total_mxn, 2),
        "period_key": payload.period_key or None, "description": payload.description.strip(),
        "concept_id": payload.concept_id, "group_id": payload.group_id, "facility_id": payload.facility_id,
        "expense_zone_id": payload.expense_zone_id,
        "payment_target": payload.payment_target,
        "reimbursement_recipient_id": payload.reimbursement_recipient_id,
        "reimbursement_account_id": payload.reimbursement_account_id,
        "created_by_type": "admin", "created_by": ctx["actor_id"],
        "status": "sent_to_accountant", "sent_to_accountant_at": _now(),
        "observation": "Alerta: " + " ".join(alerts) if alerts else "",
    }
    created_rows = _insert_expense_invoices(ctx, row)
    if not created_rows:
        raise HTTPException(500, "No se pudo confirmar la captura.")
    created = created_rows[0]
    _audit(ctx, "invoice", int(created["id"]), "direct_created", after=created)
    return {"item": created, "alerts": alerts}


@router.post("/gastos/invoices/direct/batch", status_code=201)
def create_direct_invoice_batch(payload: DirectInvoiceBatchCreate, token: str = Query(default=""),
                                authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    """Create several invoices atomically with one shared supplier and accounting context."""
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Gastos y pagos puede registrar gastos directos.")
    if payload.payment_target == "reimbursement" and not payload.reimbursement_recipient_id:
        raise HTTPException(400, "Selecciona a la persona que recibirá el reembolso.")
    if payload.payment_target == "reimbursement" and not payload.reimbursement_account_id:
        raise HTTPException(400, "Selecciona si el reembolso será a nómina o tarjeta de crédito.")
    if payload.payment_target == "supplier" and (payload.reimbursement_recipient_id or payload.reimbursement_account_id):
        raise HTTPException(400, "Un pago al proveedor no debe incluir una persona a reembolsar.")
    selected_zone_count = sum(value is not None for value in (payload.group_id, payload.facility_id, payload.expense_zone_id))
    if selected_zone_count > 1:
        raise HTTPException(400, "Selecciona una sola zona o centro de costo.")
    if payload.group_id and not any(int(row["id"]) == payload.group_id for row in _profile_expense_groups(ctx)):
        raise HTTPException(400, "La zona seleccionada no pertenece a esta empresa.")
    if payload.facility_id and not any(int(row.get("id") or 0) == payload.facility_id for row in _profile_facilities(ctx)):
        raise HTTPException(400, "La zona seleccionada no pertenece a esta empresa.")
    if payload.expense_zone_id and not any(int(row.get("id") or 0) == payload.expense_zone_id for row in _expense_zones(ctx)):
        raise HTTPException(400, "La zona interna seleccionada no pertenece a esta empresa.")
    if payload.reimbursement_recipient_id:
        recipients = (_base_query(ctx, "gas_lp_expense_recipients")
                      .eq("id", payload.reimbursement_recipient_id).eq("status", "active").limit(1).execute().data or [])
        if not recipients:
            raise HTTPException(400, "La persona a reembolsar no está disponible.")
        accounts = (_base_query(ctx, "gas_lp_expense_recipient_accounts")
                    .eq("id", payload.reimbursement_account_id).eq("recipient_id", payload.reimbursement_recipient_id)
                    .eq("status", "active").limit(1).execute().data or [])
        if not accounts:
            raise HTTPException(400, "El destino de reembolso no pertenece a esta persona.")
    supplier = (_base_query(ctx, "gas_lp_expense_suppliers").eq("id", payload.supplier_id)
                .eq("validation_status", "validated").eq("status", "active").limit(1).execute().data or [])
    if not supplier:
        raise HTTPException(400, "Proveedor no disponible en esta empresa.")
    concept = (_base_query(ctx, "gas_lp_expense_concepts").eq("id", payload.concept_id)
               .eq("status", "active").limit(1).execute().data or [])
    if not concept:
        raise HTTPException(400, "Concepto no disponible en esta empresa.")

    normalized_numbers = [
        line.invoice_number.strip().casefold()
        for line in payload.invoices
        if not _is_without_folio(line.invoice_number)
    ]
    if len(set(normalized_numbers)) != len(normalized_numbers):
        raise HTTPException(400, "Hay folios repetidos dentro de la captura múltiple.")

    created_at = _now()
    rows: list[dict[str, Any]] = []
    alerts_by_invoice: list[dict[str, Any]] = []
    skipped_duplicates: list[dict[str, Any]] = []
    for line in payload.invoices:
        alerts = _invoice_alerts(
            ctx, supplier_id=payload.supplier_id, invoice_number=line.invoice_number,
            invoice_date=line.invoice_date, total_mxn=line.total_mxn,
        )
        if alerts:
            skipped_duplicates.append({"invoice_number": line.invoice_number.strip(), "alerts": alerts})
            alerts_by_invoice.append({"invoice_number": line.invoice_number.strip(), "alerts": alerts, "skipped": True})
            continue
        rows.append({
            "tenant_id": ctx["tenant_id"], "profile_id": ctx["perfil_id"], "supplier_id": payload.supplier_id,
            "expense_type": "direct", "invoice_number": line.invoice_number.strip(),
            "invoice_date": line.invoice_date.isoformat(), "total_mxn": round(line.total_mxn, 2),
            "period_key": payload.period_key or None, "description": payload.description.strip(),
            "concept_id": payload.concept_id, "group_id": payload.group_id, "facility_id": payload.facility_id,
            "expense_zone_id": payload.expense_zone_id, "payment_target": payload.payment_target,
            "reimbursement_recipient_id": payload.reimbursement_recipient_id,
            "reimbursement_account_id": payload.reimbursement_account_id,
            "created_by_type": "admin", "created_by": ctx["actor_id"],
            "status": "sent_to_accountant", "sent_to_accountant_at": created_at,
            "observation": "Alerta: " + " ".join(alerts) if alerts else "",
        })
        alerts_by_invoice.append({"invoice_number": line.invoice_number.strip(), "alerts": alerts})
    if not rows:
        return {"items": [], "count": 0, "alerts": alerts_by_invoice, "skipped_duplicates": skipped_duplicates}
    # PostgREST executes a multi-row insert as one SQL statement: all rows are
    # persisted together or none are persisted.
    created = _insert_expense_invoices(ctx, rows)
    if len(created) != len(rows):
        raise HTTPException(500, "No se pudo confirmar la captura completa.")
    for item in created:
        _audit(ctx, "invoice", int(item["id"]), "direct_batch_created", after=item)
    return {"items": created, "count": len(created), "alerts": alerts_by_invoice, "skipped_duplicates": skipped_duplicates}


@router.get("/gastos/invoices")
def list_invoices(status: str = Query(default=""), search: str = Query(default="", max_length=100),
                  invoice_date_from: date | None = Query(default=None),
                  invoice_date_to: date | None = Query(default=None),
                  capture_date_from: date | None = Query(default=None),
                  capture_date_to: date | None = Query(default=None),
                  limit: int = Query(default=200, ge=1, le=500), token: str = Query(default=""),
                  authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    query = _base_query(ctx, "gas_lp_expense_invoices")
    if ctx["is_manager"]:
        query = query.eq("created_by_type", "manager").eq("created_by", ctx["actor_id"])
    if status:
        query = query.eq("status", status)
    if invoice_date_from:
        query = query.gte("invoice_date", invoice_date_from.isoformat())
    if invoice_date_to:
        query = query.lte("invoice_date", invoice_date_to.isoformat())
    if capture_date_from:
        query = query.gte("created_at", capture_date_from.isoformat())
    if capture_date_to:
        query = query.lt("created_at", (capture_date_to + timedelta(days=1)).isoformat())
    if search.strip():
        query = query.ilike("invoice_number", f"%{search.strip()}%")
    items = query.order("created_at", desc=True).limit(limit).execute().data or []
    invoice_ids = [int(row["id"]) for row in items]
    if not invoice_ids:
        return {"items": items}
    links = (ctx["sb"].table("gas_lp_expense_invoice_vouchers")
             .select("invoice_id,voucher_id,amount_mxn")
             .in_("invoice_id", invoice_ids).execute().data or [])
    voucher_ids = sorted({int(link["voucher_id"]) for link in links})
    voucher_rows = (
        _base_query(ctx, "gas_lp_expense_vouchers").in_("id", voucher_ids).execute().data or []
    ) if voucher_ids else []
    vouchers = {int(row["id"]): row for row in voucher_rows}
    links_by_invoice: defaultdict[int, list[dict[str, Any]]] = defaultdict(list)
    for link in links:
        voucher = vouchers.get(int(link["voucher_id"]))
        if voucher:
            links_by_invoice[int(link["invoice_id"])].append({
                "id": int(voucher["id"]), "folio": voucher.get("folio") or "",
                "amount_mxn": float(link.get("amount_mxn") or 0),
                "group_id": voucher.get("group_id"), "vehicle_id": voucher.get("vehicle_id"),
                "concept_id": voucher.get("concept_id"), "driver_name": voucher.get("driver_name") or "",
                "created_by_name": voucher.get("created_by_name") or "",
            })
    for item in items:
        item["vouchers"] = links_by_invoice.get(int(item["id"]), [])
    folio_counts: defaultdict[tuple[int, str], int] = defaultdict(int)
    for item in items:
        if item.get("status") not in {"rejected", "cancelled"} and not _is_without_folio(item.get("invoice_number")):
            folio_counts[(int(item.get("supplier_id") or 0), _normalize(item.get("invoice_number")))] += 1
    for item in items:
        key = (int(item.get("supplier_id") or 0), _normalize(item.get("invoice_number")))
        item["duplicate_invoice_number"] = not _is_without_folio(item.get("invoice_number")) and folio_counts[key] > 1
    allocations = (ctx["sb"].table("gas_lp_expense_payment_allocations")
                   .select("invoice_id,amount_mxn").in_("invoice_id", invoice_ids).execute().data or [])
    advance_applications = (ctx["sb"].table("gas_lp_expense_advance_applications")
                            .select("invoice_id,amount_mxn").in_("invoice_id", invoice_ids)
                            .execute().data or [])
    paid_by_invoice: defaultdict[int, float] = defaultdict(float)
    for allocation in [*allocations, *advance_applications]:
        paid_by_invoice[int(allocation["invoice_id"])] += float(allocation.get("amount_mxn") or 0)
    for item in items:
        paid = round(paid_by_invoice.get(int(item["id"]), float(item.get("paid_amount_mxn") or 0)), 2)
        item["applied_amount_mxn"] = paid
        raw_balance = float(item.get("total_mxn") or 0) - paid
        item["balance_mxn"] = 0.0 if abs(raw_balance) < PAYMENT_BALANCE_TOLERANCE else round(max(0, raw_balance), 2)
    return {"items": items}


@router.post("/gastos/payments", status_code=201)
def create_expense_payment(payload: ExpensePaymentCreate, token: str = Query(default=""),
                           authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Gastos y pagos puede registrar pagos.")
    allocation_map = {int(row.invoice_id): round(float(row.amount_mxn), 2) for row in payload.invoice_allocations}
    if len(allocation_map) != len(payload.invoice_allocations):
        raise HTTPException(400, "Una factura no puede repetirse dentro del mismo pago.")
    allocation_total = round(sum(allocation_map.values()), 2)
    paid_total = round(float(payload.amount_mxn), 2)
    # amount_mxn is the real bank transfer. Allocations are capped at invoice
    # balances, so a small bank overpayment remains an explicit difference.
    if round(allocation_total * 100) > round(paid_total * 100):
        raise HTTPException(400, "Lo aplicado a las facturas no puede superar el monto realmente pagado.")
    payment_difference = round(paid_total - allocation_total, 2)
    invoices = (_base_query(ctx, "gas_lp_expense_invoices").in_("id", list(allocation_map)).execute().data or [])
    if len(invoices) != len(allocation_map):
        raise HTTPException(400, "Una o más facturas no pertenecen a esta empresa.")
    if any(row.get("status") not in {"accepted", "sent_to_accountant", "paid"} for row in invoices):
        raise HTTPException(409, "Todas las facturas deben estar aceptadas o en contabilidad.")
    target_keys = {
        (row.get("payment_target") or "supplier",
         int(row.get("reimbursement_recipient_id") or row.get("supplier_id") or 0),
         int(row.get("reimbursement_account_id") or 0))
        for row in invoices
    }
    if len(target_keys) != 1:
        raise HTTPException(400, "Un pago agrupado debe corresponder al mismo proveedor o persona a reembolsar.")
    invoice_ids = list(allocation_map)
    credit_notes = []
    if payload.credit_note_ids:
        credit_notes = (_base_query(ctx, "gas_lp_expense_invoices")
                        .in_("id", payload.credit_note_ids).execute().data or [])
        if len(credit_notes) != len(set(payload.credit_note_ids)):
            raise HTTPException(400, "Una o más notas de crédito no pertenecen a esta empresa.")
        if any(row.get("expense_type") != "credit_note" or row.get("status") in {"paid", "cancelled"}
               for row in credit_notes):
            raise HTTPException(409, "Una nota de crédito ya no está disponible para aplicar.")
        credit_keys = {
            (row.get("payment_target") or "supplier",
             int(row.get("reimbursement_recipient_id") or row.get("supplier_id") or 0),
             int(row.get("reimbursement_account_id") or 0))
            for row in credit_notes
        }
        if len(credit_keys) != 1:
            raise HTTPException(400, "Las notas de crédito deben corresponder al mismo destinatario.")
        if next(iter(credit_keys)) != next(iter(target_keys)):
            raise HTTPException(400, "La nota de crédito debe corresponder al mismo destinatario del pago.")
    old_allocations = (ctx["sb"].table("gas_lp_expense_payment_allocations")
                       .select("invoice_id,amount_mxn").in_("invoice_id", invoice_ids).execute().data or [])
    old_advance_applications = (ctx["sb"].table("gas_lp_expense_advance_applications")
                                .select("invoice_id,amount_mxn").in_("invoice_id", invoice_ids)
                                .execute().data or [])
    already_paid: defaultdict[int, float] = defaultdict(float)
    for row in [*old_allocations, *old_advance_applications]:
        already_paid[int(row["invoice_id"])] += float(row.get("amount_mxn") or 0)
    credit_remaining = round(sum(float(row.get("total_mxn") or 0) for row in credit_notes), 2)
    credit_applied_by_invoice: dict[int, float] = {}
    for invoice in invoices:
        invoice_id = int(invoice["id"])
        outstanding = round(float(invoice["total_mxn"]) - already_paid[invoice_id], 2)
        if allocation_map[invoice_id] > outstanding + MONEY_TOLERANCE:
            raise HTTPException(400, f"La aplicación a {invoice['invoice_number']} supera su saldo de ${outstanding:,.2f}.")
        credit_applied = min(max(0.0, outstanding - allocation_map[invoice_id]), credit_remaining)
        credit_applied_by_invoice[invoice_id] = round(credit_applied, 2)
        credit_remaining = round(credit_remaining - credit_applied, 2)
    target, target_id, target_account_id = next(iter(target_keys))
    payment_row = {
        "tenant_id": ctx["tenant_id"], "profile_id": ctx["perfil_id"], "payment_target": target,
        "supplier_id": target_id if target == "supplier" else None,
        "reimbursement_recipient_id": target_id if target == "reimbursement" else None,
        "reimbursement_account_id": target_account_id if target == "reimbursement" else None,
        "paid_on": payload.paid_on.isoformat(), "amount_mxn": paid_total,
        "method": payload.method.strip(), "reference": payload.reference.strip(),
        "notes": payload.notes.strip(), "created_by": ctx["actor_id"],
    }
    payment = ctx["sb"].table("gas_lp_expense_payments").insert(payment_row).execute().data[0]
    ctx["sb"].table("gas_lp_expense_payment_allocations").insert([
        {"payment_id": payment["id"], "invoice_id": invoice_id, "amount_mxn": amount}
        for invoice_id, amount in allocation_map.items()
    ]).execute()
    for invoice in invoices:
        invoice_id = int(invoice["id"])
        applied = round(already_paid[invoice_id] + allocation_map[invoice_id] + credit_applied_by_invoice.get(invoice_id, 0), 2)
        complete = applied + PAYMENT_BALANCE_TOLERANCE >= float(invoice["total_mxn"])
        update = {"paid_amount_mxn": applied, "paid_on": payload.paid_on.isoformat(),
                  "status": "paid" if complete else "sent_to_accountant", "updated_at": _now()}
        if complete:
            update["paid_at"] = _now()
        ctx["sb"].table("gas_lp_expense_invoices").update(update).eq("id", invoice_id).execute()
    for credit in credit_notes:
        ctx["sb"].table("gas_lp_expense_invoices").update({
            "paid_amount_mxn": round(float(credit.get("total_mxn") or 0), 2),
            "paid_on": payload.paid_on.isoformat(), "paid_at": _now(), "status": "paid",
            "observation": "Nota de crédito aplicada en pago.", "updated_at": _now(),
        }).eq("id", int(credit["id"])).execute()
    if target == "reimbursement":
        party = (_base_query(ctx, "gas_lp_expense_recipients").eq("id", target_id).limit(1).execute().data or [{}])[0]
        party_email, party_name = party.get("email"), party.get("name") or "Persona a reembolsar"
    else:
        party = (_base_query(ctx, "gas_lp_expense_suppliers").eq("id", target_id).limit(1).execute().data or [{}])[0]
        party_email, party_name = party.get("payment_email"), party.get("commercial_name") or "Proveedor"
    delivery = send_gas_lp_expense_payment_email(
        to_email=party_email, supplier_name=party_name, company_name=_profile(ctx).get("nombre") or "",
        invoice_number=", ".join(str(row.get("invoice_number") or "") for row in invoices),
        paid_on=payload.paid_on.isoformat(), amount=payload.amount_mxn,
        invoices=[{
            "invoice_number": row.get("invoice_number") or "",
            "invoice_date": row.get("invoice_date") or "",
            "total_mxn": row.get("total_mxn") or 0,
            "amount_paid_mxn": allocation_map[int(row["id"])],
        } for row in invoices],
        idempotency_key=f"gas-lp-expense-payment-{ctx['tenant_id']}-{payment['id']}",
    )
    email_update = {"email_status": "sent" if delivery.ok else ("skipped" if delivery.skipped else "failed"),
                    "email_metadata": delivery.as_metadata()}
    ctx["sb"].table("gas_lp_expense_payments").update(email_update).eq("id", payment["id"]).execute()
    _audit(ctx, "payment", int(payment["id"]), "created",
           after={**payment, **email_update, "applied_amount_mxn": allocation_total,
                  "difference_mxn": payment_difference})
    return {"item": {**payment, **email_update, "applied_amount_mxn": allocation_total,
                     "difference_mxn": payment_difference}}


@router.get("/gastos/payments")
def list_expense_payments(limit: int = Query(default=200, ge=1, le=500), token: str = Query(default=""),
                          authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Gastos y pagos puede consultar pagos.")
    payments = (_base_query(ctx, "gas_lp_expense_payments")
                .order("paid_on", desc=True).order("created_at", desc=True)
                .limit(limit).execute().data or [])
    # Un anticipo ya es una salida de dinero. Se integra al historial de pagos
    # como movimiento de consulta, sin crear un segundo pago ni duplicar el egreso.
    advances = (_base_query(ctx, "gas_lp_expense_advances")
                .in_("status", ["pending", "partial", "applied"])
                .order("paid_on", desc=True).order("created_at", desc=True)
                .limit(limit).execute().data or [])
    advance_ids = [int(row["id"]) for row in advances]
    advance_applications = (ctx["sb"].table("gas_lp_expense_advance_applications")
                            .select("advance_id,amount_mxn").in_("advance_id", advance_ids)
                            .execute().data or []) if advance_ids else []
    applied_by_advance: defaultdict[int, float] = defaultdict(float)
    for application in advance_applications:
        applied_by_advance[int(application["advance_id"])] += float(application.get("amount_mxn") or 0)
    for advance in advances:
        applied = round(applied_by_advance[int(advance["id"])], 2)
        advance["is_advance"] = True
        advance["applied_amount_mxn"] = applied
        advance["available_amount_mxn"] = round(max(0, float(advance.get("amount_mxn") or 0) - applied), 2)
        advance["allocations"] = []
    payment_ids = [int(row["id"]) for row in payments]
    if not payment_ids:
        return {"items": advances[:limit]}
    allocations = (ctx["sb"].table("gas_lp_expense_payment_allocations")
                   .select("payment_id,invoice_id,amount_mxn")
                   .in_("payment_id", payment_ids).execute().data or [])
    invoice_ids = sorted({int(row["invoice_id"]) for row in allocations})
    invoices = (_base_query(ctx, "gas_lp_expense_invoices")
                .in_("id", invoice_ids).execute().data or []) if invoice_ids else []
    invoices_by_id = {int(row["id"]): row for row in invoices}
    all_invoice_allocations = (ctx["sb"].table("gas_lp_expense_payment_allocations")
                               .select("invoice_id,amount_mxn").in_("invoice_id", invoice_ids)
                               .execute().data or []) if invoice_ids else []
    total_applied_by_invoice: defaultdict[int, float] = defaultdict(float)
    for row in all_invoice_allocations:
        total_applied_by_invoice[int(row["invoice_id"])] += float(row.get("amount_mxn") or 0)
    allocations_by_payment: defaultdict[int, list[dict[str, Any]]] = defaultdict(list)
    for allocation in allocations:
        invoice = invoices_by_id.get(int(allocation["invoice_id"]))
        if invoice:
            total_applied = round(total_applied_by_invoice[int(invoice["id"])], 2)
            allocations_by_payment[int(allocation["payment_id"])].append({
                "amount_mxn": float(allocation.get("amount_mxn") or 0),
                "invoice": {**invoice, "applied_amount_mxn": total_applied,
                            "balance_mxn": (0.0 if abs(float(invoice.get("total_mxn") or 0) - total_applied) < PAYMENT_BALANCE_TOLERANCE else round(max(0, float(invoice.get("total_mxn") or 0) - total_applied), 2))},
            })
    for payment in payments:
        payment["allocations"] = allocations_by_payment.get(int(payment["id"]), [])
        payment["applied_amount_mxn"] = round(sum(
            float(row.get("amount_mxn") or 0) for row in payment["allocations"]
        ), 2)
        payment["difference_mxn"] = round(
            float(payment.get("amount_mxn") or 0) - payment["applied_amount_mxn"], 2
        )
    movements = sorted(
        [*payments, *advances],
        key=lambda row: (str(row.get("paid_on") or ""), str(row.get("created_at") or "")),
        reverse=True,
    )
    return {"items": movements[:limit]}


@router.put("/gastos/payments/{payment_id}/paid-date")
def update_expense_payment_date(payment_id: int, payload: ExpensePaymentDateUpdate,
                                token: str = Query(default=""),
                                authorization: str = Header(default=""),
                                x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                                x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    """Correct the payment date and keep linked invoices consistent."""
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Gastos y pagos puede corregir la fecha de un pago.")
    rows = _base_query(ctx, "gas_lp_expense_payments").eq("id", payment_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Pago no encontrado.")
    payment = rows[0]
    next_date = payload.paid_on.isoformat()
    if str(payment.get("paid_on") or "")[:10] == next_date:
        return {"ok": True, "item": payment, "unchanged": True}

    allocations = (ctx["sb"].table("gas_lp_expense_payment_allocations")
                   .select("invoice_id").eq("payment_id", payment_id).execute().data or [])
    invoice_ids = sorted({int(row["invoice_id"]) for row in allocations})
    update = {"paid_on": next_date, "updated_at": _now()}
    ctx["sb"].table("gas_lp_expense_payments").update(update).eq(
        "tenant_id", ctx["tenant_id"]
    ).eq("profile_id", ctx["perfil_id"]).eq("id", payment_id).execute()

    # paid_on on an invoice is derived from every linked payment/advance. This
    # keeps partial payments correct after changing one payment's date.
    for invoice_id in invoice_ids:
        payment_links = (ctx["sb"].table("gas_lp_expense_payment_allocations")
                         .select("payment_id").eq("invoice_id", invoice_id).execute().data or [])
        payment_ids = sorted({int(row["payment_id"]) for row in payment_links})
        linked_payments = (_base_query(ctx, "gas_lp_expense_payments")
                           .select("id,paid_on").in_("id", payment_ids).execute().data or []) if payment_ids else []
        advance_links = (ctx["sb"].table("gas_lp_expense_advance_applications")
                         .select("advance_id").eq("invoice_id", invoice_id).execute().data or [])
        advance_ids = sorted({int(row["advance_id"]) for row in advance_links})
        linked_advances = (_base_query(ctx, "gas_lp_expense_advances")
                           .select("id,paid_on").in_("id", advance_ids).execute().data or []) if advance_ids else []
        linked_dates = [str(row.get("paid_on") or "")[:10]
                        for row in [*linked_payments, *linked_advances] if row.get("paid_on")]
        ctx["sb"].table("gas_lp_expense_invoices").update({
            "paid_on": max(linked_dates) if linked_dates else None,
            "updated_at": _now(),
        }).eq("tenant_id", ctx["tenant_id"]).eq("profile_id", ctx["perfil_id"]).eq(
            "id", invoice_id
        ).execute()

    _audit(ctx, "payment", payment_id, "payment_date_corrected", before=payment, after={**payment, **update})
    return {"ok": True, "item": {**payment, **update}, "updated_invoice_ids": invoice_ids}


@router.delete("/gastos/payments/{payment_id}")
def delete_expense_payment(payment_id: int, token: str = Query(default=""),
                           authorization: str = Header(default=""),
                           x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                           x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    """Reverse an erroneously captured payment and restore invoice balances."""
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Gastos y pagos puede eliminar un pago.")
    rows = _base_query(ctx, "gas_lp_expense_payments").eq("id", payment_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Pago no encontrado.")
    payment = rows[0]
    allocations = (ctx["sb"].table("gas_lp_expense_payment_allocations")
                   .select("payment_id,invoice_id,amount_mxn").eq("payment_id", payment_id)
                   .execute().data or [])
    invoice_ids = sorted({int(row["invoice_id"]) for row in allocations})

    ctx["sb"].table("gas_lp_expense_payment_allocations").delete().eq("payment_id", payment_id).execute()
    ctx["sb"].table("gas_lp_expense_payments").delete().eq("tenant_id", ctx["tenant_id"]).eq(
        "profile_id", ctx["perfil_id"]
    ).eq("id", payment_id).execute()

    for invoice_id in invoice_ids:
        invoice_rows = _base_query(ctx, "gas_lp_expense_invoices").eq("id", invoice_id).limit(1).execute().data or []
        if not invoice_rows:
            continue
        invoice = invoice_rows[0]
        remaining = (ctx["sb"].table("gas_lp_expense_payment_allocations").select("amount_mxn")
                     .eq("invoice_id", invoice_id).execute().data or [])
        remaining_advances = (ctx["sb"].table("gas_lp_expense_advance_applications").select("amount_mxn")
                              .eq("invoice_id", invoice_id).execute().data or [])
        applied = round(sum(float(row.get("amount_mxn") or 0) for row in [*remaining, *remaining_advances]), 2)
        complete = abs(float(invoice.get("total_mxn") or 0) - applied) < PAYMENT_BALANCE_TOLERANCE
        update = {
            "paid_amount_mxn": applied,
            "paid_on": invoice.get("paid_on") if applied else None,
            "paid_at": invoice.get("paid_at") if complete else None,
            "status": "paid" if complete else "sent_to_accountant",
            "updated_at": _now(),
        }
        ctx["sb"].table("gas_lp_expense_invoices").update(update).eq(
            "tenant_id", ctx["tenant_id"]
        ).eq("profile_id", ctx["perfil_id"]).eq("id", invoice_id).execute()

    _audit(ctx, "payment", payment_id, "deleted_payment_error", before={
        **payment, "allocations": allocations,
    })
    return {"ok": True, "deleted": True, "restored_invoice_ids": invoice_ids}


@router.get("/gastos/payments/export.xlsx")
def export_expense_payments(token: str = Query(default=""), month: str = Query(default=""),
                            authorization: str = Header(default=""),
                            x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                            x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    """Monthly accountant relationship modeled after PARADOR PROVEEDOR.xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if month and not re.fullmatch(r"\d{4}-\d{2}", month):
        raise HTTPException(400, "El mes debe tener formato AAAA-MM.")
    period = month or datetime.now(timezone(timedelta(hours=-6))).strftime("%Y-%m")
    year, month_number = (int(value) for value in period.split("-"))
    start = date(year, month_number, 1)
    end = date(year + (month_number == 12), 1 if month_number == 12 else month_number + 1, 1)
    payments = (_base_query(ctx, "gas_lp_expense_payments").gte("paid_on", start.isoformat())
                .lt("paid_on", end.isoformat()).order("paid_on").order("id").execute().data or [])
    advances = (_base_query(ctx, "gas_lp_expense_advances").gte("paid_on", start.isoformat())
                .lt("paid_on", end.isoformat()).in_("status", ["pending", "partial", "applied"])
                .order("paid_on").order("id").execute().data or [])
    advance_ids = [int(row["id"]) for row in advances]
    advance_links = (ctx["sb"].table("gas_lp_expense_advance_applications")
                     .select("advance_id,amount_mxn").in_("advance_id", advance_ids)
                     .execute().data or []) if advance_ids else []
    advance_applied: defaultdict[int, float] = defaultdict(float)
    for link in advance_links:
        advance_applied[int(link["advance_id"])] += float(link.get("amount_mxn") or 0)
    payment_ids = [int(row["id"]) for row in payments]
    allocations = (ctx["sb"].table("gas_lp_expense_payment_allocations").select("*")
                   .in_("payment_id", payment_ids).execute().data or []) if payment_ids else []
    invoice_ids = sorted({int(row["invoice_id"]) for row in allocations})
    invoices = (_base_query(ctx, "gas_lp_expense_invoices").in_("id", invoice_ids)
                .execute().data or []) if invoice_ids else []
    invoices_by_id = {int(row["id"]): row for row in invoices}
    by_payment: defaultdict[int, list[dict[str, Any]]] = defaultdict(list)
    for allocation in allocations:
        by_payment[int(allocation["payment_id"])].append(allocation)
    suppliers = {int(row["id"]): row for row in (_base_query(ctx, "gas_lp_expense_suppliers").execute().data or [])}
    recipients = {int(row["id"]): row for row in (_base_query(ctx, "gas_lp_expense_recipients").execute().data or [])}
    recipient_accounts = {int(row["id"]): row for row in (
        _base_query(ctx, "gas_lp_expense_recipient_accounts").execute().data or []
    )}
    concepts = {int(row["id"]): row.get("name") or "Sin concepto" for row in (
        _base_query(ctx, "gas_lp_expense_concepts").execute().data or []
    )}

    wb = Workbook(); provider_ws = wb.active; provider_ws.title = "Pagos proveedores"
    reimbursement_ws = wb.create_sheet("Reembolsos"); summary_ws = wb.create_sheet("Resumen por destinatario")
    headers = ["ID DE PAGO", "FACTURA EN EL PAGO", "FOLIO DE FACTURA", "FECHA DE FACTURA",
               "CONCEPTO DE GASTO", "PROVEEDOR / EMISOR", "REEMBOLSADO A", "DESTINO DEL PAGO",
               "MONTO DE LA FACTURA", "MONTO PAGADO", "FECHA DE PAGO", "TOTAL TRANSFERIDO",
               "DIFERENCIA", "REFERENCIA", "NOTAS"]
    green, burgundy, thin = "A9D18E", "7A1E2C", Side(style="thin", color="E7E3DC")
    for worksheet in (provider_ws, reimbursement_ws):
        worksheet.append(headers); worksheet.freeze_panes = "A2"; worksheet.auto_filter.ref = "A1:O1"
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=green)
            cell.alignment = Alignment(horizontal="center")
    totals: defaultdict[tuple[str, str], dict[str, Any]] = defaultdict(lambda: {"invoices": 0, "payments": set(), "paid": 0.0})
    for payment in payments:
        target = payment.get("payment_target") or "supplier"
        party = recipients.get(int(payment.get("reimbursement_recipient_id") or 0), {}) if target == "reimbursement" else suppliers.get(int(payment.get("supplier_id") or 0), {})
        party_name = party.get("name") or party.get("commercial_name") or "Sin destinatario"
        account = recipient_accounts.get(int(payment.get("reimbursement_account_id") or 0), {}) if target == "reimbursement" else party
        if target == "reimbursement":
            destination = (f"Tarjeta •••• {account.get('card_last_four') or ''}"
                           if account.get("account_type") == "credit_card"
                           else " · ".join(filter(None, [account.get("bank_name"), account.get("account_number")])))
        else:
            destination = " · ".join(filter(None, [party.get("bank_name"), party.get("account_number")]))
        rows = by_payment.get(int(payment["id"]), [])
        worksheet = reimbursement_ws if target == "reimbursement" else provider_ws
        for index, allocation in enumerate(rows):
            invoice = invoices_by_id.get(int(allocation["invoice_id"]), {})
            invoice_supplier = suppliers.get(int(invoice.get("supplier_id") or 0), {})
            issuer_name = invoice_supplier.get("commercial_name") or invoice_supplier.get("legal_name") or "Sin proveedor"
            concept_name = concepts.get(int(invoice.get("concept_id") or 0), "Sin concepto")
            applied = float(allocation.get("amount_mxn") or 0)
            notes = []
            if payment.get("notes"): notes.append(str(payment["notes"]))
            if applied + MONEY_TOLERANCE < float(invoice.get("total_mxn") or 0): notes.append("PAGO PARCIAL")
            transferred = float(payment.get("amount_mxn") or 0)
            applied_payment = sum(float(item.get("amount_mxn") or 0) for item in rows)
            difference = round(transferred - applied_payment, 2)
            worksheet.append([
                f"P-{payment['id']}", f"{index + 1} de {len(rows)}", invoice.get("invoice_number") or "S/F",
                invoice.get("invoice_date"), concept_name, issuer_name,
                party_name if target == "reimbursement" else "", destination,
                float(invoice.get("total_mxn") or 0), applied, payment.get("paid_on"),
                transferred if index == 0 else None, difference if index == 0 else None,
                payment.get("reference") or "", " · ".join(dict.fromkeys(notes)),
            ])
            bucket = totals[("Reembolso" if target == "reimbursement" else "Proveedor", party_name)]
            bucket["invoices"] += 1; bucket["payments"].add(int(payment["id"])); bucket["paid"] += applied
    for advance in advances:
        party = suppliers.get(int(advance.get("supplier_id") or 0), {})
        party_name = party.get("commercial_name") or "Sin proveedor"
        amount = float(advance.get("amount_mxn") or 0)
        applied = round(advance_applied[int(advance["id"])], 2)
        available = round(max(0, amount - applied), 2)
        provider_ws.append([
            f"A-{advance['id']}", "ANTICIPO", "S/F", None,
            concepts.get(int(advance.get("concept_id") or 0), "Anticipo"), party_name, "",
            " · ".join(filter(None, [party.get("bank_name"), party.get("account_number")])),
            amount, amount, advance.get("paid_on"), amount, 0, advance.get("reference") or "",
            f"ANTICIPO PAGADO · {'PENDIENTE DE FACTURA' if available > 0 else 'CON FACTURA'}"
            + (f" · {advance.get('description')}" if advance.get("description") else ""),
        ])
        bucket = totals[("Proveedor", party_name)]
        bucket["invoices"] += 0; bucket["payments"].add(f"A-{advance['id']}"); bucket["paid"] += amount
    for worksheet in (provider_ws, reimbursement_ws):
        for row in worksheet.iter_rows(min_row=2):
            for cell in row: cell.border = Border(bottom=thin)
            row[3].number_format = 'dd/mm/yyyy'; row[8].number_format = '$#,##0.00'
            row[9].number_format = '$#,##0.00'; row[10].number_format = 'dd/mm/yyyy'
            row[11].number_format = '$#,##0.00'; row[12].number_format = '$#,##0.00'
        for column, width in enumerate((14, 18, 20, 18, 26, 34, 28, 32, 20, 18, 18, 20, 16, 24, 34), 1):
            worksheet.column_dimensions[get_column_letter(column)].width = width
    summary_ws.append(["TIPO", "PROVEEDOR O PERSONA", "FACTURAS", "TRANSFERENCIAS", "TOTAL PAGADO"])
    for cell in summary_ws[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=burgundy)
    for (kind, party_name), values in sorted(totals.items()):
        summary_ws.append([kind, party_name, values["invoices"], len(values["payments"]), values["paid"]])
    last_detail = summary_ws.max_row
    summary_ws.append(["", "TOTAL DEL MES", "", "", f"=SUM(E2:E{max(2, last_detail)})"])
    for cell in summary_ws["E"][1:]: cell.number_format = '$#,##0.00'
    summary_ws.freeze_panes = "A2"; summary_ws.auto_filter.ref = "A1:E1"
    for column, width in enumerate((16, 38, 14, 18, 20), 1): summary_ws.column_dimensions[get_column_letter(column)].width = width
    company = _profile(ctx).get("nombre") or ""
    for worksheet in wb.worksheets:
        worksheet.sheet_view.showGridLines = False; worksheet.oddFooter.center.text = f"{company} · {period}"
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="relacion-pagos-{period}.xlsx"'})


def _legacy_export_expense_payments(token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    invoices = (_base_query(ctx, "gas_lp_expense_invoices")
                .in_("status", ["accepted", "sent_to_accountant"])
                .order("invoice_date", desc=True).execute().data or [])
    invoice_ids = [int(row["id"]) for row in invoices]
    allocations = (ctx["sb"].table("gas_lp_expense_payment_allocations").select("*")
                   .in_("invoice_id", invoice_ids).execute().data or []) if invoice_ids else []
    suppliers = _base_query(ctx, "gas_lp_expense_suppliers").execute().data or []
    recipients = _base_query(ctx, "gas_lp_expense_recipients").execute().data or []
    recipient_accounts = _base_query(ctx, "gas_lp_expense_recipient_accounts").eq("status", "active").execute().data or []
    supplier_names = {int(row["id"]): row.get("commercial_name") or "" for row in suppliers}
    group_names = {int(row["id"]): row.get("name") or "" for row in _profile_expense_groups(ctx)}
    facility_names = {int(row["id"]): row.get("nombre") or row.get("clave_instalacion") or "" for row in _profile_facilities(ctx)}
    expense_zone_names = {int(row["id"]): row.get("name") or "" for row in _expense_zones(ctx)}
    allocations_by_invoice: defaultdict[int, list[dict[str, Any]]] = defaultdict(list)
    for allocation in allocations:
        allocations_by_invoice[int(allocation["invoice_id"])].append(allocation)

    wb = Workbook(); supplier_ws = wb.active; supplier_ws.title = "Pagos a proveedores"
    reimbursement_ws = wb.create_sheet("Reembolsos")
    summary_ws = wb.create_sheet("Resumen")
    headers = ["Empresa", "Zona", "Tipo de pago", "Destinatario", "Correo", "Banco",
               "Fecha factura", "Proveedor original", "Factura", "Total factura",
               "Pagado previamente", "Saldo a pagar", "Método sugerido", "Cuenta / destino"]
    for worksheet in (supplier_ws, reimbursement_ws):
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="7A1E2C")
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:N1"
    company = _profile(ctx).get("nombre") or ""
    supplier_by_id = {int(row["id"]): row for row in suppliers}
    recipient_by_id = {int(row["id"]): row for row in recipients}
    recipient_account_by_id = {int(row["id"]): row for row in recipient_accounts}
    totals_by_party: defaultdict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {"invoices": 0, "total": 0.0, "paid": 0.0, "balance": 0.0}
    )
    for invoice in invoices:
        total_applied = round(sum(float(row.get("amount_mxn") or 0) for row in allocations_by_invoice.get(int(invoice["id"]), [])), 2)
        target = invoice.get("payment_target") or "supplier"
        supplier = supplier_by_id.get(int(invoice.get("supplier_id") or 0), {})
        recipient = recipient_by_id.get(int(invoice.get("reimbursement_recipient_id") or 0), {})
        recipient_account = recipient_account_by_id.get(int(invoice.get("reimbursement_account_id") or 0), {})
        destination = recipient if target == "reimbursement" else supplier
        destination_account = (
            (f"Tarjeta •••• {recipient_account.get('card_last_four') or ''}"
             if recipient_account.get("account_type") == "credit_card"
             else recipient_account.get("account_number") or "")
            if target == "reimbursement" else supplier.get("account_number") or ""
        )
        row = [
            company, (group_names.get(int(invoice.get("group_id") or 0))
                      or facility_names.get(int(invoice.get("facility_id") or 0))
                      or expense_zone_names.get(int(invoice.get("expense_zone_id") or 0))
                      or "General de la empresa"),
            "Reembolso" if target == "reimbursement" else "Proveedor",
            destination.get("name") or destination.get("commercial_name") or "",
            destination.get("email") or destination.get("payment_email") or "",
            (recipient_account.get("bank_name") if target == "reimbursement" else destination.get("bank_name")) or "",
            str(invoice.get("invoice_date") or ""), supplier_names.get(int(invoice.get("supplier_id") or 0), ""),
            invoice.get("invoice_number") or "", float(invoice.get("total_mxn") or 0), total_applied,
            max(0, float(invoice.get("total_mxn") or 0) - total_applied), "Transferencia", destination_account,
        ]
        (reimbursement_ws if target == "reimbursement" else supplier_ws).append(row)
        party_name = str(row[3] or "Sin destinatario")
        summary = totals_by_party[("Reembolso" if target == "reimbursement" else "Proveedor", party_name)]
        summary["invoices"] += 1; summary["total"] += row[9]; summary["paid"] += row[10]; summary["balance"] += row[11]
    for worksheet in (supplier_ws, reimbursement_ws):
        for column in (10, 11, 12):
            for cell in worksheet[get_column_letter(column)][1:]: cell.number_format = '$#,##0.00'
        for index, header in enumerate(headers, 1):
            worksheet.column_dimensions[get_column_letter(index)].width = min(34, max(12, len(header) + 3))
    summary_headers = ["Tipo", "Proveedor o persona", "Facturas", "Total facturado", "Pagado previamente", "Total pendiente"]
    summary_ws.append(summary_headers)
    for cell in summary_ws[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="7A1E2C")
    for (kind, party_name), totals in sorted(totals_by_party.items()):
        summary_ws.append([kind, party_name, totals["invoices"], totals["total"], totals["paid"], totals["balance"]])
    for column in (4, 5, 6):
        for cell in summary_ws[get_column_letter(column)][1:]: cell.number_format = '$#,##0.00'
    summary_ws.freeze_panes = "A2"; summary_ws.auto_filter.ref = "A1:F1"
    for index, header in enumerate(summary_headers, 1):
        summary_ws.column_dimensions[get_column_letter(index)].width = min(36, max(14, len(header) + 3))
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": 'attachment; filename="pagos-y-reembolsos.xlsx"'})


@router.post("/gastos/invoices/{invoice_id}/transition")
def transition_invoice(invoice_id: int, payload: InvoiceTransition, token: str = Query(default=""),
                       authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo la asistente de gastos administra la factura.")
    rows = _base_query(ctx, "gas_lp_expense_invoices").eq("id", invoice_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Factura no encontrada.")
    row = rows[0]
    # A double click or a stale browser tab must not fail after the first
    # request already moved the expense to payments.
    if payload.action == "accept" and row.get("status") == "sent_to_accountant":
        return {"ok": True, "item": row, "already_applied": True}
    transitions = {
        # Aceptar significa que el gasto ya fue revisado y queda listo para pago.
        # Conservamos send_to_accountant para registros/clientes antiguos.
        "accept": ({"pending_review", "observed"}, "sent_to_accountant"),
        "observe": ({"pending_review"}, "observed"),
        "reject": ({"pending_review", "observed"}, "rejected"),
        "send_to_accountant": ({"accepted"}, "sent_to_accountant"),
        "mark_paid": ({"sent_to_accountant"}, "paid"),
        "cancel": ({"pending_review", "observed", "accepted"}, "cancelled"),
        "withdraw_from_payments": ({"sent_to_accountant"}, "pending_review"),
    }
    allowed, new_status = transitions[payload.action]
    if row["status"] not in allowed:
        raise HTTPException(409, f"No se puede aplicar esta acción desde {row['status']}.")
    if payload.action in {"observe", "reject", "cancel", "withdraw_from_payments"} and len(payload.observation.strip()) < 3:
        raise HTTPException(400, "Captura el motivo de esta acción.")
    update: dict[str, Any] = {"status": new_status, "updated_at": _now()}
    if payload.observation:
        update["observation"] = payload.observation.strip()
    if payload.action in {"accept", "observe", "reject"}:
        update.update({"reviewed_by": ctx["actor_id"], "reviewed_at": _now()})
    if payload.action in {"accept", "send_to_accountant"}:
        update["sent_to_accountant_at"] = _now()
    if payload.action == "withdraw_from_payments":
        update.update({"sent_to_accountant_at": None, "reviewed_by": ctx["actor_id"], "reviewed_at": _now()})
    if payload.action == "mark_paid":
        if not payload.paid_on or payload.paid_amount_mxn is None:
            raise HTTPException(400, "Confirma fecha y monto pagado.")
        if abs(round(payload.paid_amount_mxn, 2) - float(row["total_mxn"])) > MONEY_TOLERANCE:
            raise HTTPException(400, "El monto pagado debe coincidir con el total de la factura.")
        update.update({"paid_at": _now(), "paid_on": payload.paid_on.isoformat(),
                       "paid_amount_mxn": round(payload.paid_amount_mxn, 2),
                       "payment_email_status": "not_sent"})
        update.update(_send_payment_notification(ctx, {**row, **update}))
    ctx["sb"].table("gas_lp_expense_invoices").update(update).eq("id", invoice_id).execute()
    _audit(ctx, "invoice", invoice_id, payload.action, before=row, after=update)
    return {"ok": True, "item": {**row, **update}}


@router.put("/gastos/invoices/{invoice_id}")
def update_direct_invoice(invoice_id: int, payload: DirectInvoiceUpdate, token: str = Query(default=""),
                          authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Gastos y pagos puede editar una captura.")
    rows = _base_query(ctx, "gas_lp_expense_invoices").eq("id", invoice_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Gasto no encontrado.")
    row = rows[0]
    if row.get("expense_type") not in {"direct", "credit_note"}:
        raise HTTPException(409, "Las facturas relacionadas con vales no se editan desde esta pantalla.")
    if row.get("status") in {"paid", "cancelled"}:
        raise HTTPException(409, "Un gasto pagado o eliminado ya no se puede editar.")
    allocations = (ctx["sb"].table("gas_lp_expense_payment_allocations").select("payment_id")
                   .eq("invoice_id", invoice_id).limit(1).execute().data or [])
    if allocations:
        raise HTTPException(409, "Este gasto tiene pagos relacionados y ya no se puede editar.")
    advance_links = (ctx["sb"].table("gas_lp_expense_advance_applications").select("advance_id")
                     .eq("invoice_id", invoice_id).limit(1).execute().data or [])
    if advance_links:
        raise HTTPException(409, "Esta factura tiene anticipos aplicados y ya no se puede editar.")
    alerts = _invoice_alerts(
        ctx, supplier_id=int(row["supplier_id"]), invoice_number=payload.invoice_number,
        invoice_date=payload.invoice_date, total_mxn=payload.total_mxn,
        exclude_invoice_id=invoice_id,
    )
    observation = payload.observation.strip()
    previous_observation = str(row.get("observation") or "").strip()
    previous_was_automatic_alert = previous_observation.startswith("Alerta: ")
    if alerts and (not observation or observation == previous_observation and previous_was_automatic_alert):
        observation = "Alerta: " + " ".join(alerts)
    elif not alerts and observation == previous_observation and previous_was_automatic_alert:
        observation = ""
    update = {
        "invoice_number": payload.invoice_number.strip(),
        "invoice_date": payload.invoice_date.isoformat(),
        "total_mxn": round(payload.total_mxn, 2),
        "description": payload.description.strip(),
        "observation": observation,
        "updated_at": _now(),
    }
    ctx["sb"].table("gas_lp_expense_invoices").update(update).eq(
        "tenant_id", ctx["tenant_id"]
    ).eq("profile_id", ctx["perfil_id"]).eq("id", invoice_id).execute()
    _audit(ctx, "invoice", invoice_id, "edited", before=row, after=update)
    return {"ok": True, "item": {**row, **update}, "alerts": alerts}


@router.put("/gastos/invoices/{invoice_id}/paid-date")
def update_paid_invoice_date(invoice_id: int, payload: PaidInvoiceDateUpdate, token: str = Query(default=""),
                             authorization: str = Header(default=""),
                             x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                             x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    """Correct only the documentary invoice date after a payment exists."""
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Gastos y pagos puede corregir la fecha de una factura.")
    rows = _base_query(ctx, "gas_lp_expense_invoices").eq("id", invoice_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Factura no encontrada.")
    row = rows[0]
    if row.get("status") != "paid":
        raise HTTPException(409, "Esta corrección está disponible únicamente para facturas pagadas.")
    if row.get("expense_type") not in {"direct", "credit_note"}:
        raise HTTPException(409, "Las facturas relacionadas con vales se corrigen desde su módulo de origen.")
    update = {"invoice_date": payload.invoice_date.isoformat(), "updated_at": _now()}
    ctx["sb"].table("gas_lp_expense_invoices").update(update).eq(
        "tenant_id", ctx["tenant_id"]
    ).eq("profile_id", ctx["perfil_id"]).eq("id", invoice_id).execute()
    _audit(ctx, "invoice", invoice_id, "paid_invoice_date_corrected", before=row, after=update)
    return {"ok": True, "item": {**row, **update}}


@router.delete("/gastos/invoices/{invoice_id}")
def delete_direct_invoice(invoice_id: int, token: str = Query(default=""),
                          authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo la asistente de gastos puede eliminar una captura.")
    rows = _base_query(ctx, "gas_lp_expense_invoices").eq("id", invoice_id).limit(1).execute().data or []
    if not rows:
        raise HTTPException(404, "Gasto no encontrado.")
    row = rows[0]
    if row.get("expense_type") not in {"direct", "credit_note"}:
        raise HTTPException(409, "Las facturas relacionadas con vales no se eliminan desde esta pantalla.")
    if row.get("status") not in {"pending_review", "observed", "rejected", "accepted", "sent_to_accountant"}:
        raise HTTPException(409, "Este gasto ya avanzó en el proceso y no se puede eliminar.")
    allocations = (ctx["sb"].table("gas_lp_expense_payment_allocations").select("payment_id")
                   .eq("invoice_id", invoice_id).limit(1).execute().data or [])
    if allocations:
        raise HTTPException(409, "Este gasto tiene pagos relacionados y no se puede eliminar.")
    advance_links = (ctx["sb"].table("gas_lp_expense_advance_applications").select("advance_id")
                     .eq("invoice_id", invoice_id).limit(1).execute().data or [])
    if advance_links:
        raise HTTPException(409, "Esta factura tiene anticipos aplicados y no se puede eliminar.")
    update = {
        "status": "cancelled", "observation": "Captura eliminada por error.",
        "reviewed_by": ctx["actor_id"], "reviewed_at": _now(), "updated_at": _now(),
    }
    ctx["sb"].table("gas_lp_expense_invoices").update(update).eq(
        "tenant_id", ctx["tenant_id"]
    ).eq("profile_id", ctx["perfil_id"]).eq("id", invoice_id).execute()
    _audit(ctx, "invoice", invoice_id, "deleted_capture_error", before=row, after=update)
    return {"ok": True, "deleted": True}


@router.post("/gastos/invoices/{invoice_id}/payment-email")
def resend_payment_email(invoice_id: int, token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_admin"]:
        raise HTTPException(403, "Solo Gastos y pagos puede enviar esta notificación.")
    rows = _base_query(ctx, "gas_lp_expense_invoices").eq("id", invoice_id).limit(1).execute().data or []
    if not rows or rows[0].get("status") != "paid":
        raise HTTPException(409, "La factura debe estar pagada.")
    row = rows[0]
    update = _send_payment_notification(ctx, row)
    ctx["sb"].table("gas_lp_expense_invoices").update(update).eq("id", invoice_id).execute()
    _audit(ctx, "invoice", invoice_id, "payment_email_retried", before=row, after=update)
    return {"ok": update["payment_email_status"] == "sent", **update}


@router.put("/gastos/invoices/{invoice_id}/correct")
def correct_observed_invoice(invoice_id: int, payload: InvoiceCorrection, token: str = Query(default=""),
                             authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    if not ctx["is_manager"]:
        raise HTTPException(403, "La corrección corresponde al gerente.")
    rows = (_base_query(ctx, "gas_lp_expense_invoices").eq("id", invoice_id)
            .eq("created_by_type", "manager").eq("created_by", ctx["actor_id"]).limit(1).execute().data or [])
    if not rows:
        raise HTTPException(404, "Factura no encontrada.")
    row = rows[0]
    if row["status"] != "observed":
        raise HTTPException(409, "Solo una factura observada puede corregirse y reenviarse.")
    links = (ctx["sb"].table("gas_lp_expense_invoice_vouchers").select("amount_mxn")
             .eq("invoice_id", invoice_id).execute().data or [])
    expected = round(sum(float(link.get("amount_mxn") or 0) for link in links), 2)
    if abs(expected - round(payload.total_mxn, 2)) > MONEY_TOLERANCE:
        raise HTTPException(400, f"El total debe coincidir con los vales: ${expected:,.2f}.")
    alerts = _invoice_alerts(
        ctx, supplier_id=int(row["supplier_id"]), invoice_number=payload.invoice_number,
        invoice_date=payload.invoice_date, total_mxn=expected, exclude_invoice_id=invoice_id,
    )
    update = {
        "invoice_number": payload.invoice_number.strip(), "invoice_date": payload.invoice_date.isoformat(),
        "total_mxn": expected, "status": "pending_review",
        "observation": "Alerta: " + " ".join(alerts) if alerts else "",
        "reviewed_by": None, "reviewed_at": None, "updated_at": _now(),
    }
    ctx["sb"].table("gas_lp_expense_invoices").update(update).eq("id", invoice_id).execute()
    _audit(ctx, "invoice", invoice_id, "corrected_and_resubmitted", before=row, after=update)
    return {"ok": True, "item": {**row, **update}}


@router.get("/gastos/analytics")
def analytics(token: str = Query(default=""), authorization: str = Header(default=""),
                  x_flotilla_access: str = Header(default="", alias="X-Flotilla-Access"),
                  x_perfil_id: str = Header(default="", alias="X-Perfil-ID"),
                  period: str = Query(default=""), zone_id: int | None = Query(default=None),
                  supplier_id: int | None = Query(default=None), concept_id: int | None = Query(default=None),
                  status: str = Query(default="")):
    ctx = _ctx(authorization, x_flotilla_access, token, x_perfil_id)
    invoices = _base_query(ctx, "gas_lp_expense_invoices").execute().data or []
    active_invoices = [
        row for row in invoices if row.get("status") not in {"rejected", "cancelled"}
    ]
    suppliers = _base_query(ctx, "gas_lp_expense_suppliers").execute().data or []
    concepts = _base_query(ctx, "gas_lp_expense_concepts").execute().data or []
    vouchers = _base_query(ctx, "gas_lp_expense_vouchers").execute().data or []
    invoice_ids = [int(row["id"]) for row in active_invoices]
    links = (ctx["sb"].table("gas_lp_expense_invoice_vouchers").select(
        "invoice_id,voucher_id,amount_mxn"
    ).in_("invoice_id", invoice_ids).execute().data or []) if invoice_ids else []
    groups = ctx["sb"].table("fleet_groups").select("id,name").eq("tenant_id", ctx["tenant_id"]).execute().data or []
    facilities = _profile_facilities(ctx)
    expense_zones = _expense_zones(ctx)
    vehicles = ctx["sb"].table("fleet_vehicles").select(
        "id,vehicle_number"
    ).eq("tenant_id", ctx["tenant_id"]).execute().data or []
    supplier_names = {int(row["id"]): row["commercial_name"] for row in suppliers}
    concept_names = {int(row["id"]): row["name"] for row in concepts}
    group_names = {int(row["id"]): row["name"] for row in groups}
    facility_names = {int(row["id"]): row.get("nombre") or row.get("clave_instalacion") or "Zona" for row in facilities}
    expense_zone_names = {int(row["id"]): row.get("name") or "Zona" for row in expense_zones}
    vehicle_names = {int(row["id"]): row["vehicle_number"] for row in vehicles}
    voucher_by_id = {int(row["id"]): row for row in vouchers}
    invoice_by_id = {int(row["id"]): row for row in active_invoices}
    def invoice_zone_id(row):
        return row.get("facility_id") or row.get("expense_zone_id") or row.get("group_id")
    if period:
        active_invoices = [row for row in active_invoices if str(row.get("invoice_date") or "").startswith(period)]
    if supplier_id is not None:
        active_invoices = [row for row in active_invoices if int(row.get("supplier_id") or 0) == supplier_id]
    if concept_id is not None:
        active_invoices = [row for row in active_invoices if int(row.get("concept_id") or 0) == concept_id]
    if zone_id is not None:
        active_invoices = [row for row in active_invoices if int(invoice_zone_id(row) or 0) == zone_id]
    if status == "paid":
        active_invoices = [row for row in active_invoices if row.get("status") == "paid"]
    elif status == "pending":
        active_invoices = [row for row in active_invoices if row.get("status") != "paid"]
    invoice_by_id = {int(row["id"]): row for row in active_invoices}
    dimensions: dict[str, defaultdict[str, float]] = {
        key: defaultdict(float) for key in
        ("status", "supplier", "type", "month", "concept", "zone", "unit", "manager")
    }
    for row in active_invoices:
        amount = float(row.get("total_mxn") or 0) * (-1 if row.get("expense_type") == "credit_note" else 1)
        dimensions["status"][row["status"]] += amount
        dimensions["supplier"][supplier_names.get(int(row["supplier_id"]), "Proveedor")] += amount
        dimensions["type"]["Con vales" if row["expense_type"] == "voucher" else ("Nota de crédito" if row["expense_type"] == "credit_note" else "Gasto directo")] += amount
        dimensions["month"][str(row.get("invoice_date") or "")[:7] or "Sin fecha"] += amount
        if row.get("concept_id"):
            dimensions["concept"][concept_names.get(int(row["concept_id"]), "Concepto")] += amount
        if row["expense_type"] == "direct":
            if row.get("facility_id"):
                dimensions["zone"][facility_names.get(int(row["facility_id"]), "Zona")] += amount
            elif row.get("expense_zone_id"):
                dimensions["zone"][expense_zone_names.get(int(row["expense_zone_id"]), "Zona")] += amount
            elif row.get("group_id"):
                dimensions["zone"][group_names.get(int(row["group_id"]), "Zona")] += amount
            else:
                dimensions["zone"]["General de la empresa"] += amount
    for link in links:
        invoice = invoice_by_id.get(int(link["invoice_id"]))
        voucher = voucher_by_id.get(int(link["voucher_id"]))
        if not invoice or not voucher:
            continue
        amount = float(link.get("amount_mxn") or 0)
        dimensions["concept"][concept_names.get(int(voucher["concept_id"]), "Concepto")] += amount
        dimensions["zone"][group_names.get(int(voucher["group_id"]), "Zona")] += amount
        dimensions["unit"][vehicle_names.get(int(voucher["vehicle_id"]), "Unidad")] += amount
        dimensions["manager"][voucher.get("created_by_name") or "Gerente"] += amount
    number_counts: defaultdict[tuple[int, str], int] = defaultdict(int)
    supplier_months: defaultdict[int, defaultdict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in active_invoices:
        if not _is_without_folio(row.get("invoice_number")):
            number_counts[(int(row["supplier_id"]), _normalize(row["invoice_number"]))] += 1
        supplier_months[int(row["supplier_id"])][str(row.get("invoice_date") or "")[:7]] += float(row.get("total_mxn") or 0) * (-1 if row.get("expense_type") == "credit_note" else 1)
    today = date.today()
    stale_amount = sum(
        row["status"] == "amount_pending"
        and str(row.get("issued_on") or "")[:10] < (today - timedelta(days=7)).isoformat()
        for row in vouchers
    )
    stale_ready = sum(
        row["status"] == "ready_to_invoice"
        and str(row.get("issued_on") or "")[:10] < (today - timedelta(days=15)).isoformat()
        for row in vouchers
    )
    stale_accounting = sum(
        row["status"] == "sent_to_accountant"
        and str(row.get("sent_to_accountant_at") or "")[:10] < (today - timedelta(days=7)).isoformat()
        for row in active_invoices
    )
    alerts = {
        "duplicate_invoice_numbers": sum(count - 1 for count in number_counts.values() if count > 1),
        # Distinct folios can legitimately share supplier, date and amount.
        # Treating that signature as a duplicate created noisy false positives.
        "similar_invoices": 0,
        "pending_suppliers": sum(row["validation_status"] == "pending" for row in suppliers),
        "vouchers_without_amount": stale_amount, "vouchers_not_invoiced": stale_ready,
        "accounting_payment_overdue": stale_accounting,
    }
    anomalies: list[dict[str, Any]] = []
    for supplier_id, monthly in supplier_months.items():
        ordered = sorted((month, amount) for month, amount in monthly.items() if month)
        if len(ordered) < 2:
            continue
        current_month, current_amount = ordered[-1]
        previous_values = [amount for _, amount in ordered[:-1][-6:]]
        baseline = sum(previous_values) / len(previous_values)
        if baseline > 0 and current_amount >= baseline * 3 and current_amount - baseline >= 5000:
            anomalies.append({
                "severity": "high", "type": "supplier_spike",
                "label": supplier_names.get(supplier_id, "Proveedor"),
                "message": f"Subió de un promedio de ${baseline:,.2f} a ${current_amount:,.2f} en {current_month}.",
            })
    def ranked(key: str, limit: int = 20) -> list[dict[str, Any]]:
        return [{"label": label, "amount": round(amount, 2)}
                for label, amount in sorted(dimensions[key].items(), key=lambda item: item[1], reverse=True)[:limit]]
    detail = [{"invoice_date": row.get("invoice_date"), "supplier": supplier_names.get(int(row["supplier_id"]), "Proveedor"), "concept": concept_names.get(int(row.get("concept_id") or 0), "—"), "zone": expense_zone_names.get(int(invoice_zone_id(row) or 0), "General de la empresa"), "total_mxn": round(float(row.get("total_mxn") or 0), 2), "status": "pending" if row.get("status") != "paid" else "paid"} for row in active_invoices]
    return {
        "totals": {"all": round(sum(float(row.get("total_mxn") or 0) for row in active_invoices), 2),
                   "paid": round(dimensions["status"].get("paid", 0), 2),
                   "pending": round(sum(value for key, value in dimensions["status"].items() if key != "paid"), 2)},
        **{f"by_{key}": ranked(key) for key in dimensions},
        "alerts": alerts,
        "alert_total": sum(alerts.values()), "invoice_count": len(active_invoices), "detail": detail,
        "anomalies": anomalies,
    }
